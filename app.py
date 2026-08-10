from __future__ import annotations

import csv
import json
import shutil
import time
import threading
from calendar import monthrange
from datetime import date, datetime, timedelta
from functools import lru_cache
from io import StringIO
from pathlib import Path
from typing import Any
from urllib.parse import quote
from uuid import uuid4

from flask import Flask, Response, jsonify, redirect, render_template, request, send_from_directory, url_for
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)


class WorkbookWriteError(RuntimeError):
    """Raised when the workbook cannot be updated because it is locked or busy."""


@app.after_request
def no_cache(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    response.headers["Surrogate-Control"] = "no-store"
    return response


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK_NAME = "H-Queex_Financial_Control V7.0.xlsm"
WORKBOOK_PATH = BASE_DIR / DEFAULT_WORKBOOK_NAME
SUBSCRIPTIONS_PATH = BASE_DIR / "subscriptions.json"
ARCHIVE_PATH = BASE_DIR / "archives.json"
AUDIT_LOG_PATH = BASE_DIR / "audit-log.json"
BUSINESS_PROFILE_PATH = BASE_DIR / "business-profile.json"
CHART_OF_ACCOUNTS_PATH = BASE_DIR / "chart-of-accounts.json"
LEDGER_JOURNAL_PATH = BASE_DIR / "ledger-journal.json"
CAPITAL_ASSETS_PATH = BASE_DIR / "capital-assets.json"
PAYROLL_PATH = BASE_DIR / "payroll-register.json"
BANK_STATEMENTS_PATH = BASE_DIR / "bank-statements.json"
INCOME_PATH = BASE_DIR / "income.json"
EXPENSES_PATH = BASE_DIR / "expenses.json"
INVOICES_PATH = BASE_DIR / "invoices.json"
CLIENTS_PATH = BASE_DIR / "clients.json"
SUPPLIERS_PATH = BASE_DIR / "suppliers.json"
SERVICES_PATH = BASE_DIR / "services.json"
TAX_RULES_PATH = BASE_DIR / "tax-rules.json"
BACKUPS_DIR = BASE_DIR / "backups"
RECEIPTS_DIR = BASE_DIR / "receipts"
RECEIPTS_DIR.mkdir(parents=True, exist_ok=True)
ALLOWED_RECEIPT_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "gif", "heic", "webp"}
COMPANY_DOCUMENTS_PATH = BASE_DIR / "documents.json"
COMPLIANCE_CALENDAR_PATH = BASE_DIR / "compliance-calendar.json"
COMPANY_DOCUMENTS_DIR = BASE_DIR / "documents"
COMPANY_DOCUMENTS_DIR.mkdir(parents=True, exist_ok=True)
ALLOWED_DOCUMENT_EXTENSIONS = {"pdf", "docx", "png", "jpg", "jpeg"}
MAX_DOCUMENT_SIZE_BYTES = 10 * 1024 * 1024
DOCUMENT_CATEGORIES = (
    "Compliance",
    "Insurance",
    "Business Planning",
    "Banking",
    "Legal",
    "Templates",
    "Client Agreements",
    "Other",
)
DOCUMENT_STATUSES = ("active", "archived")
COMPLIANCE_REPEAT_FREQUENCIES = ("", "monthly", "quarterly", "annual")
COMPLIANCE_STATUSES = ("pending", "complete")
DOCUMENT_EXPIRY_WARNING_DAYS = 30
PROJECTS_PATH = BASE_DIR / "projects.json"
DELIVERY_LOG_PATH = BASE_DIR / "delivery-log.json"
SOPS_PATH = BASE_DIR / "sops.json"
SOP_FILES_DIR = BASE_DIR / "sops"
SOP_FILES_DIR.mkdir(parents=True, exist_ok=True)
ALLOWED_SOP_EXTENSIONS = {"pdf", "docx", "png", "jpg", "jpeg"}
DELIVERY_FILES_DIR = BASE_DIR / "delivery-files"
DELIVERY_FILES_DIR.mkdir(parents=True, exist_ok=True)
ALLOWED_DELIVERY_EXTENSIONS = {"pdf", "docx", "png", "jpg", "jpeg", "xlsx", "csv"}
PROJECT_STATUSES = ("Enquiry", "Proposed", "Active", "On Hold", "Completed", "Cancelled")
PROJECT_KANBAN_STATUSES = ("Enquiry", "Proposed", "Active", "On Hold", "Completed")
DMAIC_PHASES = ("Define", "Measure", "Analyse", "Improve", "Control")
DMAIC_PHASE_STATUSES = ("Not Started", "In Progress", "Complete")
DELIVERY_SERVICE_TYPES = (
    "Process Review",
    "SOP Creation",
    "Workshop",
    "Report",
    "Advisory Call",
    "KPI Review",
    "Implementation",
    "Other",
)
SOP_STATUSES = ("Draft", "Review", "Approved", "Superseded")
SOP_STATUS_WORKFLOW = ("Draft", "Review", "Approved")
PROJECT_DEADLINE_WARNING_DAYS = 14
GDRIVE_BACKUP_DIR = Path("G:/My Drive/H-Queex — Working Documents/H-Queex Hub/Backups")
BACKUP_STATUS_PATH = BASE_DIR / "backup-status.json"
BACKUP_RETENTION_DAYS = 30
SHEET_JSON_PATHS = {
    "Income": INCOME_PATH,
    "Expenses": EXPENSES_PATH,
    "Invoices": INVOICES_PATH,
    "Clients": CLIENTS_PATH,
    "Suppliers": SUPPLIERS_PATH,
}

if not SUBSCRIPTIONS_PATH.exists():
    SUBSCRIPTIONS_PATH.write_text("[]", encoding="utf-8")

if not SERVICES_PATH.exists():
    SERVICES_PATH.write_text("[]", encoding="utf-8")

_sheet_write_lock = threading.Lock()
SUBSCRIPTION_FREQUENCIES = {"monthly": 1, "quarterly": 3, "yearly": 12}
SUBSCRIPTION_STATUSES = ("active", "paused", "cancelled")
SERVICE_TIERS = ("core", "addon")
SERVICE_GROUPS = (
    "Documentation and Knowledge Assets",
    "Data and Reporting Tools",
    "Communication and Strategic Assets",
)
SERVICE_PRICE_TYPES = ("fixed", "from", "hourly", "retainer")
SERVICE_BILLING_FREQUENCIES = ("one-off", "monthly", "quarterly", "annual")
SERVICE_STATUSES = ("active", "archived")
CLIENT_SERVICE_TIERS = ("None", "Clarity Base", "Clarity Plus", "Clarity Partner")
CLIENT_RETAINER_FREQUENCIES = ("monthly", "quarterly", "annual")
BUSINESS_STRUCTURES = ("sole_trader", "limited_company")
INCOME_PAYMENT_METHODS = {
    "sole_trader": ["Business Bank Account", "Stripe", "PayPal", "Cash", "Proprietor Capital"],
    "limited_company": ["Business Bank Account", "Stripe", "PayPal", "Cash", "Director Loan Account"],
}
EXPENSE_PAYMENT_METHODS = {
    "sole_trader": ["Business Bank", "Credit Card", "Cash", "Proprietor Contribution"],
    "limited_company": ["Business Bank", "Credit Card", "Cash", "Director Contribution"],
}
VAT_RATE_OPTIONS = ["0%", "4.8%", "9%", "13.5%", "23%", "Exempt"]
VAT_TREATMENT_OPTIONS = [
    {"value": "standard", "label": "Standard"},
    {"value": "zero_rated", "label": "Zero-rated"},
    {"value": "exempt", "label": "Exempt"},
    {"value": "reverse_charge", "label": "Reverse charge"},
]
SUPPLY_TYPE_OPTIONS = [
    {"value": "services", "label": "Services"},
    {"value": "goods", "label": "Goods"},
]
INVOICE_STATUS_OPTIONS = ["Draft", "Issued", "Paid", "Partially Paid", "Overdue", "Bad Debt", "Cancelled"]
INCOME_STATUS_OPTIONS = ["Received", "Pending", "Cancelled"]
INCOME_SOURCES = ("manual", "invoiced")
EXPENSE_STATUS_OPTIONS = ["Pending", "Approved", "Paid", "Auto-posted", "Cancelled"]
EXPENSE_INPUT_VAT_OPTIONS = ["Yes", "No", "Partial"]
EXPENSE_DEDUCTIBILITY_OPTIONS = ["Fully Deductible", "Partially Deductible", "Non-Deductible"]
RECONCILIATION_OPTIONS = ["Reconciled", "Unreconciled"]
YES_NO_OPTIONS = ["Yes", "No"]
PAYROLL_STATUS_OPTIONS = ["Draft", "Approved", "Paid", "Filed"]
RECONCILIATION_MATCH_DAYS = 3
VAT_TURNOVER_THRESHOLDS = {
    "services": {"label": "Services", "annual_limit": 42000.0},
    "goods": {"label": "Goods", "annual_limit": 85000.0},
}
VAT_THRESHOLD_WARNING_RATIO = 0.8
PHASE_POLICY = {
    "sole_trader": {
        "tax_regime": "Income Tax (Form 11)",
        "estimated_tax_rate": 0.20,
        "report_template": "Form 11 outputs",
        "owner_account_label": "Proprietor Capital Account",
        "next_filing_deadline": "Preliminary tax and Form 11 due by 31 October (ROS extension may apply)",
    },
    "limited_company": {
        "tax_regime": "Corporation Tax (CT1)",
        "estimated_tax_rate": 0.125,
        "report_template": "CT1 outputs",
        "owner_account_label": "Director Loan Account",
        "next_filing_deadline": "CT1 due 9 months after accounting year end",
    },
}
DEFAULT_CHART_OF_ACCOUNTS = [
    {"code": "1000", "name": "Cash at Bank", "type": "Asset", "tax_treatment": "n/a", "active": True},
    {"code": "1100", "name": "Accounts Receivable", "type": "Asset", "tax_treatment": "n/a", "active": True},
    {"code": "1200", "name": "Input VAT Control", "type": "Asset", "tax_treatment": "vat", "active": True},
    {"code": "2000", "name": "Accounts Payable", "type": "Liability", "tax_treatment": "n/a", "active": True},
    {"code": "2100", "name": "Output VAT Control", "type": "Liability", "tax_treatment": "vat", "active": True},
    {"code": "2200", "name": "PAYE / USC / PRSI Control", "type": "Liability", "tax_treatment": "payroll", "active": True},
    {"code": "3000", "name": "Owner / Director Account", "type": "Equity", "tax_treatment": "entity", "active": True},
    {"code": "4000", "name": "Consulting / Project Fees", "type": "Income", "tax_treatment": "trading", "active": True},
    {"code": "4010", "name": "Retainer / Advisory Fees", "type": "Income", "tax_treatment": "trading", "active": True},
    {"code": "4020", "name": "Service Add-ons", "type": "Income", "tax_treatment": "trading", "active": True},
    {"code": "4030", "name": "Grant Income", "type": "Income", "tax_treatment": "non-trading", "active": True},
    {"code": "4040", "name": "BTWEA / Welfare Support", "type": "Income", "tax_treatment": "personal-excluded", "active": True},
    {"code": "4900", "name": "Other Income", "type": "Income", "tax_treatment": "trading", "active": True},
    {"code": "5000", "name": "Software and Subscriptions", "type": "Expense", "tax_treatment": "deductible", "active": True},
    {"code": "5010", "name": "Domain / Hosting / Website", "type": "Expense", "tax_treatment": "deductible", "active": True},
    {"code": "5020", "name": "Professional Fees", "type": "Expense", "tax_treatment": "deductible", "active": True},
    {"code": "5030", "name": "Marketing and Advertising", "type": "Expense", "tax_treatment": "deductible", "active": True},
    {"code": "5040", "name": "Bank and Transaction Fees", "type": "Expense", "tax_treatment": "deductible", "active": True},
    {"code": "5200", "name": "Equipment and Hardware", "type": "Expense", "tax_treatment": "capital-check", "active": True},
    {"code": "5210", "name": "Home Office Expenses", "type": "Expense", "tax_treatment": "partial", "active": True},
    {"code": "5220", "name": "Motor Expenses", "type": "Expense", "tax_treatment": "partial", "active": True},
    {"code": "5230", "name": "Phone and Communications", "type": "Expense", "tax_treatment": "partial", "active": True},
    {"code": "5240", "name": "Travel and Subsistence", "type": "Expense", "tax_treatment": "partial", "active": True},
    {"code": "5300", "name": "Salaries and Wages", "type": "Expense", "tax_treatment": "deductible", "active": True},
    {"code": "5310", "name": "Employer PRSI", "type": "Expense", "tax_treatment": "deductible", "active": True},
    {"code": "5800", "name": "Entertainment", "type": "Expense", "tax_treatment": "non-deductible", "active": True},
    {"code": "5810", "name": "Personal Expenses", "type": "Expense", "tax_treatment": "non-deductible", "active": True},
    {"code": "5820", "name": "Fines and Penalties", "type": "Expense", "tax_treatment": "non-deductible", "active": True},
    {"code": "5900", "name": "Non-Deductible Items", "type": "Expense", "tax_treatment": "non-deductible", "active": False},
]
INCOME_CATEGORY_ACCOUNT_MAP = {
    "consulting / project fees": "4000",
    "retainer / advisory fees": "4010",
    "service add-ons": "4020",
    "grant income": "4030",
    "btwea / welfare support": "4040",
    "other income": "4900",
}
EXPENSE_CATEGORY_ACCOUNT_MAP = {
    "software and subscriptions": "5000",
    "domain / hosting / website": "5010",
    "professional fees": "5020",
    "marketing and advertising": "5030",
    "bank and transaction fees": "5040",
    "equipment and hardware": "5200",
    "home office expenses": "5210",
    "motor expenses": "5220",
    "phone and communications": "5230",
    "travel and subsistence": "5240",
    "entertainment": "5800",
    "personal expenses": "5810",
    "fines and penalties": "5820",
    "non-deductible items": "5900",
}
ENTITY_ROUTE_MAP = {
    "income": "income_view",
    "expense": "expenses_view",
    "invoice": "invoices_view",
    "client": "clients_view",
    "supplier": "suppliers_view",
    "subscription": "subscriptions_view",
    "payroll": "payroll_view",
}
WORKBOOK_ENTITY_CONFIG = {
    "income": {"sheet": "Income", "audit_type": "income"},
    "expense": {"sheet": "Expenses", "audit_type": "expense"},
    "invoice": {"sheet": "Invoices", "audit_type": "invoice"},
    "client": {"sheet": "Clients", "audit_type": "client"},
    "supplier": {"sheet": "Suppliers", "audit_type": "supplier"},
}


def _default_business_profile() -> dict[str, Any]:
    return {
        "business_name": "H-Queex",
        "owner_name": "Hevandro Martire",
        "cro_number": "790968",
        "registration_date": "2026-08-04",
        "structure": "sole_trader",
        "vat_registered": True,
        "vat_threshold_basis": "services",
        "transition_date": "",
        "pre_trading_start_date": "",
        "trading_start_date": "",
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }


def _normalize_business_structure(value: Any) -> str:
    structure = str(value or "").strip().lower()
    return structure if structure in BUSINESS_STRUCTURES else "sole_trader"


def _load_business_profile() -> dict[str, Any]:
    if not BUSINESS_PROFILE_PATH.exists():
        return _default_business_profile()

    try:
        payload = json.loads(BUSINESS_PROFILE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return _default_business_profile()

    if not isinstance(payload, dict):
        return _default_business_profile()

    defaults = _default_business_profile()
    return {
        "business_name": str(payload.get("business_name") or defaults["business_name"]).strip(),
        "owner_name": str(payload.get("owner_name") or defaults["owner_name"]).strip(),
        "cro_number": str(payload.get("cro_number") or defaults["cro_number"]).strip(),
        "registration_date": str(payload.get("registration_date") or defaults["registration_date"]).strip(),
        "structure": _normalize_business_structure(payload.get("structure")),
        "vat_registered": bool(payload.get("vat_registered", defaults["vat_registered"])),
        "vat_threshold_basis": _normalize_vat_threshold_basis(payload.get("vat_threshold_basis")),
        "transition_date": str(payload.get("transition_date") or "").strip(),
        "pre_trading_start_date": str(payload.get("pre_trading_start_date") or "").strip(),
        "trading_start_date": str(payload.get("trading_start_date") or "").strip(),
        "updated_at": str(payload.get("updated_at") or defaults["updated_at"]).strip(),
    }


def _save_business_profile(profile: dict[str, Any]) -> None:
    normalized = {
        "business_name": str(profile.get("business_name") or "H-Queex").strip(),
        "owner_name": str(profile.get("owner_name") or "Hevandro Martire").strip(),
        "cro_number": str(profile.get("cro_number") or "790968").strip(),
        "registration_date": str(profile.get("registration_date") or "2026-08-04").strip(),
        "structure": _normalize_business_structure(profile.get("structure")),
        "vat_registered": bool(profile.get("vat_registered", True)),
        "vat_threshold_basis": _normalize_vat_threshold_basis(profile.get("vat_threshold_basis")),
        "transition_date": str(profile.get("transition_date") or "").strip(),
        "pre_trading_start_date": str(profile.get("pre_trading_start_date") or "").strip(),
        "trading_start_date": str(profile.get("trading_start_date") or "").strip(),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    BUSINESS_PROFILE_PATH.write_text(json.dumps(normalized, indent=2), encoding="utf-8")


def _normalize_vat_threshold_basis(value: Any) -> str:
    basis = str(value or "").strip().lower()
    return basis if basis in VAT_TURNOVER_THRESHOLDS else "services"


def _parse_transaction_date(value: Any) -> date | None:
    if isinstance(value, date):
        return value
    if not value:
        return None

    text = str(value).strip()
    if not text:
        return None

    for date_format in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    return None


def _resolve_phase_tag(transaction_date: Any) -> str:
    profile = _load_business_profile()
    structure = _normalize_business_structure(profile.get("structure"))
    transition_date = _parse_transaction_date(profile.get("transition_date"))
    pre_trading_start = _parse_transaction_date(profile.get("pre_trading_start_date"))
    record_date = _parse_transaction_date(transaction_date)

    trading_start = transition_date
    if pre_trading_start is not None and record_date is not None:
        if record_date >= pre_trading_start and (trading_start is None or record_date < trading_start):
            return "Pre-Trading"

    if structure == "sole_trader":
        return "Phase 1"

    if transition_date is None:
        return "Phase 2"

    if record_date is not None and record_date < transition_date:
        return "Phase 1"
    return "Phase 2"


def _phase_label_for_structure(structure: Any) -> str:
    if _normalize_business_structure(structure) == "limited_company":
        return "Phase 2 - Private Limited Company"
    return "Phase 1 - Sole Trader / Business Name"


def _append_message_to_path(path: str, message: str) -> str:
    separator = "&" if "?" in path else "?"
    return f"{path}{separator}message={quote(message)}"


def _build_phase_policy(summary: dict[str, Any], structure: str) -> dict[str, Any]:
    structure_key = _normalize_business_structure(structure)
    config = PHASE_POLICY.get(structure_key, PHASE_POLICY["sole_trader"])
    net_profit = _coerce_number(summary.get("net_cashflow", 0))
    taxable_profit = max(net_profit, 0.0)
    estimated_tax_due = round(taxable_profit * float(config.get("estimated_tax_rate", 0)), 2)
    return {
        "tax_regime": config.get("tax_regime", "Income Tax (Form 11)"),
        "estimated_tax_rate": float(config.get("estimated_tax_rate", 0.0)),
        "estimated_tax_due": estimated_tax_due,
        "report_template": config.get("report_template", "Form 11 outputs"),
        "owner_account_label": config.get("owner_account_label", "Proprietor Capital Account"),
        "next_filing_deadline": config.get("next_filing_deadline", ""),
        "taxable_profit_basis": taxable_profit,
    }


def _resolve_workbook_path() -> Path:
    candidates = [
        Path(WORKBOOK_PATH),
        BASE_DIR / DEFAULT_WORKBOOK_NAME,
    ]

    matching_workbooks = sorted(
        [
            path
            for path in BASE_DIR.glob("H-Queex_Financial_Control*.xls*")
            if path.is_file() and ".tmp-" not in path.name
        ],
        key=lambda path: path.name,
    )
    candidates.extend(matching_workbooks)

    if not matching_workbooks:
        legacy_candidates = [
            BASE_DIR / "H-Queex_Financial_Control by Claude V6.0 for App.xlsx",
            BASE_DIR / "H-Queex_Financial_Control by Claude V6.0 for App.xlsm",
        ]
        candidates.extend(legacy_candidates)

    for candidate in candidates:
        if candidate.exists():
            return candidate

    searched = ", ".join(str(candidate) for candidate in candidates)
    raise FileNotFoundError(
        f"Could not locate the finance workbook. Looked in: {searched}"
    )


SHEET_HEADERS = {
    "Income": ["Date", "Description", "Client / Source", "Category", "Invoice #", "Invoice ID", "Source", "Amount (€)", "Status", "Total incl. VAT (€)", "Payment Method", "Payment Date", "Notes"],
    "Expenses": [
        "Date (Registered)",
        "Supplier / Payee",
        "Supplier VAT Number",
        "Receipt / Invoice Ref",
        "Title",
        "Description",
        "Category",
        "Base Net Amount (€)",
        "Delivery (€)",
        "Fees (€)",
        "Other Charges (€)",
        "Discount Type",
        "Discount Value",
        "Discount (€)",
        "Net Amount (€)",
        "VAT Rate",
        "VAT Amount (€)",
        "Total (€)",
        "Input VAT Reclaimable",
        "Payment Method",
        "Deductibility Status",
        "Capital Expenditure Flag",
        "Receipt Attached",
        "Bank Reconciliation",
        "Status",
        "Notes",
    ],
    "Invoices": [
        "Invoice #",
        "Issue Date",
        "Due Date",
        "Client Name",
        "Client VAT Number",
        "Client Address",
        "Service / Product",
        "Base Net Amount (€)",
        "Delivery (€)",
        "Fees (€)",
        "Other Charges (€)",
        "Discount Type",
        "Discount Value",
        "Discount (€)",
        "Net (€)",
        "VAT Rate",
        "VAT Amount (€)",
        "Total (€)",
        "Balance Due (€)",
        "Status",
        "Payment Method",
        "Payment Date",
        "Bank Reconciliation",
        "Notes",
    ],
    "AR": ["Client", "Invoice #", "Issue Date", "Due Date", "Total (€)", "Paid (€)", "Balance (€)", "Status"],
    "AP": ["Ref #", "Supplier Name", "Description", "Invoice Date", "Due Date", "Net (€)", "Total (€)", "Paid (€)", "Balance Due (€)", "Status"],
    "VAT": ["Period", "Output VAT — Sales (€)", "Input VAT — Purchases (€)", "Net VAT Due (€)", "VAT Paid (€)", "Balance (€)", "Due Date", "Status"],
    "Clients": ["Client Name", "Contact Person", "Email", "Phone", "Country", "Service Tier", "Retainer Frequency", "Retainer Amount (€)"],
    "Suppliers": ["Supplier Name", "Contact Person", "Email", "Phone", "Country", "Default VAT Treatment", "Needs Completion"],
}

HEADER_ALIASES = {
    "Income": {
        "Amount": "Amount (€)",
        "Amount (€)": "Amount (€)",
        "Total": "Amount (€)",
        "Client Source": "Client / Source",
        "Client / Source": "Client / Source",
        "Client": "Client / Source",
        "Invoice Number": "Invoice #",
        "Invoice #": "Invoice #",
    },
    "Expenses": {
        "Amount": "Total (€)",
        "Base Net": "Base Net Amount (€)",
        "Base Net Amount": "Base Net Amount (€)",
        "Base Net Amount (€)": "Base Net Amount (€)",
        "Delivery": "Delivery (€)",
        "Delivery (€)": "Delivery (€)",
        "Fees": "Fees (€)",
        "Fees (€)": "Fees (€)",
        "Other Charges": "Other Charges (€)",
        "Other Charges (€)": "Other Charges (€)",
        "Discount": "Discount (€)",
        "Discount (€)": "Discount (€)",
        "Net Amount": "Net Amount (€)",
        "Net Amount (€)": "Net Amount (€)",
        "Total Amount": "Total (€)",
        "Supplier": "Supplier / Payee",
        "Supplier / Payee": "Supplier / Payee",
        "Supplier VAT": "Supplier VAT Number",
        "Supplier VAT Number": "Supplier VAT Number",
        "Receipt Ref": "Receipt / Invoice Ref",
        "Reference": "Receipt / Invoice Ref",
        "Input VAT": "Input VAT Reclaimable",
        "Deductibility": "Deductibility Status",
        "Capex": "Capital Expenditure Flag",
        "Receipt Attached": "Receipt Attached",
        "Bank Reconciliation": "Bank Reconciliation",
        "Notes": "Notes",
    },
    "Clients": {
        "Name": "Client Name",
        "Client": "Client Name",
        "Client Name": "Client Name",
    },
    "Suppliers": {
        "Name": "Supplier Name",
        "Supplier": "Supplier Name",
        "Supplier Name": "Supplier Name",
    },
    "Invoices": {
        "Invoice Number": "Invoice #",
        "Invoice #": "Invoice #",
        "Client VAT": "Client VAT Number",
        "VAT Number": "Client VAT Number",
        "Client Address": "Client Address",
        "Payment Date": "Payment Date",
        "Bank Reconciliation": "Bank Reconciliation",
        "Notes": "Notes",
    },
}


def _coerce_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("€", "").strip()
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def _coerce_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        return round(value, 2)
    return value


def _parse_vat_rate(value: Any) -> float:
    text = str(value or "").strip().lower()
    if not text or text == "exempt":
        return 0.0
    cleaned = text.replace("%", "")
    try:
        return float(cleaned) / 100.0
    except ValueError:
        return 0.0


def _normalize_vat_treatment(value: Any, vat_rate: Any = None) -> str:
    text = str(value or "").strip().lower().replace("-", "_").replace(" ", "_")
    if text in {"standard", "zero_rated", "exempt", "reverse_charge"}:
        return text

    vat_rate_text = str(vat_rate or "").strip().lower()
    if vat_rate_text == "exempt":
        return "exempt"
    if vat_rate_text in {"0%", "0", "0.0", "0.00"}:
        return "zero_rated"
    return "standard"


def _normalize_supply_type(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"services", "service"}:
        return "services"
    if text in {"goods", "good"}:
        return "goods"
    return "services"


def _normalize_input_vat_reclaimable(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text == "partial":
        return "Partial"
    if text in {"yes", "y", "true", "1"}:
        return "Yes"
    if text in {"no", "n", "false", "0"}:
        return "No"
    return "Yes"


NON_DEDUCTIBLE_CATEGORIES = {"entertainment", "personal expenses", "fines and penalties"}
PARTIALLY_DEDUCTIBLE_CATEGORIES = {"home office expenses", "motor expenses", "phone and communications"}


def _deductibility_locked_for_category(category: Any) -> str | None:
    """Return the forced Deductibility Status for a category, or None if the field is user-editable."""
    category_key = _normalize_category_key(category)
    if category_key in NON_DEDUCTIBLE_CATEGORIES:
        return "Non-Deductible"
    if category_key in PARTIALLY_DEDUCTIBLE_CATEGORIES:
        return "Partially Deductible"
    return None


def _normalize_deductibility_status(value: Any, category: Any) -> str:
    locked = _deductibility_locked_for_category(category)
    if locked is not None:
        return locked
    if str(value or "").strip() in EXPENSE_DEDUCTIBILITY_OPTIONS:
        return str(value).strip()
    return "Fully Deductible"


def _normalize_yes_no(value: Any, *, default_yes: bool = False) -> str:
    text = str(value or "").strip().lower()
    if text in {"yes", "y", "true", "1"}:
        return "Yes"
    if text in {"no", "n", "false", "0"}:
        return "No"
    return "Yes" if default_yes else "No"


def _normalize_reconciliation(value: Any) -> str:
    text = str(value or "").strip().lower()
    return "Reconciled" if text == "reconciled" else "Unreconciled"


def _normalize_payroll_status(value: Any) -> str:
    text = str(value or "").strip().title()
    return text if text in PAYROLL_STATUS_OPTIONS else "Draft"


def _normalize_payroll_payload(payload: dict[str, Any]) -> dict[str, Any]:
    gross_pay = round(_coerce_number(payload.get("Gross Pay (€)")), 2)
    paye_amount = round(max(_coerce_number(payload.get("PAYE (€)")), 0.0), 2)
    usc_amount = round(max(_coerce_number(payload.get("USC (€)")), 0.0), 2)
    employee_prsi_amount = round(max(_coerce_number(payload.get("Employee PRSI (€)")), 0.0), 2)
    employer_prsi_amount = round(max(_coerce_number(payload.get("Employer PRSI (€)")), 0.0), 2)
    deductions_total = round(paye_amount + usc_amount + employee_prsi_amount, 2)
    net_pay = round(max(gross_pay - deductions_total, 0.0), 2)
    employer_cost = round(gross_pay + employer_prsi_amount, 2)

    payload["Gross Pay (€)"] = f"{gross_pay:.2f}"
    payload["PAYE (€)"] = f"{paye_amount:.2f}"
    payload["USC (€)"] = f"{usc_amount:.2f}"
    payload["Employee PRSI (€)"] = f"{employee_prsi_amount:.2f}"
    payload["Employer PRSI (€)"] = f"{employer_prsi_amount:.2f}"
    payload["Net Pay (€)"] = f"{net_pay:.2f}"
    payload["Employer Cost (€)"] = f"{employer_cost:.2f}"
    payload["Status"] = _normalize_payroll_status(payload.get("Status"))
    payload["Bank Reconciliation"] = _normalize_reconciliation(payload.get("Bank Reconciliation"))
    return payload


def _load_payroll_entries() -> list[dict[str, Any]]:
    records = _load_json_records(PAYROLL_PATH)
    normalized: list[dict[str, Any]] = []
    for record in records:
        if not isinstance(record, dict):
            continue
        item = {
            "id": str(record.get("id") or uuid4()),
            "Pay Date": str(record.get("Pay Date") or "").strip(),
            "Payroll Period": str(record.get("Payroll Period") or "").strip(),
            "Employee Name": str(record.get("Employee Name") or "").strip(),
            "Gross Pay (€)": record.get("Gross Pay (€)", "0.00"),
            "PAYE (€)": record.get("PAYE (€)", "0.00"),
            "USC (€)": record.get("USC (€)", "0.00"),
            "Employee PRSI (€)": record.get("Employee PRSI (€)", "0.00"),
            "Employer PRSI (€)": record.get("Employer PRSI (€)", "0.00"),
            "Net Pay (€)": record.get("Net Pay (€)", "0.00"),
            "Employer Cost (€)": record.get("Employer Cost (€)", "0.00"),
            "Status": record.get("Status", "Draft"),
            "Payment Method": str(record.get("Payment Method") or "").strip(),
            "Payment Date": str(record.get("Payment Date") or "").strip(),
            "Bank Reconciliation": record.get("Bank Reconciliation", "Unreconciled"),
            "Notes": str(record.get("Notes") or "").strip(),
            "Phase Tag": str(record.get("Phase Tag") or "").strip(),
            "created_at": str(record.get("created_at") or datetime.now().isoformat(timespec="seconds")),
            "last_updated_at": str(record.get("last_updated_at") or datetime.now().isoformat(timespec="seconds")),
        }
        _normalize_payroll_payload(item)
        normalized.append(item)
    normalized.sort(key=lambda item: (item.get("Pay Date") or "", item.get("Employee Name") or ""), reverse=True)
    return normalized


def _save_payroll_entries(entries: list[dict[str, Any]]) -> None:
    payload: list[dict[str, Any]] = []
    for entry in entries:
        item = dict(entry)
        _normalize_payroll_payload(item)
        payload.append(item)
    _save_json_records(PAYROLL_PATH, payload)


def _find_payroll_by_id(entries: list[dict[str, Any]], payroll_id: Any) -> dict[str, Any] | None:
    target_id = str(payroll_id or "").strip()
    if not target_id:
        return None
    for entry in entries:
        if str(entry.get("id") or "") == target_id:
            return entry
    return None


def _summarize_payroll_entries(entries: list[dict[str, Any]]) -> dict[str, Any]:
    liabilities_total = 0.0
    for entry in entries:
        liabilities_total += _coerce_number(entry.get("PAYE (€)"))
        liabilities_total += _coerce_number(entry.get("USC (€)"))
        liabilities_total += _coerce_number(entry.get("Employee PRSI (€)"))
        liabilities_total += _coerce_number(entry.get("Employer PRSI (€)"))

    return {
        "count": len(entries),
        "gross_total": round(sum(_coerce_number(entry.get("Gross Pay (€)")) for entry in entries), 2),
        "net_total": round(sum(_coerce_number(entry.get("Net Pay (€)")) for entry in entries), 2),
        "employer_cost_total": round(sum(_coerce_number(entry.get("Employer Cost (€)")) for entry in entries), 2),
        "liabilities_total": round(liabilities_total, 2),
        "paid_count": sum(1 for entry in entries if _normalize_payroll_status(entry.get("Status")) == "Paid"),
    }


def _export_payroll_csv(entries: list[dict[str, Any]]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "pay_date",
            "payroll_period",
            "employee_name",
            "gross_pay_eur",
            "paye_eur",
            "usc_eur",
            "employee_prsi_eur",
            "employer_prsi_eur",
            "net_pay_eur",
            "employer_cost_eur",
            "status",
            "payment_method",
            "bank_reconciliation",
            "notes",
            "phase_tag",
        ]
    )
    for entry in sorted(entries, key=lambda item: (str(item.get("Pay Date") or ""), str(item.get("Employee Name") or "")), reverse=True):
        writer.writerow(
            [
                entry.get("Pay Date", ""),
                entry.get("Payroll Period", ""),
                entry.get("Employee Name", ""),
                round(_coerce_number(entry.get("Gross Pay (€)")), 2),
                round(_coerce_number(entry.get("PAYE (€)")), 2),
                round(_coerce_number(entry.get("USC (€)")), 2),
                round(_coerce_number(entry.get("Employee PRSI (€)")), 2),
                round(_coerce_number(entry.get("Employer PRSI (€)")), 2),
                round(_coerce_number(entry.get("Net Pay (€)")), 2),
                round(_coerce_number(entry.get("Employer Cost (€)")), 2),
                _normalize_payroll_status(entry.get("Status")),
                entry.get("Payment Method", ""),
                _normalize_reconciliation(entry.get("Bank Reconciliation")),
                entry.get("Notes", ""),
                entry.get("Phase Tag", ""),
            ]
        )
    return buffer.getvalue()


def _is_paid_status(entity_type: str, status: Any) -> bool:
    normalized = str(status or "").strip().lower()
    if entity_type == "expense":
        return normalized in {"paid", "auto-posted", "auto_posted"}
    if entity_type == "income":
        return normalized in {"received", "partially received", "partially_received"}
    if entity_type == "invoice":
        return normalized in {"paid", "partially paid", "partially_paid"}
    if entity_type == "payroll":
        return normalized in {"paid", "filed"}
    return False


def _apply_default_payment_date_for_paid(payload: dict[str, Any], entity_type: str, date_key: str) -> bool:
    if not _is_paid_status(entity_type, payload.get("Status")):
        return False
    if str(payload.get("Payment Date") or "").strip():
        return False
    fallback_date = str(payload.get(date_key) or "").strip()
    if _parse_iso_date(fallback_date) is not None:
        payload["Payment Date"] = fallback_date
        return True
    return False


def _parse_bank_statement_date(value: Any) -> str:
    parsed = _parse_iso_date(value)
    if parsed:
        return parsed.isoformat()
    raw = str(value or "").strip()
    if not raw:
        return ""
    for pattern in ["%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"]:
        try:
            return datetime.strptime(raw, pattern).date().isoformat()
        except ValueError:
            continue
    return ""


def _load_bank_statement_lines() -> list[dict[str, Any]]:
    lines = _load_json_records(BANK_STATEMENTS_PATH)
    normalized: list[dict[str, Any]] = []
    for line in lines:
        if not isinstance(line, dict):
            continue
        normalized.append(
            {
                "id": str(line.get("id") or uuid4()),
                "date": _parse_bank_statement_date(line.get("date")),
                "description": str(line.get("description") or "").strip(),
                "reference": str(line.get("reference") or "").strip(),
                "amount_eur": round(_coerce_number(line.get("amount_eur")), 2),
                "balance_eur": round(_coerce_number(line.get("balance_eur")), 2),
                "payment_method": str(line.get("payment_method") or "").strip(),
                "source_filename": str(line.get("source_filename") or "").strip(),
                "uploaded_at": str(line.get("uploaded_at") or ""),
            }
        )
    normalized.sort(key=lambda item: (item.get("date") or "", item.get("id") or ""), reverse=True)
    return normalized


def _save_bank_statement_lines(lines: list[dict[str, Any]]) -> None:
    payload: list[dict[str, Any]] = []
    for line in lines:
        payload.append(
            {
                "id": str(line.get("id") or uuid4()),
                "date": _parse_bank_statement_date(line.get("date")),
                "description": str(line.get("description") or "").strip(),
                "reference": str(line.get("reference") or "").strip(),
                "amount_eur": round(_coerce_number(line.get("amount_eur")), 2),
                "balance_eur": round(_coerce_number(line.get("balance_eur")), 2),
                "payment_method": str(line.get("payment_method") or "").strip(),
                "source_filename": str(line.get("source_filename") or "").strip(),
                "uploaded_at": str(line.get("uploaded_at") or datetime.now().isoformat(timespec="seconds")),
            }
        )
    _save_json_records(BANK_STATEMENTS_PATH, payload)


def _ingest_bank_statement_csv(content: str, *, source_filename: str) -> dict[str, int]:
    existing_lines = _load_bank_statement_lines()
    existing_keys = {
        f"{line.get('date')}|{line.get('description')}|{line.get('reference')}|{round(_coerce_number(line.get('amount_eur')), 2):.2f}"
        for line in existing_lines
    }
    reader = csv.DictReader(StringIO(content))
    imported_count = 0
    skipped_count = 0

    for raw_row in reader:
        if not isinstance(raw_row, dict):
            skipped_count += 1
            continue
        row = {str(key or "").strip().lower(): value for key, value in raw_row.items()}
        date_text = _parse_bank_statement_date(
            row.get("date")
            or row.get("transaction date")
            or row.get("posted date")
            or row.get("value date")
        )
        description = str(row.get("description") or row.get("narrative") or row.get("details") or "").strip()
        reference = str(row.get("reference") or row.get("ref") or row.get("transaction id") or "").strip()
        payment_method = str(row.get("payment method") or row.get("account") or "").strip()

        amount_value = _try_parse_number(row.get("amount"))
        if amount_value is None:
            debit_value = _try_parse_number(row.get("debit") or row.get("withdrawal"))
            credit_value = _try_parse_number(row.get("credit") or row.get("deposit"))
            if debit_value is not None and credit_value is None:
                amount_value = -abs(debit_value)
            elif credit_value is not None and debit_value is None:
                amount_value = abs(credit_value)
            elif credit_value is not None and debit_value is not None:
                amount_value = abs(credit_value) - abs(debit_value)

        if not date_text or amount_value is None or abs(amount_value) < 0.005:
            skipped_count += 1
            continue

        rounded_amount = round(float(amount_value), 2)
        key = f"{date_text}|{description}|{reference}|{rounded_amount:.2f}"
        if key in existing_keys:
            skipped_count += 1
            continue

        existing_lines.append(
            {
                "id": str(uuid4()),
                "date": date_text,
                "description": description,
                "reference": reference,
                "amount_eur": rounded_amount,
                "balance_eur": round(_coerce_number(row.get("balance") or row.get("running balance")), 2),
                "payment_method": payment_method,
                "source_filename": source_filename,
                "uploaded_at": datetime.now().isoformat(timespec="seconds"),
            }
        )
        existing_keys.add(key)
        imported_count += 1

    _save_bank_statement_lines(existing_lines)
    return {"imported_count": imported_count, "skipped_count": skipped_count}


def _match_bank_statement_lines(reconciliation_rows: list[dict[str, Any]], statement_lines: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    candidates = [
        row
        for row in reconciliation_rows
        if bool(row.get("is_paid")) and row.get("bank_reconciliation") == "Unreconciled" and _coerce_number(row.get("amount_eur")) > 0
    ]
    used_candidate_indexes: set[int] = set()
    matched_statement_ids: set[str] = set()

    for row in reconciliation_rows:
        row["statement_match_count"] = 0
        row["statement_match_ids"] = []

    enriched_statement_lines: list[dict[str, Any]] = []
    for statement_line in statement_lines:
        line = dict(statement_line)
        line_date = _parse_iso_date(line.get("date"))
        line_amount = abs(round(_coerce_number(line.get("amount_eur")), 2))
        best_index: int | None = None
        best_score: tuple[int, float] | None = None

        for index, candidate in enumerate(candidates):
            if index in used_candidate_indexes:
                continue
            candidate_amount = round(_coerce_number(candidate.get("amount_eur")), 2)
            if abs(candidate_amount - line_amount) > 0.009:
                continue

            candidate_date = _parse_iso_date(candidate.get("date"))
            if line_date is None or candidate_date is None:
                continue

            date_diff = abs((candidate_date - line_date).days)
            if date_diff > RECONCILIATION_MATCH_DAYS:
                continue

            candidate_method = str(candidate.get("payment_method") or "").strip().lower()
            statement_method = str(line.get("payment_method") or "").strip().lower()
            method_score = 1 if candidate_method and statement_method and candidate_method == statement_method else 0
            score = (method_score, -float(date_diff))
            if best_score is None or score > best_score:
                best_score = score
                best_index = index

        if best_index is not None:
            used_candidate_indexes.add(best_index)
            matched_statement_ids.add(str(line.get("id") or ""))
            matched_row = candidates[best_index]
            matched_row["statement_match_count"] = int(matched_row.get("statement_match_count") or 0) + 1
            current_ids = matched_row.get("statement_match_ids") if isinstance(matched_row.get("statement_match_ids"), list) else []
            current_ids.append(str(line.get("id") or ""))
            matched_row["statement_match_ids"] = current_ids
            line["matched_entity_type"] = matched_row.get("entity_type")
            line["matched_reference"] = matched_row.get("reference")
        else:
            line["matched_entity_type"] = ""
            line["matched_reference"] = ""
        enriched_statement_lines.append(line)

    unmatched_lines = [line for line in enriched_statement_lines if str(line.get("id") or "") not in matched_statement_ids]

    for row in reconciliation_rows:
        if not bool(row.get("is_paid")):
            continue
        if row.get("bank_reconciliation") != "Unreconciled":
            continue
        if int(row.get("statement_match_count") or 0) == 0:
            reasons = row.get("exception_reasons") if isinstance(row.get("exception_reasons"), list) else []
            if "no_bank_statement_match" not in reasons:
                reasons.append("no_bank_statement_match")
            row["exception_reasons"] = reasons

    return enriched_statement_lines, unmatched_lines


def _export_bank_statement_csv(lines: list[dict[str, Any]], *, unmatched_only: bool = False) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "date",
            "description",
            "reference",
            "amount_eur",
            "balance_eur",
            "payment_method",
            "matched_entity_type",
            "matched_reference",
            "source_filename",
            "uploaded_at",
        ]
    )
    for line in lines:
        if unmatched_only and str(line.get("matched_entity_type") or "").strip():
            continue
        writer.writerow(
            [
                line.get("date", ""),
                line.get("description", ""),
                line.get("reference", ""),
                round(_coerce_number(line.get("amount_eur")), 2),
                round(_coerce_number(line.get("balance_eur")), 2),
                line.get("payment_method", ""),
                line.get("matched_entity_type", ""),
                line.get("matched_reference", ""),
                line.get("source_filename", ""),
                line.get("uploaded_at", ""),
            ]
        )
    return buffer.getvalue()


def _build_reconciliation_rows(data: dict[str, Any], payroll_entries: list[dict[str, Any]], *, today: date | None = None) -> list[dict[str, Any]]:
    current_day = today or date.today()
    rows: list[dict[str, Any]] = []

    for expense in data.get("sheets", {}).get("Expenses", []):
        transaction_date = _parse_iso_date(expense.get("Date (Registered)"))
        status = str(expense.get("Status") or "").strip()
        reconciliation = _normalize_reconciliation(expense.get("Bank Reconciliation"))
        is_paid = _is_paid_status("expense", status)
        amount = round(_coerce_number(expense.get("Total (€)")), 2)
        age_days = (current_day - transaction_date).days if transaction_date else None
        reasons: list[str] = []
        if is_paid and reconciliation == "Unreconciled" and age_days is not None and age_days > 7:
            reasons.append("paid_unreconciled_over_7_days")
        if is_paid and not str(expense.get("Payment Method") or "").strip():
            reasons.append("missing_payment_method")
        if is_paid and not str(expense.get("Supplier / Payee") or "").strip():
            reasons.append("missing_counterparty")
        rows.append(
            {
                "entity_type": "expense",
                "row_number": expense.get("__row_number"),
                "payroll_id": "",
                "date": transaction_date.isoformat() if transaction_date else str(expense.get("Date (Registered)") or ""),
                "counterparty": str(expense.get("Supplier / Payee") or ""),
                "reference": str(expense.get("Receipt / Invoice Ref") or expense.get("Title") or ""),
                "amount_eur": amount,
                "status": status,
                "is_paid": is_paid,
                "bank_reconciliation": reconciliation,
                "payment_method": str(expense.get("Payment Method") or ""),
                "age_days": age_days,
                "exception_reasons": reasons,
                "matching_key": "",
                "matching_group_size": 1,
                "statement_match_count": 0,
                "statement_match_ids": [],
            }
        )

    for invoice in data.get("sheets", {}).get("Invoices", []):
        payment_date = _parse_iso_date(invoice.get("Payment Date"))
        issue_date = _parse_iso_date(invoice.get("Issue Date"))
        effective_date = payment_date or issue_date
        status = _normalize_invoice_status(invoice.get("Status"))
        reconciliation = _normalize_reconciliation(invoice.get("Bank Reconciliation"))
        is_paid = _is_paid_status("invoice", status)
        amount = round(_coerce_number(invoice.get("Total (€)")), 2)
        age_days = (current_day - effective_date).days if effective_date else None
        reasons: list[str] = []
        if is_paid and not payment_date:
            reasons.append("missing_payment_date")
        if is_paid and reconciliation == "Unreconciled" and age_days is not None and age_days > 7:
            reasons.append("paid_unreconciled_over_7_days")
        if is_paid and not str(invoice.get("Payment Method") or "").strip():
            reasons.append("missing_payment_method")
        rows.append(
            {
                "entity_type": "invoice",
                "row_number": invoice.get("__row_number"),
                "payroll_id": "",
                "date": effective_date.isoformat() if effective_date else str(invoice.get("Issue Date") or ""),
                "counterparty": str(invoice.get("Client Name") or ""),
                "reference": str(invoice.get("Invoice #") or ""),
                "amount_eur": amount,
                "status": status,
                "is_paid": is_paid,
                "bank_reconciliation": reconciliation,
                "payment_method": str(invoice.get("Payment Method") or ""),
                "age_days": age_days,
                "exception_reasons": reasons,
                "matching_key": "",
                "matching_group_size": 1,
                "statement_match_count": 0,
                "statement_match_ids": [],
            }
        )

    for payroll in payroll_entries:
        pay_date = _parse_iso_date(payroll.get("Pay Date"))
        status = _normalize_payroll_status(payroll.get("Status"))
        reconciliation = _normalize_reconciliation(payroll.get("Bank Reconciliation"))
        is_paid = _is_paid_status("payroll", status)
        amount = round(_coerce_number(payroll.get("Net Pay (€)")), 2)
        age_days = (current_day - pay_date).days if pay_date else None
        reasons: list[str] = []
        if is_paid and reconciliation == "Unreconciled" and age_days is not None and age_days > 7:
            reasons.append("paid_unreconciled_over_7_days")
        if is_paid and not str(payroll.get("Payment Method") or "").strip():
            reasons.append("missing_payment_method")
        rows.append(
            {
                "entity_type": "payroll",
                "row_number": None,
                "payroll_id": str(payroll.get("id") or ""),
                "date": pay_date.isoformat() if pay_date else str(payroll.get("Pay Date") or ""),
                "counterparty": str(payroll.get("Employee Name") or ""),
                "reference": str(payroll.get("Payroll Period") or ""),
                "amount_eur": amount,
                "status": status,
                "is_paid": is_paid,
                "bank_reconciliation": reconciliation,
                "payment_method": str(payroll.get("Payment Method") or ""),
                "age_days": age_days,
                "exception_reasons": reasons,
                "matching_key": "",
                "matching_group_size": 1,
                "statement_match_count": 0,
                "statement_match_ids": [],
            }
        )

    groups: dict[str, int] = {}
    for row in rows:
        payment_method = str(row.get("payment_method") or "").strip().lower()
        date_text = str(row.get("date") or "")
        amount_value = round(_coerce_number(row.get("amount_eur")), 2)
        if not date_text or amount_value <= 0 or not payment_method:
            continue
        key = f"{date_text}|{amount_value:.2f}|{payment_method}"
        row["matching_key"] = key
        groups[key] = groups.get(key, 0) + 1

    for row in rows:
        key = str(row.get("matching_key") or "")
        row["matching_group_size"] = groups.get(key, 1) if key else 1

    rows.sort(key=lambda item: (str(item.get("date") or ""), str(item.get("entity_type") or ""), str(item.get("reference") or "")), reverse=True)
    return rows


def _summarize_reconciliation(rows: list[dict[str, Any]]) -> dict[str, Any]:
    paid_rows = [row for row in rows if bool(row.get("is_paid"))]
    unreconciled_paid_rows = [row for row in paid_rows if row.get("bank_reconciliation") == "Unreconciled"]
    reconciled_paid_rows = [row for row in paid_rows if row.get("bank_reconciliation") == "Reconciled"]
    exception_rows = [row for row in paid_rows if row.get("exception_reasons")]
    return {
        "tracked_rows": len(rows),
        "paid_rows": len(paid_rows),
        "reconciled_paid_rows": len(reconciled_paid_rows),
        "unreconciled_paid_rows": len(unreconciled_paid_rows),
        "exception_rows": len(exception_rows),
        "reconciled_amount_eur": round(sum(_coerce_number(row.get("amount_eur")) for row in reconciled_paid_rows), 2),
        "unreconciled_amount_eur": round(sum(_coerce_number(row.get("amount_eur")) for row in unreconciled_paid_rows), 2),
    }


def _export_reconciliation_csv(rows: list[dict[str, Any]], *, exceptions_only: bool = False) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "entity_type",
            "reference",
            "counterparty",
            "date",
            "amount_eur",
            "status",
            "bank_reconciliation",
            "payment_method",
            "is_paid",
            "age_days",
            "exception_reasons",
            "matching_group_size",
        ]
    )
    for row in rows:
        if exceptions_only and not row.get("exception_reasons"):
            continue
        writer.writerow(
            [
                row.get("entity_type", ""),
                row.get("reference", ""),
                row.get("counterparty", ""),
                row.get("date", ""),
                round(_coerce_number(row.get("amount_eur")), 2),
                row.get("status", ""),
                row.get("bank_reconciliation", ""),
                row.get("payment_method", ""),
                "Yes" if row.get("is_paid") else "No",
                row.get("age_days", ""),
                "|".join(row.get("exception_reasons", [])),
                row.get("matching_group_size", 1),
            ]
        )
    return buffer.getvalue()


def _is_capital_expense(payload: dict[str, Any]) -> bool:
    explicit_flag = str(payload.get("Capital Expenditure Flag") or "").strip().lower()
    if explicit_flag in {"yes", "y", "true", "1"}:
        return True
    total_amount = _coerce_number(payload.get("Total (€)"))
    return total_amount > 1000.0


def _process_compliance_flags(form) -> tuple[list[dict[str, Any]], bool, str]:
    """Parse the Layer 1/2 compliance flags a form submitted. Returns (flags, acknowledged, acknowledged_at)."""
    try:
        flags = json.loads(form.get("compliance_flags_json") or "[]")
        if not isinstance(flags, list):
            flags = []
    except (TypeError, ValueError):
        flags = []
    flags = [flag for flag in flags if isinstance(flag, dict) and flag.get("message")]
    acknowledged = form.get("flags_acknowledged") == "1"
    acknowledged_at = str(form.get("flags_acknowledged_at") or "").strip()
    return flags, acknowledged, acknowledged_at


def _apply_compliance_flags_to_payload(payload: dict[str, Any], flags: list[dict[str, Any]], acknowledged: bool, acknowledged_at: str) -> dict[str, str]:
    """Stamp compliance flag fields onto a payload. Returns validation errors if a red flag wasn't acknowledged."""
    errors: dict[str, str] = {}
    has_red_flag = any(str(flag.get("severity")) == "danger" for flag in flags)
    if has_red_flag and not acknowledged:
        errors["compliance_flags"] = "This transaction has compliance flags that need your attention — acknowledge them or fix the flagged fields before saving"
    payload["Compliance Flags"] = json.dumps(flags)
    payload["Flags Acknowledged"] = "Yes" if (has_red_flag and acknowledged) else "No"
    payload["Flags Acknowledged At"] = acknowledged_at if (has_red_flag and acknowledged) else ""
    return errors


def _save_uploaded_receipt(file_storage: Any) -> str:
    """Save an uploaded receipt file into RECEIPTS_DIR and return the stored filename, or '' if no valid file."""
    if not file_storage or not getattr(file_storage, "filename", ""):
        return ""
    original_name = secure_filename(file_storage.filename)
    if not original_name or "." not in original_name:
        return ""
    extension = original_name.rsplit(".", 1)[-1].lower()
    if extension not in ALLOWED_RECEIPT_EXTENSIONS:
        return ""
    stored_name = f"{uuid4().hex[:10]}_{original_name}"
    file_storage.save(RECEIPTS_DIR / stored_name)
    return stored_name


PRE_TRADING_CAPEX_THRESHOLD = 1000.0
PRE_TRADING_CAPEX_MESSAGE = "Pre-trading capital expenditure — confirm this qualifies as a pre-trading expense with your accountant before claiming."


def _apply_pretrading_compliance_flag(payload: dict[str, Any], flags: list[dict[str, Any]]) -> None:
    if str(payload.get("Phase Tag") or "") != "Pre-Trading":
        return
    if _coerce_number(payload.get("Total (€)")) <= PRE_TRADING_CAPEX_THRESHOLD:
        return
    if any(flag.get("key") == "pretrading_capex" for flag in flags if isinstance(flag, dict)):
        return
    flags.append({"key": "pretrading_capex", "severity": "warning", "message": PRE_TRADING_CAPEX_MESSAGE})


def _apply_expense_compliance_fields(payload: dict[str, Any]) -> None:
    if _normalize_category_key(payload.get("Category")) == "entertainment":
        payload["Input VAT Reclaimable"] = "No"
    else:
        payload["Input VAT Reclaimable"] = _normalize_input_vat_reclaimable(payload.get("Input VAT Reclaimable"))
    payload["Deductibility Status"] = _normalize_deductibility_status(payload.get("Deductibility Status"), payload.get("Category"))
    if payload["Deductibility Status"] == "Non-Deductible":
        payload["Net Amount (€)"] = f"{0.0:.2f}"
    payload["Receipt Attached"] = _normalize_yes_no(payload.get("Receipt Attached"), default_yes=False)
    payload["Bank Reconciliation"] = _normalize_reconciliation(payload.get("Bank Reconciliation"))
    payload["Capital Expenditure Flag"] = "Yes" if _is_capital_expense(payload) else "No"


def _load_capital_assets() -> list[dict[str, Any]]:
    return _load_json_records(CAPITAL_ASSETS_PATH)


def _save_capital_assets(assets: list[dict[str, Any]]) -> None:
    _save_json_records(CAPITAL_ASSETS_PATH, assets)


def _upsert_capital_asset_from_expense(payload: dict[str, Any], row_number: int, *, active: bool) -> None:
    assets = _load_capital_assets()
    target_id = f"expense-{row_number}"
    remaining = [asset for asset in assets if str(asset.get("id") or "") != target_id]

    if active:
        total_amount = round(_coerce_number(payload.get("Total (€)")), 2)
        annual_allowance = round(total_amount * 0.125, 2)
        remaining.append(
            {
                "id": target_id,
                "source": "expense",
                "expense_row_number": row_number,
                "acquisition_date": str(payload.get("Date (Registered)") or ""),
                "supplier": str(payload.get("Supplier / Payee") or ""),
                "description": str(payload.get("Description") or payload.get("Title") or ""),
                "category": str(payload.get("Category") or ""),
                "cost_eur": total_amount,
                "allowance_rate": 0.125,
                "allowance_years": 8,
                "annual_allowance_eur": annual_allowance,
                "phase_tag": str(payload.get("Phase Tag") or ""),
                "active": True,
                "updated_at": datetime.now().isoformat(timespec="seconds"),
            }
        )

    _save_capital_assets(remaining)


def _summarize_capital_assets(assets: list[dict[str, Any]]) -> dict[str, Any]:
    active_assets = [asset for asset in assets if bool(asset.get("active", True))]
    return {
        "asset_count": len(active_assets),
        "total_cost": round(sum(_coerce_number(asset.get("cost_eur")) for asset in active_assets), 2),
        "annual_allowance_total": round(sum(_coerce_number(asset.get("annual_allowance_eur")) for asset in active_assets), 2),
    }


def _export_capital_allowances_csv(assets: list[dict[str, Any]]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow([
        "asset_id",
        "acquisition_date",
        "supplier",
        "description",
        "category",
        "cost_eur",
        "allowance_rate",
        "allowance_years",
        "annual_allowance_eur",
        "phase_tag",
        "active",
    ])
    for asset in assets:
        writer.writerow([
            asset.get("id", ""),
            asset.get("acquisition_date", ""),
            asset.get("supplier", ""),
            asset.get("description", ""),
            asset.get("category", ""),
            round(_coerce_number(asset.get("cost_eur")), 2),
            asset.get("allowance_rate", 0.125),
            asset.get("allowance_years", 8),
            round(_coerce_number(asset.get("annual_allowance_eur")), 2),
            asset.get("phase_tag", ""),
            "Yes" if bool(asset.get("active", True)) else "No",
        ])
    return buffer.getvalue()


def _normalize_invoice_status(value: Any) -> str:
    status = str(value or "").strip().lower()
    status_map = {
        "draft": "Draft",
        "issued": "Issued",
        "paid": "Paid",
        "partially paid": "Partially Paid",
        "partially_paid": "Partially Paid",
        "overdue": "Overdue",
        "bad debt": "Bad Debt",
        "bad_debt": "Bad Debt",
        "cancelled": "Cancelled",
    }
    return status_map.get(status, "Draft")


def _auto_flag_overdue_invoices(invoices: list[dict[str, Any]]) -> bool:
    """Flip Issued invoices past their due date to Overdue in place. Returns True if anything changed."""
    changed = False
    today = date.today()
    for row in invoices:
        if _normalize_invoice_status(row.get("Status")) != "Issued":
            continue
        due_date = _parse_iso_date(row.get("Due Date"))
        if due_date and due_date < today:
            row["Status"] = "Overdue"
            changed = True
    return changed


def _sync_invoice_income_entry(invoice: dict[str, Any]) -> None:
    """Create/update/remove the Income entry linked to an invoice so it never needs manual duplication.

    Paid/Partially Paid invoices get exactly one linked Income row (Source=invoiced) reflecting the
    cumulative amount actually received. Any other status removes the linked row — Draft/Issued/Overdue
    haven't been received yet, and Bad Debt/Cancelled never will be.
    """
    invoice_number = str(invoice.get("Invoice #") or "").strip()
    if not invoice_number:
        return

    status = _normalize_invoice_status(invoice.get("Status"))
    income_records = _load_sheet_records_raw("Income")
    existing_index = next(
        (index for index, record in enumerate(income_records) if str(record.get("Invoice ID") or "") == invoice_number),
        None,
    )

    if status in ("Paid", "Partially Paid"):
        total = _coerce_number(invoice.get("Total (€)"))
        balance = _coerce_number(invoice.get("Balance Due (€)"))
        amount_received = round(max(total - balance, 0.0), 2)
        ratio = (amount_received / total) if total > 0 else 1.0
        line_item_names = ", ".join(str(item.get("name") or "").strip() for item in invoice.get("line_items") or [] if str(item.get("name") or "").strip())
        record = {
            "Date": invoice.get("Payment Date") or invoice.get("Issue Date") or "",
            "Description": line_item_names or invoice.get("Service / Product") or f"Invoice {invoice_number}",
            "Client / Source": invoice.get("Client Name") or "",
            "Category": "Consulting / Project Fees",
            "Invoice #": invoice_number,
            "Invoice ID": invoice_number,
            "Source": "invoiced",
            "Amount (€)": f"{round(_coerce_number(invoice.get('Net (€)')) * ratio, 2):.2f}",
            "Total incl. VAT (€)": f"{amount_received:.2f}",
            "VAT Rate": invoice.get("VAT Rate") or "0%",
            "VAT Amount (€)": f"{round(_coerce_number(invoice.get('VAT Amount (€)')) * ratio, 2):.2f}",
            "VAT Treatment": invoice.get("VAT Treatment") or "standard",
            "Supply Type": invoice.get("Supply Type") or "services",
            "Status": "Received",
            "Payment Method": invoice.get("Payment Method") or "",
            "Phase Tag": invoice.get("Phase Tag") or "",
        }
        if existing_index is not None:
            income_records[existing_index] = record
        else:
            income_records.append(record)
        _save_sheet_records_raw("Income", income_records)
    elif existing_index is not None:
        del income_records[existing_index]
        _save_sheet_records_raw("Income", income_records)


def _migrate_income_invoice_linkage() -> None:
    """Startup consistency pass: stamp Source/Invoice ID on legacy Income rows and make sure every
    Paid/Partially Paid invoice has exactly one linked Income row. Safe to run on every startup."""
    income_records = _load_sheet_records_raw("Income")
    changed = False
    for record in income_records:
        if "Source" not in record:
            record["Source"] = "manual"
            changed = True
        if "Invoice ID" not in record:
            record["Invoice ID"] = ""
            changed = True
    if changed:
        _save_sheet_records_raw("Income", income_records)

    for invoice in _load_sheet_records_raw("Invoices"):
        if _normalize_invoice_status(invoice.get("Status")) in ("Paid", "Partially Paid"):
            _sync_invoice_income_entry(invoice)


def _normalize_expense_status(value: Any) -> str:
    status = str(value or "").strip().lower()
    status_map = {
        "pending": "Pending",
        "approved": "Approved",
        "paid": "Paid",
        "auto-posted": "Auto-posted",
        "auto_posted": "Auto-posted",
        "autoposted": "Auto-posted",
        "cancelled": "Cancelled",
    }
    return status_map.get(status, "Pending")


def _next_invoice_number(issue_date_value: Any, existing_invoices: list[dict[str, Any]]) -> str:
    issue_date = _parse_transaction_date(issue_date_value) or date.today()
    year_token = str(issue_date.year)
    prefix = f"HQ-{year_token}-"
    max_seq = 0
    for row in existing_invoices:
        invoice_number = str(row.get("Invoice #") or "").strip().upper()
        if not invoice_number.startswith(prefix):
            continue
        suffix = invoice_number.replace(prefix, "", 1)
        if suffix.isdigit():
            max_seq = max(max_seq, int(suffix))
    return f"{prefix}{max_seq + 1:03d}"


def _apply_vat_classification(
    payload: dict[str, Any],
    *,
    vat_rate_key: str,
    vat_treatment_key: str = "VAT Treatment",
    supply_type_key: str = "Supply Type",
) -> None:
    payload[vat_treatment_key] = _normalize_vat_treatment(payload.get(vat_treatment_key), payload.get(vat_rate_key))
    payload[supply_type_key] = _normalize_supply_type(payload.get(supply_type_key))


def _is_vat_registered() -> bool:
    profile = _load_business_profile()
    return bool(profile.get("vat_registered", True))


def _calculate_threshold_status(turnover: float, annual_threshold: float) -> tuple[str, str, float, float, float]:
    warning_threshold = round(annual_threshold * VAT_THRESHOLD_WARNING_RATIO, 2)
    remaining_before_limit = round(max(annual_threshold - turnover, 0.0), 2)
    progress_pct = round((turnover / annual_threshold) * 100, 1) if annual_threshold > 0 else 0.0

    if annual_threshold <= 0:
        return "normal", "No VAT turnover threshold configured.", warning_threshold, remaining_before_limit, progress_pct
    if turnover >= annual_threshold:
        return "exceeded", "Annual turnover has exceeded the selected VAT registration threshold.", warning_threshold, remaining_before_limit, progress_pct
    if turnover >= warning_threshold:
        return "warning", "Annual turnover is above 80% of the selected VAT registration threshold.", warning_threshold, remaining_before_limit, progress_pct
    return "normal", "Annual turnover is below the VAT registration warning threshold.", warning_threshold, remaining_before_limit, progress_pct


def _collect_turnover_streams(income_rows: list[dict[str, Any]], invoice_rows: list[dict[str, Any]]) -> dict[str, float]:
    income_stream_totals = {"services": 0.0, "goods": 0.0}
    invoice_stream_totals = {"services": 0.0, "goods": 0.0}

    for row in income_rows:
        supply_type = _normalize_supply_type(row.get("Supply Type"))
        amount = _coerce_number(row.get("Total incl. VAT (€)", row.get("Amount (€)", 0)))
        income_stream_totals[supply_type] += amount

    for row in invoice_rows:
        supply_type = _normalize_supply_type(row.get("Supply Type"))
        amount = _coerce_number(row.get("Total (€)", row.get("Amount (€)", 0)))
        invoice_stream_totals[supply_type] += amount

    return {
        "services": round(max(income_stream_totals["services"], invoice_stream_totals["services"]), 2),
        "goods": round(max(income_stream_totals["goods"], invoice_stream_totals["goods"]), 2),
    }


def _compute_vat_threshold_summary(income_rows: list[dict[str, Any]], invoice_rows: list[dict[str, Any]], basis: str) -> dict[str, Any]:
    normalized_basis = _normalize_vat_threshold_basis(basis)
    threshold_config = VAT_TURNOVER_THRESHOLDS.get(normalized_basis, VAT_TURNOVER_THRESHOLDS["services"])
    annual_threshold = float(threshold_config.get("annual_limit") or 0.0)

    cash_turnover = round(sum(_coerce_number(row.get("Total incl. VAT (€)", row.get("Amount (€)", 0))) for row in income_rows), 2)
    invoiced_turnover = round(sum(_coerce_number(row.get("Total (€)", row.get("Amount (€)", 0))) for row in invoice_rows), 2)

    stream_turnovers = _collect_turnover_streams(income_rows, invoice_rows)
    taxable_turnover = round(stream_turnovers.get(normalized_basis, 0.0), 2)
    status, message, warning_threshold, remaining_before_limit, progress_pct = _calculate_threshold_status(taxable_turnover, annual_threshold)

    stream_trackers: list[dict[str, Any]] = []
    for stream_basis, stream_config in VAT_TURNOVER_THRESHOLDS.items():
        stream_annual_threshold = float(stream_config.get("annual_limit") or 0.0)
        stream_turnover = round(stream_turnovers.get(stream_basis, 0.0), 2)
        stream_status, _, stream_warning_threshold, stream_remaining_before_limit, stream_progress_pct = _calculate_threshold_status(stream_turnover, stream_annual_threshold)
        stream_trackers.append(
            {
                "basis": stream_basis,
                "basis_label": str(stream_config.get("label") or stream_basis.title()),
                "taxable_turnover": stream_turnover,
                "annual_threshold": round(stream_annual_threshold, 2),
                "warning_threshold": stream_warning_threshold,
                "remaining_before_limit": stream_remaining_before_limit,
                "progress_pct": stream_progress_pct,
                "status": stream_status,
                "is_selected": stream_basis == normalized_basis,
            }
        )

    return {
        "basis": normalized_basis,
        "basis_label": str(threshold_config.get("label") or normalized_basis.title()),
        "annual_threshold": round(annual_threshold, 2),
        "warning_threshold": warning_threshold,
        "warning_ratio_pct": int(VAT_THRESHOLD_WARNING_RATIO * 100),
        "taxable_turnover": taxable_turnover,
        "cash_turnover": cash_turnover,
        "invoiced_turnover": invoiced_turnover,
        "remaining_before_limit": remaining_before_limit,
        "progress_pct": progress_pct,
        "status": status,
        "message": message,
        "stream_trackers": stream_trackers,
    }


def _normalize_vat_fields(payload: dict[str, Any], *, net_key: str, total_key: str, vat_rate_key: str, vat_amount_key: str, vat_registered: bool) -> None:
    raw_net = _coerce_number(payload.get(net_key))
    raw_total = _coerce_number(payload.get(total_key))
    rate_value = payload.get(vat_rate_key) or "0%"
    rate_ratio = _parse_vat_rate(rate_value)
    explicit_vat = _coerce_number(payload.get(vat_amount_key))

    if raw_total <= 0 and raw_net > 0:
        raw_total = raw_net
    if raw_net <= 0 and raw_total > 0:
        raw_net = raw_total

    if not vat_registered:
        vat_amount = 0.0
        if raw_total <= 0:
            raw_total = raw_net
        raw_net = raw_total
        rate_value = "0%"
    else:
        vat_amount = 0.0
        if explicit_vat > 0:
            vat_amount = explicit_vat
        elif raw_total > 0 and raw_total >= raw_net:
            vat_amount = raw_total - raw_net
        elif raw_net > 0 and rate_ratio > 0:
            vat_amount = raw_net * rate_ratio

        if raw_total <= 0 and raw_net > 0:
            raw_total = raw_net + vat_amount
        if raw_net <= 0 and raw_total > 0:
            raw_net = max(raw_total - vat_amount, 0.0)

    payload[net_key] = f"{round(raw_net, 2):.2f}"
    payload[total_key] = f"{round(raw_total, 2):.2f}"
    payload[vat_rate_key] = str(rate_value)
    payload[vat_amount_key] = f"{round(vat_amount, 2):.2f}"


def _resolve_discount_amount(payload: dict[str, Any], base_net: float) -> float:
    discount_type = str(payload.get("Discount Type") or "€").strip()
    discount_value = _coerce_number(payload.get("Discount Value"))
    if discount_type == "%":
        return max(base_net * (discount_value / 100.0), 0.0)
    return max(discount_value, 0.0)


def _apply_expense_amount_breakdown(payload: dict[str, Any]) -> None:
    base_net = _coerce_number(payload.get("Base Net Amount (€)"))
    delivery = _coerce_number(payload.get("Delivery (€)"))
    fees = _coerce_number(payload.get("Fees (€)"))
    other_charges = _coerce_number(payload.get("Other Charges (€)"))
    discount = _resolve_discount_amount(payload, base_net)

    subtotal_before_discount = base_net + delivery + fees + other_charges
    taxable_net = max(subtotal_before_discount - discount, 0.0)

    payload["Base Net Amount (€)"] = f"{round(base_net, 2):.2f}"
    payload["Delivery (€)"] = f"{round(delivery, 2):.2f}"
    payload["Fees (€)"] = f"{round(fees, 2):.2f}"
    payload["Other Charges (€)"] = f"{round(other_charges, 2):.2f}"
    payload["Discount Type"] = "%" if str(payload.get("Discount Type") or "€").strip() == "%" else "€"
    payload["Discount Value"] = f"{round(_coerce_number(payload.get('Discount Value')), 2):.2f}"
    payload["Discount (€)"] = f"{round(discount, 2):.2f}"
    payload["Net Amount (€)"] = f"{round(taxable_net, 2):.2f}"


def _normalize_invoice_balance(payload: dict[str, Any]) -> None:
    status = _normalize_invoice_status(payload.get("Status"))
    total = _coerce_number(payload.get("Total (€)"))
    if status == "Paid":
        payload["Balance Due (€)"] = f"{0:.2f}"
    elif status in ("Draft", "Issued", "Overdue") and not str(payload.get("Balance Due (€)") or "").strip():
        payload["Balance Due (€)"] = f"{round(total, 2):.2f}"


def _apply_invoice_amount_breakdown(payload: dict[str, Any]) -> None:
    base_net = _coerce_number(payload.get("Base Net Amount (€)"))
    delivery = _coerce_number(payload.get("Delivery (€)"))
    fees = _coerce_number(payload.get("Fees (€)"))
    other_charges = _coerce_number(payload.get("Other Charges (€)"))
    discount = _resolve_discount_amount(payload, base_net)

    subtotal_before_discount = base_net + delivery + fees + other_charges
    taxable_net = max(subtotal_before_discount - discount, 0.0)

    payload["Base Net Amount (€)"] = f"{round(base_net, 2):.2f}"
    payload["Delivery (€)"] = f"{round(delivery, 2):.2f}"
    payload["Fees (€)"] = f"{round(fees, 2):.2f}"
    payload["Other Charges (€)"] = f"{round(other_charges, 2):.2f}"
    payload["Discount Type"] = "%" if str(payload.get("Discount Type") or "€").strip() == "%" else "€"
    payload["Discount Value"] = f"{round(_coerce_number(payload.get('Discount Value')), 2):.2f}"
    payload["Discount (€)"] = f"{round(discount, 2):.2f}"
    payload["Net (€)"] = f"{round(taxable_net, 2):.2f}"


def _compute_invoice_line_item(item: dict[str, Any]) -> dict[str, Any]:
    quantity = _coerce_number(item.get("quantity")) or 1.0
    unit_price = round(_coerce_number(item.get("unit_price")), 2)
    base_amount = round(quantity * unit_price, 2)

    discount_type = "%" if str(item.get("discount_type") or "€").strip() == "%" else "€"
    discount_value = round(_coerce_number(item.get("discount_value")), 2)
    if discount_type == "%":
        discount_amount = round(base_amount * (discount_value / 100.0), 2)
    else:
        discount_amount = round(discount_value, 2)
    discount_amount = max(min(discount_amount, base_amount), 0.0)

    net_amount = round(max(base_amount - discount_amount, 0.0), 2)
    vat_rate_value = str(item.get("vat_rate") or "0%")
    vat_amount = round(net_amount * _parse_vat_rate(vat_rate_value), 2) if _is_vat_registered() else 0.0
    total = round(net_amount + vat_amount, 2)

    return {
        "service_id": str(item.get("service_id") or "").strip(),
        "name": str(item.get("name") or "").strip(),
        "description": str(item.get("description") or "").strip(),
        "quantity": quantity,
        "unit_price": unit_price,
        "discount_type": discount_type,
        "discount_value": discount_value,
        "discount_amount": discount_amount,
        "net_amount": net_amount,
        "vat_rate": vat_rate_value,
        "vat_amount": vat_amount,
        "total": total,
    }


def _apply_invoice_line_items(payload: dict[str, Any], raw_line_items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Compute and store authoritative line items on an invoice payload, then roll them up into the
    invoice's flat aggregate fields (Base Net/Discount/Net/VAT Amount/Total/Service-Product) so the
    rest of the invoice pipeline — validation, balance normalization, income sync — needs no changes."""
    computed = [_compute_invoice_line_item(item) for item in raw_line_items if str(item.get("name") or "").strip()]
    payload["line_items"] = computed

    base_total = round(sum(item["quantity"] * item["unit_price"] for item in computed), 2)
    discount_total = round(sum(item["discount_amount"] for item in computed), 2)
    net_total = round(sum(item["net_amount"] for item in computed), 2)
    vat_total = round(sum(item["vat_amount"] for item in computed), 2)
    grand_total = round(sum(item["total"] for item in computed), 2)

    payload["Base Net Amount (€)"] = f"{base_total:.2f}"
    payload["Discount Type"] = "€"
    payload["Discount Value"] = f"{discount_total:.2f}"
    payload["Discount (€)"] = f"{discount_total:.2f}"
    payload["Net (€)"] = f"{net_total:.2f}"
    payload["VAT Amount (€)"] = f"{vat_total:.2f}"
    payload["Total (€)"] = f"{grand_total:.2f}"
    payload["VAT Rate"] = computed[0]["vat_rate"] if len(computed) == 1 else (computed[0]["vat_rate"] if computed and all(item["vat_rate"] == computed[0]["vat_rate"] for item in computed) else "Mixed")
    if computed:
        payload["Service / Product"] = ", ".join(item["name"] for item in computed if item["name"])
    return computed


def _migrate_invoice_line_items() -> None:
    """Give every legacy invoice (pre-line-items) a single line item derived from its flat fields,
    so line_items is always the source of truth going forward. The flat Service/Product field is
    kept untouched for backwards compatibility."""
    invoices = _load_sheet_records_raw("Invoices")
    changed = False
    for invoice in invoices:
        if invoice.get("line_items"):
            continue
        net_amount = round(_coerce_number(invoice.get("Net (€)")), 2)
        vat_amount = round(_coerce_number(invoice.get("VAT Amount (€)")), 2)
        total = round(_coerce_number(invoice.get("Total (€)")), 2)
        name = str(invoice.get("Service / Product") or "Service").strip() or "Service"
        invoice["line_items"] = [
            {
                "service_id": "",
                "name": name,
                "description": name,
                "quantity": 1.0,
                "unit_price": net_amount,
                "discount_type": "€",
                "discount_value": 0.0,
                "discount_amount": 0.0,
                "net_amount": net_amount,
                "vat_rate": str(invoice.get("VAT Rate") or "0%"),
                "vat_amount": vat_amount,
                "total": total,
            }
        ]
        changed = True
    if changed:
        _save_sheet_records_raw("Invoices", invoices)


def _format_currency(value: float) -> str:
    try:
        numeric_value = float(value)
    except (TypeError, ValueError):
        numeric_value = _coerce_number(value)
    return f"€{numeric_value:,.2f}"


def _normalize_header_name(name: Any, sheet_name: str) -> str:
    if name is None:
        return ""
    text = str(name).strip()
    if not text:
        return ""
    aliases = HEADER_ALIASES.get(sheet_name, {})
    return aliases.get(text, text)


def _get_header_row(ws, sheet_name: str) -> list[Any]:
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        if _is_header_row([value for value in row if value is not None], sheet_name):
            return [value for value in row if value is not None]
    return list(SHEET_HEADERS.get(sheet_name, []))


def _is_header_row(values: list[Any], sheet_name: str) -> bool:
    expected = SHEET_HEADERS.get(sheet_name, [])
    if not expected:
        return False

    normalized = [_normalize_header_name(value, sheet_name) for value in values]
    matches = sum(1 for value in normalized if value in expected)

    if matches >= 3:
        return True

    if sheet_name == "Income" and any("Income Register" in str(value) for value in values):
        return False

    return False


def _find_header_row_number(ws, sheet_name: str) -> int | None:
    for row in ws.iter_rows(min_row=1, max_row=12):
        values = [cell.value for cell in row if cell.value is not None]
        if _is_header_row(values, sheet_name):
            return row[0].row
    return None


def _save_workbook_atomic(wb, resolved_path: Path) -> None:
    temp_path = resolved_path.with_name(f"{resolved_path.stem}.tmp-{uuid4().hex}{resolved_path.suffix}")
    try:
        wb.save(temp_path)

        last_error: PermissionError | None = None
        for _ in range(10):
            try:
                temp_path.replace(resolved_path)
                return
            except PermissionError as exc:
                last_error = exc
                time.sleep(0.2)

        raise WorkbookWriteError("Workbook is busy. Close Excel for this file and retry.") from last_error
    finally:
        if temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


def _load_sheet_records_raw(sheet_name: str) -> list[dict[str, Any]]:
    path = SHEET_JSON_PATHS[sheet_name]
    return _load_json_records(path)


def _save_sheet_records_raw(sheet_name: str, records: list[dict[str, Any]]) -> None:
    path = SHEET_JSON_PATHS[sheet_name]
    _save_json_records(path, records)


def _load_sheet_rows_with_row_numbers(sheet_name: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index, record in enumerate(_load_sheet_records_raw(sheet_name), start=1):
        row = dict(record) if isinstance(record, dict) else {}
        row["__row_number"] = index
        rows.append(row)
    return rows


def _append_row_to_sheet(sheet_name: str, values: dict[str, Any]) -> int:
    with _sheet_write_lock:
        records = _load_sheet_records_raw(sheet_name)
        record = {k: v for k, v in values.items() if k != "__row_number"}
        records.append(record)
        _save_sheet_records_raw(sheet_name, records)
        return len(records)


def _delete_row_from_sheet(sheet_name: str, row_number: int) -> None:
    with _sheet_write_lock:
        records = _load_sheet_records_raw(sheet_name)
        if row_number < 1 or row_number > len(records):
            raise ValueError(f"Invalid row number for {sheet_name}: {row_number}")
        del records[row_number - 1]
        _save_sheet_records_raw(sheet_name, records)


def _update_row_in_sheet(sheet_name: str, row_number: int, values: dict[str, Any]) -> None:
    with _sheet_write_lock:
        records = _load_sheet_records_raw(sheet_name)
        if row_number < 1 or row_number > len(records):
            raise ValueError(f"Invalid row number for {sheet_name}: {row_number}")
        records[row_number - 1] = {k: v for k, v in values.items() if k != "__row_number"}
        _save_sheet_records_raw(sheet_name, records)


def _load_json_records(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    try:
        records = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    return records if isinstance(records, list) else []


def _load_backup_status() -> dict[str, Any]:
    if not BACKUP_STATUS_PATH.exists():
        return {}
    try:
        status = json.loads(BACKUP_STATUS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return status if isinstance(status, dict) else {}


def _save_backup_status(status: dict[str, Any]) -> None:
    try:
        BACKUP_STATUS_PATH.write_text(json.dumps(status, indent=2), encoding="utf-8")
    except OSError:
        pass


def _backup_json_file(path: Path) -> None:
    """Immediately mirror a just-written JSON data file into today's local and Google Drive
    backup folders. Never raises — a backup failure must not break the save it's backing up."""
    if not path.exists():
        return
    today_str = date.today().isoformat()
    status = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "local_ok": False,
        "gdrive_ok": False,
        "error": "",
        "last_file": path.name,
    }
    try:
        local_dir = BACKUPS_DIR / today_str
        local_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, local_dir / path.name)
        status["local_ok"] = True
    except OSError as exc:
        status["error"] = f"Local backup failed: {exc}"

    try:
        gdrive_dir = GDRIVE_BACKUP_DIR / today_str
        gdrive_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, gdrive_dir / path.name)
        status["gdrive_ok"] = True
    except OSError as exc:
        status["error"] = (status["error"] + "; " if status["error"] else "") + f"Google Drive backup failed: {exc}"

    _save_backup_status(status)


def _prune_old_backups() -> None:
    """Keep only the last BACKUP_RETENTION_DAYS days of dated backup folders. Best-effort —
    a pruning failure (e.g. Google Drive not mounted) is silently skipped."""
    cutoff = date.today() - timedelta(days=BACKUP_RETENTION_DAYS)
    for root in (BACKUPS_DIR, GDRIVE_BACKUP_DIR):
        try:
            if not root.exists():
                continue
            for entry in root.iterdir():
                if not entry.is_dir():
                    continue
                try:
                    entry_date = date.fromisoformat(entry.name)
                except ValueError:
                    continue
                if entry_date < cutoff:
                    shutil.rmtree(entry, ignore_errors=True)
        except OSError:
            continue


def _save_json_records(path: Path, records: list[dict[str, Any]]) -> None:
    path.write_text(json.dumps(records, indent=2), encoding="utf-8")
    _backup_json_file(path)


def _append_json_record(path: Path, record: dict[str, Any]) -> None:
    records = _load_json_records(path)
    records.append(record)
    _save_json_records(path, records)


def _pop_json_record(path: Path, record_id: str) -> dict[str, Any] | None:
    records = _load_json_records(path)
    remaining: list[dict[str, Any]] = []
    matched: dict[str, Any] | None = None
    for record in records:
        if matched is None and str(record.get("id")) == record_id:
            matched = record
            continue
        remaining.append(record)
    if matched is not None:
        _save_json_records(path, remaining)
    return matched


def _find_json_record(path: Path, record_id: str) -> dict[str, Any] | None:
    for record in _load_json_records(path):
        if str(record.get("id")) == record_id:
            return record
    return None


def _find_record_by_id(records: list[dict[str, Any]], record_id: Any) -> dict[str, Any] | None:
    if not record_id:
        return None
    for record in records:
        if str(record.get("id")) == str(record_id):
            return record
    return None


def _load_chart_of_accounts() -> list[dict[str, Any]]:
    accounts = _load_json_records(CHART_OF_ACCOUNTS_PATH)
    if accounts:
        return accounts
    return [dict(account) for account in DEFAULT_CHART_OF_ACCOUNTS]


def _save_chart_of_accounts(accounts: list[dict[str, Any]]) -> None:
    _save_json_records(CHART_OF_ACCOUNTS_PATH, accounts)


def _load_tax_rules() -> dict[str, Any]:
    if not TAX_RULES_PATH.exists():
        return {"expense_categories": {}, "income_categories": {}, "field_rules": {}}
    try:
        rules = json.loads(TAX_RULES_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {"expense_categories": {}, "income_categories": {}, "field_rules": {}}
    return rules if isinstance(rules, dict) else {"expense_categories": {}, "income_categories": {}, "field_rules": {}}


def _migrate_chart_of_accounts_categories() -> None:
    """Backfill any DEFAULT_CHART_OF_ACCOUNTS entries missing from the persisted file
    (e.g. newly added compliance-flagged categories) without disturbing existing ones, and
    retire "Non-Deductible Items" (5900) now that Entertainment / Personal Expenses / Fines
    and Penalties cover that ground individually."""
    accounts = _load_json_records(CHART_OF_ACCOUNTS_PATH)
    if not accounts:
        return
    existing_codes = {str(account.get("code") or "") for account in accounts}
    changed = False
    for default_account in DEFAULT_CHART_OF_ACCOUNTS:
        if default_account["code"] not in existing_codes:
            accounts.append(dict(default_account))
            changed = True
    for account in accounts:
        if str(account.get("code") or "") == "5900" and account.get("active", True):
            account["active"] = False
            changed = True
    if changed:
        _save_json_records(CHART_OF_ACCOUNTS_PATH, accounts)


def _migrate_flag_retired_non_deductible_category() -> None:
    """Amber-flag any existing expense still posted to the retired 'Non-Deductible Items'
    category so it surfaces for re-categorisation, without touching anything else about it."""
    records = _load_sheet_records_raw("Expenses")
    changed = False
    for record in records:
        if _normalize_category_key(record.get("Category")) != "non-deductible items":
            continue
        try:
            existing_flags = json.loads(record.get("Compliance Flags") or "[]")
            if not isinstance(existing_flags, list):
                existing_flags = []
        except (TypeError, ValueError):
            existing_flags = []
        if any(flag.get("key") == "retired_category" for flag in existing_flags if isinstance(flag, dict)):
            continue
        existing_flags.append({
            "key": "retired_category",
            "severity": "warning",
            "message": "\"Non-Deductible Items\" has been retired — please recategorise this expense as Entertainment, Personal Expenses, Fines and Penalties, or another appropriate category.",
        })
        record["Compliance Flags"] = json.dumps(existing_flags)
        changed = True
    if changed:
        _save_sheet_records_raw("Expenses", records)


def _ensure_chart_of_accounts() -> list[dict[str, Any]]:
    accounts = _load_chart_of_accounts()
    if not CHART_OF_ACCOUNTS_PATH.exists():
        _save_chart_of_accounts(accounts)
    return accounts


def _normalize_category_key(value: Any) -> str:
    return str(value or "").strip().lower()


def _find_account_by_code(accounts: list[dict[str, Any]], code: str) -> dict[str, Any] | None:
    for account in accounts:
        if str(account.get("code") or "") == code:
            return account
    return None


def _resolve_account_for_entity(entity_type: str, record: dict[str, Any], accounts: list[dict[str, Any]]) -> dict[str, Any]:
    if entity_type == "income":
        category_key = _normalize_category_key(record.get("Category"))
        account_code = INCOME_CATEGORY_ACCOUNT_MAP.get(category_key, "4900")
    elif entity_type == "expense":
        category_key = _normalize_category_key(record.get("Category"))
        account_code = EXPENSE_CATEGORY_ACCOUNT_MAP.get(category_key, "5000")
    elif entity_type == "invoice":
        account_code = "4000"
    elif entity_type == "payroll":
        account_code = "5300"
    else:
        account_code = "4900"

    account = _find_account_by_code(accounts, account_code)
    if account is not None:
        return account
    return {"code": account_code, "name": "Unmapped Account", "tax_treatment": "review"}


def _extract_transaction_amount(entity_type: str, record: dict[str, Any]) -> float:
    components = _extract_amount_components(entity_type, record)
    return components["total"]


def _extract_amount_components(entity_type: str, record: dict[str, Any]) -> dict[str, float]:
    if entity_type == "income":
        net_amount = _coerce_number(record.get("Amount (€)", 0))
        total_amount = _coerce_number(record.get("Total incl. VAT (€)") or net_amount)
        if total_amount <= 0:
            total_amount = net_amount
        explicit_vat_amount = _coerce_number(record.get("VAT Amount (€)", 0))
        vat_amount = explicit_vat_amount if explicit_vat_amount > 0 else max(total_amount - net_amount, 0.0)
        return {
            "net": round(net_amount, 2),
            "vat": round(vat_amount, 2),
            "total": round(total_amount, 2),
        }

    if entity_type == "expense":
        net_amount = _coerce_number(record.get("Net Amount (€)", 0))
        total_amount = _coerce_number(record.get("Total (€)") or net_amount)
        if total_amount <= 0:
            total_amount = net_amount
        explicit_vat_amount = _coerce_number(record.get("VAT Amount (€)", 0))
        vat_amount = explicit_vat_amount if explicit_vat_amount > 0 else max(total_amount - net_amount, 0.0)
        return {
            "net": round(net_amount, 2),
            "vat": round(vat_amount, 2),
            "total": round(total_amount, 2),
        }
    if entity_type == "invoice":
        net_amount = _coerce_number(record.get("Net (€)", 0))
        total_amount = _coerce_number(record.get("Total (€)") or net_amount)
        if total_amount <= 0:
            total_amount = net_amount
        explicit_vat_amount = _coerce_number(record.get("VAT Amount (€)", 0))
        vat_amount = explicit_vat_amount if explicit_vat_amount > 0 else max(total_amount - net_amount, 0.0)
        return {
            "net": round(net_amount, 2),
            "vat": round(vat_amount, 2),
            "total": round(total_amount, 2),
        }
    if entity_type == "payroll":
        gross_pay = _coerce_number(record.get("Gross Pay (€)", 0))
        employer_prsi = _coerce_number(record.get("Employer PRSI (€)", 0))
        employer_cost = _coerce_number(record.get("Employer Cost (€)", gross_pay + employer_prsi))
        return {
            "net": round(gross_pay, 2),
            "vat": 0.0,
            "total": round(employer_cost, 2),
        }
    return {"net": 0.0, "vat": 0.0, "total": 0.0}


def _extract_transaction_date(entity_type: str, record: dict[str, Any]) -> str:
    if entity_type == "income":
        return str(record.get("Date") or "")
    if entity_type == "expense":
        return str(record.get("Date (Registered)") or "")
    if entity_type == "invoice":
        return str(record.get("Issue Date") or "")
    if entity_type == "payroll":
        return str(record.get("Pay Date") or "")
    return ""


def _extract_transaction_description(entity_type: str, record: dict[str, Any]) -> str:
    if entity_type == "income":
        return str(record.get("Description") or record.get("Client / Source") or "Income")
    if entity_type == "expense":
        return str(record.get("Description") or record.get("Title") or "Expense")
    if entity_type == "invoice":
        return str(record.get("Service / Product") or record.get("Invoice #") or "Invoice")
    if entity_type == "payroll":
        return str(record.get("Employee Name") or "Payroll")
    return str(record.get("title") or record.get("Description") or entity_type.title())


def _build_journal_lines(action: str, entity_type: str, account_code: str, total_amount: float, net_amount: float, vat_amount: float, *, record: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    if total_amount <= 0:
        return []

    if entity_type == "income":
        if action in {"archive", "cancel"}:
            if vat_amount > 0:
                return [
                    {"account_code": account_code, "debit": net_amount, "credit": 0.0},
                    {"account_code": "2100", "debit": vat_amount, "credit": 0.0},
                    {"account_code": "1000", "debit": 0.0, "credit": total_amount},
                ]
            return [
                {"account_code": account_code, "debit": total_amount, "credit": 0.0},
                {"account_code": "1000", "debit": 0.0, "credit": total_amount},
            ]
        if vat_amount > 0:
            return [
                {"account_code": "1000", "debit": total_amount, "credit": 0.0},
                {"account_code": account_code, "debit": 0.0, "credit": net_amount},
                {"account_code": "2100", "debit": 0.0, "credit": vat_amount},
            ]
        return [
            {"account_code": "1000", "debit": total_amount, "credit": 0.0},
            {"account_code": account_code, "debit": 0.0, "credit": total_amount},
        ]

    if entity_type == "expense":
        if action in {"archive", "cancel"}:
            if vat_amount > 0:
                return [
                    {"account_code": "1000", "debit": total_amount, "credit": 0.0},
                    {"account_code": account_code, "debit": 0.0, "credit": net_amount},
                    {"account_code": "1200", "debit": 0.0, "credit": vat_amount},
                ]
            return [
                {"account_code": "1000", "debit": total_amount, "credit": 0.0},
                {"account_code": account_code, "debit": 0.0, "credit": total_amount},
            ]
        if vat_amount > 0:
            return [
                {"account_code": account_code, "debit": net_amount, "credit": 0.0},
                {"account_code": "1200", "debit": vat_amount, "credit": 0.0},
                {"account_code": "1000", "debit": 0.0, "credit": total_amount},
            ]
        return [
            {"account_code": account_code, "debit": total_amount, "credit": 0.0},
            {"account_code": "1000", "debit": 0.0, "credit": total_amount},
        ]

    if entity_type == "invoice":
        if action in {"archive", "cancel"}:
            if vat_amount > 0:
                return [
                    {"account_code": account_code, "debit": net_amount, "credit": 0.0},
                    {"account_code": "2100", "debit": vat_amount, "credit": 0.0},
                    {"account_code": "1100", "debit": 0.0, "credit": total_amount},
                ]
            return [
                {"account_code": account_code, "debit": total_amount, "credit": 0.0},
                {"account_code": "1100", "debit": 0.0, "credit": total_amount},
            ]
        if vat_amount > 0:
            return [
                {"account_code": "1100", "debit": total_amount, "credit": 0.0},
                {"account_code": account_code, "debit": 0.0, "credit": net_amount},
                {"account_code": "2100", "debit": 0.0, "credit": vat_amount},
            ]
        return [
            {"account_code": "1100", "debit": total_amount, "credit": 0.0},
            {"account_code": account_code, "debit": 0.0, "credit": total_amount},
        ]

    if entity_type == "payroll":
        payroll_record = record or {}
        gross_pay = round(_coerce_number(payroll_record.get("Gross Pay (€)", net_amount)), 2)
        paye_amount = round(_coerce_number(payroll_record.get("PAYE (€)", 0)), 2)
        usc_amount = round(_coerce_number(payroll_record.get("USC (€)", 0)), 2)
        employee_prsi_amount = round(_coerce_number(payroll_record.get("Employee PRSI (€)", 0)), 2)
        employer_prsi_amount = round(_coerce_number(payroll_record.get("Employer PRSI (€)", 0)), 2)
        net_pay = round(_coerce_number(payroll_record.get("Net Pay (€)", gross_pay - paye_amount - usc_amount - employee_prsi_amount)), 2)
        liabilities = round(paye_amount + usc_amount + employee_prsi_amount + employer_prsi_amount, 2)

        if action in {"archive", "cancel"}:
            lines: list[dict[str, Any]] = [{"account_code": "1000", "debit": net_pay, "credit": 0.0}]
            if liabilities > 0:
                lines.append({"account_code": "2200", "debit": liabilities, "credit": 0.0})
            lines.append({"account_code": account_code, "debit": 0.0, "credit": gross_pay})
            if employer_prsi_amount > 0:
                lines.append({"account_code": "5310", "debit": 0.0, "credit": employer_prsi_amount})
            return lines

        lines = [{"account_code": account_code, "debit": gross_pay, "credit": 0.0}]
        if employer_prsi_amount > 0:
            lines.append({"account_code": "5310", "debit": employer_prsi_amount, "credit": 0.0})
        if liabilities > 0:
            lines.append({"account_code": "2200", "debit": 0.0, "credit": liabilities})
        lines.append({"account_code": "1000", "debit": 0.0, "credit": net_pay})
        return lines

    return []


def _find_account_name(accounts: list[dict[str, Any]], account_code: str) -> str:
    account = _find_account_by_code(accounts, account_code)
    if account is None:
        return "Unmapped Account"
    return str(account.get("name") or "Unmapped Account")


def _compute_trial_balance(ledger_entries: list[dict[str, Any]], accounts: list[dict[str, Any]]) -> dict[str, Any]:
    balances: dict[str, dict[str, Any]] = {}

    def ensure_row(code: str) -> dict[str, Any]:
        if code not in balances:
            balances[code] = {
                "account_code": code,
                "account_name": _find_account_name(accounts, code),
                "debit": 0.0,
                "credit": 0.0,
            }
        return balances[code]

    for entry in ledger_entries:
        lines = entry.get("journal_lines") if isinstance(entry.get("journal_lines"), list) else []
        if not lines:
            account_code = str(entry.get("account_code") or "")
            amount = _coerce_number(entry.get("amount_eur"))
            if not account_code or amount <= 0:
                continue
            fallback_lines = _build_journal_lines(
                str(entry.get("action") or "create"),
                str(entry.get("entity_type") or ""),
                account_code,
                amount,
                amount,
                0.0,
            )
            for line in fallback_lines:
                row = ensure_row(str(line.get("account_code") or ""))
                row["debit"] += _coerce_number(line.get("debit"))
                row["credit"] += _coerce_number(line.get("credit"))
            continue

        for line in lines:
            code = str(line.get("account_code") or "")
            if not code:
                continue
            row = ensure_row(code)
            row["debit"] += _coerce_number(line.get("debit"))
            row["credit"] += _coerce_number(line.get("credit"))

    rows = []
    total_debit = 0.0
    total_credit = 0.0
    for code in sorted(balances.keys()):
        row = balances[code]
        row["debit"] = round(row["debit"], 2)
        row["credit"] = round(row["credit"], 2)
        row["net"] = round(row["debit"] - row["credit"], 2)
        total_debit += row["debit"]
        total_credit += row["credit"]
        rows.append(row)

    total_debit = round(total_debit, 2)
    total_credit = round(total_credit, 2)
    return {
        "rows": rows,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "difference": round(total_debit - total_credit, 2),
        "is_balanced": abs(total_debit - total_credit) < 0.005,
    }


def _load_ledger_entries() -> list[dict[str, Any]]:
    return _load_json_records(LEDGER_JOURNAL_PATH)


def _record_ledger_entry(action: str, entity_type: str, record: dict[str, Any], *, source: str, row_number: int | None = None) -> None:
    if entity_type not in {"income", "expense", "invoice", "payroll"}:
        return

    accounts = _ensure_chart_of_accounts()
    account = _resolve_account_for_entity(entity_type, record, accounts)
    account_code = str(account.get("code") or "")
    amount_components = _extract_amount_components(entity_type, record)
    net_amount = round(amount_components["net"], 2)
    vat_amount = round(amount_components["vat"], 2)
    amount = round(amount_components["total"], 2)
    if not _is_vat_registered():
        vat_amount = 0.0
        net_amount = amount
    journal_lines = _build_journal_lines(action, entity_type, account_code, amount, net_amount, vat_amount, record=record)
    entry = {
        "id": str(uuid4()),
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "action": action,
        "entity_type": entity_type,
        "source": source,
        "row_number": row_number,
        "transaction_date": _extract_transaction_date(entity_type, record),
        "description": _extract_transaction_description(entity_type, record),
        "category": str(record.get("Category") or record.get("Payroll Period") or ""),
        "account_code": account_code,
        "account_name": str(account.get("name") or ""),
        "tax_treatment": str(account.get("tax_treatment") or ""),
        "vat_rate": str(record.get("VAT Rate") or "0%"),
        "vat_treatment": _normalize_vat_treatment(record.get("VAT Treatment"), record.get("VAT Rate")),
        "supply_type": _normalize_supply_type(record.get("Supply Type")),
        "net_amount_eur": net_amount,
        "vat_amount_eur": vat_amount,
        "total_amount_eur": amount,
        "amount_eur": amount,
        "journal_lines": journal_lines,
        "debit_total": round(sum(_coerce_number(line.get("debit")) for line in journal_lines), 2),
        "credit_total": round(sum(_coerce_number(line.get("credit")) for line in journal_lines), 2),
        "entry_balanced": abs(
            sum(_coerce_number(line.get("debit")) for line in journal_lines)
            - sum(_coerce_number(line.get("credit")) for line in journal_lines)
        )
        < 0.005,
        "phase_tag": str(record.get("Phase Tag") or ""),
    }
    _append_json_record(LEDGER_JOURNAL_PATH, entry)


def _vat_period_bounds(target_date: date) -> tuple[date, date, str]:
    if target_date.month in (1, 2):
        start_month = 1
    elif target_date.month in (3, 4):
        start_month = 3
    elif target_date.month in (5, 6):
        start_month = 5
    elif target_date.month in (7, 8):
        start_month = 7
    elif target_date.month in (9, 10):
        start_month = 9
    else:
        start_month = 11

    end_month = start_month + 1
    start_date = date(target_date.year, start_month, 1)
    end_day = monthrange(target_date.year, end_month)[1]
    end_date = date(target_date.year, end_month, end_day)
    label = f"{start_date.strftime('%b')} - {end_date.strftime('%b %Y')}"
    return start_date, end_date, label


def _compute_vat_control_summary(ledger_entries: list[dict[str, Any]]) -> dict[str, Any]:
    current_day = date.today()
    period_start, period_end, period_label = _vat_period_bounds(current_day)
    if period_end.month == 12:
        due_year = period_end.year + 1
        due_month = 1
    else:
        due_year = period_end.year
        due_month = period_end.month + 1
    due_date = date(due_year, due_month, 23).isoformat()
    summary = {
        "period_label": period_label,
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "t1_output_vat": 0.0,
        "t2_input_vat": 0.0,
        "t3_net_vat": 0.0,
        "t4_refund": 0.0,
        "due_date": due_date,
        "zero_rated_sales": 0.0,
        "exempt_sales": 0.0,
        "reverse_charge_purchases": 0.0,
        "treatment_notes": "",
    }

    if not _is_vat_registered():
        return summary

    for entry in ledger_entries:
        transaction_date = _parse_transaction_date(entry.get("transaction_date"))
        if transaction_date is None or transaction_date < period_start or transaction_date > period_end:
            continue

        lines = entry.get("journal_lines") if isinstance(entry.get("journal_lines"), list) else []
        for line in lines:
            code = str(line.get("account_code") or "")
            debit = _coerce_number(line.get("debit"))
            credit = _coerce_number(line.get("credit"))
            if code == "2100":
                summary["t1_output_vat"] += max(credit - debit, 0.0)
            elif code == "1200":
                summary["t2_input_vat"] += max(debit - credit, 0.0)

        vat_treatment = _normalize_vat_treatment(entry.get("vat_treatment"), entry.get("vat_rate"))
        total_amount = _coerce_number(entry.get("total_amount_eur"))
        entity_type = str(entry.get("entity_type") or "")
        if entity_type in {"income", "invoice"} and vat_treatment == "zero_rated":
            summary["zero_rated_sales"] += total_amount
        elif entity_type in {"income", "invoice"} and vat_treatment == "exempt":
            summary["exempt_sales"] += total_amount
        elif entity_type == "expense" and vat_treatment == "reverse_charge":
            summary["reverse_charge_purchases"] += total_amount

    summary["t1_output_vat"] = round(summary["t1_output_vat"], 2)
    summary["t2_input_vat"] = round(summary["t2_input_vat"], 2)
    summary["t3_net_vat"] = round(max(summary["t1_output_vat"] - summary["t2_input_vat"], 0.0), 2)
    summary["t4_refund"] = round(max(summary["t2_input_vat"] - summary["t1_output_vat"], 0.0), 2)
    summary["zero_rated_sales"] = round(summary["zero_rated_sales"], 2)
    summary["exempt_sales"] = round(summary["exempt_sales"], 2)
    summary["reverse_charge_purchases"] = round(summary["reverse_charge_purchases"], 2)
    summary["treatment_notes"] = (
        f"Zero-rated sales €{summary['zero_rated_sales']:.2f}; "
        f"Exempt sales €{summary['exempt_sales']:.2f}; "
        f"Reverse-charge purchases €{summary['reverse_charge_purchases']:.2f}"
    )
    return summary


def _export_trial_balance_csv(trial_balance: dict[str, Any]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["account_code", "account_name", "debit_eur", "credit_eur", "net_eur"])
    for row in trial_balance.get("rows", []):
        writer.writerow([
            row.get("account_code", ""),
            row.get("account_name", ""),
            row.get("debit", 0),
            row.get("credit", 0),
            row.get("net", 0),
        ])
    writer.writerow([
        "TOTAL",
        "",
        trial_balance.get("total_debit", 0),
        trial_balance.get("total_credit", 0),
        trial_balance.get("difference", 0),
    ])
    return buffer.getvalue()


def _export_vat3_csv(vat_summary: dict[str, Any]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["VAT3_Period", "T1", "T2", "T3", "T4", "Due_Date", "ZeroRatedSales", "ExemptSales", "ReverseChargePurchases", "Treatment_Notes"])
    writer.writerow([
        vat_summary.get("period_label", ""),
        vat_summary.get("t1_output_vat", 0),
        vat_summary.get("t2_input_vat", 0),
        vat_summary.get("t3_net_vat", 0),
        vat_summary.get("t4_refund", 0),
        vat_summary.get("due_date", ""),
        vat_summary.get("zero_rated_sales", 0),
        vat_summary.get("exempt_sales", 0),
        vat_summary.get("reverse_charge_purchases", 0),
        vat_summary.get("treatment_notes", ""),
    ])
    return buffer.getvalue()


def _entry_vat_anomaly_flags(entry: dict[str, Any]) -> list[str]:
    vat_treatment = _normalize_vat_treatment(entry.get("vat_treatment"), entry.get("vat_rate"))
    vat_amount = _coerce_number(entry.get("vat_amount_eur"))
    total_amount = _coerce_number(entry.get("total_amount_eur"))
    vat_rate_ratio = _parse_vat_rate(entry.get("vat_rate"))
    flags: list[str] = []

    if vat_treatment in {"zero_rated", "exempt"} and vat_amount > 0.009:
        flags.append("non_zero_vat_with_zero_or_exempt_treatment")
    if vat_treatment == "reverse_charge" and vat_amount > 0.009:
        flags.append("reverse_charge_should_not_post_local_vat_amount")
    if vat_treatment == "standard" and total_amount > 0 and vat_rate_ratio > 0 and vat_amount <= 0.009:
        flags.append("missing_vat_amount_for_standard_rate")

    return flags


def _detect_vat_anomalies(ledger_entries: list[dict[str, Any]], *, limit: int = 20) -> list[dict[str, Any]]:
    anomalies: list[dict[str, Any]] = []
    for entry in sorted(ledger_entries, key=lambda item: str(item.get("timestamp") or ""), reverse=True):
        flags = _entry_vat_anomaly_flags(entry)
        if not flags:
            continue
        anomalies.append(
            {
                "timestamp": str(entry.get("timestamp") or ""),
                "entity_type": str(entry.get("entity_type") or ""),
                "description": str(entry.get("description") or ""),
                "transaction_date": str(entry.get("transaction_date") or ""),
                "amount_eur": round(_coerce_number(entry.get("amount_eur")), 2),
                "vat_rate": str(entry.get("vat_rate") or "0%"),
                "vat_treatment": _normalize_vat_treatment(entry.get("vat_treatment"), entry.get("vat_rate")),
                "supply_type": _normalize_supply_type(entry.get("supply_type")),
                "flags": flags,
            }
        )
        if len(anomalies) >= limit:
            break
    return anomalies


def _export_ledger_journal_csv(ledger_entries: list[dict[str, Any]]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "timestamp",
            "action",
            "entity_type",
            "transaction_date",
            "description",
            "account_code",
            "account_name",
            "amount_eur",
            "debit_total",
            "credit_total",
            "phase_tag",
            "vat_rate",
            "vat_treatment",
            "supply_type",
            "vat_amount_eur",
            "net_amount_eur",
            "total_amount_eur",
            "anomaly_flags",
        ]
    )

    for entry in sorted(ledger_entries, key=lambda item: str(item.get("timestamp") or ""), reverse=True):
        flags = _entry_vat_anomaly_flags(entry)
        writer.writerow(
            [
                entry.get("timestamp", ""),
                entry.get("action", ""),
                entry.get("entity_type", ""),
                entry.get("transaction_date", ""),
                entry.get("description", ""),
                entry.get("account_code", ""),
                entry.get("account_name", ""),
                round(_coerce_number(entry.get("amount_eur")), 2),
                round(_coerce_number(entry.get("debit_total")), 2),
                round(_coerce_number(entry.get("credit_total")), 2),
                entry.get("phase_tag", ""),
                str(entry.get("vat_rate") or "0%"),
                _normalize_vat_treatment(entry.get("vat_treatment"), entry.get("vat_rate")),
                _normalize_supply_type(entry.get("supply_type")),
                round(_coerce_number(entry.get("vat_amount_eur")), 2),
                round(_coerce_number(entry.get("net_amount_eur")), 2),
                round(_coerce_number(entry.get("total_amount_eur")), 2),
                "|".join(flags),
            ]
        )
    return buffer.getvalue()


def _summarize_chart_of_accounts(accounts: list[dict[str, Any]]) -> dict[str, Any]:
    summary = {
        "total_accounts": len(accounts),
        "income_accounts": 0,
        "expense_accounts": 0,
        "asset_accounts": 0,
        "liability_accounts": 0,
        "equity_accounts": 0,
    }
    for account in accounts:
        account_type = str(account.get("type") or "").strip().lower()
        if account_type == "income":
            summary["income_accounts"] += 1
        elif account_type == "expense":
            summary["expense_accounts"] += 1
        elif account_type == "asset":
            summary["asset_accounts"] += 1
        elif account_type == "liability":
            summary["liability_accounts"] += 1
        elif account_type == "equity":
            summary["equity_accounts"] += 1
    return summary


def _record_audit(action: str, entity_type: str, details: dict[str, Any]) -> None:
    _append_json_record(
        AUDIT_LOG_PATH,
        {
            "id": str(uuid4()),
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "action": action,
            "entity_type": entity_type,
            "details": details,
        },
    )


def _archive_record(entity_type: str, record: dict[str, Any], *, source: str) -> None:
    sanitized_record = {key: value for key, value in record.items() if not key.startswith("__")}
    archive_entry = {
        "id": str(uuid4()),
        "archived_at": datetime.now().isoformat(timespec="seconds"),
        "entity_type": entity_type,
        "source": source,
        "record": sanitized_record,
    }
    _append_json_record(ARCHIVE_PATH, archive_entry)
    _record_audit("archive", entity_type, {"source": source, "record": sanitized_record})


def _load_archives() -> list[dict[str, Any]]:
    return _load_json_records(ARCHIVE_PATH)


def _load_audit_entries() -> list[dict[str, Any]]:
    return _load_json_records(AUDIT_LOG_PATH)


def _parse_validation_query(raw: Any) -> dict[str, str]:
    if not raw:
        return {}
    try:
        parsed = json.loads(str(raw))
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    if not isinstance(parsed, dict):
        return {}
    return {str(key): str(value) for key, value in parsed.items()}


def _parse_form_query(raw: Any) -> dict[str, Any]:
    if not raw:
        return {}
    try:
        parsed = json.loads(str(raw))
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _build_validation_message(errors: dict[str, str]) -> str:
    return f"Validation: {'; '.join(errors.values())}"


def _redirect_with_form_errors(route_name: str, form_data: dict[str, Any], validation_errors: dict[str, str], **extra_params: Any):
    return redirect(
        url_for(
            route_name,
            message=_build_validation_message(validation_errors),
            validation_errors=json.dumps(validation_errors),
            form_data=json.dumps(form_data),
            **extra_params,
        )
    )


def _validate_positive_amount(value: Any, field_name: str, label: str, errors: dict[str, str], *, allow_zero: bool = False) -> None:
    if value in (None, ""):
        return
    numeric_value = _coerce_number(value)
    if numeric_value < 0 or (not allow_zero and numeric_value == 0):
        comparator = "zero or greater" if allow_zero else "greater than zero"
        errors[field_name] = f"{label} must be {comparator}"


def _validate_required_text(value: Any, field_name: str, label: str, errors: dict[str, str]) -> None:
    if not str(value or "").strip():
        errors[field_name] = f"{label} is required"


def _validate_income_payload(payload: dict[str, Any]) -> dict[str, str]:
    errors: dict[str, str] = {}
    _validate_required_text(payload.get("Date"), "date", "Income date", errors)
    _validate_required_text(payload.get("Description"), "description", "Income description", errors)
    _validate_positive_amount(payload.get("Amount (€)"), "amount", "Income amount", errors)
    return errors


def _validate_expense_payload(payload: dict[str, Any]) -> dict[str, str]:
    errors: dict[str, str] = {}
    _validate_required_text(payload.get("Date (Registered)"), "date", "Expense date", errors)
    if not str(payload.get("Title") or "").strip() and not str(payload.get("Description") or "").strip():
        errors["title"] = "Expense title or description is required"
        errors["description"] = "Expense title or description is required"
    _validate_positive_amount(payload.get("Net Amount (€)"), "net_amount", "Expense net amount", errors, allow_zero=True)
    _validate_positive_amount(payload.get("Total (€)"), "total_amount", "Expense total amount", errors)
    _validate_positive_amount(payload.get("Base Net Amount (€)"), "base_net_amount", "Expense base net amount", errors, allow_zero=True)
    _validate_positive_amount(payload.get("Delivery (€)"), "delivery_amount", "Expense delivery amount", errors, allow_zero=True)
    _validate_positive_amount(payload.get("Fees (€)"), "fees_amount", "Expense fees amount", errors, allow_zero=True)
    _validate_positive_amount(payload.get("Other Charges (€)"), "other_charges_amount", "Expense other charges amount", errors, allow_zero=True)
    _validate_positive_amount(payload.get("Discount (€)"), "discount_amount", "Expense discount amount", errors, allow_zero=True)
    subtotal_before_discount = (
        _coerce_number(payload.get("Base Net Amount (€)"))
        + _coerce_number(payload.get("Delivery (€)"))
        + _coerce_number(payload.get("Fees (€)"))
        + _coerce_number(payload.get("Other Charges (€)"))
    )
    if _coerce_number(payload.get("Discount (€)")) > subtotal_before_discount:
        errors["discount_amount"] = "Expense discount cannot exceed base net plus add-on charges"
    input_vat_reclaimable = _normalize_input_vat_reclaimable(payload.get("Input VAT Reclaimable"))
    if _is_vat_registered() and input_vat_reclaimable in {"Yes", "Partial"} and not str(payload.get("Supplier VAT Number") or "").strip():
        errors["supplier_vat_number"] = "Supplier VAT number is required when claiming input VAT"
    if str(payload.get("Deductibility Status") or "").strip() not in EXPENSE_DEDUCTIBILITY_OPTIONS:
        errors["deductibility_status"] = "Deductibility status is invalid"
    if str(payload.get("Receipt Attached") or "").strip() not in YES_NO_OPTIONS:
        errors["receipt_attached"] = "Receipt attached must be Yes or No"
    if str(payload.get("Bank Reconciliation") or "").strip() not in RECONCILIATION_OPTIONS:
        errors["bank_reconciliation"] = "Bank reconciliation must be Reconciled or Unreconciled"
    status = _normalize_expense_status(payload.get("Status"))
    if status not in EXPENSE_STATUS_OPTIONS:
        errors["status"] = "Expense status is invalid"
    if _is_paid_status("expense", status) and not str(payload.get("Payment Method") or "").strip():
        errors["payment_method"] = "Payment method is required when status is Paid"
    if _coerce_number(payload.get("Net Amount (€)")) > _coerce_number(payload.get("Total (€)")) and payload.get("Net Amount (€)") not in (None, ""):
        errors["total_amount"] = "Expense total amount must be at least the net amount"
    return errors


def _validate_invoice_payload(payload: dict[str, Any]) -> dict[str, str]:
    errors: dict[str, str] = {}
    _validate_required_text(payload.get("Invoice #"), "invoice_number", "Invoice number", errors)
    _validate_required_text(payload.get("Client Name"), "client_name", "Invoice client", errors)
    _validate_positive_amount(payload.get("Net (€)"), "net_amount", "Invoice net amount", errors, allow_zero=True)
    _validate_positive_amount(payload.get("Total (€)"), "total_amount", "Invoice total amount", errors)
    _validate_positive_amount(payload.get("Balance Due (€)"), "balance_due", "Invoice balance due", errors, allow_zero=True)
    issue_date = _parse_iso_date(payload.get("Issue Date"))
    due_date = _parse_iso_date(payload.get("Due Date"))
    if issue_date and due_date and due_date < issue_date:
        errors["due_date"] = "Invoice due date cannot be before the issue date"
    total_amount = _coerce_number(payload.get("Total (€)"))
    balance_due = _coerce_number(payload.get("Balance Due (€)"))
    if total_amount and balance_due > total_amount:
        errors["balance_due"] = "Invoice balance due cannot exceed the total amount"
    status = _normalize_invoice_status(payload.get("Status"))
    if status not in INVOICE_STATUS_OPTIONS:
        errors["status"] = "Invoice status is invalid"
    if _is_paid_status("invoice", status) and not str(payload.get("Payment Method") or "").strip():
        errors["payment_method"] = "Payment method is required when status is Paid"
    if _is_paid_status("invoice", status) and _parse_iso_date(payload.get("Payment Date")) is None:
        errors["payment_date"] = "Payment date is required when status is Paid"
    if status != "Draft" and not str(payload.get("Client VAT Number") or "").strip() and _is_vat_registered():
        errors["client_vat_number"] = "Client VAT number is required for VAT invoices"
    if status != "Draft" and not str(payload.get("Client Address") or "").strip() and _is_vat_registered():
        errors["client_address"] = "Client address is required for VAT invoices"
    return errors


def _validate_client_payload(payload: dict[str, Any]) -> dict[str, str]:
    errors: dict[str, str] = {}
    _validate_required_text(payload.get("Client Name"), "client_name", "Client name", errors)
    if str(payload.get("Service Tier") or "None") == "Clarity Partner":
        _validate_positive_amount(payload.get("Retainer Amount (€)"), "retainer_amount", "Retainer amount", errors, allow_zero=True)
    return errors


def _validate_supplier_payload(payload: dict[str, Any]) -> dict[str, str]:
    errors: dict[str, str] = {}
    _validate_required_text(payload.get("Supplier Name"), "supplier_name", "Supplier name", errors)
    return errors


def _validate_subscription_payload(payload: dict[str, Any]) -> dict[str, str]:
    errors: dict[str, str] = {}
    _validate_required_text(payload.get("title"), "title", "Subscription name", errors)
    _validate_positive_amount(payload.get("net_amount"), "net_amount", "Subscription net amount", errors, allow_zero=True)
    _validate_positive_amount(payload.get("total_amount"), "total_amount", "Subscription total amount", errors)
    frequency = str(payload.get("frequency") or "").strip().lower()
    if frequency not in SUBSCRIPTION_FREQUENCIES:
        errors["frequency"] = "Subscription frequency is invalid"
    status = str(payload.get("status") or "").strip().lower()
    if status not in SUBSCRIPTION_STATUSES:
        errors["status"] = "Subscription status is invalid"
    start_date = _parse_iso_date(payload.get("start_date"))
    next_charge_date = _parse_iso_date(payload.get("next_charge_date") or payload.get("start_date"))
    end_date = _parse_iso_date(payload.get("end_date"))
    if start_date is None:
        errors["start_date"] = "Subscription start date is required"
    if next_charge_date is None:
        errors["next_charge_date"] = "Subscription next charge date is required"
    if start_date and next_charge_date and next_charge_date < start_date:
        errors["next_charge_date"] = "Subscription next charge date cannot be before the start date"
    if start_date and end_date and end_date < start_date:
        errors["end_date"] = "Subscription end date cannot be before the start date"
    return errors


def _validate_payroll_payload(payload: dict[str, Any]) -> dict[str, str]:
    errors: dict[str, str] = {}
    _validate_required_text(payload.get("Pay Date"), "pay_date", "Payroll pay date", errors)
    _validate_required_text(payload.get("Employee Name"), "employee_name", "Employee name", errors)
    _validate_positive_amount(payload.get("Gross Pay (€)"), "gross_pay", "Gross pay", errors)
    _validate_positive_amount(payload.get("PAYE (€)"), "paye", "PAYE", errors, allow_zero=True)
    _validate_positive_amount(payload.get("USC (€)"), "usc", "USC", errors, allow_zero=True)
    _validate_positive_amount(payload.get("Employee PRSI (€)"), "employee_prsi", "Employee PRSI", errors, allow_zero=True)
    _validate_positive_amount(payload.get("Employer PRSI (€)"), "employer_prsi", "Employer PRSI", errors, allow_zero=True)

    gross_pay = _coerce_number(payload.get("Gross Pay (€)"))
    deductions = (
        _coerce_number(payload.get("PAYE (€)"))
        + _coerce_number(payload.get("USC (€)"))
        + _coerce_number(payload.get("Employee PRSI (€)"))
    )
    if deductions > gross_pay + 0.0001:
        errors["gross_pay"] = "Gross pay must be at least employee deductions total"

    pay_date = _parse_iso_date(payload.get("Pay Date"))
    payment_date = _parse_iso_date(payload.get("Payment Date"))
    if pay_date and payment_date and payment_date < pay_date:
        errors["payment_date"] = "Payment date cannot be before pay date"

    status = _normalize_payroll_status(payload.get("Status"))
    if status not in PAYROLL_STATUS_OPTIONS:
        errors["status"] = "Payroll status is invalid"
    if _is_paid_status("payroll", status) and not str(payload.get("Payment Method") or "").strip():
        errors["payment_method"] = "Payment method is required when status is Paid"
    if _is_paid_status("payroll", status) and payment_date is None:
        errors["payment_date"] = "Payment date is required when status is Paid"
    if str(payload.get("Bank Reconciliation") or "").strip() not in RECONCILIATION_OPTIONS:
        errors["bank_reconciliation"] = "Bank reconciliation must be Reconciled or Unreconciled"

    return errors


def _build_validation_state(active_tab: str) -> tuple[dict[str, str], dict[str, Any]]:
    expected_tab = request.args.get("validation_tab")
    if expected_tab != active_tab:
        return {}, {}
    return _parse_validation_query(request.args.get("validation_errors")), _parse_form_query(request.args.get("form_data"))


def _build_workbook_form_data(payload: dict[str, Any], field_map: dict[str, str]) -> dict[str, Any]:
    return {form_field: payload.get(sheet_field, "") for form_field, sheet_field in field_map.items()}


def _restore_workbook_archive(entity_type: str, archive_entry: dict[str, Any]) -> None:
    config = WORKBOOK_ENTITY_CONFIG[entity_type]
    row_number = _append_row_to_sheet(config["sheet"], archive_entry.get("record", {}))
    load_finance_data.cache_clear()
    _record_audit("restore", config["audit_type"], {"archive_id": archive_entry.get("id"), "row_number": row_number, "record": archive_entry.get("record", {})})
    _record_ledger_entry("restore", entity_type, archive_entry.get("record", {}), source="archive", row_number=row_number)


def _restore_subscription_archive(archive_entry: dict[str, Any]) -> None:
    subscriptions = _load_subscriptions()
    record = dict(archive_entry.get("record", {}))
    if not record.get("id"):
        record["id"] = str(uuid4())
    record["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
    subscriptions.append(record)
    _save_subscriptions(subscriptions)
    _record_audit("restore", "subscription", {"archive_id": archive_entry.get("id"), "subscription_id": record.get("id"), "record": record})


def _restore_payroll_archive(archive_entry: dict[str, Any]) -> None:
    payroll_entries = _load_payroll_entries()
    record = dict(archive_entry.get("record", {}))
    if not record.get("id"):
        record["id"] = str(uuid4())
    record["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
    _normalize_payroll_payload(record)
    payroll_entries.append(record)
    _save_payroll_entries(payroll_entries)
    _record_audit("restore", "payroll", {"archive_id": archive_entry.get("id"), "payroll_id": record.get("id"), "record": record})
    _record_ledger_entry("restore", "payroll", record, source="archive", row_number=None)


def _collect_select_options(data: dict[str, Any], sheet_name: str, field_name: str, base_options: list[str] | None = None) -> list[str]:
    options: list[str] = list(base_options) if base_options else []
    rows = data.get("sheets", {}).get(sheet_name, [])
    for row in rows:
        value = row.get(field_name)
        if not value:
            continue
        value_str = str(value).strip()
        if value_str and value_str not in options:
            options.append(value_str)
    return options


def _chart_of_accounts_category_options(account_type: str) -> list[str]:
    return [account["name"] for account in DEFAULT_CHART_OF_ACCOUNTS if account.get("type") == account_type and account.get("active", True)]


def _parse_iso_date(value: Any) -> date | None:
    if isinstance(value, date):
        return value
    if not value:
        return None
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except ValueError:
        return None


def _add_months(base_date: date, months: int) -> date:
    total_month = (base_date.month - 1) + months
    year = base_date.year + (total_month // 12)
    month = (total_month % 12) + 1
    day = min(base_date.day, monthrange(year, month)[1])
    return date(year, month, day)


def _load_subscriptions() -> list[dict[str, Any]]:
    if not SUBSCRIPTIONS_PATH.exists():
        return []

    try:
        records = json.loads(SUBSCRIPTIONS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []

    today_iso = date.today().isoformat()
    subscriptions: list[dict[str, Any]] = []
    for record in records if isinstance(records, list) else []:
        if not isinstance(record, dict):
            continue
        frequency = str(record.get("frequency") or "monthly").strip().lower()
        status = str(record.get("status") or "active").strip().lower()
        start_date = _parse_iso_date(record.get("start_date"))
        next_charge_date = _parse_iso_date(record.get("next_charge_date"))
        last_posted_date = _parse_iso_date(record.get("last_posted_date"))
        end_date = _parse_iso_date(record.get("end_date"))
        normalized_start = (start_date or next_charge_date or date.today()).isoformat()
        subscriptions.append(
            {
                "id": str(record.get("id") or uuid4()),
                "title": str(record.get("title") or "").strip(),
                "description": str(record.get("description") or "").strip(),
                "supplier": str(record.get("supplier") or "").strip(),
                "category": str(record.get("category") or "").strip(),
                "net_amount": round(_coerce_number(record.get("net_amount")), 2),
                "total_amount": round(_coerce_number(record.get("total_amount")), 2),
                "frequency": frequency if frequency in SUBSCRIPTION_FREQUENCIES else "monthly",
                "start_date": normalized_start,
                "next_charge_date": (next_charge_date or start_date or date.today()).isoformat(),
                "last_posted_date": last_posted_date.isoformat() if last_posted_date else "",
                "end_date": end_date.isoformat() if end_date else "",
                "status": status if status in SUBSCRIPTION_STATUSES else "active",
                "notes": str(record.get("notes") or "").strip(),
                "created_at": str(record.get("created_at") or datetime.now().isoformat(timespec="seconds")),
                "last_updated_at": str(record.get("last_updated_at") or datetime.now().isoformat(timespec="seconds")),
            }
        )

    return subscriptions


def _save_subscriptions(subscriptions: list[dict[str, Any]]) -> None:
    payload = []
    for subscription in subscriptions:
        payload.append(
            {
                "id": subscription.get("id") or str(uuid4()),
                "title": str(subscription.get("title") or "").strip(),
                "description": str(subscription.get("description") or "").strip(),
                "supplier": str(subscription.get("supplier") or "").strip(),
                "category": str(subscription.get("category") or "").strip(),
                "net_amount": round(_coerce_number(subscription.get("net_amount")), 2),
                "total_amount": round(_coerce_number(subscription.get("total_amount")), 2),
                "frequency": subscription.get("frequency") if subscription.get("frequency") in SUBSCRIPTION_FREQUENCIES else "monthly",
                "start_date": (_parse_iso_date(subscription.get("start_date")) or date.today()).isoformat(),
                "next_charge_date": (_parse_iso_date(subscription.get("next_charge_date")) or date.today()).isoformat(),
                "last_posted_date": (_parse_iso_date(subscription.get("last_posted_date")) or None).isoformat() if _parse_iso_date(subscription.get("last_posted_date")) else "",
                "end_date": (_parse_iso_date(subscription.get("end_date")) or None).isoformat() if _parse_iso_date(subscription.get("end_date")) else "",
                "status": subscription.get("status") if subscription.get("status") in SUBSCRIPTION_STATUSES else "active",
                "notes": str(subscription.get("notes") or "").strip(),
                "created_at": str(subscription.get("created_at") or datetime.now().isoformat(timespec="seconds")),
                "last_updated_at": str(subscription.get("last_updated_at") or datetime.now().isoformat(timespec="seconds")),
            }
        )

    _save_json_records(SUBSCRIPTIONS_PATH, payload)


def _normalize_document_category(value: Any) -> str:
    category = str(value or "").strip()
    return category if category in DOCUMENT_CATEGORIES else "Other"


def _default_company_documents() -> list[dict[str, Any]]:
    now = datetime.now().isoformat(timespec="seconds")
    return [
        {
            "id": str(uuid4()),
            "name": "CRO Certificate of Registration",
            "category": "Compliance",
            "description": "",
            "filename": "",
            "file_path": "",
            "date_added": "2026-08-04",
            "expiry_date": "",
            "status": "active",
            "notes": "H-Queex business name registration, CRO No. 790968",
            "created_at": now,
            "last_updated_at": now,
        }
    ]


def _load_company_documents() -> list[dict[str, Any]]:
    if not COMPANY_DOCUMENTS_PATH.exists():
        documents = _default_company_documents()
        _save_json_records(COMPANY_DOCUMENTS_PATH, documents)
        return documents

    records = _load_json_records(COMPANY_DOCUMENTS_PATH)
    documents: list[dict[str, Any]] = []
    for record in records:
        if not isinstance(record, dict):
            continue
        status = str(record.get("status") or "active").strip().lower()
        documents.append(
            {
                "id": str(record.get("id") or uuid4()),
                "name": str(record.get("name") or "").strip(),
                "category": _normalize_document_category(record.get("category")),
                "description": str(record.get("description") or "").strip(),
                "filename": str(record.get("filename") or "").strip(),
                "file_path": str(record.get("file_path") or "").strip(),
                "date_added": str(record.get("date_added") or "").strip(),
                "expiry_date": str(record.get("expiry_date") or "").strip(),
                "status": status if status in DOCUMENT_STATUSES else "active",
                "notes": str(record.get("notes") or "").strip(),
                "created_at": str(record.get("created_at") or datetime.now().isoformat(timespec="seconds")),
                "last_updated_at": str(record.get("last_updated_at") or datetime.now().isoformat(timespec="seconds")),
            }
        )
    return documents


def _save_company_documents(documents: list[dict[str, Any]]) -> None:
    _save_json_records(COMPANY_DOCUMENTS_PATH, documents)


def _save_uploaded_document(file_storage: Any) -> tuple[str, str]:
    """Returns (stored_filename, error_message). stored_filename is '' on failure or no file."""
    if not file_storage or not getattr(file_storage, "filename", ""):
        return "", ""

    original_name = secure_filename(file_storage.filename)
    if not original_name or "." not in original_name:
        return "", "File must have a valid name and extension"

    extension = original_name.rsplit(".", 1)[-1].lower()
    if extension not in ALLOWED_DOCUMENT_EXTENSIONS:
        return "", "File type not allowed. Accepted: PDF, DOCX, PNG, JPG"

    file_storage.seek(0, 2)
    size_bytes = file_storage.tell()
    file_storage.seek(0)
    if size_bytes > MAX_DOCUMENT_SIZE_BYTES:
        return "", "File exceeds the 10MB maximum size"

    stored_name = f"{uuid4().hex[:10]}_{original_name}"
    file_storage.save(COMPANY_DOCUMENTS_DIR / stored_name)
    return stored_name, ""


def _document_expiry_severity(expiry_date: Any, *, today: date | None = None) -> str:
    parsed = _parse_iso_date(expiry_date)
    if parsed is None:
        return ""
    current_day = today or date.today()
    if parsed < current_day:
        return "expired"
    if parsed <= current_day + timedelta(days=DOCUMENT_EXPIRY_WARNING_DAYS):
        return "soon"
    return "ok"


def _documents_expiring_soon(documents: list[dict[str, Any]], *, today: date | None = None) -> list[dict[str, Any]]:
    current_day = today or date.today()
    expiring = []
    for document in documents:
        if document.get("status") != "active":
            continue
        severity = _document_expiry_severity(document.get("expiry_date"), today=current_day)
        if severity in ("expired", "soon"):
            expiring.append({**document, "expiry_severity": severity})
    return sorted(expiring, key=lambda item: item.get("expiry_date") or "")


def _load_compliance_entries() -> list[dict[str, Any]]:
    records = _load_json_records(COMPLIANCE_CALENDAR_PATH)
    entries: list[dict[str, Any]] = []
    for record in records:
        if not isinstance(record, dict):
            continue
        status = str(record.get("status") or "pending").strip().lower()
        frequency = str(record.get("repeat_frequency") or "").strip().lower()
        entries.append(
            {
                "id": str(record.get("id") or uuid4()),
                "name": str(record.get("name") or "").strip(),
                "due_date": str(record.get("due_date") or "").strip(),
                "description": str(record.get("description") or "").strip(),
                "repeat_frequency": frequency if frequency in COMPLIANCE_REPEAT_FREQUENCIES else "",
                "linked_document_id": str(record.get("linked_document_id") or "").strip(),
                "status": status if status in COMPLIANCE_STATUSES else "pending",
                "created_at": str(record.get("created_at") or datetime.now().isoformat(timespec="seconds")),
                "last_updated_at": str(record.get("last_updated_at") or datetime.now().isoformat(timespec="seconds")),
            }
        )
    return entries


def _save_compliance_entries(entries: list[dict[str, Any]]) -> None:
    _save_json_records(COMPLIANCE_CALENDAR_PATH, entries)


def _compliance_deadline_severity(due_date: date, *, today: date | None = None) -> str:
    current_day = today or date.today()
    if due_date < current_day:
        return "red"
    if due_date <= current_day + timedelta(days=DOCUMENT_EXPIRY_WARNING_DAYS):
        return "amber"
    return "green"


def _vat3_period_due_dates(year: int) -> list[tuple[date, date, date]]:
    """Returns (period_start, period_end, due_date) for each of the 6 VAT3 bi-monthly periods in a year."""
    periods = []
    for start_month in (1, 3, 5, 7, 9, 11):
        period_start = date(year, start_month, 1)
        end_month = start_month + 1
        end_day = monthrange(year, end_month)[1]
        period_end = date(year, end_month, end_day)
        due_month = end_month + 1
        due_year = year
        if due_month > 12:
            due_month -= 12
            due_year += 1
        due_date = date(due_year, due_month, 23)
        periods.append((period_start, period_end, due_date))
    return periods


def _build_compliance_deadlines(
    business_profile: dict[str, Any],
    data: dict[str, Any],
    manual_entries: list[dict[str, Any]],
    *,
    today: date | None = None,
) -> list[dict[str, Any]]:
    current_day = today or date.today()
    structure = _normalize_business_structure(business_profile.get("structure"))
    phase_policy = _build_phase_policy(data.get("summary", {}), structure)
    deadlines: list[dict[str, Any]] = []

    if business_profile.get("vat_registered"):
        for period_start, period_end, due_date in _vat3_period_due_dates(current_day.year) + _vat3_period_due_dates(current_day.year + 1):
            if due_date < current_day - timedelta(days=90):
                continue
            deadlines.append(
                {
                    "id": f"vat3-{period_start.isoformat()}",
                    "name": f"VAT 3 return — {period_start.strftime('%b')} to {period_end.strftime('%b %Y')}",
                    "due_date": due_date.isoformat(),
                    "description": "Bi-monthly VAT 3 return and payment due to Revenue.",
                    "category": "auto",
                    "severity": _compliance_deadline_severity(due_date, today=current_day),
                    "link": "/ledger",
                    "editable": False,
                }
            )

    if structure == "sole_trader":
        form11_due = date(current_day.year, 10, 31)
        if form11_due < current_day:
            form11_due = date(current_day.year + 1, 10, 31)
        deadlines.append(
            {
                "id": f"form11-{form11_due.isoformat()}",
                "name": "Form 11 (Phase 1)",
                "due_date": form11_due.isoformat(),
                "description": "Annual Income Tax return due via ROS.",
                "category": "auto",
                "severity": _compliance_deadline_severity(form11_due, today=current_day),
                "link": "/company/profile",
                "editable": False,
            }
        )
        deadlines.append(
            {
                "id": f"prelim-tax-{form11_due.isoformat()}",
                "name": "Preliminary tax (Phase 1)",
                "due_date": form11_due.isoformat(),
                "description": f"Estimated amount: {_format_currency(phase_policy.get('estimated_tax_due', 0))}, based on current year profit.",
                "category": "auto",
                "severity": _compliance_deadline_severity(form11_due, today=current_day),
                "link": "/company/profile",
                "editable": False,
            }
        )

    if structure == "limited_company":
        transition_date = _parse_transaction_date(business_profile.get("transition_date"))
        in_phase_2 = transition_date is not None and current_day >= transition_date
        if in_phase_2:
            year_end = date(current_day.year, 12, 31)
            ct1_due = date(year_end.year + 1, 9, 30)
            if ct1_due < current_day:
                ct1_due = date(year_end.year + 2, 9, 30)
            deadlines.append(
                {
                    "id": f"ct1-{ct1_due.isoformat()}",
                    "name": "CT1 (Phase 2)",
                    "due_date": ct1_due.isoformat(),
                    "description": "Corporation Tax return due 9 months after accounting year end (calendar year end assumed).",
                    "category": "auto",
                    "severity": _compliance_deadline_severity(ct1_due, today=current_day),
                    "link": "/company/profile",
                    "editable": False,
                }
            )

        registration_date = _parse_transaction_date(business_profile.get("registration_date"))
        if registration_date is not None:
            anniversary = date(current_day.year, registration_date.month, registration_date.day)
            cro_due = anniversary + timedelta(days=56)
            if cro_due < current_day:
                anniversary = date(current_day.year + 1, registration_date.month, registration_date.day)
                cro_due = anniversary + timedelta(days=56)
            deadlines.append(
                {
                    "id": f"cro-annual-return-{cro_due.isoformat()}",
                    "name": "CRO annual return",
                    "due_date": cro_due.isoformat(),
                    "description": "Annual return (B1) and financial statements due to the CRO.",
                    "category": "auto",
                    "severity": _compliance_deadline_severity(cro_due, today=current_day),
                    "link": "/company/profile",
                    "editable": False,
                }
            )

    payroll_entries = _load_payroll_entries()
    if payroll_entries:
        for month_offset in range(0, 3):
            month = current_day.month + month_offset
            year = current_day.year
            while month > 12:
                month -= 12
                year += 1
            p30_due = date(year, month, 23)
            if p30_due < current_day - timedelta(days=30):
                continue
            deadlines.append(
                {
                    "id": f"p30-{p30_due.isoformat()}",
                    "name": "P30",
                    "due_date": p30_due.isoformat(),
                    "description": "Monthly payroll tax return and payment due to Revenue.",
                    "category": "auto",
                    "severity": _compliance_deadline_severity(p30_due, today=current_day),
                    "link": "/payroll",
                    "editable": False,
                }
            )

    for entry in manual_entries:
        due_date = _parse_iso_date(entry.get("due_date"))
        if due_date is None or entry.get("status") == "complete":
            continue
        deadlines.append(
            {
                "id": entry["id"],
                "name": entry.get("name") or "Untitled deadline",
                "due_date": due_date.isoformat(),
                "description": entry.get("description") or "",
                "category": "manual",
                "severity": _compliance_deadline_severity(due_date, today=current_day),
                "link": "/company/compliance",
                "editable": True,
                "linked_document_id": entry.get("linked_document_id") or "",
                "repeat_frequency": entry.get("repeat_frequency") or "",
            }
        )

    return sorted(deadlines, key=lambda item: item["due_date"])


# --- Operations: Projects ---------------------------------------------------

def _default_dmaic_phase() -> dict[str, Any]:
    return {
        "status": "Not Started",
        "start_date": "",
        "completion_date": "",
        "notes": "",
        "deliverables": [],
        "key_findings": "",
    }


def _default_dmaic() -> dict[str, Any]:
    return {phase: _default_dmaic_phase() for phase in DMAIC_PHASES}


def _normalize_dmaic_phase(raw: Any) -> dict[str, Any]:
    if not isinstance(raw, dict):
        raw = {}
    status = str(raw.get("status") or "Not Started").strip()
    deliverables = raw.get("deliverables")
    if not isinstance(deliverables, list):
        deliverables = []
    return {
        "status": status if status in DMAIC_PHASE_STATUSES else "Not Started",
        "start_date": str(raw.get("start_date") or "").strip(),
        "completion_date": str(raw.get("completion_date") or "").strip(),
        "notes": str(raw.get("notes") or "").strip(),
        "deliverables": [str(item).strip() for item in deliverables if str(item).strip()],
        "key_findings": str(raw.get("key_findings") or "").strip(),
    }


def _normalize_dmaic(raw: Any) -> dict[str, Any]:
    if not isinstance(raw, dict):
        raw = {}
    return {phase: _normalize_dmaic_phase(raw.get(phase)) for phase in DMAIC_PHASES}


def _dmaic_completion_percentage(dmaic: dict[str, Any]) -> int:
    complete_count = sum(1 for phase in DMAIC_PHASES if dmaic.get(phase, {}).get("status") == "Complete")
    return round((complete_count / len(DMAIC_PHASES)) * 100)


def _validate_dmaic_transition(dmaic: dict[str, Any], phase: str, new_status: str) -> str:
    """Returns an error message if the transition is not allowed, else ''."""
    if phase not in DMAIC_PHASES:
        return "Unknown DMAIC phase"
    if new_status not in DMAIC_PHASE_STATUSES:
        return "Unknown status"
    if new_status != "Complete":
        return ""
    phase_index = DMAIC_PHASES.index(phase)
    if phase_index == 0:
        return ""
    previous_phase = DMAIC_PHASES[phase_index - 1]
    if dmaic.get(previous_phase, {}).get("status") != "Complete":
        return f"{previous_phase} must be marked Complete before {phase} can be completed"
    return ""


def _generate_project_number(existing_projects: list[dict[str, Any]], *, year: int | None = None) -> str:
    target_year = year or date.today().year
    prefix = f"HQ-PRJ-{target_year}-"
    highest = 0
    for project in existing_projects:
        number = str(project.get("project_number") or "")
        if number.startswith(prefix):
            try:
                sequence = int(number[len(prefix):])
            except ValueError:
                continue
            highest = max(highest, sequence)
    return f"{prefix}{highest + 1:03d}"


def _normalize_project_line_item(item: dict[str, Any]) -> dict[str, Any]:
    quantity = _coerce_number(item.get("quantity")) or 1.0
    unit_price = round(_coerce_number(item.get("unit_price")), 2)
    return {
        "service_id": str(item.get("service_id") or "").strip(),
        "name": str(item.get("name") or "").strip(),
        "quantity": quantity,
        "unit_price": unit_price,
        "total": round(quantity * unit_price, 2),
    }


def _normalize_project(record: dict[str, Any]) -> dict[str, Any]:
    status = str(record.get("status") or "Enquiry").strip()
    tier = str(record.get("service_tier") or "None").strip()
    linked_invoice_ids = record.get("linked_invoice_ids")
    if not isinstance(linked_invoice_ids, list):
        linked_invoice_ids = []
    raw_line_items = record.get("line_items")
    if not isinstance(raw_line_items, list):
        raw_line_items = []
    line_items = [_normalize_project_line_item(item) for item in raw_line_items if isinstance(item, dict)]
    total_value = record.get("total_value")
    now = datetime.now().isoformat(timespec="seconds")
    return {
        "id": str(record.get("id") or uuid4()),
        "project_number": str(record.get("project_number") or "").strip(),
        "client_id": str(record.get("client_id") or "").strip(),
        "client_name": str(record.get("client_name") or "").strip(),
        "service_tier": tier if tier in CLIENT_SERVICE_TIERS else "None",
        "title": str(record.get("title") or "").strip(),
        "description": str(record.get("description") or "").strip(),
        "status": status if status in PROJECT_STATUSES else "Enquiry",
        "start_date": str(record.get("start_date") or "").strip(),
        "target_end_date": str(record.get("target_end_date") or "").strip(),
        "actual_end_date": str(record.get("actual_end_date") or "").strip(),
        "linked_invoice_ids": [str(item) for item in linked_invoice_ids],
        "line_items": line_items,
        "total_value": round(_coerce_number(total_value), 2) if total_value not in (None, "") else round(sum(item["total"] for item in line_items), 2),
        "notes": str(record.get("notes") or "").strip(),
        "dmaic": _normalize_dmaic(record.get("dmaic")),
        "created_at": str(record.get("created_at") or now),
        "updated_at": str(record.get("updated_at") or now),
    }


def _load_projects() -> list[dict[str, Any]]:
    return [_normalize_project(record) for record in _load_json_records(PROJECTS_PATH) if isinstance(record, dict)]


def _save_projects(projects: list[dict[str, Any]]) -> None:
    _save_json_records(PROJECTS_PATH, [_normalize_project(project) for project in projects])


def _projects_with_deadline_within(projects: list[dict[str, Any]], days: int, *, today: date | None = None) -> list[dict[str, Any]]:
    current_day = today or date.today()
    upcoming = []
    for project in projects:
        if project.get("status") in ("Completed", "Cancelled"):
            continue
        target = _parse_iso_date(project.get("target_end_date"))
        if target and current_day <= target <= current_day + timedelta(days=days):
            upcoming.append(project)
    return sorted(upcoming, key=lambda item: item["target_end_date"])


# --- Operations: Delivery Log ------------------------------------------------

def _normalize_delivery_entry(record: dict[str, Any]) -> dict[str, Any]:
    service_type = str(record.get("service_type") or "Other").strip()
    now = datetime.now().isoformat(timespec="seconds")
    hours_spent = record.get("hours_spent")
    return {
        "id": str(record.get("id") or uuid4()),
        "date": str(record.get("date") or date.today().isoformat()).strip(),
        "project_id": str(record.get("project_id") or "").strip(),
        "client_id": str(record.get("client_id") or "").strip(),
        "client_name": str(record.get("client_name") or "").strip(),
        "service_type": service_type if service_type in DELIVERY_SERVICE_TYPES else "Other",
        "description": str(record.get("description") or "").strip(),
        "hours_spent": round(_coerce_number(hours_spent), 2) if hours_spent not in (None, "") else None,
        "deliverable_filename": str(record.get("deliverable_filename") or "").strip(),
        "billing_period": str(record.get("billing_period") or "").strip(),
        "invoiced": bool(record.get("invoiced", False)),
        "invoice_id": str(record.get("invoice_id") or "").strip(),
        "created_at": str(record.get("created_at") or now),
        "last_updated_at": str(record.get("last_updated_at") or now),
    }


def _load_delivery_log() -> list[dict[str, Any]]:
    return [_normalize_delivery_entry(record) for record in _load_json_records(DELIVERY_LOG_PATH) if isinstance(record, dict)]


def _save_delivery_log(entries: list[dict[str, Any]]) -> None:
    _save_json_records(DELIVERY_LOG_PATH, [_normalize_delivery_entry(entry) for entry in entries])


def _save_uploaded_delivery_file(file_storage: Any) -> tuple[str, str]:
    if not file_storage or not getattr(file_storage, "filename", ""):
        return "", ""
    original_name = secure_filename(file_storage.filename)
    if not original_name or "." not in original_name:
        return "", "File must have a valid name and extension"
    extension = original_name.rsplit(".", 1)[-1].lower()
    if extension not in ALLOWED_DELIVERY_EXTENSIONS:
        return "", "File type not allowed"
    file_storage.seek(0, 2)
    size_bytes = file_storage.tell()
    file_storage.seek(0)
    if size_bytes > MAX_DOCUMENT_SIZE_BYTES:
        return "", "File exceeds the 10MB maximum size"
    stored_name = f"{uuid4().hex[:10]}_{original_name}"
    file_storage.save(DELIVERY_FILES_DIR / stored_name)
    return stored_name, ""


def _clarity_partner_pending_billing(delivery_entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for entry in delivery_entries:
        if entry.get("invoiced") or not entry.get("billing_period"):
            continue
        key = (entry.get("client_name") or "", entry.get("billing_period") or "")
        groups.setdefault(key, []).append(entry)
    pending = []
    for (client_name, billing_period), entries in groups.items():
        if not client_name:
            continue
        pending.append({
            "client_name": client_name,
            "billing_period": billing_period,
            "entry_count": len(entries),
            "total_hours": round(sum(_coerce_number(entry.get("hours_spent")) for entry in entries), 2),
        })
    return sorted(pending, key=lambda item: (item["client_name"], item["billing_period"]))


# --- Operations: SOP Library -------------------------------------------------

def _normalize_sop(record: dict[str, Any]) -> dict[str, Any]:
    status = str(record.get("status") or "Draft").strip()
    now = datetime.now().isoformat(timespec="seconds")
    return {
        "id": str(record.get("id") or uuid4()),
        "title": str(record.get("title") or "").strip(),
        "client_id": str(record.get("client_id") or "").strip(),
        "client_name": str(record.get("client_name") or "").strip(),
        "project_id": str(record.get("project_id") or "").strip(),
        "version": str(record.get("version") or "V1.0").strip(),
        "status": status if status in SOP_STATUSES else "Draft",
        "process_area": str(record.get("process_area") or "").strip(),
        "description": str(record.get("description") or "").strip(),
        "filename": str(record.get("filename") or "").strip(),
        "date_created": str(record.get("date_created") or date.today().isoformat()).strip(),
        "date_approved": str(record.get("date_approved") or "").strip(),
        "approved_by": str(record.get("approved_by") or "").strip(),
        "notes": str(record.get("notes") or "").strip(),
        "created_at": str(record.get("created_at") or now),
        "last_updated_at": str(record.get("last_updated_at") or now),
    }


def _load_sops() -> list[dict[str, Any]]:
    return [_normalize_sop(record) for record in _load_json_records(SOPS_PATH) if isinstance(record, dict)]


def _save_sops(sops: list[dict[str, Any]]) -> None:
    _save_json_records(SOPS_PATH, [_normalize_sop(sop) for sop in sops])


def _save_uploaded_sop(file_storage: Any) -> tuple[str, str]:
    if not file_storage or not getattr(file_storage, "filename", ""):
        return "", ""
    original_name = secure_filename(file_storage.filename)
    if not original_name or "." not in original_name:
        return "", "File must have a valid name and extension"
    extension = original_name.rsplit(".", 1)[-1].lower()
    if extension not in ALLOWED_SOP_EXTENSIONS:
        return "", "File type not allowed. Accepted: PDF, DOCX, PNG, JPG"
    file_storage.seek(0, 2)
    size_bytes = file_storage.tell()
    file_storage.seek(0)
    if size_bytes > MAX_DOCUMENT_SIZE_BYTES:
        return "", "File exceeds the 10MB maximum size"
    stored_name = f"{uuid4().hex[:10]}_{original_name}"
    file_storage.save(SOP_FILES_DIR / stored_name)
    return stored_name, ""


def _normalize_service(record: dict[str, Any]) -> dict[str, Any]:
    now = datetime.now().isoformat(timespec="seconds")
    tier = str(record.get("tier") or "addon").strip().lower()
    if tier not in SERVICE_TIERS:
        tier = "addon"
    price_type = str(record.get("price_type") or "fixed").strip().lower()
    if price_type not in SERVICE_PRICE_TYPES:
        price_type = "fixed"
    billing_frequency = str(record.get("billing_frequency") or "").strip().lower() or None
    if billing_frequency not in SERVICE_BILLING_FREQUENCIES:
        billing_frequency = "monthly" if price_type == "retainer" else "one-off"
    status = str(record.get("status") or "active").strip().lower()
    if status not in SERVICE_STATUSES:
        status = "active"
    group = str(record.get("group") or "").strip() or None
    if tier == "core":
        group = None
    elif group not in SERVICE_GROUPS:
        group = group or SERVICE_GROUPS[0]
    quarterly_price = record.get("quarterly_price")
    annual_price = record.get("annual_price")
    return {
        "id": str(record.get("id") or uuid4()),
        "name": str(record.get("name") or "").strip(),
        "tier": tier,
        "group": group,
        "description": str(record.get("description") or "").strip(),
        "price": round(_coerce_number(record.get("price")), 2),
        "price_type": price_type,
        "billing_frequency": billing_frequency,
        "quarterly_price": round(_coerce_number(quarterly_price), 2) if quarterly_price not in (None, "") else None,
        "annual_price": round(_coerce_number(annual_price), 2) if annual_price not in (None, "") else None,
        "status": status,
        "website_display_price": bool(record.get("website_display_price", True)),
        "website_display_label": str(record.get("website_display_label") or "").strip(),
        "date_added": str(record.get("date_added") or now),
        "date_updated": str(record.get("date_updated") or now),
    }


CORE_TIER_DISPLAY_ORDER = {"Clarity Base": 0, "Clarity Plus": 1, "Clarity Partner": 2}


def _load_services() -> list[dict[str, Any]]:
    records = _load_json_records(SERVICES_PATH)
    services = [_normalize_service(record) for record in records if isinstance(record, dict)]
    services.sort(
        key=lambda item: (
            item.get("tier") != "core",
            CORE_TIER_DISPLAY_ORDER.get(item.get("name") or "", 99),
            item.get("group") or "",
            item.get("name") or "",
        )
    )
    return services


def _save_services(services: list[dict[str, Any]]) -> None:
    _save_json_records(SERVICES_PATH, [_normalize_service(service) for service in services])


def _find_service_by_id(services: list[dict[str, Any]], service_id: Any) -> dict[str, Any] | None:
    target_id = str(service_id or "").strip()
    if not target_id:
        return None
    for service in services:
        if str(service.get("id") or "") == target_id:
            return service
    return None


def _validate_service_payload(payload: dict[str, Any]) -> dict[str, str]:
    errors: dict[str, str] = {}
    _validate_required_text(payload.get("name"), "name", "Service name", errors)
    if str(payload.get("tier")) not in SERVICE_TIERS:
        errors["tier"] = "Service tier is invalid"
    if str(payload.get("tier")) == "addon" and not str(payload.get("group") or "").strip():
        errors["group"] = "Add-on services require a group"
    _validate_positive_amount(payload.get("price"), "price", "Service price", errors, allow_zero=True)
    return errors


def _ensure_default_services() -> None:
    if _load_json_records(SERVICES_PATH):
        return

    now = datetime.now().isoformat(timespec="seconds")

    def _service(
        name: str,
        tier: str,
        price: float,
        price_type: str,
        billing_frequency: str,
        description: str,
        website_display_label: str,
        *,
        group: str | None = None,
        quarterly_price: float | None = None,
        annual_price: float | None = None,
    ) -> dict[str, Any]:
        return {
            "id": str(uuid4()),
            "name": name,
            "tier": tier,
            "group": group,
            "description": description,
            "price": price,
            "price_type": price_type,
            "billing_frequency": billing_frequency,
            "quarterly_price": quarterly_price,
            "annual_price": annual_price,
            "status": "active",
            "website_display_price": True,
            "website_display_label": website_display_label,
            "date_added": now,
            "date_updated": now,
        }

    defaults = [
        _service("Clarity Base", "core", 950.0, "fixed", "one-off", "Foundation process audit, documentation and knowledge mapping", "From €950"),
        _service("Clarity Plus", "core", 2500.0, "from", "one-off", "Workflow design, digital systems integration and implementation", "Scoped per engagement — from €2,500"),
        _service("Clarity Partner", "core", 650.0, "retainer", "monthly", "Ongoing advisory retainer with performance monitoring", "From €650/month", quarterly_price=1800.0, annual_price=6500.0),
        _service("SOP Creation (per procedure)", "addon", 150.0, "fixed", "one-off", "Documented standard operating procedure for a single process", "€150 per procedure", group="Documentation and Knowledge Assets"),
        _service("Knowledge Base Setup", "addon", 450.0, "fixed", "one-off", "Structured knowledge base configured for the business", "€450", group="Documentation and Knowledge Assets"),
        _service("Process Manual", "addon", 750.0, "fixed", "one-off", "Full process manual covering end-to-end operations", "€750", group="Documentation and Knowledge Assets"),
        _service("KPI Dashboard Design", "addon", 550.0, "fixed", "one-off", "Custom KPI dashboard design for performance tracking", "€550", group="Data and Reporting Tools"),
        _service("Reporting Template Pack", "addon", 350.0, "fixed", "one-off", "Set of reusable reporting templates", "€350", group="Data and Reporting Tools"),
        _service("Data Audit and Cleanup", "addon", 400.0, "fixed", "one-off", "Audit and cleanup of existing operational data", "€400", group="Data and Reporting Tools"),
        _service("Process Communication Pack", "addon", 300.0, "fixed", "one-off", "Materials to communicate new processes internally", "€300", group="Communication and Strategic Assets"),
        _service("Stakeholder Report Template", "addon", 250.0, "fixed", "one-off", "Template for structured stakeholder reporting", "€250", group="Communication and Strategic Assets"),
        _service("Implementation Roadmap Document", "addon", 350.0, "fixed", "one-off", "Roadmap document outlining implementation phases", "€350", group="Communication and Strategic Assets"),
    ]
    _save_json_records(SERVICES_PATH, defaults)


def _build_subscription_expense_payload(subscription: dict[str, Any], charge_date: date) -> dict[str, Any]:
    title = subscription.get("title") or "Subscription"
    description = subscription.get("description") or title
    return {
        "Date (Registered)": charge_date.isoformat(),
        "Title": title,
        "Description": f"{description} (Subscription charge)",
        "Supplier / Payee": subscription.get("supplier") or title,
        "Category": subscription.get("category") or "Subscription",
        "Net Amount (€)": subscription.get("net_amount") or subscription.get("total_amount") or 0,
        "Total (€)": subscription.get("total_amount") or subscription.get("net_amount") or 0,
        "Input VAT Reclaimable": "No",
        "Deductibility Status": _normalize_deductibility_status("", subscription.get("category") or ""),
        "Capital Expenditure Flag": "No",
        "Receipt Attached": "No",
        "Bank Reconciliation": "Unreconciled",
        "Status": "Auto-posted",
        "Phase Tag": _resolve_phase_tag(charge_date.isoformat()),
    }


def _sync_subscriptions_to_expenses(today: date | None = None) -> dict[str, int]:
    current_day = today or date.today()
    subscriptions = _load_subscriptions()
    posted_count = 0
    changed = False

    for subscription in subscriptions:
        if subscription.get("status") != "active":
            continue

        next_charge = _parse_iso_date(subscription.get("next_charge_date")) or _parse_iso_date(subscription.get("start_date"))
        if next_charge is None:
            continue

        end_date = _parse_iso_date(subscription.get("end_date"))
        frequency_months = SUBSCRIPTION_FREQUENCIES.get(str(subscription.get("frequency")), 1)

        while next_charge <= current_day and (end_date is None or next_charge <= end_date):
            payload = _build_subscription_expense_payload(subscription, next_charge)
            row_number = _append_row_to_sheet("Expenses", payload)
            subscription["last_posted_date"] = next_charge.isoformat()
            subscription["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
            next_charge = _add_months(next_charge, frequency_months)
            subscription["next_charge_date"] = next_charge.isoformat()
            posted_count += 1
            changed = True
            _record_ledger_entry("sync_post", "expense", payload, source="subscription", row_number=row_number)

    if changed:
        _save_subscriptions(subscriptions)
        load_finance_data.cache_clear()

    active_count = sum(1 for subscription in subscriptions if subscription.get("status") == "active")
    return {"posted_count": posted_count, "active_count": active_count}


def _build_subscription_rows(subscriptions: list[dict[str, Any]], today: date | None = None) -> list[dict[str, Any]]:
    current_day = today or date.today()
    rows: list[dict[str, Any]] = []
    for subscription in subscriptions:
        next_charge = _parse_iso_date(subscription.get("next_charge_date"))
        status = str(subscription.get("status") or "active")
        days_until = (next_charge - current_day).days if next_charge else None
        if status != "active":
            due_label = status.title()
        elif days_until is None:
            due_label = "Unscheduled"
        elif days_until < 0:
            due_label = "Overdue"
        elif days_until == 0:
            due_label = "Due today"
        elif days_until <= 7:
            due_label = "Due soon"
        else:
            due_label = "Scheduled"

        rows.append(
            {
                **subscription,
                "next_charge_date": next_charge.isoformat() if next_charge else "",
                "last_posted_date": (_parse_iso_date(subscription.get("last_posted_date")) or None).isoformat() if _parse_iso_date(subscription.get("last_posted_date")) else "",
                "end_date": (_parse_iso_date(subscription.get("end_date")) or None).isoformat() if _parse_iso_date(subscription.get("end_date")) else "",
                "days_until": days_until,
                "due_label": due_label,
                "monthly_equivalent": round((_coerce_number(subscription.get("total_amount")) or _coerce_number(subscription.get("net_amount"))) / SUBSCRIPTION_FREQUENCIES.get(str(subscription.get("frequency")), 1), 2),
            }
        )

    rows.sort(key=lambda row: (row.get("status") != "active", row.get("next_charge_date") or "9999-12-31", row.get("title") or ""))
    return rows


def _summarize_subscriptions(subscription_rows: list[dict[str, Any]], today: date | None = None) -> dict[str, Any]:
    current_day = today or date.today()
    active_rows = [row for row in subscription_rows if row.get("status") == "active"]
    due_rows = [row for row in active_rows if row.get("next_charge_date") and (_parse_iso_date(row.get("next_charge_date")) or current_day) <= current_day]
    upcoming_rows = [row for row in active_rows if row.get("next_charge_date") and 0 <= ((_parse_iso_date(row.get("next_charge_date")) or current_day) - current_day).days <= 30]
    monthly_commitment = sum(_coerce_number(row.get("monthly_equivalent")) for row in active_rows)
    return {
        "active_count": len(active_rows),
        "due_count": len(due_rows),
        "upcoming_count": len(upcoming_rows),
        "monthly_commitment": monthly_commitment,
    }


def _build_chart_data(summary: dict[str, Any]) -> dict[str, float]:
    income_total = _coerce_number(summary.get("income_total", 0))
    expense_total = _coerce_number(summary.get("expense_total", 0))
    net_cashflow = _coerce_number(summary.get("net_cashflow", income_total - expense_total))
    max_chart_value = max(abs(income_total), abs(expense_total), 1.0)
    return {
        "income": income_total,
        "expense": expense_total,
        "net": net_cashflow,
        "income_width": min(100, abs(income_total) / max_chart_value * 100),
        "expense_width": min(100, abs(expense_total) / max_chart_value * 100),
        "net_width": min(100, abs(net_cashflow) / max_chart_value * 100),
    }


def _next_vat_return_due_date(today: date) -> date:
    """Irish VAT is filed bi-monthly, due the 19th of the month after each period ends
    (Jan-Feb -> 19 Mar, Mar-Apr -> 19 May, ... Nov-Dec -> 19 Jan). Return whichever
    fixed due date is soonest on/after `today`."""
    candidates = []
    for year in (today.year, today.year + 1):
        for month in (3, 5, 7, 9, 11):
            candidates.append(date(year, month, 19))
        candidates.append(date(year, 1, 19))
    candidates = sorted(set(candidates))
    for candidate in candidates:
        if candidate >= today:
            return candidate
    return candidates[-1]


def _build_upcoming_actions(
    data: dict[str, Any],
    business_structure: str,
    clients_catalog: list[dict[str, Any]],
    *,
    today: date | None = None,
) -> list[dict[str, Any]]:
    current_day = today or date.today()
    actions: list[dict[str, Any]] = []
    invoices = data.get("sheets", {}).get("Invoices", [])

    overdue = [row for row in invoices if _normalize_invoice_status(row.get("Status")) == "Overdue"]
    if overdue:
        total = round(sum(_coerce_number(row.get("Balance Due (€)")) for row in overdue), 2)
        actions.append({
            "severity": "danger",
            "label": f"{len(overdue)} invoice{'s' if len(overdue) != 1 else ''} overdue",
            "detail": f"{_format_currency(total)} outstanding",
            "link": "/invoices",
        })

    due_soon = []
    for row in invoices:
        if _normalize_invoice_status(row.get("Status")) != "Issued":
            continue
        due_date = _parse_iso_date(row.get("Due Date"))
        if due_date and current_day <= due_date <= current_day + timedelta(days=14):
            due_soon.append(row)
    if due_soon:
        total = round(sum(_coerce_number(row.get("Balance Due (€)")) for row in due_soon), 2)
        actions.append({
            "severity": "warning",
            "label": f"{len(due_soon)} invoice{'s' if len(due_soon) != 1 else ''} due within 14 days",
            "detail": f"{_format_currency(total)} expected",
            "link": "/invoices",
        })

    vat_due = _next_vat_return_due_date(current_day)
    actions.append({
        "severity": "info",
        "label": "Next VAT return due",
        "detail": vat_due.strftime("%d %B %Y"),
        "link": "/ledger",
    })

    if _normalize_business_structure(business_structure) == "sole_trader":
        form11_deadline = date(current_day.year, 10, 31)
        if form11_deadline < current_day:
            form11_deadline = date(current_day.year + 1, 10, 31)
        actions.append({
            "severity": "info",
            "label": "Form 11 deadline",
            "detail": form11_deadline.strftime("%d %B %Y"),
            "link": "/settings",
        })

    monthly_partners = [client for client in clients_catalog if client.get("tier") == "Clarity Partner" and client.get("retainer_frequency") == "monthly"]
    if monthly_partners:
        total = round(sum(_coerce_number(client.get("retainer_amount")) for client in monthly_partners), 2)
        actions.append({
            "severity": "info",
            "label": f"{len(monthly_partners)} Clarity Partner client{'s' if len(monthly_partners) != 1 else ''} billing this month",
            "detail": f"{_format_currency(total)} expected",
            "link": "/clients",
        })

    return actions


def _build_monthly_net_trend(data: dict[str, Any], *, today: date | None = None, months: int = 6) -> list[dict[str, Any]]:
    current_day = today or date.today()
    period_keys = []
    year, month = current_day.year, current_day.month
    for _ in range(months):
        period_keys.append((year, month))
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    period_keys.reverse()

    income_by_period = {key: 0.0 for key in period_keys}
    expense_by_period = {key: 0.0 for key in period_keys}

    for row in data.get("sheets", {}).get("Income", []):
        if str(row.get("Status") or "").strip().lower() != "received":
            continue
        parsed = _parse_iso_date(row.get("Date"))
        if not parsed:
            continue
        key = (parsed.year, parsed.month)
        if key in income_by_period:
            income_by_period[key] += _coerce_number(row.get("Amount (€)"))

    for row in data.get("sheets", {}).get("Expenses", []):
        parsed = _parse_iso_date(row.get("Date (Registered)"))
        if not parsed:
            continue
        key = (parsed.year, parsed.month)
        if key in expense_by_period:
            expense_by_period[key] += _coerce_number(row.get("Total (€)"))

    trend = []
    for key in period_keys:
        net = round(income_by_period[key] - expense_by_period[key], 2)
        trend.append({"label": date(key[0], key[1], 1).strftime("%b"), "net": net})

    max_abs = max([abs(item["net"]) for item in trend] + [1.0])
    for item in trend:
        item["height_pct"] = round(min(100, abs(item["net"]) / max_abs * 100), 1)
        item["positive"] = item["net"] >= 0
    return trend


def _parse_row_number(value: Any) -> int | None:
    try:
        row_number = int(str(value).strip())
    except (TypeError, ValueError):
        return None
    return row_number if row_number >= 1 else None


def _find_row_by_number(rows: list[dict[str, Any]], row_number: Any) -> dict[str, Any] | None:
    resolved_row_number = _parse_row_number(row_number)
    if resolved_row_number is None:
        return None
    for row in rows:
        if row.get("__row_number") == resolved_row_number:
            return row
    return None


def _find_subscription_by_id(subscriptions: list[dict[str, Any]], subscription_id: Any) -> dict[str, Any] | None:
    resolved_id = str(subscription_id or "").strip()
    if not resolved_id:
        return None
    for subscription in subscriptions:
        if str(subscription.get("id")) == resolved_id:
            return subscription
    return None


def _find_sheet_row_or_raise(sheet_name: str, row_number: int) -> dict[str, Any]:
    data = load_finance_data()
    row = _find_row_by_number(data.get("sheets", {}).get(sheet_name, []), row_number)
    if row is None:
        raise ValueError(f"Could not find {sheet_name} row {row_number}")
    return row


def _try_parse_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("€", "").strip()
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _values_equivalent(left: Any, right: Any) -> bool:
    left_number = _try_parse_number(left)
    right_number = _try_parse_number(right)
    if left_number is not None and right_number is not None:
        return abs(left_number - right_number) < 0.000001

    left_text = str(left or "").strip()
    right_text = str(right or "").strip()
    return left_text == right_text


def _records_match(left: dict[str, Any], right: dict[str, Any], headers: list[str]) -> bool:
    for header in headers:
        if not _values_equivalent(left.get(header, ""), right.get(header, "")):
            return False
    return True


def _find_restore_conflict(entity_type: str, archive_entry: dict[str, Any]) -> dict[str, Any] | None:
    record = archive_entry.get("record", {}) if isinstance(archive_entry.get("record"), dict) else {}
    if entity_type == "subscription":
        for subscription in _load_subscriptions():
            if subscription.get("id") == record.get("id"):
                return subscription
            if all(
                str(subscription.get(field, "") or "").strip() == str(record.get(field, "") or "").strip()
                for field in ["title", "supplier", "frequency", "next_charge_date", "status"]
            ):
                return subscription
        return None

    if entity_type == "payroll":
        for payroll_entry in _load_payroll_entries():
            if payroll_entry.get("id") == record.get("id"):
                return payroll_entry
            if all(
                str(payroll_entry.get(field, "") or "").strip() == str(record.get(field, "") or "").strip()
                for field in ["Pay Date", "Employee Name", "Gross Pay (€)"]
            ):
                return payroll_entry
        return None

    config = WORKBOOK_ENTITY_CONFIG.get(entity_type)
    if not config:
        return None

    data = load_finance_data()
    rows = data.get("sheets", {}).get(config["sheet"], [])
    headers = SHEET_HEADERS.get(config["sheet"], [])
    for row in rows:
        if _records_match(row, record, headers):
            return row
    return None


def _export_audit_entries_csv(entries: list[dict[str, Any]]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["timestamp", "action", "entity_type", "details_json"])
    for entry in entries:
        writer.writerow([
            entry.get("timestamp", ""),
            entry.get("action", ""),
            entry.get("entity_type", ""),
            json.dumps(entry.get("details", {}), ensure_ascii=True),
        ])
    return buffer.getvalue()


def _summarize_archives(archive_records: list[dict[str, Any]]) -> dict[str, int]:
    summary: dict[str, int] = {}
    for record in archive_records:
        entity_type = str(record.get("entity_type") or "other")
        summary[entity_type] = summary.get(entity_type, 0) + 1
    return summary


def _build_sync_message(sync_result: dict[str, int]) -> str | None:
    posted_count = sync_result.get("posted_count", 0)
    if posted_count <= 0:
        return None
    if posted_count == 1:
        return "1 subscription charge was posted to expenses automatically."
    return f"{posted_count} subscription charges were posted to expenses automatically."


def _build_page_context(
    page_title: str,
    active_tab: str,
    data: dict[str, Any],
    *,
    income: list[dict[str, Any]] | None = None,
    expenses: list[dict[str, Any]] | None = None,
    invoices: list[dict[str, Any]] | None = None,
    clients: list[dict[str, Any]] | None = None,
    suppliers: list[dict[str, Any]] | None = None,
    subscriptions: list[dict[str, Any]] | None = None,
    payroll: list[dict[str, Any]] | None = None,
    services: list[dict[str, Any]] | None = None,
    editing_service: dict[str, Any] | None = None,
    service_form: dict[str, Any] | None = None,
    show_archived_services: bool = False,
    subscription_summary: dict[str, Any] | None = None,
    payroll_summary: dict[str, Any] | None = None,
    chart_data: dict[str, float] | None = None,
    editing_income: dict[str, Any] | None = None,
    editing_expense: dict[str, Any] | None = None,
    editing_invoice: dict[str, Any] | None = None,
    editing_client: dict[str, Any] | None = None,
    editing_supplier: dict[str, Any] | None = None,
    editing_subscription: dict[str, Any] | None = None,
    editing_payroll: dict[str, Any] | None = None,
    income_form: dict[str, Any] | None = None,
    expense_form: dict[str, Any] | None = None,
    invoice_form: dict[str, Any] | None = None,
    client_form: dict[str, Any] | None = None,
    supplier_form: dict[str, Any] | None = None,
    subscription_form: dict[str, Any] | None = None,
    payroll_form: dict[str, Any] | None = None,
    validation_errors: dict[str, str] | None = None,
    archived_records: list[dict[str, Any]] | None = None,
    archive_summary: dict[str, int] | None = None,
    audit_entries: list[dict[str, Any]] | None = None,
    message: str | None = None,
    sync_message: str | None = None,
    error: str | None = None,
    phase_filter: str | None = None,
    editing_document: dict[str, Any] | None = None,
    document_form: dict[str, Any] | None = None,
    editing_compliance_entry: dict[str, Any] | None = None,
    compliance_form: dict[str, Any] | None = None,
    editing_project: dict[str, Any] | None = None,
    project_form: dict[str, Any] | None = None,
    project_detail: dict[str, Any] | None = None,
    dmaic_project: dict[str, Any] | None = None,
    editing_delivery: dict[str, Any] | None = None,
    delivery_form: dict[str, Any] | None = None,
    editing_sop: dict[str, Any] | None = None,
    sop_form: dict[str, Any] | None = None,
    status_filter: str | None = None,
    tier_filter: str | None = None,
) -> dict[str, Any]:
    try:
        workbook_path = _resolve_workbook_path()
    except FileNotFoundError:
        workbook_path = WORKBOOK_PATH
    summary = data.get("summary", {})
    version = request.args.get("v", "20260729")
    business_profile = _load_business_profile()
    structure = _normalize_business_structure(business_profile.get("structure"))
    vat_threshold_basis = _normalize_vat_threshold_basis(business_profile.get("vat_threshold_basis"))
    income_rows_for_metrics = income if income is not None else data.get("sheets", {}).get("Income", [])
    invoice_rows_for_metrics = invoices if invoices is not None else data.get("sheets", {}).get("Invoices", [])
    phase_policy = _build_phase_policy(summary, structure)
    chart_of_accounts = _ensure_chart_of_accounts()
    ledger_entries = _load_ledger_entries()
    ledger_entries_sorted = sorted(ledger_entries, key=lambda item: str(item.get("timestamp") or ""), reverse=True)
    trial_balance = _compute_trial_balance(ledger_entries, chart_of_accounts)
    vat_control_summary = _compute_vat_control_summary(ledger_entries)
    vat_threshold_summary = _compute_vat_threshold_summary(income_rows_for_metrics, invoice_rows_for_metrics, vat_threshold_basis)
    vat_anomalies = _detect_vat_anomalies(ledger_entries)
    capital_assets = _load_capital_assets()
    capital_summary = _summarize_capital_assets(capital_assets)
    payroll_entries = payroll if payroll is not None else _load_payroll_entries()
    resolved_payroll_summary = payroll_summary or _summarize_payroll_entries(payroll_entries)
    bank_statement_lines = _load_bank_statement_lines()
    reconciliation_rows = _build_reconciliation_rows(data, payroll_entries)
    matched_statement_lines, unmatched_statement_lines = _match_bank_statement_lines(reconciliation_rows, bank_statement_lines)
    reconciliation_summary = _summarize_reconciliation(reconciliation_rows)
    reconciliation_summary["statement_line_count"] = len(matched_statement_lines)
    reconciliation_summary["unmatched_statement_count"] = len(unmatched_statement_lines)
    reconciliation_exceptions = [row for row in reconciliation_rows if row.get("exception_reasons")]
    all_services = services if services is not None else _load_services()
    active_services = [service for service in all_services if service.get("status") == "active"]
    clients_catalog = [
        {
            "name": row.get("Client Name") or "",
            "tier": row.get("Service Tier") or "None",
            "retainer_frequency": row.get("Retainer Frequency") or "",
            "retainer_amount": _coerce_number(row.get("Retainer Amount (€)")),
        }
        for row in data.get("sheets", {}).get("Clients", [])
        if row.get("Client Name")
    ]
    active_clients_count = sum(1 for client in clients_catalog if client["tier"] and client["tier"] != "None")
    company_documents = _load_company_documents()
    for document in company_documents:
        document["expiry_severity"] = _document_expiry_severity(document.get("expiry_date"))
    documents_expiring_soon = _documents_expiring_soon(company_documents)
    compliance_entries = _load_compliance_entries()
    compliance_deadlines = _build_compliance_deadlines(business_profile, data, compliance_entries)
    upcoming_actions = _build_upcoming_actions(data, structure, clients_catalog)
    for deadline in compliance_deadlines[:3]:
        upcoming_actions.append({
            "severity": {"red": "danger", "amber": "warning", "green": "info"}.get(deadline["severity"], "info"),
            "label": deadline["name"],
            "detail": f"Due {datetime.fromisoformat(deadline['due_date']).strftime('%d %B %Y')}",
            "link": deadline["link"],
        })
    if documents_expiring_soon:
        upcoming_actions.append({
            "severity": "warning" if all(doc["expiry_severity"] == "soon" for doc in documents_expiring_soon) else "danger",
            "label": f"{len(documents_expiring_soon)} document{'s' if len(documents_expiring_soon) != 1 else ''} expiring soon",
            "detail": documents_expiring_soon[0]["name"],
            "link": "/company/documents",
        })
    projects = _load_projects()
    active_projects_count = sum(1 for project in projects if project.get("status") == "Active")
    projects_due_soon = _projects_with_deadline_within(projects, PROJECT_DEADLINE_WARNING_DAYS)
    for project in projects_due_soon[:3]:
        upcoming_actions.append({
            "severity": "warning",
            "label": f"Project due soon: {project['title']}",
            "detail": f"Target end {project['target_end_date']}",
            "link": f"/operations/projects/{project['id']}",
        })
    delivery_log = _load_delivery_log()
    pending_billing = _clarity_partner_pending_billing(delivery_log)
    for pending in pending_billing[:3]:
        upcoming_actions.append({
            "severity": "info",
            "label": f"Pending billing: {pending['client_name']} ({pending['billing_period']})",
            "detail": f"{pending['entry_count']} unbilled deliver{'y' if pending['entry_count'] == 1 else 'ies'}",
            "link": "/operations/delivery",
        })
    sops = _load_sops()
    dmaic_completion_percentage = _dmaic_completion_percentage(dmaic_project["dmaic"]) if dmaic_project else 0
    monthly_net_trend = _build_monthly_net_trend(data)
    return {
        "page_title": page_title,
        "summary": summary,
        "income": income or [],
        "expenses": expenses or [],
        "invoices": invoices or [],
        "clients": clients or [],
        "suppliers": suppliers or [],
        "subscriptions": subscriptions or [],
        "payroll": payroll_entries,
        "subscription_summary": subscription_summary or {"active_count": 0, "due_count": 0, "upcoming_count": 0, "monthly_commitment": 0.0},
        "payroll_summary": resolved_payroll_summary,
        "chart_data": chart_data or _build_chart_data(summary),
        "format_currency": _format_currency,
        "raw_amount": _coerce_number,
        "version": version,
        "active_tab": active_tab,
        "phase_filter": phase_filter,
        "editing_income": editing_income,
        "editing_expense": editing_expense,
        "editing_invoice": editing_invoice,
        "editing_client": editing_client,
        "editing_supplier": editing_supplier,
        "editing_subscription": editing_subscription,
        "editing_payroll": editing_payroll,
        "income_form": income_form or {},
        "expense_form": expense_form or {},
        "invoice_form": invoice_form or {},
        "client_form": client_form or {},
        "supplier_form": supplier_form or {},
        "subscription_form": subscription_form or {},
        "payroll_form": payroll_form or {},
        "validation_errors": validation_errors or {},
        "archived_records": archived_records or [],
        "archive_summary": archive_summary or {},
        "audit_entries": audit_entries or [],
        "today_iso": date.today().isoformat(),
        "current_month_label": date.today().strftime("%B %Y"),
        "message": message,
        "sync_message": sync_message,
        "error": error,
        "income_clients": _collect_select_options(data, "Income", "Client / Source", base_options=_collect_select_options(data, "Clients", "Client Name")),
        "income_categories": _collect_select_options(data, "Income", "Category", base_options=_chart_of_accounts_category_options("Income")),
        "expense_suppliers": _collect_select_options(data, "Expenses", "Supplier / Payee", base_options=_collect_select_options(data, "Suppliers", "Supplier Name")),
        "expense_categories": _collect_select_options(data, "Expenses", "Category", base_options=_chart_of_accounts_category_options("Expense")),
        "client_names": _collect_select_options(data, "Clients", "Client Name"),
        "supplier_names": _collect_select_options(data, "Suppliers", "Supplier Name"),
        "subscription_frequencies": list(SUBSCRIPTION_FREQUENCIES.keys()),
        "subscription_statuses": list(SUBSCRIPTION_STATUSES),
        "workbook_path": str(workbook_path.name),
        "workbook_status": "connected" if workbook_path.exists() else "missing",
        "business_profile": business_profile,
        "business_structure": structure,
        "vat_registered": bool(business_profile.get("vat_registered", True)),
        "vat_threshold_options": [
            {"value": key, "label": value["label"]}
            for key, value in VAT_TURNOVER_THRESHOLDS.items()
        ],
        "vat_threshold_summary": vat_threshold_summary,
        "phase_label": _phase_label_for_structure(structure),
        "phase_policy": phase_policy,
        "income_payment_methods": INCOME_PAYMENT_METHODS.get(structure, INCOME_PAYMENT_METHODS["sole_trader"]),
        "expense_payment_methods": EXPENSE_PAYMENT_METHODS.get(structure, EXPENSE_PAYMENT_METHODS["sole_trader"]),
        "vat_rate_options": VAT_RATE_OPTIONS,
        "vat_treatment_options": VAT_TREATMENT_OPTIONS,
        "supply_type_options": SUPPLY_TYPE_OPTIONS,
        "income_status_options": INCOME_STATUS_OPTIONS,
        "invoice_status_options": INVOICE_STATUS_OPTIONS,
        "expense_status_options": EXPENSE_STATUS_OPTIONS,
        "expense_input_vat_options": EXPENSE_INPUT_VAT_OPTIONS,
        "expense_deductibility_options": EXPENSE_DEDUCTIBILITY_OPTIONS,
        "reconciliation_options": RECONCILIATION_OPTIONS,
        "yes_no_options": YES_NO_OPTIONS,
        "payroll_status_options": PAYROLL_STATUS_OPTIONS,
        "chart_of_accounts": chart_of_accounts,
        "coa_summary": _summarize_chart_of_accounts(chart_of_accounts),
        "ledger_entries": ledger_entries_sorted[:80],
        "vat_anomalies": vat_anomalies,
        "vat_anomaly_count": len(vat_anomalies),
        "capital_summary": capital_summary,
        "capital_assets": capital_assets,
        "reconciliation_rows": reconciliation_rows,
        "reconciliation_summary": reconciliation_summary,
        "reconciliation_exceptions": reconciliation_exceptions,
        "bank_statement_lines": matched_statement_lines,
        "unmatched_statement_lines": unmatched_statement_lines,
        "ledger_summary": {
            "entries_count": len(ledger_entries),
            "posted_total": round(sum(_coerce_number(entry.get("amount_eur")) for entry in ledger_entries), 2),
            "debit_total": trial_balance["total_debit"],
            "credit_total": trial_balance["total_credit"],
        },
        "trial_balance": trial_balance,
        "vat_control_summary": vat_control_summary,
        "services": all_services,
        "active_services": active_services,
        "editing_service": editing_service,
        "service_form": service_form or {},
        "show_archived_services": show_archived_services,
        "service_groups": list(SERVICE_GROUPS),
        "service_price_types": list(SERVICE_PRICE_TYPES),
        "service_billing_frequencies": list(SERVICE_BILLING_FREQUENCIES),
        "services_catalog_json": json.dumps(active_services).replace("</", "<\\/"),
        "tax_rules_json": json.dumps(_load_tax_rules()).replace("</", "<\\/"),
        "clients_catalog_json": json.dumps(clients_catalog).replace("</", "<\\/"),
        "active_clients_count": active_clients_count,
        "upcoming_actions": upcoming_actions,
        "backup_status": _load_backup_status(),
        "monthly_net_trend": monthly_net_trend,
        "client_service_tiers": list(CLIENT_SERVICE_TIERS),
        "client_retainer_frequencies": list(CLIENT_RETAINER_FREQUENCIES),
        "company_documents": company_documents,
        "documents_expiring_soon": documents_expiring_soon,
        "document_categories": list(DOCUMENT_CATEGORIES),
        "editing_document": editing_document,
        "document_form": document_form or {},
        "compliance_entries": compliance_entries,
        "compliance_deadlines": compliance_deadlines,
        "compliance_repeat_frequencies": list(COMPLIANCE_REPEAT_FREQUENCIES),
        "editing_compliance_entry": editing_compliance_entry,
        "compliance_form": compliance_form or {},
        "projects": projects,
        "active_projects_count": active_projects_count,
        "projects_due_soon": projects_due_soon,
        "project_statuses": list(PROJECT_STATUSES),
        "project_kanban_statuses": list(PROJECT_KANBAN_STATUSES),
        "dmaic_phases": list(DMAIC_PHASES),
        "dmaic_phase_statuses": list(DMAIC_PHASE_STATUSES),
        "editing_project": editing_project,
        "project_form": project_form or {},
        "project_detail": project_detail,
        "dmaic_project": dmaic_project,
        "status_filter": status_filter or "",
        "tier_filter": tier_filter or "",
        "delivery_log": delivery_log,
        "delivery_service_types": list(DELIVERY_SERVICE_TYPES),
        "pending_billing": pending_billing,
        "editing_delivery": editing_delivery,
        "delivery_form": delivery_form or {},
        "sops": sops,
        "sop_statuses": list(SOP_STATUSES),
        "sop_status_workflow": list(SOP_STATUS_WORKFLOW),
        "editing_sop": editing_sop,
        "sop_form": sop_form or {},
        "dmaic_completion_percentage": dmaic_completion_percentage,
    }


@lru_cache(maxsize=1)
def load_finance_data() -> dict[str, Any]:
    sheets: dict[str, Any] = {}
    for sheet_name in ["Income", "Expenses", "Invoices", "Clients", "Suppliers"]:
        sheets[sheet_name] = _load_sheet_rows_with_row_numbers(sheet_name)

    for row in sheets["Expenses"]:
        if not row.get("Total (€)"):
            row["Total (€)"] = round(_coerce_number(row.get("Net Amount (€)", 0)) + _coerce_number(row.get("VAT (€)", 0)) + _coerce_number(row.get("Fees (€)", 0)), 2)
    for row in sheets["Invoices"]:
        if not row.get("Total (€)"):
            row["Total (€)"] = round(_coerce_number(row.get("Net (€)", 0)) + _coerce_number(row.get("VAT (€)", 0)), 2)

    if _auto_flag_overdue_invoices(sheets["Invoices"]):
        _save_sheet_records_raw("Invoices", [{k: v for k, v in row.items() if not k.startswith("__")} for row in sheets["Invoices"]])

    income_total = sum(
        _coerce_number(row.get("Amount (€)", row.get("Total incl. VAT (€)", 0)))
        for row in sheets["Income"]
        if str(row.get("Status") or "").strip().lower() == "received"
    )
    expense_total = sum(_coerce_number(row.get("Total (€)")) for row in sheets["Expenses"])
    invoice_balance = sum(
        _coerce_number(row.get("Balance Due (€)", row.get("Balance (€)", 0)))
        for row in sheets["Invoices"]
        if _normalize_invoice_status(row.get("Status")) in ("Issued", "Overdue", "Partially Paid")
    )
    ap_balance = sum(
        _coerce_number(row.get("Total (€)", 0))
        for row in sheets["Expenses"]
        if not _is_paid_status("expense", row.get("Status"))
    )
    vat_balance = _compute_vat_control_summary(_load_ledger_entries()).get("t3_net_vat", 0.0)

    try:
        resolved_workbook_name = _resolve_workbook_path().name
    except FileNotFoundError:
        resolved_workbook_name = None

    return {
        "sheets": sheets,
        "workbook_path": resolved_workbook_name,
        "summary": {
            "income_total": income_total,
            "expense_total": expense_total,
            "net_cashflow": income_total - expense_total,
            "invoice_balance": invoice_balance,
            "ap_balance": ap_balance,
            "vat_balance": vat_balance,
        },
    }


def _read_workbook_sheet_rows(wb, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in wb:
        return []
    ws = wb[sheet_name]
    rows: list[dict[str, Any]] = []
    headers: list[Any] = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        cleaned = [_coerce_value(v) for v in row]
        if not header_found and _is_header_row(cleaned, sheet_name):
            headers = [_normalize_header_name(v, sheet_name) for v in cleaned]
            header_found = True
            continue
        if not header_found:
            continue
        if not any(v not in (None, "") for v in cleaned):
            continue
        rows.append(dict(zip(headers, cleaned)))
    return rows


def _migrate_transaction_sheets_from_workbook() -> None:
    """One-time seed of income/expenses/invoices/clients/suppliers JSON files from the
    legacy xlsm workbook, if present. Runs once at process start; after the JSON files
    exist, the workbook is never read for normal operation again."""
    missing_sheet_names = [name for name, path in SHEET_JSON_PATHS.items() if not path.exists()]
    if not missing_sheet_names:
        return

    try:
        resolved_path = _resolve_workbook_path()
    except FileNotFoundError:
        resolved_path = None

    seeded: dict[str, list[dict[str, Any]]] = {name: [] for name in missing_sheet_names}
    if resolved_path is not None and resolved_path.exists():
        try:
            wb = load_workbook(resolved_path, data_only=True, read_only=True)
            try:
                for sheet_name in missing_sheet_names:
                    seeded[sheet_name] = _read_workbook_sheet_rows(wb, sheet_name)
            finally:
                wb.close()
        except Exception:
            # Migration is best-effort; fall back to empty JSON files for any sheet
            # we couldn't read rather than blocking startup.
            pass

    for sheet_name in missing_sheet_names:
        _save_json_records(SHEET_JSON_PATHS[sheet_name], seeded[sheet_name])


@app.route("/")
def index():
    sync_result = _sync_subscriptions_to_expenses()
    data = load_finance_data()
    subscription_rows = _build_subscription_rows(_load_subscriptions())
    subscription_summary = _summarize_subscriptions(subscription_rows)
    if "error" in data:
        return render_template("index.html", **_build_page_context("H-Queex Hub", "dashboard", {}, error=data["error"], subscriptions=subscription_rows, subscription_summary=subscription_summary, sync_message=_build_sync_message(sync_result)))

    summary = data["summary"]
    income = data["sheets"].get("Income", [])[:8]
    expenses = data["sheets"].get("Expenses", [])[:8]
    invoices = data["sheets"].get("Invoices", [])[:8]
    clients = data["sheets"].get("Clients", [])[:8]
    suppliers = data["sheets"].get("Suppliers", [])[:8]
    return render_template(
        "index.html",
        **_build_page_context(
            "H-Queex Hub",
            "dashboard",
            data,
            income=income,
            expenses=expenses,
            invoices=invoices,
            clients=clients,
            suppliers=suppliers,
            subscriptions=subscription_rows,
            subscription_summary=subscription_summary,
            chart_data=_build_chart_data(summary),
            message=request.args.get("message"),
            sync_message=_build_sync_message(sync_result),
        ),
    )


@app.route("/finance")
def finance_view():
    data = load_finance_data()
    return render_template(
        "index.html",
        **_build_page_context(
            "Finance hub",
            "finance",
            data,
            message=request.args.get("message"),
        ),
    )


@app.route("/documents/<path:filename>")
def serve_company_document(filename):
    as_attachment = request.args.get("download") == "1"
    return send_from_directory(COMPANY_DOCUMENTS_DIR, filename, as_attachment=as_attachment)


@app.route("/company")
def company_view():
    data = load_finance_data()
    return render_template(
        "index.html",
        **_build_page_context(
            "Company",
            "company",
            data,
            message=request.args.get("message"),
        ),
    )


@app.route("/company/documents")
def company_documents_view():
    data = load_finance_data()
    documents = _load_company_documents()
    editing_document = _find_record_by_id(documents, request.args.get("edit_id"))
    return render_template(
        "index.html",
        **_build_page_context(
            "Documents",
            "company_documents",
            data,
            editing_document=editing_document,
            message=request.args.get("message"),
        ),
    )


@app.route("/company/documents/upload", methods=["POST"])
def upload_company_document():
    name = str(request.form.get("name") or "").strip()
    category = _normalize_document_category(request.form.get("category"))
    description = str(request.form.get("description") or "").strip()
    expiry_date = str(request.form.get("expiry_date") or "").strip()
    notes = str(request.form.get("notes") or "").strip()

    errors: dict[str, str] = {}
    _validate_required_text(name, "name", "Document name", errors)
    if expiry_date and _parse_transaction_date(expiry_date) is None:
        errors["expiry_date"] = "Expiry date must be a valid date"

    stored_filename = ""
    if not errors:
        stored_filename, upload_error = _save_uploaded_document(request.files.get("document_file"))
        if upload_error:
            errors["document_file"] = upload_error

    if errors:
        return _redirect_with_form_errors(
            "company_documents_view",
            {"name": name, "category": category, "description": description, "expiry_date": expiry_date, "notes": notes},
            errors,
        )

    now = datetime.now().isoformat(timespec="seconds")
    document = {
        "id": str(uuid4()),
        "name": name,
        "category": category,
        "description": description,
        "filename": stored_filename,
        "file_path": f"documents/{stored_filename}" if stored_filename else "",
        "date_added": date.today().isoformat(),
        "expiry_date": expiry_date,
        "status": "active",
        "notes": notes,
        "created_at": now,
        "last_updated_at": now,
    }
    documents = _load_company_documents()
    documents.append(document)
    _save_company_documents(documents)
    _record_audit("create", "company_document", {"document_id": document["id"], "record": document})
    return redirect(url_for("company_documents_view", message="Document uploaded"))


@app.route("/company/documents/update", methods=["POST"])
def update_company_document():
    document_id = str(request.form.get("document_id") or "").strip()
    name = str(request.form.get("name") or "").strip()
    category = _normalize_document_category(request.form.get("category"))
    description = str(request.form.get("description") or "").strip()
    expiry_date = str(request.form.get("expiry_date") or "").strip()
    notes = str(request.form.get("notes") or "").strip()
    status = str(request.form.get("status") or "active").strip().lower()
    status = status if status in DOCUMENT_STATUSES else "active"

    errors: dict[str, str] = {}
    _validate_required_text(name, "name", "Document name", errors)
    if expiry_date and _parse_transaction_date(expiry_date) is None:
        errors["expiry_date"] = "Expiry date must be a valid date"

    documents = _load_company_documents()
    document = _find_record_by_id(documents, document_id)
    if document is None:
        return redirect(url_for("company_documents_view", message="Document not found"))

    if errors:
        return _redirect_with_form_errors(
            "company_documents_view",
            {"name": name, "category": category, "description": description, "expiry_date": expiry_date, "notes": notes},
            errors,
            edit_id=document_id,
        )

    stored_filename, upload_error = _save_uploaded_document(request.files.get("document_file"))
    if upload_error:
        return _redirect_with_form_errors(
            "company_documents_view",
            {"name": name, "category": category, "description": description, "expiry_date": expiry_date, "notes": notes},
            {"document_file": upload_error},
            edit_id=document_id,
        )

    document["name"] = name
    document["category"] = category
    document["description"] = description
    document["expiry_date"] = expiry_date
    document["notes"] = notes
    document["status"] = status
    if stored_filename:
        document["filename"] = stored_filename
        document["file_path"] = f"documents/{stored_filename}"
    document["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
    _save_company_documents(documents)
    _record_audit("update", "company_document", {"document_id": document_id, "record": document})
    return redirect(url_for("company_documents_view", message="Document updated"))


@app.route("/company/documents/archive", methods=["POST"])
def archive_company_document():
    document_id = str(request.form.get("document_id") or "").strip()
    documents = _load_company_documents()
    document = _find_record_by_id(documents, document_id)
    if document is not None:
        document["status"] = "archived"
        document["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
        _save_company_documents(documents)
        _record_audit("archive", "company_document", {"document_id": document_id, "record": document})
    return redirect(url_for("company_documents_view", message="Document archived"))


@app.route("/company/compliance")
def company_compliance_view():
    data = load_finance_data()
    entries = _load_compliance_entries()
    editing_compliance_entry = _find_record_by_id(entries, request.args.get("edit_id"))
    return render_template(
        "index.html",
        **_build_page_context(
            "Compliance Calendar",
            "company_compliance",
            data,
            editing_compliance_entry=editing_compliance_entry,
            message=request.args.get("message"),
        ),
    )


@app.route("/company/compliance/add", methods=["POST"])
def add_compliance_entry():
    name = str(request.form.get("name") or "").strip()
    due_date = str(request.form.get("due_date") or "").strip()
    description = str(request.form.get("description") or "").strip()
    repeat_frequency = str(request.form.get("repeat_frequency") or "").strip().lower()
    linked_document_id = str(request.form.get("linked_document_id") or "").strip()

    errors: dict[str, str] = {}
    _validate_required_text(name, "name", "Deadline name", errors)
    if not due_date or _parse_transaction_date(due_date) is None:
        errors["due_date"] = "Due date is required and must be valid"

    if errors:
        return _redirect_with_form_errors(
            "company_compliance_view",
            {"name": name, "due_date": due_date, "description": description, "repeat_frequency": repeat_frequency},
            errors,
        )

    now = datetime.now().isoformat(timespec="seconds")
    entry = {
        "id": str(uuid4()),
        "name": name,
        "due_date": due_date,
        "description": description,
        "repeat_frequency": repeat_frequency if repeat_frequency in COMPLIANCE_REPEAT_FREQUENCIES else "",
        "linked_document_id": linked_document_id,
        "status": "pending",
        "created_at": now,
        "last_updated_at": now,
    }
    entries = _load_compliance_entries()
    entries.append(entry)
    _save_compliance_entries(entries)
    _record_audit("create", "compliance_entry", {"entry_id": entry["id"], "record": entry})
    return redirect(url_for("company_compliance_view", message="Deadline added"))


@app.route("/company/compliance/update", methods=["POST"])
def update_compliance_entry():
    entry_id = str(request.form.get("entry_id") or "").strip()
    name = str(request.form.get("name") or "").strip()
    due_date = str(request.form.get("due_date") or "").strip()
    description = str(request.form.get("description") or "").strip()
    repeat_frequency = str(request.form.get("repeat_frequency") or "").strip().lower()
    linked_document_id = str(request.form.get("linked_document_id") or "").strip()

    errors: dict[str, str] = {}
    _validate_required_text(name, "name", "Deadline name", errors)
    if not due_date or _parse_transaction_date(due_date) is None:
        errors["due_date"] = "Due date is required and must be valid"

    entries = _load_compliance_entries()
    entry = _find_record_by_id(entries, entry_id)
    if entry is None:
        return redirect(url_for("company_compliance_view", message="Deadline not found"))

    if errors:
        return _redirect_with_form_errors(
            "company_compliance_view",
            {"name": name, "due_date": due_date, "description": description, "repeat_frequency": repeat_frequency},
            errors,
            edit_id=entry_id,
        )

    entry["name"] = name
    entry["due_date"] = due_date
    entry["description"] = description
    entry["repeat_frequency"] = repeat_frequency if repeat_frequency in COMPLIANCE_REPEAT_FREQUENCIES else ""
    entry["linked_document_id"] = linked_document_id
    entry["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
    _save_compliance_entries(entries)
    _record_audit("update", "compliance_entry", {"entry_id": entry_id, "record": entry})
    return redirect(url_for("company_compliance_view", message="Deadline updated"))


@app.route("/company/compliance/complete", methods=["POST"])
def complete_compliance_entry():
    entry_id = str(request.form.get("entry_id") or "").strip()
    entries = _load_compliance_entries()
    entry = _find_record_by_id(entries, entry_id)
    if entry is not None:
        entry["status"] = "complete"
        entry["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
        _save_compliance_entries(entries)
        _record_audit("complete", "compliance_entry", {"entry_id": entry_id, "record": entry})
    return redirect(url_for("company_compliance_view", message="Deadline marked complete"))


@app.route("/company/profile")
def company_profile_view():
    data = load_finance_data()
    return render_template(
        "index.html",
        **_build_page_context(
            "Business Profile",
            "company_profile",
            data,
            message=request.args.get("message"),
        ),
    )


@app.route("/company/profile/update", methods=["POST"])
def update_company_profile():
    business_name = str(request.form.get("business_name") or "").strip()
    owner_name = str(request.form.get("owner_name") or "").strip()
    cro_number = str(request.form.get("cro_number") or "").strip()
    registration_date = str(request.form.get("registration_date") or "").strip()
    structure = _normalize_business_structure(request.form.get("structure"))
    trading_start_date = str(request.form.get("trading_start_date") or "").strip()
    pre_trading_start_date = str(request.form.get("pre_trading_start_date") or "").strip()
    vat_registered = str(request.form.get("vat_registered") or "").strip().lower() in {"1", "true", "yes", "on"}
    vat_threshold_basis = _normalize_vat_threshold_basis(request.form.get("vat_threshold_basis"))
    transition_date = str(request.form.get("transition_date") or "").strip()

    errors: dict[str, str] = {}
    _validate_required_text(business_name, "business_name", "Business name", errors)
    for field_name, label, value in (
        ("registration_date", "Registration date", registration_date),
        ("trading_start_date", "Trading start date", trading_start_date),
        ("pre_trading_start_date", "Pre-trading start date", pre_trading_start_date),
        ("transition_date", "Transition date", transition_date),
    ):
        if value and _parse_transaction_date(value) is None:
            errors[field_name] = f"{label} must be a valid date"

    if errors:
        return _redirect_with_form_errors(
            "company_profile_view",
            {
                "business_name": business_name,
                "owner_name": owner_name,
                "cro_number": cro_number,
                "registration_date": registration_date,
                "structure": structure,
                "trading_start_date": trading_start_date,
                "pre_trading_start_date": pre_trading_start_date,
                "transition_date": transition_date,
            },
            errors,
        )

    profile = _load_business_profile()
    profile["business_name"] = business_name
    profile["owner_name"] = owner_name
    profile["cro_number"] = cro_number
    profile["registration_date"] = registration_date
    profile["structure"] = structure
    profile["trading_start_date"] = trading_start_date
    profile["pre_trading_start_date"] = pre_trading_start_date
    profile["vat_registered"] = vat_registered
    profile["vat_threshold_basis"] = vat_threshold_basis
    profile["transition_date"] = transition_date
    _save_business_profile(profile)
    _record_audit("update", "business_profile", {"record": profile})
    return redirect(url_for("company_profile_view", message="Business profile updated"))


@app.route("/company/settings")
def company_settings_view():
    data = load_finance_data()
    return render_template(
        "index.html",
        **_build_page_context(
            "Settings",
            "settings",
            data,
            message=request.args.get("message"),
        ),
    )


@app.route("/settings")
def settings_view():
    return redirect(url_for("company_settings_view", **request.args))


# --- Operations: Projects ----------------------------------------------------

def _clients_catalog_for_operations(data: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {"name": row.get("Client Name") or "", "tier": row.get("Service Tier") or "None"}
        for row in data.get("sheets", {}).get("Clients", [])
        if row.get("Client Name")
    ]


def _client_tier_for_name(data: dict[str, Any], client_name: str) -> str:
    for client in _clients_catalog_for_operations(data):
        if client["name"] == client_name:
            return client["tier"]
    return "None"


@app.route("/operations")
def operations_view():
    data = load_finance_data()
    return render_template(
        "index.html",
        **_build_page_context(
            "Operations",
            "operations",
            data,
            message=request.args.get("message"),
        ),
    )


@app.route("/operations/projects")
def operations_projects_view():
    data = load_finance_data()
    projects = _load_projects()
    editing_project = _find_record_by_id(projects, request.args.get("edit_id"))
    status_filter = str(request.args.get("status_filter") or "")
    tier_filter = str(request.args.get("tier_filter") or "")
    return render_template(
        "index.html",
        **_build_page_context(
            "Projects",
            "operations_projects",
            data,
            editing_project=editing_project,
            status_filter=status_filter,
            tier_filter=tier_filter,
            message=request.args.get("message"),
        ),
    )


@app.route("/operations/projects/<project_id>")
def operations_project_detail_view(project_id):
    data = load_finance_data()
    projects = _load_projects()
    project = _find_record_by_id(projects, project_id)
    if project is None:
        return redirect(url_for("operations_projects_view", message="Project not found"))
    invoices = data.get("sheets", {}).get("Invoices", [])
    linked_invoices = [row for row in invoices if str(row.get("Invoice #")) in project.get("linked_invoice_ids", [])]
    delivery_entries = [entry for entry in _load_delivery_log() if entry.get("project_id") == project_id]
    sops = [sop for sop in _load_sops() if sop.get("project_id") == project_id]
    return render_template(
        "index.html",
        **_build_page_context(
            f"Project: {project.get('title') or project.get('project_number')}",
            "operations_project_detail",
            data,
            project_detail={"project": project, "linked_invoices": linked_invoices, "delivery_entries": delivery_entries, "sops": sops},
            message=request.args.get("message"),
        ),
    )


def _project_payload_from_form(data: dict[str, Any]) -> dict[str, Any]:
    client_name = str(request.form.get("client_name") or "").strip()
    try:
        raw_line_items = json.loads(request.form.get("line_items_json") or "[]")
        if not isinstance(raw_line_items, list):
            raw_line_items = []
    except (TypeError, ValueError):
        raw_line_items = []
    return {
        "client_id": client_name,
        "client_name": client_name,
        "service_tier": _client_tier_for_name(data, client_name),
        "title": str(request.form.get("title") or "").strip(),
        "description": str(request.form.get("description") or "").strip(),
        "status": str(request.form.get("status") or "Enquiry").strip(),
        "start_date": str(request.form.get("start_date") or "").strip(),
        "target_end_date": str(request.form.get("target_end_date") or "").strip(),
        "actual_end_date": str(request.form.get("actual_end_date") or "").strip(),
        "notes": str(request.form.get("notes") or "").strip(),
        "line_items": raw_line_items,
    }


def _validate_project_payload(payload: dict[str, Any]) -> dict[str, str]:
    errors: dict[str, str] = {}
    _validate_required_text(payload.get("title"), "title", "Project title", errors)
    _validate_required_text(payload.get("client_name"), "client_name", "Client", errors)
    if payload.get("status") not in PROJECT_STATUSES:
        errors["status"] = "Invalid status"
    for field_name, label in (("start_date", "Start date"), ("target_end_date", "Target end date"), ("actual_end_date", "Actual end date")):
        value = payload.get(field_name)
        if value and _parse_transaction_date(value) is None:
            errors[field_name] = f"{label} must be a valid date"
    return errors


@app.route("/operations/projects/add", methods=["POST"])
def add_project():
    data = load_finance_data()
    payload = _project_payload_from_form(data)
    errors = _validate_project_payload(payload)
    if errors:
        return _redirect_with_form_errors("operations_projects_view", {**payload, "line_items_json": request.form.get("line_items_json", "")}, errors)

    projects = _load_projects()
    now = datetime.now().isoformat(timespec="seconds")
    project = _normalize_project({
        **payload,
        "id": str(uuid4()),
        "project_number": _generate_project_number(projects),
        "linked_invoice_ids": [],
        "created_at": now,
        "updated_at": now,
    })
    projects.append(project)
    _save_projects(projects)
    _record_audit("create", "project", {"project_id": project["id"], "record": project})
    return redirect(url_for("operations_projects_view", message="Project added"))


@app.route("/operations/projects/update", methods=["POST"])
def update_project():
    project_id = str(request.form.get("project_id") or "").strip()
    data = load_finance_data()
    payload = _project_payload_from_form(data)
    errors = _validate_project_payload(payload)

    projects = _load_projects()
    project = _find_record_by_id(projects, project_id)
    if project is None:
        return redirect(url_for("operations_projects_view", message="Project not found"))

    if errors:
        return _redirect_with_form_errors(
            "operations_projects_view",
            {**payload, "line_items_json": request.form.get("line_items_json", "")},
            errors,
            edit_id=project_id,
        )

    project.update(payload)
    project["updated_at"] = datetime.now().isoformat(timespec="seconds")
    _save_projects(projects)
    _record_audit("update", "project", {"project_id": project_id, "record": project})
    return redirect(url_for("operations_projects_view", message="Project updated"))


@app.route("/operations/projects/archive", methods=["POST"])
def archive_project():
    project_id = str(request.form.get("project_id") or "").strip()
    projects = _load_projects()
    project = _find_record_by_id(projects, project_id)
    if project is not None:
        project["status"] = "Cancelled"
        project["updated_at"] = datetime.now().isoformat(timespec="seconds")
        _save_projects(projects)
        _record_audit("archive", "project", {"project_id": project_id, "record": project})
    return redirect(url_for("operations_projects_view", message="Project archived"))


@app.route("/operations/projects/status", methods=["POST"])
def update_project_status():
    project_id = str(request.form.get("project_id") or "").strip()
    new_status = str(request.form.get("status") or "").strip()
    next_page = str(request.args.get("next") or url_for("operations_projects_view"))
    if new_status not in PROJECT_STATUSES:
        return redirect(_append_message_to_path(next_page, "Invalid status"))

    projects = _load_projects()
    project = _find_record_by_id(projects, project_id)
    if project is not None:
        project["status"] = new_status
        project["updated_at"] = datetime.now().isoformat(timespec="seconds")
        _save_projects(projects)
        _record_audit("update", "project", {"project_id": project_id, "field": "status", "value": new_status})
    return redirect(_append_message_to_path(next_page, "Project status updated"))


# --- Operations: DMAIC Tracker ------------------------------------------------

@app.route("/operations/dmaic")
def operations_dmaic_view():
    data = load_finance_data()
    projects = _load_projects()
    project_id = str(request.args.get("project_id") or "")
    dmaic_project = _find_record_by_id(projects, project_id) if project_id else None
    return render_template(
        "index.html",
        **_build_page_context(
            "DMAIC Tracker",
            "operations_dmaic",
            data,
            dmaic_project=dmaic_project,
            message=request.args.get("message"),
        ),
    )


@app.route("/operations/dmaic/<project_id>")
def operations_dmaic_project_view(project_id):
    data = load_finance_data()
    projects = _load_projects()
    dmaic_project = _find_record_by_id(projects, project_id)
    if dmaic_project is None:
        return redirect(url_for("operations_dmaic_view", message="Project not found"))
    return render_template(
        "index.html",
        **_build_page_context(
            f"DMAIC Tracker: {dmaic_project.get('title')}",
            "operations_dmaic",
            data,
            dmaic_project=dmaic_project,
            message=request.args.get("message"),
        ),
    )


@app.route("/operations/dmaic/update", methods=["POST"])
def update_dmaic_phase():
    project_id = str(request.form.get("project_id") or "").strip()
    phase = str(request.form.get("phase") or "").strip()
    new_status = str(request.form.get("status") or "").strip()

    projects = _load_projects()
    project = _find_record_by_id(projects, project_id)
    if project is None:
        return redirect(url_for("operations_dmaic_view", message="Project not found"))

    if phase not in DMAIC_PHASES:
        return redirect(url_for("operations_dmaic_project_view", project_id=project_id, message="Unknown DMAIC phase"))

    transition_error = _validate_dmaic_transition(project["dmaic"], phase, new_status)
    if transition_error:
        return redirect(url_for("operations_dmaic_project_view", project_id=project_id, message=transition_error))

    deliverables_raw = request.form.get("deliverables") or ""
    deliverables = [line.strip() for line in deliverables_raw.splitlines() if line.strip()]

    project["dmaic"][phase] = {
        "status": new_status if new_status in DMAIC_PHASE_STATUSES else project["dmaic"][phase]["status"],
        "start_date": str(request.form.get("start_date") or "").strip(),
        "completion_date": str(request.form.get("completion_date") or "").strip(),
        "notes": str(request.form.get("notes") or "").strip(),
        "deliverables": deliverables,
        "key_findings": str(request.form.get("key_findings") or "").strip(),
    }
    project["updated_at"] = datetime.now().isoformat(timespec="seconds")
    _save_projects(projects)
    _record_audit("update", "project_dmaic", {"project_id": project_id, "phase": phase, "record": project["dmaic"][phase]})
    return redirect(url_for("operations_dmaic_project_view", project_id=project_id, message=f"{phase} updated"))


# --- Operations: Delivery Log -------------------------------------------------

@app.route("/operations/delivery")
def operations_delivery_view():
    data = load_finance_data()
    entries = _load_delivery_log()
    client_filter = str(request.args.get("client_filter") or "")
    project_filter = str(request.args.get("project_filter") or "")
    billing_period_filter = str(request.args.get("billing_period_filter") or "")
    if client_filter:
        entries = [entry for entry in entries if entry.get("client_name") == client_filter]
    if project_filter:
        entries = [entry for entry in entries if entry.get("project_id") == project_filter]
    if billing_period_filter:
        entries = [entry for entry in entries if entry.get("billing_period") == billing_period_filter]
    clarity_partner_client_names = [client["name"] for client in _clients_catalog_for_operations(data) if client["tier"] == "Clarity Partner"]
    return render_template(
        "index.html",
        **_build_page_context(
            "Delivery Log",
            "operations_delivery",
            data,
            delivery_form={"client_filter": client_filter, "project_filter": project_filter, "billing_period_filter": billing_period_filter},
            message=request.args.get("message"),
        ),
        delivery_entries_filtered=entries,
        clarity_partner_client_names=clarity_partner_client_names,
    )


@app.route("/operations/delivery/add", methods=["POST"])
def add_delivery_entry():
    data = load_finance_data()
    project_id = str(request.form.get("project_id") or "").strip()
    projects = _load_projects()
    project = _find_record_by_id(projects, project_id)
    client_name = project.get("client_name") if project else str(request.form.get("client_name") or "").strip()

    payload = {
        "date": str(request.form.get("date") or date.today().isoformat()).strip(),
        "project_id": project_id,
        "client_id": client_name,
        "client_name": client_name,
        "service_type": str(request.form.get("service_type") or "Other").strip(),
        "description": str(request.form.get("description") or "").strip(),
        "hours_spent": request.form.get("hours_spent") or None,
        "billing_period": str(request.form.get("billing_period") or "").strip(),
    }

    errors: dict[str, str] = {}
    _validate_required_text(payload.get("description"), "description", "Description", errors)
    _validate_required_text(client_name, "client_name", "Client", errors)
    if not _parse_transaction_date(payload.get("date")):
        errors["date"] = "Date must be valid"

    if errors:
        return _redirect_with_form_errors("operations_delivery_view", payload, errors)

    stored_filename, upload_error = _save_uploaded_delivery_file(request.files.get("deliverable_file"))
    if upload_error:
        return _redirect_with_form_errors("operations_delivery_view", payload, {"deliverable_file": upload_error})

    now = datetime.now().isoformat(timespec="seconds")
    entry = _normalize_delivery_entry({
        **payload,
        "id": str(uuid4()),
        "deliverable_filename": stored_filename,
        "invoiced": False,
        "invoice_id": "",
        "created_at": now,
        "last_updated_at": now,
    })
    entries = _load_delivery_log()
    entries.append(entry)
    _save_delivery_log(entries)
    _record_audit("create", "delivery_entry", {"entry_id": entry["id"], "record": entry})
    return redirect(url_for("operations_delivery_view", message="Delivery logged"))


@app.route("/operations/delivery/generate-invoice/<client_id>/<period>", methods=["POST"])
def generate_delivery_invoice(client_id, period):
    client_name = client_id
    billing_period = period
    entries = _load_delivery_log()
    unbilled = [entry for entry in entries if entry.get("client_name") == client_name and entry.get("billing_period") == billing_period and not entry.get("invoiced")]
    if not unbilled:
        return redirect(url_for("operations_delivery_view", message="No unbilled delivery entries found for that client and period"))

    existing_invoices = load_finance_data().get("sheets", {}).get("Invoices", [])
    issue_date = date.today().isoformat()
    generated_invoice_number = _next_invoice_number(issue_date, existing_invoices)

    raw_line_items = []
    for entry in unbilled:
        hours = entry.get("hours_spent")
        quantity = hours if hours else 1.0
        unit_price = 0.0
        raw_line_items.append({
            "service_id": "",
            "name": entry.get("service_type") or "Delivery",
            "description": entry.get("description") or "",
            "quantity": quantity,
            "unit_price": unit_price,
            "discount_type": "€",
            "discount_value": 0,
            "vat_rate": "0%",
        })

    client_rows = load_finance_data().get("sheets", {}).get("Clients", [])
    client_row = next((row for row in client_rows if row.get("Client Name") == client_name), {})
    retainer_amount = _coerce_number(client_row.get("Retainer Amount (€)"))
    if not retainer_amount:
        return redirect(url_for("operations_delivery_view", message=f"{client_name} has no retainer amount set — add one on the Client record before generating this invoice"))
    raw_line_items[0]["unit_price"] = retainer_amount
    for item in raw_line_items[1:]:
        item["unit_price"] = 0.0

    payload = {
        "Invoice #": generated_invoice_number,
        "Issue Date": issue_date,
        "Due Date": (date.today() + timedelta(days=14)).isoformat(),
        "Client Name": client_name,
        "Client VAT Number": "",
        "Client Address": "",
        "Service / Product": "",
        "VAT Treatment": "standard",
        "Supply Type": "services",
        "Balance Due (€)": "",
        "Status": "Draft",
        "Payment Method": "",
        "Payment Date": "",
        "Bank Reconciliation": "Unreconciled",
        "Notes": f"Generated from Delivery Log — billing period {billing_period}",
        "Phase Tag": _resolve_phase_tag(issue_date),
    }
    _apply_invoice_line_items(payload, raw_line_items)
    _apply_vat_classification(payload, vat_rate_key="VAT Rate")
    _normalize_invoice_balance(payload)
    validation_errors = _validate_invoice_payload(payload)
    if validation_errors:
        return redirect(url_for("operations_delivery_view", message=_build_validation_message(validation_errors)))

    _append_row_to_sheet("Invoices", payload)
    load_finance_data.cache_clear()
    _record_audit("create", "invoice", {"record": payload, "source": "delivery_log"})

    for entry in entries:
        if entry.get("client_name") == client_name and entry.get("billing_period") == billing_period and not entry.get("invoiced"):
            entry["invoiced"] = True
            entry["invoice_id"] = generated_invoice_number
            entry["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
    _save_delivery_log(entries)

    project_ids = {entry.get("project_id") for entry in unbilled if entry.get("project_id")}
    if project_ids:
        projects = _load_projects()
        for project in projects:
            if project["id"] in project_ids and generated_invoice_number not in project["linked_invoice_ids"]:
                project["linked_invoice_ids"].append(generated_invoice_number)
                project["updated_at"] = datetime.now().isoformat(timespec="seconds")
        _save_projects(projects)

    return redirect(url_for("operations_delivery_view", message=f"Invoice {generated_invoice_number} generated"))


# --- Operations: SOP Library -------------------------------------------------

@app.route("/operations/sops")
def operations_sops_view():
    data = load_finance_data()
    sops = _load_sops()
    editing_sop = _find_record_by_id(sops, request.args.get("edit_id"))
    client_filter = str(request.args.get("client_filter") or "")
    status_filter = str(request.args.get("status_filter") or "")
    process_area_filter = str(request.args.get("process_area_filter") or "")
    return render_template(
        "index.html",
        **_build_page_context(
            "SOP Library",
            "operations_sops",
            data,
            editing_sop=editing_sop,
            sop_form={"client_filter": client_filter, "status_filter": status_filter, "process_area_filter": process_area_filter},
            message=request.args.get("message"),
        ),
    )


@app.route("/operations/sops/add", methods=["POST"])
def add_sop():
    client_name = str(request.form.get("client_name") or "").strip()
    title = str(request.form.get("title") or "").strip()
    supersedes_id = str(request.form.get("supersedes_id") or "").strip()

    errors: dict[str, str] = {}
    _validate_required_text(title, "title", "SOP title", errors)
    _validate_required_text(client_name, "client_name", "Client", errors)

    if errors:
        return _redirect_with_form_errors("operations_sops_view", {"title": title, "client_name": client_name}, errors)

    stored_filename, upload_error = _save_uploaded_sop(request.files.get("sop_file"))
    if upload_error:
        return _redirect_with_form_errors("operations_sops_view", {"title": title, "client_name": client_name}, {"sop_file": upload_error})

    sops = _load_sops()
    version = str(request.form.get("version") or "V1.0").strip()
    if supersedes_id:
        previous = _find_record_by_id(sops, supersedes_id)
        if previous is not None:
            previous["status"] = "Superseded"
            previous["last_updated_at"] = datetime.now().isoformat(timespec="seconds")

    now = datetime.now().isoformat(timespec="seconds")
    sop = _normalize_sop({
        "id": str(uuid4()),
        "title": title,
        "client_id": client_name,
        "client_name": client_name,
        "project_id": str(request.form.get("project_id") or "").strip(),
        "version": version,
        "status": "Draft",
        "process_area": str(request.form.get("process_area") or "").strip(),
        "description": str(request.form.get("description") or "").strip(),
        "filename": stored_filename,
        "date_created": date.today().isoformat(),
        "notes": str(request.form.get("notes") or "").strip(),
        "created_at": now,
        "last_updated_at": now,
    })
    sops.append(sop)
    _save_sops(sops)
    _record_audit("create", "sop", {"sop_id": sop["id"], "record": sop})
    return redirect(url_for("operations_sops_view", message="SOP added"))


@app.route("/operations/sops/update", methods=["POST"])
def update_sop():
    sop_id = str(request.form.get("sop_id") or "").strip()
    new_status = str(request.form.get("status") or "").strip()

    sops = _load_sops()
    sop = _find_record_by_id(sops, sop_id)
    if sop is None:
        return redirect(url_for("operations_sops_view", message="SOP not found"))

    if new_status and new_status != sop["status"]:
        if new_status not in SOP_STATUSES:
            return redirect(url_for("operations_sops_view", message="Invalid status"))
        if new_status in SOP_STATUS_WORKFLOW and sop["status"] in SOP_STATUS_WORKFLOW:
            current_index = SOP_STATUS_WORKFLOW.index(sop["status"])
            new_index = SOP_STATUS_WORKFLOW.index(new_status)
            if new_index != current_index + 1:
                return redirect(url_for("operations_sops_view", message="SOPs must move through Draft → Review → Approved in order"))
        sop["status"] = new_status
        if new_status == "Approved":
            sop["date_approved"] = date.today().isoformat()
            sop["approved_by"] = str(request.form.get("approved_by") or "").strip()

    for field_name, form_key in (
        ("title", "title"),
        ("process_area", "process_area"),
        ("description", "description"),
        ("notes", "notes"),
    ):
        if form_key in request.form:
            sop[field_name] = str(request.form.get(form_key) or "").strip()

    sop["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
    _save_sops(sops)
    _record_audit("update", "sop", {"sop_id": sop_id, "record": sop})
    return redirect(url_for("operations_sops_view", message="SOP updated"))


@app.route("/sops/<path:filename>")
def serve_sop_file(filename):
    as_attachment = request.args.get("download") == "1"
    return send_from_directory(SOP_FILES_DIR, filename, as_attachment=as_attachment)


@app.route("/delivery-files/<path:filename>")
def serve_delivery_file(filename):
    as_attachment = request.args.get("download") == "1"
    return send_from_directory(DELIVERY_FILES_DIR, filename, as_attachment=as_attachment)


def _backup_eligible_files() -> list[Path]:
    return list(SHEET_JSON_PATHS.values()) + [
        SUBSCRIPTIONS_PATH,
        ARCHIVE_PATH,
        AUDIT_LOG_PATH,
        BUSINESS_PROFILE_PATH,
        CHART_OF_ACCOUNTS_PATH,
        LEDGER_JOURNAL_PATH,
        CAPITAL_ASSETS_PATH,
        PAYROLL_PATH,
        BANK_STATEMENTS_PATH,
        SERVICES_PATH,
        COMPANY_DOCUMENTS_PATH,
        COMPLIANCE_CALENDAR_PATH,
        PROJECTS_PATH,
        DELIVERY_LOG_PATH,
        SOPS_PATH,
    ]


def _list_available_backups() -> list[dict[str, Any]]:
    if not BACKUPS_DIR.exists():
        return []
    entries = []
    for date_dir in sorted(BACKUPS_DIR.iterdir(), key=lambda entry: entry.name, reverse=True):
        if not date_dir.is_dir():
            continue
        try:
            date.fromisoformat(date_dir.name)
        except ValueError:
            continue
        files = sorted(f.name for f in date_dir.iterdir() if f.is_file())
        if files:
            entries.append({"date": date_dir.name, "files": files})
    return entries


@app.route("/settings/backups")
def backups_view():
    data = load_finance_data()
    return render_template(
        "index.html",
        **_build_page_context(
            "Restore Backups",
            "settings",
            data,
            message=request.args.get("message"),
        ),
        available_backups=_list_available_backups(),
    )


@app.route("/settings/backups/restore", methods=["POST"])
def restore_backup():
    backup_date = str(request.form.get("backup_date") or "").strip()
    filename = str(request.form.get("filename") or "").strip()
    destination_by_name = {path.name: path for path in _backup_eligible_files()}
    if filename not in destination_by_name:
        return redirect(url_for("backups_view", message="Invalid backup file"))
    try:
        date.fromisoformat(backup_date)
    except ValueError:
        return redirect(url_for("backups_view", message="Invalid backup date"))
    source = BACKUPS_DIR / backup_date / filename
    if not source.exists():
        return redirect(url_for("backups_view", message="Backup not found"))
    destination = destination_by_name[filename]
    try:
        shutil.copy2(source, destination)
    except OSError as exc:
        return redirect(url_for("backups_view", message=f"Restore failed: {exc}"))
    load_finance_data.cache_clear()
    _record_audit("restore_backup", "system", {"filename": filename, "backup_date": backup_date})
    return redirect(url_for("backups_view", message=f"Restored {filename} from {backup_date}"))


@app.route("/healthz")
def healthz():
    return Response("ok", mimetype="text/plain")


@app.route("/income")
def income_view():
    _sync_subscriptions_to_expenses()
    data = load_finance_data()
    rows = data["sheets"].get("Income", [])
    invoiced_income = [row for row in rows if str(row.get("Source") or "manual") == "invoiced"]
    manual_income = [row for row in rows if str(row.get("Source") or "manual") != "invoiced"]
    invoiced_income_total = round(sum(_coerce_number(row.get("Total incl. VAT (€)")) for row in invoiced_income), 2)
    manual_income_received_total = round(
        sum(_coerce_number(row.get("Amount (€)")) for row in manual_income if str(row.get("Status") or "").strip().lower() == "received"),
        2,
    )
    validation_errors, income_form = _build_validation_state("income")
    editing_income = _find_row_by_number(rows, request.args.get("edit_row"))
    if editing_income is not None and str(editing_income.get("Source") or "manual") == "invoiced":
        editing_income = None
    return render_template(
        "index.html",
        **_build_page_context("Income", "income", data, income=rows, editing_income=editing_income, income_form=income_form, validation_errors=validation_errors, message=request.args.get("message")),
        invoiced_income=invoiced_income,
        manual_income=manual_income,
        invoiced_income_total=invoiced_income_total,
        manual_income_received_total=manual_income_received_total,
    )


@app.route("/expenses")
def expenses_view():
    sync_result = _sync_subscriptions_to_expenses()
    data = load_finance_data()
    rows = data["sheets"].get("Expenses", [])
    editing_expense = _find_row_by_number(rows, request.args.get("edit_row"))
    phase_filter = str(request.args.get("phase_filter") or "All").strip()
    if phase_filter not in {"All", "Pre-Trading", "Phase 1", "Phase 2"}:
        phase_filter = "All"
    visible_rows = rows if phase_filter == "All" else [row for row in rows if str(row.get("Phase Tag") or "") == phase_filter]
    subscription_rows = _build_subscription_rows(_load_subscriptions())
    validation_errors, expense_form = _build_validation_state("expenses")
    return render_template("index.html", **_build_page_context("Expenses", "expenses", data, expenses=visible_rows, phase_filter=phase_filter, subscription_summary=_summarize_subscriptions(subscription_rows), editing_expense=editing_expense, expense_form=expense_form, validation_errors=validation_errors, message=request.args.get("message"), sync_message=_build_sync_message(sync_result)))


@app.route("/invoices")
def invoices_view():
    _sync_subscriptions_to_expenses()
    data = load_finance_data()
    rows = data["sheets"].get("Invoices", [])
    validation_errors, invoice_form = _build_validation_state("invoices")
    editing_invoice = _find_row_by_number(rows, request.args.get("edit_row"))
    return render_template("index.html", **_build_page_context("Invoices", "invoices", data, invoices=rows, editing_invoice=editing_invoice, invoice_form=invoice_form, validation_errors=validation_errors, message=request.args.get("message")))


@app.route("/services")
def services_view():
    data = load_finance_data()
    services = _load_services()
    validation_errors, service_form = _build_validation_state("services")
    editing_service = _find_service_by_id(services, request.args.get("edit_id"))
    show_archived_services = request.args.get("show_archived") == "1"
    return render_template(
        "index.html",
        **_build_page_context(
            "Services",
            "services",
            data,
            services=services,
            editing_service=editing_service,
            service_form=service_form,
            show_archived_services=show_archived_services,
            validation_errors=validation_errors,
            message=request.args.get("message"),
        ),
    )


@app.route("/clients")
def clients_view():
    _sync_subscriptions_to_expenses()
    data = load_finance_data()
    rows = data["sheets"].get("Clients", [])
    validation_errors, client_form = _build_validation_state("clients")
    editing_client = _find_row_by_number(rows, request.args.get("edit_row"))
    return render_template("index.html", **_build_page_context("Clients", "clients", data, clients=rows, editing_client=editing_client, client_form=client_form, validation_errors=validation_errors, message=request.args.get("message")))


@app.route("/suppliers")
def suppliers_view():
    _sync_subscriptions_to_expenses()
    data = load_finance_data()
    rows = data["sheets"].get("Suppliers", [])
    validation_errors, supplier_form = _build_validation_state("suppliers")
    editing_supplier = _find_row_by_number(rows, request.args.get("edit_row"))
    return render_template("index.html", **_build_page_context("Suppliers", "suppliers", data, suppliers=rows, editing_supplier=editing_supplier, supplier_form=supplier_form, validation_errors=validation_errors, message=request.args.get("message")))


@app.route("/subscriptions")
def subscriptions_view():
    sync_result = _sync_subscriptions_to_expenses()
    data = load_finance_data()
    subscription_rows = _build_subscription_rows(_load_subscriptions())
    subscription_summary = _summarize_subscriptions(subscription_rows)
    validation_errors, subscription_form = _build_validation_state("subscriptions")
    editing_subscription = _find_subscription_by_id(subscription_rows, request.args.get("edit_id"))
    return render_template(
        "index.html",
        **_build_page_context(
            "Subscriptions",
            "subscriptions",
            data if "error" not in data else {},
            subscriptions=subscription_rows,
            subscription_summary=subscription_summary,
            editing_subscription=editing_subscription,
            subscription_form=subscription_form,
            validation_errors=validation_errors,
            message=request.args.get("message"),
            sync_message=_build_sync_message(sync_result),
            error=data.get("error"),
        ),
    )


@app.route("/payroll")
def payroll_view():
    _sync_subscriptions_to_expenses()
    data = load_finance_data()
    payroll_entries = _load_payroll_entries()
    validation_errors, payroll_form = _build_validation_state("payroll")
    editing_payroll = _find_payroll_by_id(payroll_entries, request.args.get("edit_id"))
    return render_template(
        "index.html",
        **_build_page_context(
            "Payroll",
            "payroll",
            data if "error" not in data else {},
            payroll=payroll_entries,
            payroll_summary=_summarize_payroll_entries(payroll_entries),
            editing_payroll=editing_payroll,
            payroll_form=payroll_form,
            validation_errors=validation_errors,
            message=request.args.get("message"),
            error=data.get("error"),
        ),
    )


@app.route("/archive")
def archive_view():
    _sync_subscriptions_to_expenses()
    data = load_finance_data()
    archive_records = sorted(_load_archives(), key=lambda record: str(record.get("archived_at") or ""), reverse=True)
    audit_entries = sorted(_load_audit_entries(), key=lambda record: str(record.get("timestamp") or ""), reverse=True)
    return render_template(
        "index.html",
        **_build_page_context(
            "Archive & Audit",
            "archive",
            data if "error" not in data else {},
            archived_records=archive_records[:50],
            archive_summary=_summarize_archives(archive_records),
            audit_entries=audit_entries[:50],
            error=data.get("error"),
            message=request.args.get("message"),
        ),
    )


@app.route("/ledger")
def ledger_view():
    _sync_subscriptions_to_expenses()
    data = load_finance_data()
    return render_template(
        "index.html",
        **_build_page_context(
            "Chart of Accounts & Ledger",
            "ledger",
            data if "error" not in data else {},
            error=data.get("error"),
            message=request.args.get("message"),
        ),
    )


@app.route("/ledger/trial-balance.csv")
def export_trial_balance_csv():
    accounts = _ensure_chart_of_accounts()
    ledger_entries = _load_ledger_entries()
    trial_balance = _compute_trial_balance(ledger_entries, accounts)
    csv_text = _export_trial_balance_csv(trial_balance)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=trial-balance.csv"},
    )


@app.route("/capital-allowances/export.csv")
def export_capital_allowances_csv():
    assets = _load_capital_assets()
    csv_text = _export_capital_allowances_csv(assets)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=capital-allowances.csv"},
    )


@app.route("/ledger/journal.csv")
def export_ledger_journal_csv():
    ledger_entries = _load_ledger_entries()
    csv_text = _export_ledger_journal_csv(ledger_entries)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=ledger-journal.csv"},
    )


@app.route("/vat3/export.csv")
def export_vat3_csv():
    vat_summary = _compute_vat_control_summary(_load_ledger_entries())
    csv_text = _export_vat3_csv(vat_summary)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=vat3-summary.csv"},
    )


@app.route("/audit/export.csv")
def export_audit_csv():
    entries = sorted(_load_audit_entries(), key=lambda record: str(record.get("timestamp") or ""), reverse=True)
    csv_text = _export_audit_entries_csv(entries)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=audit-log.csv"},
    )


@app.route("/payroll/export.csv")
def export_payroll_csv():
    entries = _load_payroll_entries()
    csv_text = _export_payroll_csv(entries)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=payroll-register.csv"},
    )


@app.route("/reconciliation/export.csv")
def export_reconciliation_csv():
    data = load_finance_data()
    payroll_entries = _load_payroll_entries()
    rows = _build_reconciliation_rows(data, payroll_entries)
    _match_bank_statement_lines(rows, _load_bank_statement_lines())
    csv_text = _export_reconciliation_csv(rows)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=reconciliation-queue.csv"},
    )


@app.route("/reconciliation/exceptions.csv")
def export_reconciliation_exceptions_csv():
    data = load_finance_data()
    payroll_entries = _load_payroll_entries()
    rows = _build_reconciliation_rows(data, payroll_entries)
    _match_bank_statement_lines(rows, _load_bank_statement_lines())
    csv_text = _export_reconciliation_csv(rows, exceptions_only=True)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=reconciliation-exceptions.csv"},
    )


@app.route("/reconciliation/bank-statements.csv")
def export_bank_statements_csv():
    statement_lines = _load_bank_statement_lines()
    data = load_finance_data()
    payroll_entries = _load_payroll_entries()
    reconciliation_rows = _build_reconciliation_rows(data, payroll_entries)
    matched_lines, _ = _match_bank_statement_lines(reconciliation_rows, statement_lines)
    csv_text = _export_bank_statement_csv(matched_lines)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=bank-statements.csv"},
    )


@app.route("/reconciliation/unmatched-bank-statements.csv")
def export_unmatched_bank_statements_csv():
    statement_lines = _load_bank_statement_lines()
    data = load_finance_data()
    payroll_entries = _load_payroll_entries()
    reconciliation_rows = _build_reconciliation_rows(data, payroll_entries)
    matched_lines, _ = _match_bank_statement_lines(reconciliation_rows, statement_lines)
    csv_text = _export_bank_statement_csv(matched_lines, unmatched_only=True)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=unmatched-bank-statements.csv"},
    )


@app.route("/archive/restore", methods=["POST"])
def restore_archive_record():
    archive_id = str(request.form.get("archive_id") or "").strip()
    force_restore = str(request.form.get("force_restore") or "").strip().lower() in {"1", "true", "yes"}
    archive_entry = _find_json_record(ARCHIVE_PATH, archive_id)
    if archive_entry is None:
        return redirect(url_for("archive_view", message="Archive record could not be restored"))

    entity_type = str(archive_entry.get("entity_type") or "").strip()
    conflict = _find_restore_conflict(entity_type, archive_entry)
    if conflict is not None and not force_restore:
        _record_audit("restore_conflict", entity_type or "archive", {"archive_id": archive_id, "record": archive_entry.get("record", {}), "conflict": conflict})
        return redirect(url_for("archive_view", message=f"Restore conflict detected for {entity_type}. Use Force Restore if you want to restore it anyway."))

    archive_entry = _pop_json_record(ARCHIVE_PATH, archive_id)
    if archive_entry is None:
        return redirect(url_for("archive_view", message="Archive record could not be restored"))

    if entity_type == "subscription":
        _restore_subscription_archive(archive_entry)
    elif entity_type == "payroll":
        _restore_payroll_archive(archive_entry)
    elif entity_type in WORKBOOK_ENTITY_CONFIG:
        _restore_workbook_archive(entity_type, archive_entry)
    else:
        _append_json_record(ARCHIVE_PATH, archive_entry)
        return redirect(url_for("archive_view", message="Archive record type is not supported for restore"))

    return redirect(url_for("archive_view", message=f"{entity_type.title()} restored"))


@app.route("/refresh", methods=["POST"])
def refresh_workbook():
    load_finance_data.cache_clear()
    next_page = request.args.get("next", "/")
    return redirect(next_page or "/")


@app.route("/business-structure/update", methods=["POST"])
def update_business_structure():
    structure = _normalize_business_structure(request.form.get("structure"))
    transition_date = str(request.form.get("transition_date") or "").strip()
    pre_trading_start_date = str(request.form.get("pre_trading_start_date") or "").strip()
    vat_registered = str(request.form.get("vat_registered") or "").strip().lower() in {"1", "true", "yes", "on"}
    vat_threshold_basis = _normalize_vat_threshold_basis(request.form.get("vat_threshold_basis"))
    if transition_date and _parse_transaction_date(transition_date) is None:
        next_page = str(request.args.get("next") or "/")
        return redirect(_append_message_to_path(next_page, "Invalid transition date format"))
    if pre_trading_start_date and _parse_transaction_date(pre_trading_start_date) is None:
        next_page = str(request.args.get("next") or "/")
        return redirect(_append_message_to_path(next_page, "Invalid pre-trading start date format"))

    profile = _load_business_profile()
    previous_profile = dict(profile)
    profile["structure"] = structure
    profile["transition_date"] = transition_date
    profile["pre_trading_start_date"] = pre_trading_start_date
    profile["vat_registered"] = vat_registered
    profile["vat_threshold_basis"] = vat_threshold_basis
    _save_business_profile(profile)
    _record_audit(
        "business_structure_update",
        "settings",
        {
            "from": {
                "structure": previous_profile.get("structure"),
                "transition_date": previous_profile.get("transition_date"),
                "vat_registered": bool(previous_profile.get("vat_registered", False)),
                "vat_threshold_basis": _normalize_vat_threshold_basis(previous_profile.get("vat_threshold_basis")),
            },
            "to": {
                "structure": structure,
                "transition_date": transition_date,
                "vat_registered": vat_registered,
                "vat_threshold_basis": vat_threshold_basis,
            },
        },
    )

    next_page = str(request.args.get("next") or "/")
    return redirect(_append_message_to_path(next_page, "Business structure updated"))


@app.route("/subscriptions/sync", methods=["POST"])
def sync_subscriptions():
    sync_result = _sync_subscriptions_to_expenses()
    posted_count = sync_result.get("posted_count", 0)
    if posted_count == 1:
        message = "1 subscription charge posted to expenses"
    elif posted_count > 1:
        message = f"{posted_count} subscription charges posted to expenses"
    else:
        message = "No subscription charges were due"
    _record_audit("sync", "subscriptions", {"posted_count": posted_count})
    return redirect(url_for("subscriptions_view", message=message))


@app.route("/reconciliation/import-statement", methods=["POST"])
def import_bank_statement():
    return_to = str(request.form.get("return_to") or "/ledger")
    statement_file = request.files.get("statement_file")
    if statement_file is None or not str(statement_file.filename or "").strip():
        return redirect(_append_message_to_path(return_to, "Bank statement import failed: file is required"))

    try:
        content = statement_file.read().decode("utf-8-sig", errors="ignore")
    except OSError:
        return redirect(_append_message_to_path(return_to, "Bank statement import failed: could not read file"))

    result = _ingest_bank_statement_csv(content, source_filename=str(statement_file.filename or "statement.csv"))
    imported_count = result.get("imported_count", 0)
    skipped_count = result.get("skipped_count", 0)
    _record_audit(
        "import",
        "bank_statement",
        {
            "source_filename": str(statement_file.filename or "statement.csv"),
            "imported_count": imported_count,
            "skipped_count": skipped_count,
        },
    )
    return redirect(_append_message_to_path(return_to, f"Bank statement imported: {imported_count} new lines, {skipped_count} skipped"))


@app.route("/reconciliation/apply-suggested", methods=["POST"])
def apply_suggested_reconciliation():
    return_to = str(request.form.get("return_to") or "/ledger")
    data = load_finance_data()
    payroll_entries = _load_payroll_entries()
    reconciliation_rows = _build_reconciliation_rows(data, payroll_entries)
    _match_bank_statement_lines(reconciliation_rows, _load_bank_statement_lines())

    duplicate_key_counts: dict[tuple[str, float, str], int] = {}
    for item in reconciliation_rows:
        if not bool(item.get("is_paid")):
            continue
        if item.get("bank_reconciliation") != "Unreconciled":
            continue
        item_key = (
            str(item.get("date") or ""),
            round(_coerce_number(item.get("amount_eur")), 2),
            str(item.get("payment_method") or "").strip().lower(),
        )
        if not item_key[0] or item_key[1] <= 0:
            continue
        duplicate_key_counts[item_key] = duplicate_key_counts.get(item_key, 0) + 1

    expenses_marked = 0
    invoices_marked = 0
    payroll_marked = 0
    workbook_changed = False
    payroll_changed = False

    for row in reconciliation_rows:
        if not bool(row.get("is_paid")):
            continue
        if row.get("bank_reconciliation") == "Reconciled":
            continue
        if int(row.get("statement_match_count") or 0) <= 0:
            continue
        row_key = (
            str(row.get("date") or ""),
            round(_coerce_number(row.get("amount_eur")), 2),
            str(row.get("payment_method") or "").strip().lower(),
        )
        if duplicate_key_counts.get(row_key, 0) > 1:
            continue
        # Batch apply should only reconcile unambiguous suggestions.
        if int(row.get("matching_group_size") or 1) > 1:
            continue
        if int(row.get("statement_match_count") or 0) != 1:
            continue

        entity_type = str(row.get("entity_type") or "")
        if entity_type in {"expense", "invoice"}:
            row_number = _parse_row_number(row.get("row_number"))
            if row_number is None:
                continue
            sheet_name = "Expenses" if entity_type == "expense" else "Invoices"
            record = _find_sheet_row_or_raise(sheet_name, row_number)
            if _normalize_reconciliation(record.get("Bank Reconciliation")) == "Reconciled":
                continue
            record["Bank Reconciliation"] = "Reconciled"
            _update_row_in_sheet(sheet_name, row_number, record)
            workbook_changed = True
            if entity_type == "expense":
                expenses_marked += 1
            else:
                invoices_marked += 1
            continue

        if entity_type == "payroll":
            payroll_id = str(row.get("payroll_id") or "")
            payroll_entry = _find_payroll_by_id(payroll_entries, payroll_id)
            if payroll_entry is None:
                continue
            if _normalize_reconciliation(payroll_entry.get("Bank Reconciliation")) == "Reconciled":
                continue
            payroll_entry["Bank Reconciliation"] = "Reconciled"
            payroll_entry["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
            payroll_changed = True
            payroll_marked += 1

    if workbook_changed:
        load_finance_data.cache_clear()
    if payroll_changed:
        _save_payroll_entries(payroll_entries)

    total_marked = expenses_marked + invoices_marked + payroll_marked
    _record_audit(
        "reconcile_batch",
        "reconciliation",
        {
            "expenses_marked": expenses_marked,
            "invoices_marked": invoices_marked,
            "payroll_marked": payroll_marked,
            "total_marked": total_marked,
        },
    )
    if total_marked == 0:
        return redirect(_append_message_to_path(return_to, "No suggested matches were available to apply"))
    return redirect(_append_message_to_path(return_to, f"Applied {total_marked} suggested reconciliation matches"))


@app.route("/reconciliation/mark", methods=["POST"])
def mark_reconciliation_status():
    entity_type = str(request.form.get("entity_type") or "").strip().lower()
    target_status = _normalize_reconciliation(request.form.get("bank_reconciliation"))
    return_to = str(request.form.get("return_to") or "/ledger")

    def _reconciled_mark_block_reason(record: dict[str, Any], entity: str) -> str | None:
        if target_status != "Reconciled":
            return None
        if not _is_paid_status(entity, record.get("Status")):
            return "Record must have a paid status before it can be reconciled"
        has_payment_method_column = "Payment Method" in record
        has_payment_date_column = "Payment Date" in record
        if has_payment_method_column and not str(record.get("Payment Method") or "").strip():
            return "Payment method is required before a paid record can be reconciled"
        if entity in {"invoice", "payroll"} and has_payment_date_column and _parse_iso_date(record.get("Payment Date")) is None:
            source_key = "Issue Date" if entity == "invoice" else "Pay Date"
            fallback_date = str(record.get(source_key) or "").strip()
            if _parse_iso_date(fallback_date) is None:
                return "Payment date is required before a paid record can be reconciled"
            record["Payment Date"] = fallback_date
        return None

    if entity_type in {"expense", "invoice"}:
        row_number = _parse_row_number(request.form.get("row_number"))
        if row_number is None:
            return redirect(_append_message_to_path(return_to, "Reconciliation update failed: missing row number"))
        sheet_name = "Expenses" if entity_type == "expense" else "Invoices"
        row = _find_sheet_row_or_raise(sheet_name, row_number)
        block_reason = _reconciled_mark_block_reason(row, entity_type)
        if block_reason:
            return redirect(_append_message_to_path(return_to, f"Reconciliation update failed: {block_reason}"))
        row["Bank Reconciliation"] = target_status
        _update_row_in_sheet(sheet_name, row_number, row)
        load_finance_data.cache_clear()
        _record_audit("reconcile", entity_type, {"row_number": row_number, "bank_reconciliation": target_status})
        return redirect(_append_message_to_path(return_to, f"{entity_type.title()} reconciliation updated"))

    if entity_type == "payroll":
        payroll_id = str(request.form.get("payroll_id") or "").strip()
        payroll_entries = _load_payroll_entries()
        payroll_entry = _find_payroll_by_id(payroll_entries, payroll_id)
        if payroll_entry is None:
            return redirect(_append_message_to_path(return_to, "Reconciliation update failed: payroll entry not found"))
        block_reason = _reconciled_mark_block_reason(payroll_entry, "payroll")
        if block_reason:
            return redirect(_append_message_to_path(return_to, f"Reconciliation update failed: {block_reason}"))
        payroll_entry["Bank Reconciliation"] = target_status
        payroll_entry["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
        _save_payroll_entries(payroll_entries)
        _record_audit("reconcile", "payroll", {"payroll_id": payroll_id, "bank_reconciliation": target_status})
        return redirect(_append_message_to_path(return_to, "Payroll reconciliation updated"))

    return redirect(_append_message_to_path(return_to, "Reconciliation update failed: unsupported entity type"))


@app.route("/payroll/add", methods=["POST"])
def add_payroll():
    payload = {
        "id": str(uuid4()),
        "Pay Date": request.form.get("pay_date", ""),
        "Payroll Period": request.form.get("payroll_period", ""),
        "Employee Name": request.form.get("employee_name", ""),
        "Gross Pay (€)": request.form.get("gross_pay", ""),
        "PAYE (€)": request.form.get("paye", ""),
        "USC (€)": request.form.get("usc", ""),
        "Employee PRSI (€)": request.form.get("employee_prsi", ""),
        "Employer PRSI (€)": request.form.get("employer_prsi", ""),
        "Net Pay (€)": request.form.get("net_pay", ""),
        "Employer Cost (€)": request.form.get("employer_cost", ""),
        "Status": request.form.get("status", "Draft"),
        "Payment Method": request.form.get("payment_method", ""),
        "Payment Date": request.form.get("payment_date", ""),
        "Bank Reconciliation": request.form.get("bank_reconciliation", "Unreconciled"),
        "Notes": request.form.get("notes", ""),
        "Phase Tag": _resolve_phase_tag(request.form.get("pay_date", "")),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "last_updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    _normalize_payroll_payload(payload)
    payment_date_autofilled = _apply_default_payment_date_for_paid(payload, "payroll", "Pay Date")
    validation_errors = _validate_payroll_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "payroll_view",
            {
                "pay_date": request.form.get("pay_date", ""),
                "payroll_period": request.form.get("payroll_period", ""),
                "employee_name": request.form.get("employee_name", ""),
                "gross_pay": request.form.get("gross_pay", ""),
                "paye": request.form.get("paye", ""),
                "usc": request.form.get("usc", ""),
                "employee_prsi": request.form.get("employee_prsi", ""),
                "employer_prsi": request.form.get("employer_prsi", ""),
                "status": request.form.get("status", "Draft"),
                "payment_method": request.form.get("payment_method", ""),
                "payment_date": request.form.get("payment_date", ""),
                "bank_reconciliation": request.form.get("bank_reconciliation", "Unreconciled"),
                "notes": request.form.get("notes", ""),
            },
            validation_errors,
            validation_tab="payroll",
        )

    payroll_entries = _load_payroll_entries()
    payroll_entries.append(payload)
    _save_payroll_entries(payroll_entries)
    _record_audit("create", "payroll", {"payroll_id": payload["id"], "record": payload})
    _record_ledger_entry("create", "payroll", payload, source="payroll", row_number=None)
    message = "Payroll entry added"
    if payment_date_autofilled:
        message = "Payroll entry added (payment date auto-filled from pay date)"
    return redirect(url_for("payroll_view", message=message))


@app.route("/payroll/update", methods=["POST"])
def update_payroll():
    payroll_id = str(request.form.get("payroll_id") or "").strip()
    payroll_entries = _load_payroll_entries()
    existing = _find_payroll_by_id(payroll_entries, payroll_id)
    if existing is None:
        return redirect(url_for("payroll_view", message="Payroll entry could not be updated"))

    payload = {
        "id": existing.get("id") or payroll_id,
        "Pay Date": request.form.get("pay_date", existing.get("Pay Date", "")),
        "Payroll Period": request.form.get("payroll_period", existing.get("Payroll Period", "")),
        "Employee Name": request.form.get("employee_name", existing.get("Employee Name", "")),
        "Gross Pay (€)": request.form.get("gross_pay", existing.get("Gross Pay (€)", "")),
        "PAYE (€)": request.form.get("paye", existing.get("PAYE (€)", "")),
        "USC (€)": request.form.get("usc", existing.get("USC (€)", "")),
        "Employee PRSI (€)": request.form.get("employee_prsi", existing.get("Employee PRSI (€)", "")),
        "Employer PRSI (€)": request.form.get("employer_prsi", existing.get("Employer PRSI (€)", "")),
        "Net Pay (€)": request.form.get("net_pay", existing.get("Net Pay (€)", "")),
        "Employer Cost (€)": request.form.get("employer_cost", existing.get("Employer Cost (€)", "")),
        "Status": request.form.get("status", existing.get("Status", "Draft")),
        "Payment Method": request.form.get("payment_method", existing.get("Payment Method", "")),
        "Payment Date": request.form.get("payment_date", existing.get("Payment Date", "")),
        "Bank Reconciliation": request.form.get("bank_reconciliation", existing.get("Bank Reconciliation", "Unreconciled")),
        "Notes": request.form.get("notes", existing.get("Notes", "")),
        "Phase Tag": _resolve_phase_tag(request.form.get("pay_date", existing.get("Pay Date", ""))),
        "created_at": str(existing.get("created_at") or datetime.now().isoformat(timespec="seconds")),
        "last_updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    _normalize_payroll_payload(payload)
    payment_date_autofilled = _apply_default_payment_date_for_paid(payload, "payroll", "Pay Date")
    validation_errors = _validate_payroll_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "payroll_view",
            {
                "pay_date": payload.get("Pay Date", ""),
                "payroll_period": payload.get("Payroll Period", ""),
                "employee_name": payload.get("Employee Name", ""),
                "gross_pay": payload.get("Gross Pay (€)", ""),
                "paye": payload.get("PAYE (€)", ""),
                "usc": payload.get("USC (€)", ""),
                "employee_prsi": payload.get("Employee PRSI (€)", ""),
                "employer_prsi": payload.get("Employer PRSI (€)", ""),
                "status": payload.get("Status", "Draft"),
                "payment_method": payload.get("Payment Method", ""),
                "payment_date": payload.get("Payment Date", ""),
                "bank_reconciliation": payload.get("Bank Reconciliation", "Unreconciled"),
                "notes": payload.get("Notes", ""),
            },
            validation_errors,
            validation_tab="payroll",
            edit_id=payroll_id,
        )

    existing.update(payload)
    _save_payroll_entries(payroll_entries)
    _record_audit("update", "payroll", {"payroll_id": payroll_id, "record": existing})
    _record_ledger_entry("update", "payroll", existing, source="payroll", row_number=None)
    message = "Payroll entry updated"
    if payment_date_autofilled:
        message = "Payroll entry updated (payment date auto-filled from pay date)"
    return redirect(url_for("payroll_view", message=message))


@app.route("/payroll/delete", methods=["POST"])
def delete_payroll():
    payroll_id = str(request.form.get("payroll_id") or "").strip()
    payroll_entries = _load_payroll_entries()
    existing = _find_payroll_by_id(payroll_entries, payroll_id)
    if existing is None:
        return redirect(url_for("payroll_view", message="Payroll entry could not be removed"))

    _archive_record("payroll", existing, source="payroll")
    remaining = [entry for entry in payroll_entries if str(entry.get("id") or "") != payroll_id]
    _save_payroll_entries(remaining)
    _record_ledger_entry("archive", "payroll", existing, source="payroll", row_number=None)
    return redirect(url_for("payroll_view", message="Payroll entry archived"))


@app.route("/income/add", methods=["POST"])
def add_income():
    status = request.form.get("status", "Received")
    if status not in INCOME_STATUS_OPTIONS:
        status = "Received"
    payload = {
        "Date": request.form.get("date", ""),
        "Description": request.form.get("description", ""),
        "Client / Source": request.form.get("client_source", ""),
        "Category": request.form.get("category", ""),
        "Invoice #": "",
        "Invoice ID": "",
        "Source": "manual",
        "Amount (€)": request.form.get("amount", ""),
        "Total incl. VAT (€)": request.form.get("total_incl_vat", ""),
        "VAT Rate": request.form.get("vat_rate", "0%"),
        "VAT Amount (€)": request.form.get("vat_amount", ""),
        "VAT Treatment": request.form.get("vat_treatment", "standard"),
        "Supply Type": request.form.get("supply_type", "services"),
        "Status": status,
        "Payment Method": request.form.get("payment_method", ""),
        "Payment Date": request.form.get("payment_date", ""),
        "Notes": request.form.get("notes", ""),
        "Phase Tag": _resolve_phase_tag(request.form.get("date", "")),
    }
    _normalize_vat_fields(
        payload,
        net_key="Amount (€)",
        total_key="Total incl. VAT (€)",
        vat_rate_key="VAT Rate",
        vat_amount_key="VAT Amount (€)",
        vat_registered=_is_vat_registered(),
    )
    _apply_vat_classification(payload, vat_rate_key="VAT Rate")
    compliance_flags, flags_acknowledged, flags_acknowledged_at = _process_compliance_flags(request.form)
    validation_errors = _validate_income_payload(payload)
    validation_errors.update(_apply_compliance_flags_to_payload(payload, compliance_flags, flags_acknowledged, flags_acknowledged_at))
    if validation_errors:
        return _redirect_with_form_errors(
            "income_view",
            _build_workbook_form_data(payload, {
                "date": "Date",
                "description": "Description",
                "client_source": "Client / Source",
                "category": "Category",
                "amount": "Amount (€)",
                "total_incl_vat": "Total incl. VAT (€)",
                "vat_rate": "VAT Rate",
                "vat_amount": "VAT Amount (€)",
                "vat_treatment": "VAT Treatment",
                "supply_type": "Supply Type",
                "status": "Status",
                "payment_method": "Payment Method",
                "payment_date": "Payment Date",
                "notes": "Notes",
            }),
            validation_errors,
            validation_tab="income",
        )
    row_number = _append_row_to_sheet("Income", payload)
    load_finance_data.cache_clear()
    _record_audit("create", "income", {"row_number": row_number, "record": payload})
    _record_ledger_entry("create", "income", payload, source="workbook", row_number=row_number)
    return redirect(url_for("income_view", message="Income entry added"))


@app.route("/income/update", methods=["POST"])
def update_income():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("income_view", message="Income entry could not be updated"))

    current_row = _find_sheet_row_or_raise("Income", row_number)
    if str(current_row.get("Source") or "manual") == "invoiced":
        return redirect(url_for("income_view", message="This entry is linked to an invoice — edit the invoice instead"))

    status = request.form.get("status", current_row.get("Status", "Received"))
    if status not in INCOME_STATUS_OPTIONS:
        status = "Received"
    payload = {
        "Date": request.form.get("date", ""),
        "Description": request.form.get("description", ""),
        "Client / Source": request.form.get("client_source", ""),
        "Category": request.form.get("category", ""),
        "Invoice #": "",
        "Invoice ID": "",
        "Source": "manual",
        "Amount (€)": request.form.get("amount", ""),
        "Total incl. VAT (€)": request.form.get("total_incl_vat", ""),
        "VAT Rate": request.form.get("vat_rate", "0%"),
        "VAT Amount (€)": request.form.get("vat_amount", ""),
        "VAT Treatment": request.form.get("vat_treatment", "standard"),
        "Supply Type": request.form.get("supply_type", "services"),
        "Status": status,
        "Payment Method": request.form.get("payment_method", ""),
        "Payment Date": request.form.get("payment_date", ""),
        "Notes": request.form.get("notes", ""),
        "Phase Tag": _resolve_phase_tag(request.form.get("date", "")),
    }
    _normalize_vat_fields(
        payload,
        net_key="Amount (€)",
        total_key="Total incl. VAT (€)",
        vat_rate_key="VAT Rate",
        vat_amount_key="VAT Amount (€)",
        vat_registered=_is_vat_registered(),
    )
    _apply_vat_classification(payload, vat_rate_key="VAT Rate")
    compliance_flags, flags_acknowledged, flags_acknowledged_at = _process_compliance_flags(request.form)
    validation_errors = _validate_income_payload(payload)
    validation_errors.update(_apply_compliance_flags_to_payload(payload, compliance_flags, flags_acknowledged, flags_acknowledged_at))
    if validation_errors:
        return _redirect_with_form_errors(
            "income_view",
            _build_workbook_form_data(payload, {
                "date": "Date",
                "description": "Description",
                "client_source": "Client / Source",
                "category": "Category",
                "amount": "Amount (€)",
                "total_incl_vat": "Total incl. VAT (€)",
                "vat_rate": "VAT Rate",
                "vat_amount": "VAT Amount (€)",
                "vat_treatment": "VAT Treatment",
                "supply_type": "Supply Type",
                "status": "Status",
                "payment_method": "Payment Method",
                "payment_date": "Payment Date",
                "notes": "Notes",
            }),
            validation_errors,
            validation_tab="income",
            edit_row=row_number,
        )
    _update_row_in_sheet("Income", row_number, payload)
    load_finance_data.cache_clear()
    _record_audit("update", "income", {"row_number": row_number, "record": payload})
    _record_ledger_entry("update", "income", payload, source="workbook", row_number=row_number)
    return redirect(url_for("income_view", message="Income entry updated"))


@app.route("/income/delete", methods=["POST"])
def delete_income():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("income_view", message="Income entry could not be removed"))

    row = _find_sheet_row_or_raise("Income", row_number)
    if str(row.get("Source") or "manual") == "invoiced":
        return redirect(url_for("income_view", message="This entry is linked to an invoice — edit the invoice instead"))
    _archive_record("income", row, source="workbook")
    _delete_row_from_sheet("Income", row_number)
    load_finance_data.cache_clear()
    _record_ledger_entry("archive", "income", row, source="workbook", row_number=row_number)
    return redirect(url_for("income_view", message="Income entry archived"))


@app.route("/expenses/add", methods=["POST"])
def add_expense():
    payload = {
        "Date (Registered)": request.form.get("date", ""),
        "Title": request.form.get("title", ""),
        "Description": request.form.get("description", ""),
        "Supplier / Payee": request.form.get("supplier", ""),
        "One-Off Payee": "Yes" if request.form.get("one_off_payee") == "Yes" else "No",
        "Supplier VAT Number": request.form.get("supplier_vat_number", ""),
        "Receipt / Invoice Ref": request.form.get("receipt_reference", ""),
        "Receipt Filename": "",
        "Category": request.form.get("category", ""),
        "Base Net Amount (€)": request.form.get("base_net_amount", request.form.get("net_amount", "")),
        "Delivery (€)": request.form.get("delivery_amount", ""),
        "Fees (€)": request.form.get("fees_amount", ""),
        "Other Charges (€)": request.form.get("other_charges_amount", ""),
        "Discount Type": request.form.get("discount_type", "€"),
        "Discount Value": request.form.get("discount_value", ""),
        "Net Amount (€)": request.form.get("net_amount", ""),
        "Total (€)": request.form.get("total_amount", ""),
        "VAT Rate": request.form.get("vat_rate", "0%"),
        "VAT Amount (€)": request.form.get("vat_amount", ""),
        "VAT Treatment": request.form.get("vat_treatment", "standard"),
        "Supply Type": request.form.get("supply_type", "services"),
        "Input VAT Reclaimable": request.form.get("input_vat_reclaimable", "Yes"),
        "Deductibility Status": request.form.get("deductibility_status", ""),
        "Capital Expenditure Flag": request.form.get("capital_expenditure_flag", ""),
        "Receipt Attached": request.form.get("receipt_attached", "No"),
        "Bank Reconciliation": request.form.get("bank_reconciliation", "Unreconciled"),
        "Status": _normalize_expense_status(request.form.get("status", "Pending")),
        "Payment Method": request.form.get("payment_method", ""),
        "Notes": request.form.get("notes", ""),
        "Phase Tag": _resolve_phase_tag(request.form.get("date", "")),
    }
    uploaded_receipt_name = _save_uploaded_receipt(request.files.get("receipt_file"))
    if uploaded_receipt_name:
        payload["Receipt Filename"] = uploaded_receipt_name
        payload["Receipt Attached"] = "Yes"
    _form_total = payload.get("Total (€)", "")
    _form_vat = payload.get("VAT Amount (€)", "")
    _apply_expense_amount_breakdown(payload)
    payload["Total (€)"] = _form_total
    payload["VAT Amount (€)"] = _form_vat
    _normalize_vat_fields(
        payload,
        net_key="Net Amount (€)",
        total_key="Total (€)",
        vat_rate_key="VAT Rate",
        vat_amount_key="VAT Amount (€)",
        vat_registered=_is_vat_registered(),
    )
    _apply_vat_classification(payload, vat_rate_key="VAT Rate")
    _apply_expense_compliance_fields(payload)
    compliance_flags, flags_acknowledged, flags_acknowledged_at = _process_compliance_flags(request.form)
    _apply_pretrading_compliance_flag(payload, compliance_flags)
    validation_errors = _validate_expense_payload(payload)
    validation_errors.update(_apply_compliance_flags_to_payload(payload, compliance_flags, flags_acknowledged, flags_acknowledged_at))
    if validation_errors:
        return _redirect_with_form_errors(
            "expenses_view",
            _build_workbook_form_data(payload, {
                "date": "Date (Registered)",
                "title": "Title",
                "description": "Description",
                "supplier": "Supplier / Payee",
                "supplier_vat_number": "Supplier VAT Number",
                "receipt_reference": "Receipt / Invoice Ref",
                "category": "Category",
                "base_net_amount": "Base Net Amount (€)",
                "delivery_amount": "Delivery (€)",
                "fees_amount": "Fees (€)",
                "other_charges_amount": "Other Charges (€)",
                "discount_type": "Discount Type",
                "discount_value": "Discount Value",
                "net_amount": "Net Amount (€)",
                "total_amount": "Total (€)",
                "vat_rate": "VAT Rate",
                "vat_amount": "VAT Amount (€)",
                "vat_treatment": "VAT Treatment",
                "supply_type": "Supply Type",
                "input_vat_reclaimable": "Input VAT Reclaimable",
                "deductibility_status": "Deductibility Status",
                "capital_expenditure_flag": "Capital Expenditure Flag",
                "receipt_attached": "Receipt Attached",
                "bank_reconciliation": "Bank Reconciliation",
                "status": "Status",
                "payment_method": "Payment Method",
                "notes": "Notes",
            }),
            validation_errors,
            validation_tab="expenses",
        )
    row_number = _append_row_to_sheet("Expenses", payload)
    _upsert_capital_asset_from_expense(payload, row_number, active=payload.get("Capital Expenditure Flag") == "Yes")
    load_finance_data.cache_clear()
    _record_audit("create", "expense", {"row_number": row_number, "record": payload})
    _record_ledger_entry("create", "expense", payload, source="workbook", row_number=row_number)
    return redirect(url_for("expenses_view", message="Expense entry added"))


@app.route("/expenses/update", methods=["POST"])
def update_expense():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("expenses_view", message="Expense could not be updated"))

    existing_expense_row = _find_row_by_number(_load_sheet_rows_with_row_numbers("Expenses"), row_number) or {}
    payload = {
        "Date (Registered)": request.form.get("date", ""),
        "Title": request.form.get("title", ""),
        "Description": request.form.get("description", ""),
        "Supplier / Payee": request.form.get("supplier", ""),
        "One-Off Payee": "Yes" if request.form.get("one_off_payee") == "Yes" else "No",
        "Supplier VAT Number": request.form.get("supplier_vat_number", ""),
        "Receipt / Invoice Ref": request.form.get("receipt_reference", ""),
        "Receipt Filename": existing_expense_row.get("Receipt Filename", ""),
        "Category": request.form.get("category", ""),
        "Base Net Amount (€)": request.form.get("base_net_amount", request.form.get("net_amount", "")),
        "Delivery (€)": request.form.get("delivery_amount", ""),
        "Fees (€)": request.form.get("fees_amount", ""),
        "Other Charges (€)": request.form.get("other_charges_amount", ""),
        "Discount Type": request.form.get("discount_type", "€"),
        "Discount Value": request.form.get("discount_value", ""),
        "Net Amount (€)": request.form.get("net_amount", ""),
        "Total (€)": request.form.get("total_amount", ""),
        "VAT Rate": request.form.get("vat_rate", "0%"),
        "VAT Amount (€)": request.form.get("vat_amount", ""),
        "VAT Treatment": request.form.get("vat_treatment", "standard"),
        "Supply Type": request.form.get("supply_type", "services"),
        "Input VAT Reclaimable": request.form.get("input_vat_reclaimable", "Yes"),
        "Deductibility Status": request.form.get("deductibility_status", ""),
        "Capital Expenditure Flag": request.form.get("capital_expenditure_flag", ""),
        "Receipt Attached": request.form.get("receipt_attached", "No"),
        "Bank Reconciliation": request.form.get("bank_reconciliation", "Unreconciled"),
        "Status": _normalize_expense_status(request.form.get("status", "Pending")),
        "Payment Method": request.form.get("payment_method", ""),
        "Notes": request.form.get("notes", ""),
        "Phase Tag": _resolve_phase_tag(request.form.get("date", "")),
    }
    uploaded_receipt_name = _save_uploaded_receipt(request.files.get("receipt_file"))
    if uploaded_receipt_name:
        payload["Receipt Filename"] = uploaded_receipt_name
        payload["Receipt Attached"] = "Yes"
    _form_total = payload.get("Total (€)", "")
    _form_vat = payload.get("VAT Amount (€)", "")
    _apply_expense_amount_breakdown(payload)
    payload["Total (€)"] = _form_total
    payload["VAT Amount (€)"] = _form_vat
    _normalize_vat_fields(
        payload,
        net_key="Net Amount (€)",
        total_key="Total (€)",
        vat_rate_key="VAT Rate",
        vat_amount_key="VAT Amount (€)",
        vat_registered=_is_vat_registered(),
    )
    _apply_vat_classification(payload, vat_rate_key="VAT Rate")
    _apply_expense_compliance_fields(payload)
    compliance_flags, flags_acknowledged, flags_acknowledged_at = _process_compliance_flags(request.form)
    _apply_pretrading_compliance_flag(payload, compliance_flags)
    validation_errors = _validate_expense_payload(payload)
    validation_errors.update(_apply_compliance_flags_to_payload(payload, compliance_flags, flags_acknowledged, flags_acknowledged_at))
    if validation_errors:
        return _redirect_with_form_errors(
            "expenses_view",
            _build_workbook_form_data(payload, {
                "date": "Date (Registered)",
                "title": "Title",
                "description": "Description",
                "supplier": "Supplier / Payee",
                "supplier_vat_number": "Supplier VAT Number",
                "receipt_reference": "Receipt / Invoice Ref",
                "category": "Category",
                "base_net_amount": "Base Net Amount (€)",
                "delivery_amount": "Delivery (€)",
                "fees_amount": "Fees (€)",
                "other_charges_amount": "Other Charges (€)",
                "discount_type": "Discount Type",
                "discount_value": "Discount Value",
                "net_amount": "Net Amount (€)",
                "total_amount": "Total (€)",
                "vat_rate": "VAT Rate",
                "vat_amount": "VAT Amount (€)",
                "vat_treatment": "VAT Treatment",
                "supply_type": "Supply Type",
                "input_vat_reclaimable": "Input VAT Reclaimable",
                "deductibility_status": "Deductibility Status",
                "capital_expenditure_flag": "Capital Expenditure Flag",
                "receipt_attached": "Receipt Attached",
                "bank_reconciliation": "Bank Reconciliation",
                "status": "Status",
                "payment_method": "Payment Method",
                "notes": "Notes",
            }),
            validation_errors,
            validation_tab="expenses",
            edit_row=row_number,
        )
    _update_row_in_sheet("Expenses", row_number, payload)
    _upsert_capital_asset_from_expense(payload, row_number, active=payload.get("Capital Expenditure Flag") == "Yes")
    load_finance_data.cache_clear()

    _record_audit("update", "expense", {"row_number": row_number, "record": payload})
    _record_ledger_entry("update", "expense", payload, source="workbook", row_number=row_number)
    return Response(
        """
<!doctype html>
<html lang=\"en\">
    <head>
        <meta charset=\"utf-8\">
        <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">
        <title>Expense Updated</title>
        <style>
            body { font-family: Segoe UI, Arial, sans-serif; margin: 28px; color: #0f172a; }
            .ok { color: #065f46; font-weight: 600; margin-bottom: 10px; }
            a { color: #1d4ed8; text-decoration: none; }
            a:hover { text-decoration: underline; }
        </style>
    </head>
    <body>
        <div class=\"ok\">Expense updated in app view.</div>
        <div><a href=\"/expenses\">Return to Expenses</a></div>
    </body>
</html>
        """,
        mimetype="text/html",
    )


@app.route("/expenses/delete", methods=["POST"])
def delete_expense():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("expenses_view", message="Expense could not be removed"))

    row = _find_sheet_row_or_raise("Expenses", row_number)
    try:
        _archive_record("expense", row, source="workbook")
        _delete_row_from_sheet("Expenses", row_number)
        _upsert_capital_asset_from_expense(row, row_number, active=False)
        load_finance_data.cache_clear()
        _record_ledger_entry("archive", "expense", row, source="workbook", row_number=row_number)
        return redirect(url_for("expenses_view", message="Expense archived"))
    except WorkbookWriteError as exc:
        return redirect(url_for("expenses_view", message=str(exc)))


@app.route("/invoices/add", methods=["POST"])
def add_invoice():
    existing_invoices = load_finance_data().get("sheets", {}).get("Invoices", [])
    generated_invoice_number = _next_invoice_number(request.form.get("issue_date", ""), existing_invoices)
    payload = {
        "Invoice #": generated_invoice_number,
        "Issue Date": request.form.get("issue_date", ""),
        "Due Date": request.form.get("due_date", ""),
        "Client Name": request.form.get("client_name", ""),
        "Client VAT Number": request.form.get("client_vat_number", ""),
        "Client Address": request.form.get("client_address", ""),
        "Service / Product": request.form.get("service_product", ""),
        "VAT Treatment": request.form.get("vat_treatment", "standard"),
        "Supply Type": request.form.get("supply_type", "services"),
        "Balance Due (€)": request.form.get("balance_due", ""),
        "Status": _normalize_invoice_status(request.form.get("status", "Draft")),
        "Payment Method": request.form.get("payment_method", ""),
        "Payment Date": request.form.get("payment_date", ""),
        "Bank Reconciliation": request.form.get("bank_reconciliation", "Unreconciled"),
        "Notes": request.form.get("notes", ""),
        "Phase Tag": _resolve_phase_tag(request.form.get("issue_date", "")),
    }
    try:
        raw_line_items = json.loads(request.form.get("line_items_json") or "[]")
        if not isinstance(raw_line_items, list):
            raw_line_items = []
    except (TypeError, ValueError):
        raw_line_items = []
    line_items = _apply_invoice_line_items(payload, raw_line_items)
    _apply_vat_classification(payload, vat_rate_key="VAT Rate")
    _normalize_invoice_balance(payload)
    payment_date_autofilled = _apply_default_payment_date_for_paid(payload, "invoice", "Issue Date")
    validation_errors = _validate_invoice_payload(payload)
    if not line_items:
        validation_errors["line_items"] = "At least one line item is required"
    if validation_errors:
        return _redirect_with_form_errors(
            "invoices_view",
            {
                "invoice_number": request.form.get("invoice_number", ""),
                "issue_date": request.form.get("issue_date", ""),
                "due_date": request.form.get("due_date", ""),
                "client_name": request.form.get("client_name", ""),
                "client_vat_number": request.form.get("client_vat_number", ""),
                "client_address": request.form.get("client_address", ""),
                "service_product": request.form.get("service_product", ""),
                "vat_treatment": request.form.get("vat_treatment", "standard"),
                "supply_type": request.form.get("supply_type", "services"),
                "balance_due": request.form.get("balance_due", ""),
                "status": _normalize_invoice_status(request.form.get("status", "Draft")),
                "payment_method": request.form.get("payment_method", ""),
                "payment_date": request.form.get("payment_date", ""),
                "bank_reconciliation": request.form.get("bank_reconciliation", "Unreconciled"),
                "notes": request.form.get("notes", ""),
                "line_items_json": request.form.get("line_items_json", ""),
            },
            validation_errors,
            validation_tab="invoices",
        )
    row_number = _append_row_to_sheet("Invoices", payload)
    _sync_invoice_income_entry(payload)
    load_finance_data.cache_clear()
    _record_audit("create", "invoice", {"row_number": row_number, "record": payload})
    _record_ledger_entry("create", "invoice", payload, source="workbook", row_number=row_number)
    message = "Invoice added"
    if payment_date_autofilled:
        message = "Invoice added (payment date auto-filled from issue date)"
    return redirect(url_for("invoices_view", message=message))


@app.route("/invoices/update", methods=["POST"])
def update_invoice():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("invoices_view", message="Invoice could not be updated"))

    current_row = _find_sheet_row_or_raise("Invoices", row_number)
    payload = {
        "Invoice #": str(current_row.get("Invoice #") or request.form.get("invoice_number", "")),
        "Issue Date": request.form.get("issue_date", ""),
        "Due Date": request.form.get("due_date", ""),
        "Client Name": request.form.get("client_name", ""),
        "Client VAT Number": request.form.get("client_vat_number", ""),
        "Client Address": request.form.get("client_address", ""),
        "Service / Product": request.form.get("service_product", ""),
        "VAT Treatment": request.form.get("vat_treatment", "standard"),
        "Supply Type": request.form.get("supply_type", "services"),
        "Balance Due (€)": request.form.get("balance_due", ""),
        "Status": _normalize_invoice_status(request.form.get("status", "Draft")),
        "Payment Method": request.form.get("payment_method", ""),
        "Payment Date": request.form.get("payment_date", ""),
        "Bank Reconciliation": request.form.get("bank_reconciliation", "Unreconciled"),
        "Notes": request.form.get("notes", ""),
        "Phase Tag": _resolve_phase_tag(request.form.get("issue_date", "")),
    }
    try:
        raw_line_items = json.loads(request.form.get("line_items_json") or "[]")
        if not isinstance(raw_line_items, list):
            raw_line_items = []
    except (TypeError, ValueError):
        raw_line_items = []
    line_items = _apply_invoice_line_items(payload, raw_line_items)
    _apply_vat_classification(payload, vat_rate_key="VAT Rate")
    _normalize_invoice_balance(payload)
    payment_date_autofilled = _apply_default_payment_date_for_paid(payload, "invoice", "Issue Date")
    validation_errors = _validate_invoice_payload(payload)
    if not line_items:
        validation_errors["line_items"] = "At least one line item is required"
    if validation_errors:
        return _redirect_with_form_errors(
            "invoices_view",
            {
                "invoice_number": request.form.get("invoice_number", ""),
                "issue_date": request.form.get("issue_date", ""),
                "due_date": request.form.get("due_date", ""),
                "client_name": request.form.get("client_name", ""),
                "client_vat_number": request.form.get("client_vat_number", ""),
                "client_address": request.form.get("client_address", ""),
                "service_product": request.form.get("service_product", ""),
                "vat_treatment": request.form.get("vat_treatment", "standard"),
                "supply_type": request.form.get("supply_type", "services"),
                "balance_due": request.form.get("balance_due", ""),
                "status": _normalize_invoice_status(request.form.get("status", "Draft")),
                "payment_method": request.form.get("payment_method", ""),
                "payment_date": request.form.get("payment_date", ""),
                "bank_reconciliation": request.form.get("bank_reconciliation", "Unreconciled"),
                "notes": request.form.get("notes", ""),
                "line_items_json": request.form.get("line_items_json", ""),
            },
            validation_errors,
            validation_tab="invoices",
            edit_row=row_number,
        )
    _update_row_in_sheet("Invoices", row_number, payload)
    _sync_invoice_income_entry(payload)
    load_finance_data.cache_clear()
    _record_audit("update", "invoice", {"row_number": row_number, "record": payload})
    _record_ledger_entry("update", "invoice", payload, source="workbook", row_number=row_number)
    message = "Invoice updated"
    if payment_date_autofilled:
        message = "Invoice updated (payment date auto-filled from issue date)"
    return redirect(url_for("invoices_view", message=message))


@app.route("/invoices/delete", methods=["POST"])
def delete_invoice():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("invoices_view", message="Invoice could not be removed"))

    row = _find_sheet_row_or_raise("Invoices", row_number)
    if _normalize_invoice_status(row.get("Status")) != "Cancelled":
        row["Status"] = "Cancelled"
    _update_row_in_sheet("Invoices", row_number, row)
    _sync_invoice_income_entry(row)
    load_finance_data.cache_clear()
    _record_audit("cancel", "invoice", {"row_number": row_number, "record": row})
    _record_ledger_entry("cancel", "invoice", row, source="workbook", row_number=row_number)
    return redirect(url_for("invoices_view", message="Invoice cancelled and retained for audit trail"))


@app.route("/invoices/record-payment", methods=["POST"])
def record_invoice_payment():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("invoices_view", message="Invoice could not be updated"))

    row = _find_sheet_row_or_raise("Invoices", row_number)
    total = _coerce_number(row.get("Total (€)"))
    current_balance = _coerce_number(row.get("Balance Due (€)", total))
    payment_amount = _coerce_number(request.form.get("amount_received"))
    if payment_amount <= 0:
        return redirect(url_for("invoices_view", message="Amount received must be greater than zero"))

    new_balance = round(max(current_balance - payment_amount, 0.0), 2)
    row["Balance Due (€)"] = f"{new_balance:.2f}"
    row["Payment Date"] = request.form.get("payment_date") or date.today().isoformat()
    row["Payment Method"] = request.form.get("payment_method") or row.get("Payment Method", "")
    row["Status"] = "Paid" if new_balance <= 0.01 else "Partially Paid"
    _update_row_in_sheet("Invoices", row_number, row)
    _sync_invoice_income_entry(row)
    load_finance_data.cache_clear()
    _record_audit("payment", "invoice", {"row_number": row_number, "record": row})
    _record_ledger_entry("update", "invoice", row, source="workbook", row_number=row_number)
    message = "Invoice marked as paid — income entry created" if row["Status"] == "Paid" else "Partial payment recorded — income entry updated"
    return redirect(url_for("invoices_view", message=message))


@app.route("/subscriptions/add", methods=["POST"])
def add_subscription():
    start_date_value = request.form.get("start_date", "")
    start_date = _parse_iso_date(start_date_value) or date.today()
    subscription = {
        "id": str(uuid4()),
        "title": request.form.get("title", "").strip(),
        "description": request.form.get("description", "").strip(),
        "supplier": request.form.get("supplier", "").strip(),
        "category": request.form.get("category", "").strip(),
        "net_amount": request.form.get("net_amount", ""),
        "total_amount": request.form.get("total_amount", ""),
        "frequency": request.form.get("frequency", "monthly").strip().lower(),
        "start_date": start_date.isoformat(),
        "next_charge_date": start_date.isoformat(),
        "last_posted_date": "",
        "end_date": request.form.get("end_date", "").strip(),
        "status": request.form.get("status", "active").strip().lower(),
        "notes": request.form.get("notes", "").strip(),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "last_updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    subscription["next_charge_date"] = request.form.get("next_charge_date", "").strip() or start_date.isoformat()
    validation_errors = _validate_subscription_payload(subscription)
    if validation_errors:
        return _redirect_with_form_errors(
            "subscriptions_view",
            {
                "title": request.form.get("title", "").strip(),
                "description": request.form.get("description", "").strip(),
                "supplier": request.form.get("supplier", "").strip(),
                "category": request.form.get("category", "").strip(),
                "frequency": request.form.get("frequency", "monthly").strip().lower(),
                "start_date": subscription["start_date"],
                "next_charge_date": subscription["next_charge_date"],
                "end_date": request.form.get("end_date", "").strip(),
                "net_amount": request.form.get("net_amount", ""),
                "total_amount": request.form.get("total_amount", ""),
                "status": request.form.get("status", "active").strip().lower(),
                "notes": request.form.get("notes", "").strip(),
            },
            validation_errors,
            validation_tab="subscriptions",
        )
    subscriptions = _load_subscriptions()
    subscriptions.append(subscription)
    _save_subscriptions(subscriptions)
    _record_audit("create", "subscription", {"subscription_id": subscription["id"], "record": subscription})
    return redirect(url_for("subscriptions_view", message="Subscription added"))


@app.route("/subscriptions/update", methods=["POST"])
def update_subscription():
    subscription_id = request.form.get("subscription_id", "").strip()
    subscriptions = _load_subscriptions()
    existing = _find_subscription_by_id(subscriptions, subscription_id)
    if existing is None:
        return redirect(url_for("subscriptions_view", message="Subscription could not be updated"))

    start_date = _parse_iso_date(request.form.get("start_date", "")) or _parse_iso_date(existing.get("start_date")) or date.today()
    next_charge_date = _parse_iso_date(request.form.get("next_charge_date", "")) or _parse_iso_date(existing.get("next_charge_date")) or start_date
    end_date = _parse_iso_date(request.form.get("end_date", ""))

    payload = {
        "title": request.form.get("title", "").strip(),
        "description": request.form.get("description", "").strip(),
        "supplier": request.form.get("supplier", "").strip(),
        "category": request.form.get("category", "").strip(),
        "net_amount": request.form.get("net_amount", ""),
        "total_amount": request.form.get("total_amount", ""),
        "frequency": request.form.get("frequency", "monthly").strip().lower(),
        "start_date": start_date.isoformat(),
        "next_charge_date": next_charge_date.isoformat(),
        "end_date": end_date.isoformat() if end_date else "",
        "status": request.form.get("status", "active").strip().lower(),
        "notes": request.form.get("notes", "").strip(),
    }
    validation_errors = _validate_subscription_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "subscriptions_view",
            {
                "title": payload["title"],
                "description": payload["description"],
                "supplier": payload["supplier"],
                "category": payload["category"],
                "frequency": payload["frequency"],
                "start_date": payload["start_date"],
                "next_charge_date": payload["next_charge_date"],
                "end_date": payload["end_date"],
                "net_amount": payload["net_amount"],
                "total_amount": payload["total_amount"],
                "status": payload["status"],
                "notes": payload["notes"],
            },
            validation_errors,
            validation_tab="subscriptions",
            edit_id=subscription_id,
        )

    existing.update(
        {
            **payload,
            "last_updated_at": datetime.now().isoformat(timespec="seconds"),
        }
    )
    _save_subscriptions(subscriptions)
    _record_audit("update", "subscription", {"subscription_id": subscription_id, "record": existing})
    return redirect(url_for("subscriptions_view", message="Subscription updated"))


@app.route("/subscriptions/delete", methods=["POST"])
def delete_subscription():
    subscription_id = request.form.get("subscription_id", "").strip()
    subscriptions = _load_subscriptions()
    existing = _find_subscription_by_id(subscriptions, subscription_id)
    if existing is None:
        return redirect(url_for("subscriptions_view", message="Subscription could not be removed"))

    _archive_record("subscription", existing, source="subscriptions")
    remaining = [subscription for subscription in subscriptions if str(subscription.get("id")) != subscription_id]
    _save_subscriptions(remaining)
    return redirect(url_for("subscriptions_view", message="Subscription archived"))


@app.route("/clients/add", methods=["POST"])
def add_client():
    payload = {
        "Client Name": request.form.get("client_name", ""),
        "Contact Person": request.form.get("contact_person", ""),
        "Email": request.form.get("email", ""),
        "Phone": request.form.get("phone", ""),
        "Country": request.form.get("country", ""),
        "Service Tier": request.form.get("service_tier", "None"),
        "Retainer Frequency": request.form.get("retainer_frequency", ""),
        "Retainer Amount (€)": request.form.get("retainer_amount", ""),
    }
    validation_errors = _validate_client_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "clients_view",
            {
                "client_name": request.form.get("client_name", ""),
                "contact_person": request.form.get("contact_person", ""),
                "email": request.form.get("email", ""),
                "phone": request.form.get("phone", ""),
                "country": request.form.get("country", ""),
                "service_tier": request.form.get("service_tier", "None"),
                "retainer_frequency": request.form.get("retainer_frequency", ""),
                "retainer_amount": request.form.get("retainer_amount", ""),
            },
            validation_errors,
            validation_tab="clients",
        )
    row_number = _append_row_to_sheet("Clients", payload)
    load_finance_data.cache_clear()
    _record_audit("create", "client", {"row_number": row_number, "record": payload})
    return redirect(url_for("clients_view", message="Client added"))


@app.route("/clients/update", methods=["POST"])
def update_client():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("clients_view", message="Client could not be updated"))

    payload = {
        "Client Name": request.form.get("client_name", ""),
        "Contact Person": request.form.get("contact_person", ""),
        "Email": request.form.get("email", ""),
        "Phone": request.form.get("phone", ""),
        "Country": request.form.get("country", ""),
        "Service Tier": request.form.get("service_tier", "None"),
        "Retainer Frequency": request.form.get("retainer_frequency", ""),
        "Retainer Amount (€)": request.form.get("retainer_amount", ""),
    }
    validation_errors = _validate_client_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "clients_view",
            {
                "client_name": request.form.get("client_name", ""),
                "contact_person": request.form.get("contact_person", ""),
                "email": request.form.get("email", ""),
                "phone": request.form.get("phone", ""),
                "country": request.form.get("country", ""),
                "service_tier": request.form.get("service_tier", "None"),
                "retainer_frequency": request.form.get("retainer_frequency", ""),
                "retainer_amount": request.form.get("retainer_amount", ""),
            },
            validation_errors,
            validation_tab="clients",
            edit_row=row_number,
        )
    _update_row_in_sheet("Clients", row_number, payload)
    load_finance_data.cache_clear()
    _record_audit("update", "client", {"row_number": row_number, "record": payload})
    return redirect(url_for("clients_view", message="Client updated"))


@app.route("/clients/delete", methods=["POST"])
def delete_client():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("clients_view", message="Client could not be removed"))

    row = _find_sheet_row_or_raise("Clients", row_number)
    _archive_record("client", row, source="workbook")
    _delete_row_from_sheet("Clients", row_number)
    load_finance_data.cache_clear()
    return redirect(url_for("clients_view", message="Client archived"))


def _service_form_fields(form_source) -> dict[str, str]:
    return {
        "name": form_source.get("name", ""),
        "tier": form_source.get("tier", "addon"),
        "group": form_source.get("group", ""),
        "description": form_source.get("description", ""),
        "price": form_source.get("price", ""),
        "price_type": form_source.get("price_type", "fixed"),
        "billing_frequency": form_source.get("billing_frequency", ""),
        "quarterly_price": form_source.get("quarterly_price", ""),
        "annual_price": form_source.get("annual_price", ""),
        "status": form_source.get("status", "active"),
        "website_display_price": form_source.get("website_display_price", ""),
        "website_display_label": form_source.get("website_display_label", ""),
    }


@app.route("/services/add", methods=["POST"])
def add_service():
    now = datetime.now().isoformat(timespec="seconds")
    payload = {
        "id": str(uuid4()),
        "name": request.form.get("name", ""),
        "tier": request.form.get("tier", "addon"),
        "group": request.form.get("group", ""),
        "description": request.form.get("description", ""),
        "price": request.form.get("price", "0"),
        "price_type": request.form.get("price_type", "fixed"),
        "billing_frequency": request.form.get("billing_frequency", ""),
        "quarterly_price": request.form.get("quarterly_price", ""),
        "annual_price": request.form.get("annual_price", ""),
        "status": request.form.get("status", "active"),
        "website_display_price": request.form.get("website_display_price") == "on",
        "website_display_label": request.form.get("website_display_label", ""),
        "date_added": now,
        "date_updated": now,
    }
    validation_errors = _validate_service_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "services_view",
            _service_form_fields(request.form),
            validation_errors,
            validation_tab="services",
        )
    services = _load_services()
    services.append(payload)
    _save_services(services)
    _record_audit("create", "service", {"service_id": payload["id"], "record": payload})
    return redirect(url_for("services_view", message="Service added"))


@app.route("/services/update", methods=["POST"])
def update_service():
    service_id = str(request.form.get("service_id") or "").strip()
    services = _load_services()
    existing = _find_service_by_id(services, service_id)
    if existing is None:
        return redirect(url_for("services_view", message="Service could not be updated"))

    payload = {
        "id": existing.get("id"),
        "name": request.form.get("name", existing.get("name", "")),
        "tier": request.form.get("tier", existing.get("tier", "addon")),
        "group": request.form.get("group", existing.get("group") or ""),
        "description": request.form.get("description", existing.get("description", "")),
        "price": request.form.get("price", existing.get("price", "0")),
        "price_type": request.form.get("price_type", existing.get("price_type", "fixed")),
        "billing_frequency": request.form.get("billing_frequency", existing.get("billing_frequency") or ""),
        "quarterly_price": request.form.get("quarterly_price", existing.get("quarterly_price") or ""),
        "annual_price": request.form.get("annual_price", existing.get("annual_price") or ""),
        "status": request.form.get("status", existing.get("status", "active")),
        "website_display_price": request.form.get("website_display_price") == "on",
        "website_display_label": request.form.get("website_display_label", existing.get("website_display_label", "")),
        "date_added": existing.get("date_added"),
        "date_updated": datetime.now().isoformat(timespec="seconds"),
    }
    validation_errors = _validate_service_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "services_view",
            _service_form_fields(request.form),
            validation_errors,
            validation_tab="services",
            edit_id=service_id,
        )
    existing.update(payload)
    _save_services(services)
    _record_audit("update", "service", {"service_id": service_id, "record": existing})
    return redirect(url_for("services_view", message="Service updated"))


@app.route("/services/archive", methods=["POST"])
def archive_service():
    service_id = str(request.form.get("service_id") or "").strip()
    services = _load_services()
    existing = _find_service_by_id(services, service_id)
    if existing is None:
        return redirect(url_for("services_view", message="Service not found"))
    existing["status"] = "archived"
    existing["date_updated"] = datetime.now().isoformat(timespec="seconds")
    _save_services(services)
    _record_audit("archive", "service", {"service_id": service_id, "record": existing})
    return redirect(url_for("services_view", message="Service archived", show_archived="1"))


@app.route("/services/restore", methods=["POST"])
def restore_service():
    service_id = str(request.form.get("service_id") or "").strip()
    services = _load_services()
    existing = _find_service_by_id(services, service_id)
    if existing is None:
        return redirect(url_for("services_view", message="Service not found"))
    existing["status"] = "active"
    existing["date_updated"] = datetime.now().isoformat(timespec="seconds")
    _save_services(services)
    _record_audit("restore", "service", {"service_id": service_id, "record": existing})
    return redirect(url_for("services_view", message="Service restored", show_archived="1"))


@app.route("/api/services")
def api_services():
    services = _load_services()
    payload = [
        {
            "id": service["id"],
            "name": service["name"],
            "tier": service["tier"],
            "group": service["group"],
            "description": service["description"],
            "price": service["price"],
            "price_type": service["price_type"],
            "billing_frequency": service["billing_frequency"],
            "website_display_price": service["website_display_price"],
            "website_display_label": service["website_display_label"],
        }
        for service in services
        if service["status"] == "active"
    ]
    response = jsonify(payload)
    response.headers["Access-Control-Allow-Origin"] = "*"
    return response


@app.route("/api/suppliers/search")
def api_suppliers_search():
    query = str(request.args.get("q") or "").strip().lower()
    if len(query) < 2:
        return jsonify([])
    suppliers = _load_sheet_records_raw("Suppliers")
    matches = []
    seen = set()
    for supplier in suppliers:
        name = str(supplier.get("Supplier Name") or "").strip()
        if not name or name.lower() in seen:
            continue
        if query in name.lower():
            matches.append({"name": name, "needs_completion": str(supplier.get("Needs Completion") or "No") == "Yes"})
            seen.add(name.lower())
    return jsonify(matches[:20])


@app.route("/api/suppliers/quick-add", methods=["POST"])
def api_suppliers_quick_add():
    body = request.get_json(silent=True) or {}
    name = str(body.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Supplier name is required"}), 400

    suppliers = _load_sheet_records_raw("Suppliers")
    for supplier in suppliers:
        if str(supplier.get("Supplier Name") or "").strip().lower() == name.lower():
            return jsonify({"name": supplier.get("Supplier Name"), "created": False})

    payload = {
        "Supplier Name": name,
        "Contact Person": "",
        "Email": "",
        "Phone": "",
        "Country": "",
        "Default VAT Treatment": "",
        "Needs Completion": "Yes",
    }
    row_number = _append_row_to_sheet("Suppliers", payload)
    load_finance_data.cache_clear()
    _record_audit("create", "supplier", {"row_number": row_number, "record": payload, "source": "quick_add"})
    return jsonify({"name": name, "created": True})


@app.route("/api/clients/search")
def api_clients_search():
    query = str(request.args.get("q") or "").strip().lower()
    if len(query) < 2:
        return jsonify([])
    clients = _load_sheet_records_raw("Clients")
    matches = []
    seen = set()
    for client in clients:
        name = str(client.get("Client Name") or "").strip()
        if not name or name.lower() in seen:
            continue
        if query in name.lower():
            matches.append({"name": name, "needs_completion": False})
            seen.add(name.lower())
    return jsonify(matches[:20])


@app.route("/api/clients/quick-add", methods=["POST"])
def api_clients_quick_add():
    body = request.get_json(silent=True) or {}
    name = str(body.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Client name is required"}), 400

    clients = _load_sheet_records_raw("Clients")
    for client in clients:
        if str(client.get("Client Name") or "").strip().lower() == name.lower():
            return jsonify({"name": client.get("Client Name"), "created": False})

    payload = {
        "Client Name": name,
        "Contact Person": "",
        "Email": "",
        "Phone": "",
        "Country": "",
        "Service Tier": "None",
        "Retainer Frequency": "",
        "Retainer Amount (€)": "",
    }
    row_number = _append_row_to_sheet("Clients", payload)
    load_finance_data.cache_clear()
    _record_audit("create", "client", {"row_number": row_number, "record": payload, "source": "quick_add"})
    return jsonify({"name": name, "created": True})


@app.route("/suppliers/add", methods=["POST"])
def add_supplier():
    payload = {
        "Supplier Name": request.form.get("supplier_name", ""),
        "Contact Person": request.form.get("contact_person", ""),
        "Email": request.form.get("email", ""),
        "Phone": request.form.get("phone", ""),
        "Country": request.form.get("country", ""),
        "Default VAT Treatment": request.form.get("default_vat_treatment", ""),
        "Needs Completion": "No",
    }
    validation_errors = _validate_supplier_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "suppliers_view",
            {
                "supplier_name": request.form.get("supplier_name", ""),
                "contact_person": request.form.get("contact_person", ""),
                "email": request.form.get("email", ""),
                "phone": request.form.get("phone", ""),
                "country": request.form.get("country", ""),
                "default_vat_treatment": request.form.get("default_vat_treatment", ""),
            },
            validation_errors,
            validation_tab="suppliers",
        )
    row_number = _append_row_to_sheet("Suppliers", payload)
    load_finance_data.cache_clear()
    _record_audit("create", "supplier", {"row_number": row_number, "record": payload})
    return redirect(url_for("suppliers_view", message="Supplier added"))


@app.route("/suppliers/update", methods=["POST"])
def update_supplier():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("suppliers_view", message="Supplier could not be updated"))

    payload = {
        "Supplier Name": request.form.get("supplier_name", ""),
        "Contact Person": request.form.get("contact_person", ""),
        "Email": request.form.get("email", ""),
        "Phone": request.form.get("phone", ""),
        "Country": request.form.get("country", ""),
        "Default VAT Treatment": request.form.get("default_vat_treatment", ""),
        "Needs Completion": "No",
    }
    validation_errors = _validate_supplier_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "suppliers_view",
            {
                "supplier_name": request.form.get("supplier_name", ""),
                "contact_person": request.form.get("contact_person", ""),
                "email": request.form.get("email", ""),
                "phone": request.form.get("phone", ""),
                "country": request.form.get("country", ""),
                "default_vat_treatment": request.form.get("default_vat_treatment", ""),
            },
            validation_errors,
            validation_tab="suppliers",
            edit_row=row_number,
        )
    _update_row_in_sheet("Suppliers", row_number, payload)
    load_finance_data.cache_clear()
    _record_audit("update", "supplier", {"row_number": row_number, "record": payload})
    return redirect(url_for("suppliers_view", message="Supplier updated"))


@app.route("/suppliers/delete", methods=["POST"])
def delete_supplier():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("suppliers_view", message="Supplier could not be removed"))

    row = _find_sheet_row_or_raise("Suppliers", row_number)
    _archive_record("supplier", row, source="workbook")
    _delete_row_from_sheet("Suppliers", row_number)
    load_finance_data.cache_clear()
    return redirect(url_for("suppliers_view", message="Supplier archived"))


@app.route("/export/xlsm", methods=["POST"])
def export_xlsm():
    next_page = str(request.args.get("next") or "/")
    try:
        resolved_path = _resolve_workbook_path()
    except FileNotFoundError as exc:
        return redirect(_append_message_to_path(next_page, f"Export failed: {exc}"))

    try:
        wb = load_workbook(resolved_path, data_only=False, keep_links=False)
    except Exception as exc:
        return redirect(_append_message_to_path(next_page, f"Export failed: could not open the workbook ({exc})"))

    try:
        for sheet_name in SHEET_JSON_PATHS:
            if sheet_name not in wb:
                continue
            ws = wb[sheet_name]
            header_row_number = _find_header_row_number(ws, sheet_name)
            if header_row_number is None:
                continue
            headers = _get_header_row(ws, sheet_name)
            normalized_headers = [_normalize_header_name(header, sheet_name) for header in headers]

            if ws.max_row > header_row_number:
                ws.delete_rows(header_row_number + 1, ws.max_row - header_row_number)

            for record in _load_sheet_records_raw(sheet_name):
                row_values = [record.get(header, record.get(str(header), "")) for header in normalized_headers]
                ws.append(row_values)

        _save_workbook_atomic(wb, resolved_path)
    except WorkbookWriteError as exc:
        return redirect(_append_message_to_path(next_page, f"Export failed: {exc}"))
    except Exception as exc:
        return redirect(_append_message_to_path(next_page, f"Export failed: {exc}"))
    finally:
        wb.close()

    return redirect(_append_message_to_path(next_page, "Exported current data to Excel workbook"))


_migrate_transaction_sheets_from_workbook()
_ensure_default_services()
_migrate_invoice_line_items()
_migrate_income_invoice_linkage()
_migrate_chart_of_accounts_categories()
_migrate_flag_retired_non_deductible_category()
_prune_old_backups()


if __name__ == "__main__":
    app.run(
        debug=False,
        use_reloader=False,
        host="127.0.0.1",
        port=5000,
    )
