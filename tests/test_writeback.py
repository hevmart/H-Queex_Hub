import os
import tempfile
import json
from io import BytesIO
from datetime import date, timedelta
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import app


@pytest.fixture
def workbook_copy(tmp_path):
    dst = tmp_path / "sample.xlsm"
    wb = Workbook()
    wb.remove(wb.active)

    income = wb.create_sheet("Income")
    income.append(["Date", "Description", "Client / Source", "Category", "Invoice #", "Amount (€)", "Status"])
    income.append(["2026-07-29", "Test income", "Client A", "Travel", "INV-001", 150.0, "Paid"])

    expenses = wb.create_sheet("Expenses")
    expenses.append(["Date (Registered)", "Title", "Description", "Supplier / Payee", "Category", "Net Amount (€)", "Total (€)", "Status"])
    expenses.append(["2026-07-29", "Travel", "Hotel", "Supplier A", "Travel", 100.0, 120.0, "Pending"])

    invoices = wb.create_sheet("Invoices")
    invoices.append(["Invoice #", "Issue Date", "Due Date", "Client Name", "Service / Product", "Net (€)", "Total (€)", "Balance Due (€)", "Status"])
    invoices.append(["INV-001", "2026-07-01", "2026-07-31", "Client A", "Brand Strategy", 500.0, 605.0, 605.0, "Sent"])

    clients = wb.create_sheet("Clients")
    clients.append(["Client Name", "Contact Person", "Email", "Phone", "Country"])
    clients.append(["Client A", "Jane", "jane@example.com", "123", "Belgium"])

    suppliers = wb.create_sheet("Suppliers")
    suppliers.append(["Supplier Name", "Contact Person", "Email", "Phone", "Country", "Default VAT Treatment"])
    suppliers.append(["Supplier A", "John", "john@example.com", "456", "Netherlands", "Standard"])

    wb.save(dst)
    wb.close()
    _seed_transaction_json(tmp_path)
    return dst


@pytest.fixture(autouse=True)
def isolated_subscription_file(tmp_path):
    original_path = app.SUBSCRIPTIONS_PATH
    original_archive_path = app.ARCHIVE_PATH
    original_audit_log_path = app.AUDIT_LOG_PATH
    original_business_profile_path = app.BUSINESS_PROFILE_PATH
    original_coa_path = app.CHART_OF_ACCOUNTS_PATH
    original_ledger_path = app.LEDGER_JOURNAL_PATH
    original_capital_assets_path = app.CAPITAL_ASSETS_PATH
    original_payroll_path = app.PAYROLL_PATH
    original_bank_statements_path = app.BANK_STATEMENTS_PATH
    original_income_path = app.INCOME_PATH
    original_expenses_path = app.EXPENSES_PATH
    original_invoices_path = app.INVOICES_PATH
    original_clients_path = app.CLIENTS_PATH
    original_suppliers_path = app.SUPPLIERS_PATH
    original_sheet_json_paths = dict(app.SHEET_JSON_PATHS)
    original_backups_dir = app.BACKUPS_DIR
    original_gdrive_backup_dir = app.GDRIVE_BACKUP_DIR
    original_backup_status_path = app.BACKUP_STATUS_PATH
    original_receipts_dir = app.RECEIPTS_DIR
    original_company_documents_path = app.COMPANY_DOCUMENTS_PATH
    original_compliance_calendar_path = app.COMPLIANCE_CALENDAR_PATH
    original_company_documents_dir = app.COMPANY_DOCUMENTS_DIR
    original_projects_path = app.PROJECTS_PATH
    original_delivery_log_path = app.DELIVERY_LOG_PATH
    original_sops_path = app.SOPS_PATH
    original_sop_files_dir = app.SOP_FILES_DIR
    original_delivery_files_dir = app.DELIVERY_FILES_DIR
    original_services_path = app.SERVICES_PATH
    original_leads_path = app.LEADS_PATH
    original_proposals_path = app.PROPOSALS_PATH
    original_terms_path = app.TERMS_PATH
    app.BACKUPS_DIR = tmp_path / "backups"
    app.GDRIVE_BACKUP_DIR = tmp_path / "gdrive-backups"
    app.BACKUP_STATUS_PATH = tmp_path / "backup-status.json"
    app.RECEIPTS_DIR = tmp_path / "receipts"
    app.RECEIPTS_DIR.mkdir(parents=True, exist_ok=True)
    app.SUBSCRIPTIONS_PATH = tmp_path / "subscriptions.json"
    app.ARCHIVE_PATH = tmp_path / "archives.json"
    app.AUDIT_LOG_PATH = tmp_path / "audit-log.json"
    app.BUSINESS_PROFILE_PATH = tmp_path / "business-profile.json"
    app.CHART_OF_ACCOUNTS_PATH = tmp_path / "chart-of-accounts.json"
    app.LEDGER_JOURNAL_PATH = tmp_path / "ledger-journal.json"
    app.CAPITAL_ASSETS_PATH = tmp_path / "capital-assets.json"
    app.PAYROLL_PATH = tmp_path / "payroll-register.json"
    app.BANK_STATEMENTS_PATH = tmp_path / "bank-statements.json"
    app.INCOME_PATH = tmp_path / "income.json"
    app.EXPENSES_PATH = tmp_path / "expenses.json"
    app.INVOICES_PATH = tmp_path / "invoices.json"
    app.CLIENTS_PATH = tmp_path / "clients.json"
    app.SUPPLIERS_PATH = tmp_path / "suppliers.json"
    app.SHEET_JSON_PATHS = {
        "Income": app.INCOME_PATH,
        "Expenses": app.EXPENSES_PATH,
        "Invoices": app.INVOICES_PATH,
        "Clients": app.CLIENTS_PATH,
        "Suppliers": app.SUPPLIERS_PATH,
    }
    app.COMPANY_DOCUMENTS_PATH = tmp_path / "documents.json"
    app.COMPLIANCE_CALENDAR_PATH = tmp_path / "compliance-calendar.json"
    app.COMPANY_DOCUMENTS_DIR = tmp_path / "documents"
    app.COMPANY_DOCUMENTS_DIR.mkdir(parents=True, exist_ok=True)
    app.PROJECTS_PATH = tmp_path / "projects.json"
    app.DELIVERY_LOG_PATH = tmp_path / "delivery-log.json"
    app.SOPS_PATH = tmp_path / "sops.json"
    app.SOP_FILES_DIR = tmp_path / "sops"
    app.SOP_FILES_DIR.mkdir(parents=True, exist_ok=True)
    app.DELIVERY_FILES_DIR = tmp_path / "delivery-files"
    app.DELIVERY_FILES_DIR.mkdir(parents=True, exist_ok=True)
    app.LEADS_PATH = tmp_path / "leads.json"
    app.PROPOSALS_PATH = tmp_path / "proposals.json"
    app.TERMS_PATH = tmp_path / "terms.json"
    app.SERVICES_PATH = tmp_path / "services.json"
    app.load_finance_data.cache_clear()
    yield
    app.SUBSCRIPTIONS_PATH = original_path
    app.ARCHIVE_PATH = original_archive_path
    app.AUDIT_LOG_PATH = original_audit_log_path
    app.BUSINESS_PROFILE_PATH = original_business_profile_path
    app.CHART_OF_ACCOUNTS_PATH = original_coa_path
    app.LEDGER_JOURNAL_PATH = original_ledger_path
    app.CAPITAL_ASSETS_PATH = original_capital_assets_path
    app.PAYROLL_PATH = original_payroll_path
    app.BANK_STATEMENTS_PATH = original_bank_statements_path
    app.INCOME_PATH = original_income_path
    app.EXPENSES_PATH = original_expenses_path
    app.INVOICES_PATH = original_invoices_path
    app.CLIENTS_PATH = original_clients_path
    app.SUPPLIERS_PATH = original_suppliers_path
    app.SHEET_JSON_PATHS = original_sheet_json_paths
    app.BACKUPS_DIR = original_backups_dir
    app.GDRIVE_BACKUP_DIR = original_gdrive_backup_dir
    app.BACKUP_STATUS_PATH = original_backup_status_path
    app.RECEIPTS_DIR = original_receipts_dir
    app.COMPANY_DOCUMENTS_PATH = original_company_documents_path
    app.COMPLIANCE_CALENDAR_PATH = original_compliance_calendar_path
    app.COMPANY_DOCUMENTS_DIR = original_company_documents_dir
    app.PROJECTS_PATH = original_projects_path
    app.DELIVERY_LOG_PATH = original_delivery_log_path
    app.SOPS_PATH = original_sops_path
    app.SOP_FILES_DIR = original_sop_files_dir
    app.DELIVERY_FILES_DIR = original_delivery_files_dir
    app.LEADS_PATH = original_leads_path
    app.PROPOSALS_PATH = original_proposals_path
    app.TERMS_PATH = original_terms_path
    app.SERVICES_PATH = original_services_path
    app.load_finance_data.cache_clear()


def _seed_transaction_json(tmp_path):
    (tmp_path / "income.json").write_text(
        json.dumps([
            {
                "Date": "2026-07-29",
                "Description": "Test income",
                "Client / Source": "Client A",
                "Category": "Travel",
                "Invoice #": "INV-001",
                "Amount (€)": 150.0,
                "Status": "Paid",
            }
        ]),
        encoding="utf-8",
    )
    (tmp_path / "expenses.json").write_text(
        json.dumps([
            {
                "Date (Registered)": "2026-07-29",
                "Title": "Travel",
                "Description": "Hotel",
                "Supplier / Payee": "Supplier A",
                "Category": "Travel",
                "Net Amount (€)": 100.0,
                "Total (€)": 120.0,
                "Status": "Pending",
            }
        ]),
        encoding="utf-8",
    )
    (tmp_path / "invoices.json").write_text(
        json.dumps([
            {
                "Invoice #": "INV-001",
                "Issue Date": "2026-07-01",
                "Due Date": "2026-07-31",
                "Client Name": "Client A",
                "Service / Product": "Brand Strategy",
                "Net (€)": 500.0,
                "Total (€)": 605.0,
                "Balance Due (€)": 605.0,
                "Status": "Sent",
            }
        ]),
        encoding="utf-8",
    )
    (tmp_path / "clients.json").write_text(
        json.dumps([
            {
                "Client Name": "Client A",
                "Contact Person": "Jane",
                "Email": "jane@example.com",
                "Phone": "123",
                "Country": "Belgium",
            }
        ]),
        encoding="utf-8",
    )
    (tmp_path / "suppliers.json").write_text(
        json.dumps([
            {
                "Supplier Name": "Supplier A",
                "Contact Person": "John",
                "Email": "john@example.com",
                "Phone": "456",
                "Country": "Netherlands",
                "Default VAT Treatment": "Standard",
            }
        ]),
        encoding="utf-8",
    )


def _invoice_line_items_payload(name, net_amount, *, vat_rate="23%", quantity=1):
    return json.dumps([
        {
            "service_id": "",
            "name": name,
            "description": name,
            "quantity": quantity,
            "unit_price": net_amount,
            "discount_type": "€",
            "discount_value": 0,
            "vat_rate": vat_rate,
        }
    ])


def test_append_income_row_updates_workbook(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    payload = {
        "date": "2026-07-29",
        "description": "Test income",
        "client_source": "Client A",
        "amount": "150.00",
        "status": "Paid",
    }

    response = app.app.test_client().post('/income/add', data=payload, follow_redirects=True)
    assert response.status_code == 200

    income_records = json.loads(app.INCOME_PATH.read_text(encoding="utf-8"))
    last_record = income_records[-1]
    assert last_record["Description"] == "Test income"
    assert str(last_record["Amount (€)"]) == "150.00"


def test_update_income_route_updates_existing_workbook_row(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    row_number = app.load_finance_data()["sheets"]["Income"][0]["__row_number"]
    payload = {
        "row_number": str(row_number),
        "date": "2026-07-30",
        "description": "Updated income",
        "client_source": "Client A",
        "category": "Consulting",
        "invoice_number": "INV-009",
        "amount": "250.00",
        "status": "Paid",
    }

    response = app.app.test_client().post('/income/update', data=payload, follow_redirects=True)
    assert response.status_code == 200
    assert b'Income entry updated' in response.data

    income_records = json.loads(app.INCOME_PATH.read_text(encoding="utf-8"))
    record = income_records[row_number - 1]
    assert record["Description"] == "Updated income"
    assert str(record["Amount (€)"]) == "250.00"


def test_income_validation_prevents_invalid_amount(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    payload = {
        "date": "2026-07-30",
        "description": "Invalid income",
        "client_source": "Client A",
        "amount": "-10.00",
        "status": "Paid",
    }
    response = app.app.test_client().post('/income/add', data=payload, follow_redirects=True)

    assert response.status_code == 200
    assert b'Validation:' in response.data
    assert b'Income amount must be greater than zero' in response.data
    assert b'value="-10.00"' in response.data


def test_finance_landing_page_renders_module_cards(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().get('/finance')

    assert response.status_code == 200
    assert b'Finance hub' in response.data
    assert b'/income' in response.data
    assert b'/expenses' in response.data
    assert b'/archive' in response.data


def test_refresh_route_redirects_back_to_page(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().post('/refresh?next=/income', follow_redirects=True)

    assert response.status_code == 200
    assert b'Income' in response.data


def test_business_structure_toggle_changes_dashboard_mode(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    update_response = app.app.test_client().post(
        '/business-structure/update?next=/',
        data={"structure": "limited_company", "transition_date": "2026-08-15"},
        follow_redirects=True,
    )

    assert update_response.status_code == 200
    assert 'Phase 2 — Limited Company'.encode('utf-8') in update_response.data
    assert b'Corporation Tax (CT1)' in update_response.data
    assert b'CT1 outputs' in update_response.data
    assert b'Director Loan Account' in update_response.data


def test_phase_resolution_uses_transition_date_for_limited_company(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/business-structure/update?next=/',
        data={"structure": "limited_company", "transition_date": "2026-08-15"},
    )

    assert app._resolve_phase_tag("2026-08-10") == "Phase 1"
    assert app._resolve_phase_tag("2026-08-15") == "Phase 2"
    assert app._resolve_phase_tag("2026-08-20") == "Phase 2"


def test_ledger_posts_income_with_mapped_account(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    payload = {
        "date": "2026-08-05",
        "description": "Strategy workshop",
        "client_source": "Client A",
        "category": "Consulting / Project Fees",
        "invoice_number": "INV-2026-101",
        "amount": "500.00",
        "status": "Paid",
        "payment_method": "Business Bank Account",
    }

    response = app.app.test_client().post('/income/add', data=payload, follow_redirects=True)
    assert response.status_code == 200

    ledger_entries = json.loads(app.LEDGER_JOURNAL_PATH.read_text(encoding="utf-8"))
    assert len(ledger_entries) >= 1
    latest_entry = ledger_entries[-1]
    assert latest_entry["entity_type"] == "income"
    assert latest_entry["account_code"] == "4000"
    assert latest_entry["account_name"] == "Consulting / Project Fees"
    assert latest_entry["amount_eur"] == 500.0
    assert latest_entry["entry_balanced"] is True
    assert latest_entry["debit_total"] == latest_entry["credit_total"]
    assert len(latest_entry["journal_lines"]) == 2


def test_ledger_view_renders_accounts_and_journal(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().get('/ledger')
    assert response.status_code == 200
    assert b'Chart of accounts' in response.data
    assert b'Ledger journal' in response.data
    assert b'Trial balance' in response.data
    assert b'VAT control summary' in response.data


def test_ledger_posts_expense_with_vat_control_accounts(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    payload = {
        "date": "2026-08-05",
        "title": "Cloud Hosting",
        "description": "Monthly infra",
        "supplier": "Supplier A",
        "supplier_vat_number": "IE1234567A",
        "category": "Professional Fees",
        "net_amount": "100.00",
        "total_amount": "123.00",
        "input_vat_reclaimable": "Yes",
        "status": "Paid",
        "payment_method": "Business Bank",
    }
    response = app.app.test_client().post('/expenses/add', data=payload, follow_redirects=True)
    assert response.status_code == 200

    ledger_entries = json.loads(app.LEDGER_JOURNAL_PATH.read_text(encoding="utf-8"))
    latest_entry = ledger_entries[-1]
    assert latest_entry["entity_type"] == "expense"
    assert latest_entry["vat_amount_eur"] == 23.0
    assert latest_entry["entry_balanced"] is True

    lines = latest_entry["journal_lines"]
    assert any(line["account_code"] == "1200" and line["debit"] == 23.0 for line in lines)
    assert any(line["account_code"] == "1000" and line["credit"] == 123.0 for line in lines)


def test_trial_balance_csv_export_returns_download(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/income/add',
        data={
            "date": "2026-08-05",
            "description": "Export test",
            "client_source": "Client A",
            "category": "Consulting / Project Fees",
            "invoice_number": "INV-EXP-1",
            "amount": "250.00",
            "status": "Paid",
            "payment_method": "Business Bank Account",
        },
    )

    response = app.app.test_client().get('/ledger/trial-balance.csv')
    assert response.status_code == 200
    assert response.mimetype == 'text/csv'
    assert 'attachment; filename=trial-balance.csv' in response.headers['Content-Disposition']
    assert b'account_code,account_name,debit_eur,credit_eur,net_eur' in response.data
    assert b'TOTAL' in response.data


def test_ledger_journal_csv_export_includes_vat_trace_columns(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/expenses/add',
        data={
            "date": "2026-08-05",
            "title": "RC purchase",
            "description": "Cross-border service",
            "supplier": "Supplier A",
            "supplier_vat_number": "IE1234567A",
            "category": "Professional Fees",
            "net_amount": "100.00",
            "total_amount": "100.00",
            "vat_rate": "23%",
            "vat_amount": "0.00",
            "vat_treatment": "reverse_charge",
            "supply_type": "services",
            "input_vat_reclaimable": "Yes",
            "status": "Paid",
            "payment_method": "Business Bank",
        },
    )

    response = app.app.test_client().get('/ledger/journal.csv')
    assert response.status_code == 200
    assert response.mimetype == 'text/csv'
    assert 'attachment; filename=ledger-journal.csv' in response.headers['Content-Disposition']
    assert b'vat_rate,vat_treatment,supply_type,vat_amount_eur,net_amount_eur,total_amount_eur,anomaly_flags' in response.data
    assert b'reverse_charge' in response.data


def test_vat_registration_setting_persists(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().post(
        '/business-structure/update?next=/',
        data={"structure": "sole_trader", "transition_date": "", "vat_registered": "1"},
        follow_redirects=True,
    )
    assert response.status_code == 200

    profile = json.loads(app.BUSINESS_PROFILE_PATH.read_text(encoding="utf-8"))
    assert profile["vat_registered"] is True


def test_vat_threshold_basis_setting_persists(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().post(
        '/business-structure/update?next=/',
        data={
            "structure": "sole_trader",
            "transition_date": "",
            "vat_registered": "1",
            "vat_threshold_basis": "goods",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200

    profile = json.loads(app.BUSINESS_PROFILE_PATH.read_text(encoding="utf-8"))
    assert profile["vat_threshold_basis"] == "goods"


def test_vat3_export_returns_ros_style_fields(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/business-structure/update?next=/',
        data={"structure": "sole_trader", "transition_date": "", "vat_registered": "1"},
    )
    app.app.test_client().post(
        '/expenses/add',
        data={
            "date": "2026-08-05",
            "title": "VAT Expense",
            "description": "Hosting",
            "supplier": "Supplier A",
            "supplier_vat_number": "IE1234567A",
            "category": "Professional Fees",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "vat_rate": "23%",
            "vat_amount": "23.00",
            "input_vat_reclaimable": "Yes",
            "status": "Paid",
            "payment_method": "Business Bank",
        },
    )

    response = app.app.test_client().get('/vat3/export.csv')
    assert response.status_code == 200
    assert response.mimetype == 'text/csv'
    assert 'attachment; filename=vat3-summary.csv' in response.headers['Content-Disposition']
    assert b'VAT3_Period,T1,T2,T3,T4,Due_Date' in response.data
    assert b',0.0,23.0,0.0,23.0,' in response.data
    assert b'ZeroRatedSales,ExemptSales,ReverseChargePurchases,Treatment_Notes' in response.data


def test_vat_threshold_summary_warning_and_exceeded_states():
    warning_summary = app._compute_vat_threshold_summary(
        income_rows=[{"Amount (€)": "33600.00"}],
        invoice_rows=[],
        basis="services",
    )
    assert warning_summary["status"] == "warning"
    assert warning_summary["progress_pct"] == 80.0

    exceeded_summary = app._compute_vat_threshold_summary(
        income_rows=[{"Amount (€)": "43000.00"}],
        invoice_rows=[],
        basis="services",
    )
    assert exceeded_summary["status"] == "exceeded"
    assert exceeded_summary["remaining_before_limit"] == 0.0


def test_vat_threshold_summary_splits_services_and_goods_streams():
    summary = app._compute_vat_threshold_summary(
        income_rows=[
            {"Amount (€)": "30000.00", "Supply Type": "services"},
            {"Amount (€)": "20000.00", "Supply Type": "goods"},
        ],
        invoice_rows=[
            {"Total (€)": "1000.00", "Supply Type": "services"},
            {"Total (€)": "40000.00", "Supply Type": "goods"},
        ],
        basis="services",
    )

    stream_trackers = {row["basis"]: row for row in summary["stream_trackers"]}
    assert stream_trackers["services"]["taxable_turnover"] == 30000.0
    assert stream_trackers["goods"]["taxable_turnover"] == 40000.0
    assert stream_trackers["services"]["is_selected"] is True
    assert stream_trackers["goods"]["is_selected"] is False


def test_vat_control_summary_includes_treatment_breakdown():
    period_start, _, _ = app._vat_period_bounds(date.today())
    ledger_entries = [
        {
            "transaction_date": period_start.isoformat(),
            "entity_type": "invoice",
            "vat_treatment": "zero_rated",
            "total_amount_eur": 500.0,
            "journal_lines": [],
        },
        {
            "transaction_date": period_start.isoformat(),
            "entity_type": "income",
            "vat_treatment": "exempt",
            "total_amount_eur": 250.0,
            "journal_lines": [],
        },
        {
            "transaction_date": period_start.isoformat(),
            "entity_type": "expense",
            "vat_treatment": "reverse_charge",
            "total_amount_eur": 100.0,
            "journal_lines": [],
        },
    ]

    summary = app._compute_vat_control_summary(ledger_entries)
    assert summary["zero_rated_sales"] == 500.0
    assert summary["exempt_sales"] == 250.0
    assert summary["reverse_charge_purchases"] == 100.0
    assert "Zero-rated sales" in summary["treatment_notes"]


def test_detect_vat_anomalies_flags_expected_cases():
    ledger_entries = [
        {
            "timestamp": "2026-08-05T10:00:00",
            "entity_type": "invoice",
            "description": "Exempt sale",
            "transaction_date": "2026-08-05",
            "amount_eur": 121.0,
            "total_amount_eur": 121.0,
            "vat_rate": "Exempt",
            "vat_treatment": "exempt",
            "supply_type": "services",
            "vat_amount_eur": 21.0,
        },
        {
            "timestamp": "2026-08-05T10:10:00",
            "entity_type": "expense",
            "description": "Reverse charge cost",
            "transaction_date": "2026-08-05",
            "amount_eur": 123.0,
            "total_amount_eur": 123.0,
            "vat_rate": "23%",
            "vat_treatment": "reverse_charge",
            "supply_type": "services",
            "vat_amount_eur": 23.0,
        },
    ]

    anomalies = app._detect_vat_anomalies(ledger_entries)
    assert len(anomalies) == 2
    assert any("non_zero_vat_with_zero_or_exempt_treatment" in item["flags"] for item in anomalies)
    assert any("reverse_charge_should_not_post_local_vat_amount" in item["flags"] for item in anomalies)


def test_resolve_workbook_path_finds_any_matching_financial_workbook(tmp_path, monkeypatch):
    workbook = tmp_path / "H-Queex_Financial_Control V8.0.xlsm"
    workbook.write_bytes(b"test")

    monkeypatch.setattr(app, "BASE_DIR", tmp_path)
    monkeypatch.setattr(app, "WORKBOOK_PATH", tmp_path / "missing.xlsm")

    assert app._resolve_workbook_path() == workbook


def test_header_aliases_are_resolved_when_reading_a_workbook_sheet(tmp_path):
    # Header-alias resolution is now only exercised by the one-time xlsm->JSON
    # migration and the manual xlsm export, since normal reads/writes go through
    # JSON files with canonical keys. Exercise that surviving code path directly.
    path = tmp_path / "alt.xlsm"
    wb = Workbook()
    wb.remove(wb.active)

    income = wb.create_sheet("Income")
    income.append(["Date", "Description", "Client / Source", "Amount", "Status"])
    income.append(["2026-07-29", "Renamed header", "Client A", 120.0, "Paid"])

    wb.save(path)

    try:
        rows = app._read_workbook_sheet_rows(wb, "Income")
    finally:
        wb.close()

    assert len(rows) == 1
    assert rows[0]["Description"] == "Renamed header"
    assert rows[0]["Amount (€)"] == 120.0


def test_income_add_route_does_not_touch_the_xlsm_workbook(tmp_path):
    # Normal writes are JSON-only now; the xlsm doesn't even need to exist.
    app.WORKBOOK_PATH = tmp_path / "missing.xlsm"
    app.load_finance_data.cache_clear()

    payload = {
        "date": "2026-07-30",
        "description": "Added without workbook",
        "client_source": "Client A",
        "amount": "50.00",
        "status": "Pending",
    }
    response = app.app.test_client().post('/income/add', data=payload, follow_redirects=True)
    assert response.status_code == 200
    assert not (tmp_path / "missing.xlsm").exists()

    income_records = json.loads(app.INCOME_PATH.read_text(encoding="utf-8"))
    assert income_records[-1]["Description"] == "Added without workbook"


def test_due_subscription_posts_expense_and_advances_next_charge(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    subscription_payload = [
        {
            "id": "subscription-1",
            "title": "Adobe Creative Cloud",
            "description": "Creative suite",
            "supplier": "Supplier A",
            "category": "Software",
            "net_amount": 50.0,
            "total_amount": 60.0,
            "frequency": "monthly",
            "start_date": "2026-08-05",
            "next_charge_date": "2026-08-05",
            "last_posted_date": "",
            "end_date": "",
            "status": "active",
            "notes": "Company card",
        }
    ]
    app.SUBSCRIPTIONS_PATH.write_text(json.dumps(subscription_payload), encoding="utf-8")

    result = app._sync_subscriptions_to_expenses(today=date(2026, 8, 5))

    assert result["posted_count"] == 1

    expense_records = json.loads(app.EXPENSES_PATH.read_text(encoding="utf-8"))
    last_record = expense_records[-1]
    assert last_record["Title"] == "Adobe Creative Cloud"
    assert "Subscription charge" in str(last_record.get("Description", ""))
    assert "60" in str(last_record["Total (€)"])

    subscriptions = json.loads(app.SUBSCRIPTIONS_PATH.read_text(encoding="utf-8"))
    assert subscriptions[0]["last_posted_date"] == "2026-08-05"
    assert subscriptions[0]["next_charge_date"] == "2026-09-05"


def test_add_subscription_route_persists_subscription_and_renders_register(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    payload = {
        "title": "Figma",
        "description": "Design collaboration",
        "supplier": "Supplier A",
        "category": "Software",
        "frequency": "monthly",
        "start_date": "2026-08-20",
        "net_amount": "15.00",
        "total_amount": "18.15",
        "status": "active",
        "notes": "Team plan",
    }

    response = app.app.test_client().post('/subscriptions/add', data=payload, follow_redirects=True)

    assert response.status_code == 200
    assert b'Subscription added' in response.data
    assert b'Figma' in response.data

    subscriptions = json.loads(app.SUBSCRIPTIONS_PATH.read_text(encoding="utf-8"))
    assert len(subscriptions) == 1
    assert subscriptions[0]["title"] == "Figma"
    assert subscriptions[0]["next_charge_date"] == "2026-08-20"


def test_update_and_delete_subscription_routes(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.SUBSCRIPTIONS_PATH.write_text(
        json.dumps([
            {
                "id": "sub-1",
                "title": "Adobe",
                "description": "Creative",
                "supplier": "Supplier A",
                "category": "Software",
                "net_amount": 50.0,
                "total_amount": 60.0,
                "frequency": "monthly",
                "start_date": "2026-08-01",
                "next_charge_date": "2026-08-15",
                "last_posted_date": "",
                "end_date": "",
                "status": "active",
                "notes": "note",
            }
        ]),
        encoding="utf-8",
    )

    update_payload = {
        "subscription_id": "sub-1",
        "title": "Adobe CC",
        "description": "Creative suite",
        "supplier": "Supplier A",
        "category": "Software",
        "frequency": "yearly",
        "start_date": "2026-08-01",
        "next_charge_date": "2027-08-01",
        "end_date": "",
        "net_amount": "120.00",
        "total_amount": "145.20",
        "status": "paused",
        "notes": "annual",
    }
    update_response = app.app.test_client().post('/subscriptions/update', data=update_payload, follow_redirects=True)
    assert update_response.status_code == 200
    assert b'Subscription updated' in update_response.data

    subscriptions = json.loads(app.SUBSCRIPTIONS_PATH.read_text(encoding="utf-8"))
    assert subscriptions[0]["title"] == "Adobe CC"
    assert subscriptions[0]["frequency"] == "yearly"

    delete_response = app.app.test_client().post('/subscriptions/delete', data={"subscription_id": "sub-1"}, follow_redirects=True)
    assert delete_response.status_code == 200
    assert b'Subscription archived' in delete_response.data
    assert json.loads(app.SUBSCRIPTIONS_PATH.read_text(encoding="utf-8")) == []

    archives = json.loads(app.ARCHIVE_PATH.read_text(encoding="utf-8"))
    audit_entries = json.loads(app.AUDIT_LOG_PATH.read_text(encoding="utf-8"))
    assert archives[0]["entity_type"] == "subscription"
    assert any(entry["action"] == "archive" for entry in audit_entries)


def test_delete_expense_route_removes_workbook_row(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    data = app.load_finance_data()
    expense_row_number = data["sheets"]["Expenses"][0]["__row_number"]

    response = app.app.test_client().post(
        "/expenses/delete",
        data={"row_number": str(expense_row_number)},
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"Expense archived" in response.data

    expense_records = json.loads(app.EXPENSES_PATH.read_text(encoding="utf-8"))
    assert expense_records == []

    archives = json.loads(app.ARCHIVE_PATH.read_text(encoding="utf-8"))
    assert archives[0]["entity_type"] == "expense"


def test_update_expense_route_updates_existing_workbook_row(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    row_number = app.load_finance_data()["sheets"]["Expenses"][0]["__row_number"]
    payload = {
        "row_number": str(row_number),
        "date": "2026-08-01",
        "title": "Software",
        "description": "Updated expense",
        "supplier": "Supplier A",
        "supplier_vat_number": "IE1234567A",
        "category": "Software",
        "net_amount": "90.00",
        "total_amount": "108.90",
        "input_vat_reclaimable": "Yes",
        "status": "Paid",
        "payment_method": "Business Bank",
    }

    response = app.app.test_client().post('/expenses/update', data=payload, follow_redirects=True)
    assert response.status_code == 200
    assert b'Expense updated' in response.data

    app.load_finance_data.cache_clear()
    expenses = app.load_finance_data()["sheets"]["Expenses"]
    updated = next((r for r in expenses if r.get("__row_number") == row_number), None)
    assert updated is not None
    assert updated.get("Description") == "Updated expense"
    assert str(updated.get("Total (€)")) == "108.90"


def test_expense_capex_auto_routes_to_capital_schedule(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    payload = {
        "date": "2026-08-05",
        "title": "Laptop",
        "description": "MacBook for production",
        "supplier": "Supplier A",
        "supplier_vat_number": "IE1234567A",
        "receipt_reference": "INV-CAP-1",
        "category": "Equipment and Hardware",
        "net_amount": "1500.00",
        "total_amount": "1845.00",
        "vat_rate": "23%",
        "vat_amount": "345.00",
        "input_vat_reclaimable": "Yes",
        "status": "Paid",
        "payment_method": "Business Bank",
    }

    response = app.app.test_client().post('/expenses/add', data=payload, follow_redirects=True)
    assert response.status_code == 200

    assets = json.loads(app.CAPITAL_ASSETS_PATH.read_text(encoding="utf-8"))
    assert len(assets) == 1
    assert assets[0]["source"] == "expense"
    assert assets[0]["cost_eur"] == 1845.0
    assert assets[0]["annual_allowance_eur"] == 230.62


def test_capital_allowances_export_returns_csv(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.CAPITAL_ASSETS_PATH.write_text(
        json.dumps(
            [
                {
                    "id": "expense-9",
                    "source": "expense",
                    "expense_row_number": 9,
                    "acquisition_date": "2026-08-01",
                    "supplier": "Supplier A",
                    "description": "Camera rig",
                    "category": "Equipment and Hardware",
                    "cost_eur": 1200.0,
                    "allowance_rate": 0.125,
                    "allowance_years": 8,
                    "annual_allowance_eur": 150.0,
                    "phase_tag": "Phase 1",
                    "active": True,
                }
            ]
        ),
        encoding="utf-8",
    )

    response = app.app.test_client().get('/capital-allowances/export.csv')
    assert response.status_code == 200
    assert response.mimetype == 'text/csv'
    assert 'attachment; filename=capital-allowances.csv' in response.headers['Content-Disposition']
    assert b'asset_id,acquisition_date,supplier,description,category,cost_eur,allowance_rate,allowance_years,annual_allowance_eur,phase_tag,active' in response.data
    assert b'expense-9' in response.data


def test_add_payroll_route_persists_register_and_posts_ledger_entry(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    payload = {
        "pay_date": "2026-08-05",
        "payroll_period": "2026-08",
        "employee_name": "Hev Team Member",
        "gross_pay": "3000.00",
        "paye": "450.00",
        "usc": "120.00",
        "employee_prsi": "120.00",
        "employer_prsi": "330.00",
        "status": "Paid",
        "payment_method": "Business Bank",
        "payment_date": "2026-08-05",
        "bank_reconciliation": "Reconciled",
        "notes": "Monthly payroll run",
    }

    response = app.app.test_client().post('/payroll/add', data=payload, follow_redirects=True)
    assert response.status_code == 200
    assert b'Payroll entry added' in response.data

    payroll_entries = json.loads(app.PAYROLL_PATH.read_text(encoding="utf-8"))
    assert len(payroll_entries) == 1
    assert payroll_entries[0]["Employee Name"] == "Hev Team Member"
    assert payroll_entries[0]["Net Pay (€)"] == "2310.00"
    assert payroll_entries[0]["Employer Cost (€)"] == "3330.00"

    ledger_entries = json.loads(app.LEDGER_JOURNAL_PATH.read_text(encoding="utf-8"))
    payroll_ledger = [entry for entry in ledger_entries if entry.get("entity_type") == "payroll"]
    assert payroll_ledger
    latest = payroll_ledger[-1]
    assert latest["entry_balanced"] is True
    assert any(line["account_code"] == "5300" and line["debit"] == 3000.0 for line in latest["journal_lines"])
    assert any(line["account_code"] == "2200" and line["credit"] == 1020.0 for line in latest["journal_lines"])


def test_payroll_validation_and_export_route(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    invalid_payload = {
        "pay_date": "2026-08-05",
        "employee_name": "Invalid Payroll",
        "gross_pay": "1000.00",
        "paye": "700.00",
        "usc": "200.00",
        "employee_prsi": "200.00",
        "employer_prsi": "0.00",
        "status": "Draft",
    }
    invalid_response = app.app.test_client().post('/payroll/add', data=invalid_payload, follow_redirects=True)
    assert invalid_response.status_code == 200
    assert b'Validation:' in invalid_response.data
    assert b'Gross pay must be at least employee deductions total' in invalid_response.data

    valid_payload = {
        "pay_date": "2026-08-12",
        "payroll_period": "2026-08",
        "employee_name": "Valid Payroll",
        "gross_pay": "1200.00",
        "paye": "150.00",
        "usc": "36.00",
        "employee_prsi": "48.00",
        "employer_prsi": "132.00",
        "status": "Approved",
        "bank_reconciliation": "Unreconciled",
    }
    app.app.test_client().post('/payroll/add', data=valid_payload, follow_redirects=True)

    export_response = app.app.test_client().get('/payroll/export.csv')
    assert export_response.status_code == 200
    assert export_response.mimetype == 'text/csv'
    assert 'attachment; filename=payroll-register.csv' in export_response.headers['Content-Disposition']
    assert b'pay_date,payroll_period,employee_name,gross_pay_eur,paye_eur,usc_eur,employee_prsi_eur,employer_prsi_eur,net_pay_eur,employer_cost_eur,status,payment_method,bank_reconciliation,notes,phase_tag' in export_response.data
    assert b'Valid Payroll' in export_response.data


def test_expense_paid_requires_payment_method(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().post(
        '/expenses/add',
        data={
            "date": "2026-08-05",
            "title": "Missing payment method",
            "description": "Validation check",
            "supplier": "Supplier A",
            "supplier_vat_number": "IE1234567A",
            "category": "Professional Fees",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "input_vat_reclaimable": "Yes",
            "status": "Paid",
            "bank_reconciliation": "Unreconciled",
            "payment_method": "",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b'Payment method is required when status is Paid' in response.data


def test_invoice_paid_defaults_payment_date_from_issue_date():
    payload = {
        "Status": "Paid",
        "Issue Date": "2026-08-10",
        "Payment Date": "",
    }

    applied = app._apply_default_payment_date_for_paid(payload, "invoice", "Issue Date")

    assert applied is True
    assert payload["Payment Date"] == "2026-08-10"


def test_payroll_paid_defaults_payment_date_from_pay_date(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().post(
        '/payroll/add',
        data={
            "pay_date": "2026-08-05",
            "payroll_period": "2026-08",
            "employee_name": "Default Payment Date Payroll",
            "gross_pay": "1000.00",
            "paye": "100.00",
            "usc": "30.00",
            "employee_prsi": "40.00",
            "employer_prsi": "110.00",
            "status": "Paid",
            "payment_method": "Business Bank",
            "payment_date": "",
            "bank_reconciliation": "Unreconciled",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b'Payroll entry added' in response.data
    assert b'payment date auto-filled from pay date' in response.data

    payroll_entries = json.loads(app.PAYROLL_PATH.read_text(encoding="utf-8"))
    target = next(item for item in payroll_entries if item.get("Employee Name") == "Default Payment Date Payroll")
    assert str(target.get("Payment Date") or "") == "2026-08-05"


def test_unpaid_entries_allow_blank_payment_fields(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    expense_response = app.app.test_client().post(
        '/expenses/add',
        data={
            "date": "2026-08-05",
            "title": "Draft expense",
            "description": "Blank payment fields allowed",
            "supplier": "Supplier A",
            "supplier_vat_number": "IE1234567A",
            "category": "Professional Fees",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "input_vat_reclaimable": "Yes",
            "status": "Draft",
            "payment_method": "",
            "bank_reconciliation": "Unreconciled",
        },
        follow_redirects=True,
    )
    assert expense_response.status_code == 200
    assert b'Expense entry added' in expense_response.data

    invoice_response = app.app.test_client().post(
        '/invoices/add',
        data={
            "issue_date": "2026-08-10",
            "due_date": "2026-08-20",
            "client_name": "Client A",
            "client_vat_number": "IE1234567A",
            "client_address": "1 Example Street, Dublin",
            "service_product": "Draft invoice",
            "line_items_json": _invoice_line_items_payload("Draft invoice", 100.00),
            "balance_due": "123.00",
            "status": "Draft",
            "payment_method": "",
            "payment_date": "",
            "bank_reconciliation": "Unreconciled",
        },
        follow_redirects=True,
    )
    assert invoice_response.status_code == 200
    assert b'Invoice added' in invoice_response.data

    payroll_response = app.app.test_client().post(
        '/payroll/add',
        data={
            "pay_date": "2026-08-12",
            "payroll_period": "2026-08",
            "employee_name": "Draft Payroll",
            "gross_pay": "1200.00",
            "paye": "150.00",
            "usc": "36.00",
            "employee_prsi": "48.00",
            "employer_prsi": "132.00",
            "status": "Draft",
            "payment_method": "",
            "payment_date": "",
            "bank_reconciliation": "Unreconciled",
        },
        follow_redirects=True,
    )
    assert payroll_response.status_code == 200
    assert b'Payroll entry added' in payroll_response.data


def test_reconciliation_exports_include_queue_and_exceptions(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/expenses/add',
        data={
            "date": "2026-07-01",
            "title": "Old paid expense",
            "description": "Needs reconciliation",
            "supplier": "Supplier A",
            "supplier_vat_number": "IE1234567A",
            "category": "Professional Fees",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "input_vat_reclaimable": "Yes",
            "status": "Paid",
            "bank_reconciliation": "Unreconciled",
                "payment_method": "Business Bank",
        },
    )

    app.app.test_client().post(
        '/invoices/add',
        data={
            "issue_date": "2026-07-01",
            "due_date": "2026-07-31",
            "client_name": "Client A",
            "client_vat_number": "IE1234567A",
            "client_address": "1 Example Street, Dublin",
            "service_product": "Reconciliation test",
            "line_items_json": _invoice_line_items_payload("Reconciliation test", 200.00),
            "balance_due": "0.00",
            "status": "Issued",
            "bank_reconciliation": "Unreconciled",
        },
    )
    app.load_finance_data.cache_clear()
    reconciliation_invoice = next(
        row for row in app.load_finance_data()["sheets"]["Invoices"] if row.get("Service / Product") == "Reconciliation test"
    )
    app.app.test_client().post(
        '/invoices/record-payment',
        data={
            "row_number": str(reconciliation_invoice["__row_number"]),
            "amount_received": str(reconciliation_invoice.get("Total (€)") or "246.00"),
            "payment_date": "2026-07-01",
            "payment_method": "Business Bank",
        },
    )

    app.app.test_client().post(
        '/payroll/add',
        data={
            "pay_date": "2026-07-01",
            "payroll_period": "2026-07",
            "employee_name": "Recon Payroll",
            "gross_pay": "1800.00",
            "paye": "200.00",
            "usc": "50.00",
            "employee_prsi": "70.00",
            "employer_prsi": "198.00",
            "status": "Paid",
            "bank_reconciliation": "Unreconciled",
                "payment_method": "Business Bank",
                "payment_date": "2026-07-01",
        },
    )

    # Simulate a paid invoice that legitimately has no payment date recorded
    # (e.g. edited outside the app's normal add/update validation flow), which
    # is what the missing_payment_date exception exists to catch.
    app.load_finance_data.cache_clear()
    invoice_rows = app.load_finance_data()["sheets"]["Invoices"]
    target_invoice = next(row for row in invoice_rows if row.get("Service / Product") == "Reconciliation test")
    target_invoice["Payment Date"] = ""
    app._update_row_in_sheet("Invoices", target_invoice["__row_number"], target_invoice)
    app.load_finance_data.cache_clear()

    queue_response = app.app.test_client().get('/reconciliation/export.csv')
    assert queue_response.status_code == 200
    assert queue_response.mimetype == 'text/csv'
    assert b'entity_type,reference,counterparty,date,amount_eur,status,bank_reconciliation,payment_method,is_paid,age_days,exception_reasons,matching_group_size' in queue_response.data

    exceptions_response = app.app.test_client().get('/reconciliation/exceptions.csv')
    assert exceptions_response.status_code == 200
    assert b'paid_unreconciled_over_7_days' in exceptions_response.data
    assert b'missing_payment_date' in exceptions_response.data


def test_mark_reconciliation_updates_payroll_and_expense_routes(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    expense_response = app.app.test_client().post(
        '/expenses/add',
        data={
            "date": "2026-08-05",
            "title": "Recon Expense",
            "description": "Mark endpoint check",
            "supplier": "Supplier A",
            "supplier_vat_number": "IE1234567A",
            "category": "Professional Fees",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "input_vat_reclaimable": "Yes",
            "status": "Paid",
            "bank_reconciliation": "Unreconciled",
            "payment_method": "Business Bank",
        },
        follow_redirects=True,
    )
    assert expense_response.status_code == 200

    app.load_finance_data.cache_clear()
    expense_rows = app.load_finance_data()["sheets"]["Expenses"]
    expense_row_number = expense_rows[-1]["__row_number"]
    mark_expense_response = app.app.test_client().post(
        '/reconciliation/mark',
        data={
            "entity_type": "expense",
            "row_number": str(expense_row_number),
            "bank_reconciliation": "Reconciled",
            "return_to": "/ledger",
        },
        follow_redirects=True,
    )
    assert mark_expense_response.status_code == 200
    assert b'Expense reconciliation updated' in mark_expense_response.data

    app.app.test_client().post(
        '/payroll/add',
        data={
            "pay_date": "2026-08-05",
            "payroll_period": "2026-08",
            "employee_name": "Payroll Mark",
            "gross_pay": "1000.00",
            "paye": "100.00",
            "usc": "30.00",
            "employee_prsi": "40.00",
            "employer_prsi": "110.00",
            "status": "Paid",
            "bank_reconciliation": "Unreconciled",
            "payment_method": "Business Bank",
            "payment_date": "2026-08-05",
        },
    )
    payroll_entries = json.loads(app.PAYROLL_PATH.read_text(encoding="utf-8"))
    payroll_id = payroll_entries[0]["id"]

    mark_payroll_response = app.app.test_client().post(
        '/reconciliation/mark',
        data={
            "entity_type": "payroll",
            "payroll_id": payroll_id,
            "bank_reconciliation": "Reconciled",
            "return_to": "/ledger",
        },
        follow_redirects=True,
    )
    assert mark_payroll_response.status_code == 200
    assert b'Payroll reconciliation updated' in mark_payroll_response.data

    updated_payroll_entries = json.loads(app.PAYROLL_PATH.read_text(encoding="utf-8"))
    updated = next(entry for entry in updated_payroll_entries if entry["id"] == payroll_id)
    assert updated["Bank Reconciliation"] == "Reconciled"


def test_mark_reconciliation_blocks_unpaid_expense(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/expenses/add',
        data={
            "date": "2026-08-05",
            "title": "Unpaid expense",
            "description": "Should not reconcile",
            "supplier": "Supplier A",
            "supplier_vat_number": "IE1234567A",
            "category": "Professional Fees",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "input_vat_reclaimable": "Yes",
            "status": "Draft",
            "bank_reconciliation": "Unreconciled",
            "payment_method": "",
        },
        follow_redirects=True,
    )

    app.load_finance_data.cache_clear()
    expense_row_number = app.load_finance_data()["sheets"]["Expenses"][-1]["__row_number"]
    mark_response = app.app.test_client().post(
        '/reconciliation/mark',
        data={
            "entity_type": "expense",
            "row_number": str(expense_row_number),
            "bank_reconciliation": "Reconciled",
            "return_to": "/ledger",
        },
        follow_redirects=True,
    )
    assert mark_response.status_code == 200
    assert b'Reconciliation update failed: Record must have a paid status before it can be reconciled' in mark_response.data

    app.load_finance_data.cache_clear()
    row = next(item for item in app.load_finance_data()["sheets"]["Expenses"] if item["__row_number"] == expense_row_number)
    assert str(row.get("Bank Reconciliation") or "Unreconciled") != "Reconciled"


def test_mark_reconciliation_autofills_payment_date_for_paid_payroll(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/payroll/add',
        data={
            "pay_date": "2026-08-05",
            "payroll_period": "2026-08",
            "employee_name": "Legacy Payroll",
            "gross_pay": "1000.00",
            "paye": "100.00",
            "usc": "30.00",
            "employee_prsi": "40.00",
            "employer_prsi": "110.00",
            "status": "Draft",
            "bank_reconciliation": "Unreconciled",
            "payment_method": "",
            "payment_date": "",
        },
        follow_redirects=True,
    )

    payroll_entries = json.loads(app.PAYROLL_PATH.read_text(encoding="utf-8"))
    payroll_entry = payroll_entries[0]
    payroll_entry["Status"] = "Paid"
    payroll_entry["Payment Method"] = "Business Bank"
    payroll_entry["Payment Date"] = ""
    payroll_entry["Bank Reconciliation"] = "Unreconciled"
    app._save_payroll_entries(payroll_entries)

    mark_response = app.app.test_client().post(
        '/reconciliation/mark',
        data={
            "entity_type": "payroll",
            "payroll_id": payroll_entry["id"],
            "bank_reconciliation": "Reconciled",
            "return_to": "/ledger",
        },
        follow_redirects=True,
    )
    assert mark_response.status_code == 200
    assert b'Payroll reconciliation updated' in mark_response.data

    updated_entries = json.loads(app.PAYROLL_PATH.read_text(encoding="utf-8"))
    updated = next(item for item in updated_entries if item["id"] == payroll_entry["id"])
    assert updated["Bank Reconciliation"] == "Reconciled"
    assert str(updated.get("Payment Date") or "") == "2026-08-05"


def test_mark_reconciliation_autofills_payment_date_for_paid_invoice(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    add_response = app.app.test_client().post(
        '/invoices/add',
        data={
            "issue_date": "2026-08-05",
            "due_date": "2026-08-20",
            "client_name": "Client A",
            "client_vat_number": "IE1234567A",
            "client_address": "1 Example Street, Dublin",
            "service_product": "Recon Invoice Autofill",
            "line_items_json": _invoice_line_items_payload("Recon Invoice Autofill", 100.00),
            "balance_due": "123.00",
            "status": "Draft",
            "payment_method": "",
            "payment_date": "",
            "bank_reconciliation": "Unreconciled",
        },
        follow_redirects=True,
    )
    assert add_response.status_code == 200

    app.load_finance_data.cache_clear()
    invoice_rows = app.load_finance_data()["sheets"]["Invoices"]
    target = next(row for row in invoice_rows if row.get("Service / Product") == "Recon Invoice Autofill")
    row_number = target["__row_number"]

    target["Status"] = "Paid"
    target["Payment Method"] = "Business Bank"
    target["Payment Date"] = ""
    target["Bank Reconciliation"] = "Unreconciled"
    app._update_row_in_sheet("Invoices", row_number, target)
    app.load_finance_data.cache_clear()

    mark_response = app.app.test_client().post(
        '/reconciliation/mark',
        data={
            "entity_type": "invoice",
            "row_number": str(row_number),
            "bank_reconciliation": "Reconciled",
            "return_to": "/ledger",
        },
        follow_redirects=True,
    )
    assert mark_response.status_code == 200
    assert b'Invoice reconciliation updated' in mark_response.data

    app.load_finance_data.cache_clear()
    updated = next(item for item in app.load_finance_data()["sheets"]["Invoices"] if item["__row_number"] == row_number)
    assert str(updated.get("Bank Reconciliation") or "Reconciled") == "Reconciled"
    if "Payment Date" in updated:
        assert str(updated.get("Payment Date") or "") == "2026-08-05"


def test_bank_statement_import_and_unmatched_export(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/expenses/add',
        data={
            "date": "2026-08-05",
            "description": "Should match statement",
            "supplier": "Supplier A",
            "supplier_vat_number": "IE1234567A",
            "category": "Professional Fees",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "input_vat_reclaimable": "Yes",
            "status": "Paid",
            "payment_method": "Business Bank",
            "bank_reconciliation": "Unreconciled",
        },
        follow_redirects=True,
    )

    statement_csv = "Date,Description,Reference,Amount,Payment Method\n2026-08-05,Expense Payment,EXP-1,-123.00,Business Bank\n2026-08-05,Unknown Item,UNK-1,-55.00,Business Bank\n"
    import_response = app.app.test_client().post(
        '/reconciliation/import-statement',
        data={
            "return_to": "/ledger",
            "statement_file": (BytesIO(statement_csv.encode("utf-8")), "statement.csv"),
        },
        content_type='multipart/form-data',
        follow_redirects=True,
    )
    assert import_response.status_code == 200
    assert b'Bank statement imported:' in import_response.data

    queue_response = app.app.test_client().get('/reconciliation/export.csv')
    assert queue_response.status_code == 200
    assert b'statement_match_count' not in queue_response.data
    assert b'Should match statement' in queue_response.data

    bank_lines_response = app.app.test_client().get('/reconciliation/bank-statements.csv')
    assert bank_lines_response.status_code == 200
    assert b'date,description,reference,amount_eur,balance_eur,payment_method,matched_entity_type,matched_reference,source_filename,uploaded_at' in bank_lines_response.data
    assert b'Unknown Item' in bank_lines_response.data

    unmatched_response = app.app.test_client().get('/reconciliation/unmatched-bank-statements.csv')
    assert unmatched_response.status_code == 200
    assert b'Unknown Item' in unmatched_response.data
    assert b'EXP-1' not in unmatched_response.data


def test_apply_suggested_reconciliation_marks_matched_payroll(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/payroll/add',
        data={
            "pay_date": "2026-08-05",
            "payroll_period": "2026-08",
            "employee_name": "Batch Recon Payroll",
            "gross_pay": "1000.00",
            "paye": "100.00",
            "usc": "30.00",
            "employee_prsi": "40.00",
            "employer_prsi": "110.00",
            "status": "Paid",
            "bank_reconciliation": "Unreconciled",
            "payment_method": "Business Bank",
            "payment_date": "2026-08-05",
        },
        follow_redirects=True,
    )

    statement_csv = "Date,Description,Reference,Amount,Payment Method\n2026-08-05,Payroll Payment,PAY-1,-830.00,Business Bank\n"
    app.app.test_client().post(
        '/reconciliation/import-statement',
        data={
            "return_to": "/ledger",
            "statement_file": (BytesIO(statement_csv.encode("utf-8")), "statement.csv"),
        },
        content_type='multipart/form-data',
        follow_redirects=True,
    )

    batch_response = app.app.test_client().post(
        '/reconciliation/apply-suggested',
        data={"return_to": "/ledger"},
        follow_redirects=True,
    )
    assert batch_response.status_code == 200
    assert b'Applied 1 suggested reconciliation matches' in batch_response.data

    payroll_entries = json.loads(app.PAYROLL_PATH.read_text(encoding="utf-8"))
    entry = next(item for item in payroll_entries if item["Employee Name"] == "Batch Recon Payroll")
    assert entry["Bank Reconciliation"] == "Reconciled"


def test_apply_suggested_reconciliation_skips_ambiguous_groups(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    for title in ["Ambiguous Expense A", "Ambiguous Expense B"]:
        app.app.test_client().post(
            '/expenses/add',
            data={
                "date": "2026-08-05",
                "description": title,
                "supplier": "Supplier A",
                "supplier_vat_number": "IE1234567A",
                "category": "Professional Fees",
                "net_amount": "100.00",
                "total_amount": "123.00",
                "input_vat_reclaimable": "Yes",
                "status": "Paid",
                "payment_method": "Business Bank",
                "bank_reconciliation": "Unreconciled",
            },
            follow_redirects=True,
        )

    statement_csv = "Date,Description,Reference,Amount,Payment Method\n2026-08-05,Single line,AMB-1,-123.00,Business Bank\n"
    app.app.test_client().post(
        '/reconciliation/import-statement',
        data={
            "return_to": "/ledger",
            "statement_file": (BytesIO(statement_csv.encode("utf-8")), "statement.csv"),
        },
        content_type='multipart/form-data',
        follow_redirects=True,
    )

    batch_response = app.app.test_client().post(
        '/reconciliation/apply-suggested',
        data={"return_to": "/ledger"},
        follow_redirects=True,
    )
    assert batch_response.status_code == 200

    app.load_finance_data.cache_clear()
    expenses = app.load_finance_data()["sheets"]["Expenses"]
    ambiguous = [row for row in expenses if str(row.get("Description") or "") in {"Ambiguous Expense A", "Ambiguous Expense B"}]
    assert len(ambiguous) == 2
    assert all(str(row.get("Bank Reconciliation") or "") != "Reconciled" for row in ambiguous)


def test_invoice_crud_routes_update_and_remove_rows(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    add_payload = {
        "issue_date": "2026-08-01",
        "due_date": "2026-08-31",
        "client_name": "Client A",
        "client_vat_number": "IE1234567A",
        "client_address": "1 Example Street, Dublin",
        "service_product": "Web design",
        "line_items_json": _invoice_line_items_payload("Web design", 300.00),
        "balance_due": "363.00",
        "status": "Draft",
    }
    add_response = app.app.test_client().post('/invoices/add', data=add_payload, follow_redirects=True)
    assert add_response.status_code == 200
    assert b'Invoice added' in add_response.data

    app.load_finance_data.cache_clear()
    invoice_rows = app.load_finance_data()["sheets"]["Invoices"]
    target = next(row for row in invoice_rows if str(row["Invoice #"]).startswith("HQ-2026-"))

    update_payload = {
        "row_number": str(target["__row_number"]),
        "invoice_number": str(target["Invoice #"]),
        "issue_date": "2026-08-02",
        "due_date": "2026-09-01",
        "client_name": "Client A",
        "client_vat_number": "IE1234567A",
        "client_address": "1 Example Street, Dublin",
        "service_product": "Updated service",
        "line_items_json": _invoice_line_items_payload("Updated service", 350.00),
        "balance_due": "200.00",
        "status": "Sent",
    }
    update_response = app.app.test_client().post('/invoices/update', data=update_payload, follow_redirects=True)
    assert update_response.status_code == 200
    assert b'Invoice updated' in update_response.data

    delete_response = app.app.test_client().post('/invoices/cancel', data={"row_number": str(target["__row_number"])}, follow_redirects=True)
    assert delete_response.status_code == 200
    assert b'Invoice cancelled and retained for audit trail' in delete_response.data

    app.load_finance_data.cache_clear()
    retained_invoice = next(row for row in app.load_finance_data()["sheets"]["Invoices"] if row["__row_number"] == target["__row_number"])
    assert retained_invoice["Status"] == "Cancelled"


def test_invoice_validation_blocks_invalid_due_date(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    add_payload = {
        "issue_date": "2026-08-10",
        "due_date": "2026-08-01",
        "client_name": "Client A",
        "client_vat_number": "IE1234567A",
        "client_address": "1 Example Street, Dublin",
        "service_product": "Validation",
        "net_amount": "100.00",
        "total_amount": "121.00",
        "balance_due": "121.00",
        "status": "Draft",
    }
    response = app.app.test_client().post('/invoices/add', data=add_payload, follow_redirects=True)
    assert response.status_code == 200
    assert b'Validation:' in response.data


def test_invoice_numbers_auto_generate_sequentially(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    payload = {
        "issue_date": "2026-08-10",
        "due_date": "2026-08-20",
        "client_name": "Client A",
        "client_vat_number": "IE1234567A",
        "client_address": "1 Example Street, Dublin",
        "service_product": "Service A",
        "line_items_json": _invoice_line_items_payload("Service A", 100.00),
        "balance_due": "123.00",
        "status": "Issued",
    }

    response_one = app.app.test_client().post('/invoices/add', data=payload, follow_redirects=True)
    assert response_one.status_code == 200
    response_two = app.app.test_client().post('/invoices/add', data=payload, follow_redirects=True)
    assert response_two.status_code == 200

    app.load_finance_data.cache_clear()
    rows = app.load_finance_data()["sheets"]["Invoices"]
    generated = sorted([str(row.get("Invoice #")) for row in rows if str(row.get("Invoice #", "")).startswith("HQ-2026-")])
    assert generated[-2:] == ["HQ-2026-001", "HQ-2026-002"]


def test_client_and_supplier_update_and_delete_routes(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    client_row = app.load_finance_data()["sheets"]["Clients"][0]["__row_number"]
    client_update = {
        "row_number": str(client_row),
        "client_name": "Client A Updated",
        "contact_person": "Jane Doe",
        "email": "jane.doe@example.com",
        "phone": "999",
        "country": "Belgium",
    }
    client_response = app.app.test_client().post('/clients/update', data=client_update, follow_redirects=True)
    assert client_response.status_code == 200
    assert b'Client updated' in client_response.data

    supplier_row = app.load_finance_data()["sheets"]["Suppliers"][0]["__row_number"]
    supplier_update = {
        "row_number": str(supplier_row),
        "supplier_name": "Supplier A Updated",
        "contact_person": "John Doe",
        "email": "john.doe@example.com",
        "phone": "888",
        "country": "Netherlands",
        "default_vat_treatment": "Reverse charge",
    }
    supplier_response = app.app.test_client().post('/suppliers/update', data=supplier_update, follow_redirects=True)
    assert supplier_response.status_code == 200
    assert b'Supplier updated' in supplier_response.data

    client_delete = app.app.test_client().post('/clients/delete', data={"row_number": str(client_row)}, follow_redirects=True)
    supplier_delete = app.app.test_client().post('/suppliers/delete', data={"row_number": str(supplier_row)}, follow_redirects=True)
    assert client_delete.status_code == 200
    assert supplier_delete.status_code == 200
    assert b'Client archived' in client_delete.data
    assert b'Supplier archived' in supplier_delete.data


def test_archive_view_shows_archive_and_audit_records(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app._archive_record("client", {"Client Name": "Archived Client"}, source="workbook")
    app._record_audit("update", "client", {"row_number": 2})

    response = app.app.test_client().get('/archive')
    assert response.status_code == 200
    assert b'Archive &amp; Audit' in response.data
    assert b'Archived Client' in response.data
    assert b'update' in response.data
    assert b'Restore' in response.data


def test_expense_archive_returns_message_when_workbook_is_locked(workbook_copy, monkeypatch):
    def raise_lock_error(*args, **kwargs):
        raise app.WorkbookWriteError("Workbook is locked. Close Excel and try again.")

    monkeypatch.setattr(app, "_delete_row_from_sheet", raise_lock_error)
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().post('/expenses/delete', data={'row_number': '1'}, follow_redirects=True)

    assert response.status_code == 200
    assert b'Workbook is locked' in response.data


def test_restore_archived_expense_recreates_workbook_row_and_logs_audit(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app._archive_record(
        "expense",
        {
            "Date (Registered)": "2026-08-03",
            "Title": "Recovered Expense",
            "Description": "Restored row",
            "Supplier / Payee": "Supplier A",
            "Category": "Software",
            "Net Amount (€)": "50.00",
            "Total (€)": "60.50",
            "Status": "Archived",
        },
        source="workbook",
    )
    archive_id = json.loads(app.ARCHIVE_PATH.read_text(encoding="utf-8"))[0]["id"]

    response = app.app.test_client().post('/archive/restore', data={"archive_id": archive_id}, follow_redirects=True)
    assert response.status_code == 200
    assert b'Expense restored' in response.data

    app.load_finance_data.cache_clear()
    app.load_finance_data()
    expense_records = json.loads(app.EXPENSES_PATH.read_text(encoding="utf-8"))
    assert any(record.get("Description") == "Recovered Expense — Restored row" for record in expense_records)
    assert json.loads(app.ARCHIVE_PATH.read_text(encoding="utf-8")) == []
    audit_entries = json.loads(app.AUDIT_LOG_PATH.read_text(encoding="utf-8"))
    assert any(entry["action"] == "restore" and entry["entity_type"] == "expense" for entry in audit_entries)


def test_restore_archived_subscription_recreates_subscription_and_logs_audit(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app._archive_record(
        "subscription",
        {
            "id": "restore-sub-1",
            "title": "Restored Subscription",
            "description": "Recovered",
            "supplier": "Supplier A",
            "category": "Software",
            "net_amount": 10.0,
            "total_amount": 12.1,
            "frequency": "monthly",
            "start_date": "2026-08-01",
            "next_charge_date": "2026-09-01",
            "last_posted_date": "",
            "end_date": "",
            "status": "active",
            "notes": "restored",
        },
        source="subscriptions",
    )
    archive_id = json.loads(app.ARCHIVE_PATH.read_text(encoding="utf-8"))[0]["id"]

    response = app.app.test_client().post('/archive/restore', data={"archive_id": archive_id}, follow_redirects=True)
    assert response.status_code == 200
    assert b'Subscription restored' in response.data

    subscriptions = json.loads(app.SUBSCRIPTIONS_PATH.read_text(encoding="utf-8"))
    assert subscriptions[0]["title"] == "Restored Subscription"
    audit_entries = json.loads(app.AUDIT_LOG_PATH.read_text(encoding="utf-8"))
    assert any(entry["action"] == "restore" and entry["entity_type"] == "subscription" for entry in audit_entries)


def test_restore_conflict_is_detected_until_force_restore(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.load_finance_data()  # trigger the Title-into-Description migration for row 1 first

    app._archive_record(
        "expense",
        {
            "Date (Registered)": "2026-07-29",
            "Title": "",
            "Description": "Travel — Hotel",
            "Supplier / Payee": "Supplier A",
            "Category": "Travel",
            "Net Amount (€)": 100.0,
            "Total (€)": 120.0,
            "Status": "Pending",
        },
        source="workbook",
    )
    archive_id = json.loads(app.ARCHIVE_PATH.read_text(encoding="utf-8"))[0]["id"]

    conflict_response = app.app.test_client().post('/archive/restore', data={"archive_id": archive_id}, follow_redirects=True)
    assert conflict_response.status_code == 200
    assert b'Restore conflict detected for expense' in conflict_response.data
    assert len(json.loads(app.ARCHIVE_PATH.read_text(encoding="utf-8"))) == 1

    force_response = app.app.test_client().post('/archive/restore', data={"archive_id": archive_id, "force_restore": "1"}, follow_redirects=True)
    assert force_response.status_code == 200
    assert b'Expense restored' in force_response.data
    assert json.loads(app.ARCHIVE_PATH.read_text(encoding="utf-8")) == []

    audit_entries = json.loads(app.AUDIT_LOG_PATH.read_text(encoding="utf-8"))
    assert any(entry["action"] == "restore_conflict" for entry in audit_entries)
    assert any(entry["action"] == "restore" for entry in audit_entries)


def test_audit_csv_export_returns_csv_download(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app._record_audit("update", "expense", {"row_number": 3, "status": "Paid"})

    response = app.app.test_client().get('/audit/export.csv')
    assert response.status_code == 200
    assert response.mimetype == 'text/csv'
    assert 'attachment; filename=audit-log.csv' in response.headers['Content-Disposition']
    assert b'timestamp,action,entity_type,details_json' in response.data
    assert b'update,expense' in response.data


def _add_draft_invoice(client, *, net_amount="500.00", total_amount="500.00"):
    add_payload = {
        "issue_date": "2026-08-01",
        "due_date": "2026-08-31",
        "client_name": "Client A",
        "client_vat_number": "IE1234567A",
        "client_address": "1 Example Street, Dublin",
        "service_product": "Consulting engagement",
        "line_items_json": _invoice_line_items_payload("Consulting engagement", float(net_amount), vat_rate="0%"),
        "balance_due": total_amount,
        "status": "Draft",
    }
    client.post('/invoices/add', data=add_payload, follow_redirects=True)
    app.load_finance_data.cache_clear()
    invoice_rows = app.load_finance_data()["sheets"]["Invoices"]
    return next(row for row in invoice_rows if row["Service / Product"] == "Consulting engagement")


def test_invoice_marked_paid_creates_linked_income_entry(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    invoice = _add_draft_invoice(client)
    response = client.post(
        '/invoices/record-payment',
        data={
            "row_number": str(invoice["__row_number"]),
            "amount_received": "500.00",
            "payment_date": "2026-08-05",
            "payment_method": "Business Bank Account",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b'Invoice marked as paid' in response.data

    app.load_finance_data.cache_clear()
    updated_invoice = next(row for row in app.load_finance_data()["sheets"]["Invoices"] if row["__row_number"] == invoice["__row_number"])
    assert updated_invoice["Status"] == "Paid"
    assert app._coerce_number(updated_invoice["Balance Due (€)"]) == 0

    income_rows = app.load_finance_data()["sheets"]["Income"]
    linked = [row for row in income_rows if row.get("Invoice ID") == updated_invoice["Invoice #"]]
    assert len(linked) == 1
    assert linked[0]["Source"] == "invoiced"
    assert linked[0]["Status"] == "Received"
    assert app._coerce_number(linked[0]["Total incl. VAT (€)"]) == 500.0


def test_invoice_partial_payment_creates_partial_income_entry_and_updates_on_second_payment(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    invoice = _add_draft_invoice(client, net_amount="1000.00", total_amount="1000.00")
    first_payment = client.post(
        '/invoices/record-payment',
        data={
            "row_number": str(invoice["__row_number"]),
            "amount_received": "400.00",
            "payment_date": "2026-08-05",
            "payment_method": "Cash",
        },
        follow_redirects=True,
    )
    assert b'Partial payment recorded' in first_payment.data

    app.load_finance_data.cache_clear()
    after_first = next(row for row in app.load_finance_data()["sheets"]["Invoices"] if row["__row_number"] == invoice["__row_number"])
    assert after_first["Status"] == "Partially Paid"
    assert app._coerce_number(after_first["Balance Due (€)"]) == 600.0

    income_rows = app.load_finance_data()["sheets"]["Income"]
    linked = [row for row in income_rows if row.get("Invoice ID") == after_first["Invoice #"]]
    assert len(linked) == 1
    assert app._coerce_number(linked[0]["Total incl. VAT (€)"]) == 400.0

    second_payment = client.post(
        '/invoices/record-payment',
        data={
            "row_number": str(invoice["__row_number"]),
            "amount_received": "600.00",
            "payment_date": "2026-08-10",
            "payment_method": "Cash",
        },
        follow_redirects=True,
    )
    assert b'Invoice marked as paid' in second_payment.data

    app.load_finance_data.cache_clear()
    after_second = next(row for row in app.load_finance_data()["sheets"]["Invoices"] if row["__row_number"] == invoice["__row_number"])
    assert after_second["Status"] == "Paid"
    assert app._coerce_number(after_second["Balance Due (€)"]) == 0

    income_rows_after = app.load_finance_data()["sheets"]["Income"]
    linked_after = [row for row in income_rows_after if row.get("Invoice ID") == after_second["Invoice #"]]
    assert len(linked_after) == 1, "second payment must update the existing linked row, not create a duplicate"
    assert app._coerce_number(linked_after[0]["Total incl. VAT (€)"]) == 1000.0


def test_invoice_bad_debt_removes_linked_income_entry(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    invoice = _add_draft_invoice(client)
    client.post(
        '/invoices/record-payment',
        data={"row_number": str(invoice["__row_number"]), "amount_received": "500.00", "payment_date": "2026-08-05", "payment_method": "Cash"},
        follow_redirects=True,
    )
    app.load_finance_data.cache_clear()
    paid_invoice = next(row for row in app.load_finance_data()["sheets"]["Invoices"] if row["__row_number"] == invoice["__row_number"])
    assert any(row.get("Invoice ID") == paid_invoice["Invoice #"] for row in app.load_finance_data()["sheets"]["Income"])

    # Paid invoices must have their payment reversed before a status like Bad Debt
    # can be applied — reverse first, then mark bad debt via the dedicated route
    # (Bad Debt is a system-controlled status, no longer settable via /invoices/update).
    reverse_response = client.post('/invoices/reverse-payment', data={"row_number": str(invoice["__row_number"])}, follow_redirects=True)
    assert reverse_response.status_code == 200

    response = client.post('/invoices/bad-debt', data={"row_number": str(invoice["__row_number"])}, follow_redirects=True)
    assert response.status_code == 200
    assert b'Invoice marked as bad debt' in response.data

    app.load_finance_data.cache_clear()
    remaining_links = [row for row in app.load_finance_data()["sheets"]["Income"] if row.get("Invoice ID") == paid_invoice["Invoice #"]]
    assert remaining_links == []


def test_invoice_cancelled_removes_linked_income_entry(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    invoice = _add_draft_invoice(client)
    client.post(
        '/invoices/record-payment',
        data={"row_number": str(invoice["__row_number"]), "amount_received": "500.00", "payment_date": "2026-08-05", "payment_method": "Cash"},
        follow_redirects=True,
    )
    app.load_finance_data.cache_clear()

    # Paid invoices cannot be cancelled directly — reverse the payment first.
    blocked_response = client.post('/invoices/cancel', data={"row_number": str(invoice["__row_number"])}, follow_redirects=True)
    assert b'Reverse the payments before cancelling' in blocked_response.data

    client.post('/invoices/reverse-payment', data={"row_number": str(invoice["__row_number"])}, follow_redirects=True)

    response = client.post('/invoices/cancel', data={"row_number": str(invoice["__row_number"])}, follow_redirects=True)
    assert response.status_code == 200
    assert b'Invoice cancelled' in response.data

    app.load_finance_data.cache_clear()
    invoice_row = next(row for row in app.load_finance_data()["sheets"]["Invoices"] if row["__row_number"] == invoice["__row_number"])
    assert invoice_row["Status"] == "Cancelled"
    remaining_links = [row for row in app.load_finance_data()["sheets"]["Income"] if row.get("Invoice ID") == invoice_row["Invoice #"]]
    assert remaining_links == []


def test_invoiced_income_row_cannot_be_edited_or_deleted_manually(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    invoice = _add_draft_invoice(client)
    client.post(
        '/invoices/record-payment',
        data={"row_number": str(invoice["__row_number"]), "amount_received": "500.00", "payment_date": "2026-08-05", "payment_method": "Cash"},
        follow_redirects=True,
    )
    app.load_finance_data.cache_clear()
    linked_income = next(row for row in app.load_finance_data()["sheets"]["Income"] if row.get("Source") == "invoiced")

    update_response = client.post(
        '/income/update',
        data={"row_number": str(linked_income["__row_number"]), "date": "2026-08-05", "description": "Tampered", "amount": "1.00", "status": "Received"},
        follow_redirects=True,
    )
    assert b'linked to an invoice' in update_response.data

    delete_response = client.post('/income/delete', data={"row_number": str(linked_income["__row_number"])}, follow_redirects=True)
    assert b'linked to an invoice' in delete_response.data

    app.load_finance_data.cache_clear()
    still_present = [row for row in app.load_finance_data()["sheets"]["Income"] if row.get("Source") == "invoiced"]
    assert len(still_present) == 1
    assert still_present[0]["Description"] != "Tampered"


def test_overdue_invoice_auto_flagged_and_excluded_from_income(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    add_payload = {
        "issue_date": "2026-01-01",
        "due_date": "2026-01-15",
        "client_name": "Client A",
        "client_vat_number": "IE1234567A",
        "client_address": "1 Example Street, Dublin",
        "service_product": "Overdue engagement",
        "line_items_json": _invoice_line_items_payload("Overdue engagement", 250.00, vat_rate="0%"),
        "balance_due": "",
        "status": "Issued",
    }
    add_response = client.post('/invoices/add', data=add_payload, follow_redirects=True)
    assert b'Validation:' not in add_response.data

    app.load_finance_data.cache_clear()
    data = app.load_finance_data()
    invoice = next(row for row in data["sheets"]["Invoices"] if row["Service / Product"] == "Overdue engagement")
    assert invoice["Status"] == "Overdue"
    assert app._coerce_number(invoice["Balance Due (€)"]) == 250.0
    assert invoice["Balance Due (€)"] != ""

    assert not any(row.get("Invoice ID") == invoice["Invoice #"] for row in data["sheets"]["Income"])


def test_income_total_sums_invoiced_paid_and_manual_received_only(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    baseline_income_total = app.load_finance_data()["summary"]["income_total"]

    invoice = _add_draft_invoice(client, net_amount="500.00", total_amount="500.00")
    client.post(
        '/invoices/record-payment',
        data={"row_number": str(invoice["__row_number"]), "amount_received": "500.00", "payment_date": "2026-08-05", "payment_method": "Cash"},
        follow_redirects=True,
    )

    client.post(
        '/income/add',
        data={"date": "2026-08-06", "description": "Grant", "category": "Other Income", "amount": "200.00", "status": "Received", "payment_method": "Business Bank Account"},
        follow_redirects=True,
    )
    client.post(
        '/income/add',
        data={"date": "2026-08-06", "description": "Pending grant", "category": "Other Income", "amount": "999.00", "status": "Pending", "payment_method": "Business Bank Account"},
        follow_redirects=True,
    )

    app.load_finance_data.cache_clear()
    data = app.load_finance_data()
    # Only the invoiced payment (500) and the Received manual entry (200) should count —
    # the Pending manual entry (999) must be excluded.
    assert data["summary"]["income_total"] == baseline_income_total + 500.0 + 200.0


def test_invoice_multi_line_items_stored_and_totals_computed(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    line_items = json.dumps([
        {"service_id": "svc-1", "name": "Clarity Base", "description": "Foundation audit", "quantity": 1, "unit_price": 950.00, "discount_type": "€", "discount_value": 0, "vat_rate": "0%"},
        {"service_id": "svc-2", "name": "SOP Creation", "description": "Two procedures", "quantity": 2, "unit_price": 150.00, "discount_type": "%", "discount_value": 10, "vat_rate": "0%"},
    ])
    add_payload = {
        "issue_date": "2026-08-01",
        "due_date": "2026-08-31",
        "client_name": "Client A",
        "client_vat_number": "IE1234567A",
        "client_address": "1 Example Street, Dublin",
        "line_items_json": line_items,
        "balance_due": "",
        "status": "Draft",
    }
    response = client.post('/invoices/add', data=add_payload, follow_redirects=True)
    assert response.status_code == 200
    assert b'Invoice added' in response.data

    app.load_finance_data.cache_clear()
    invoice = next(row for row in app.load_finance_data()["sheets"]["Invoices"] if row.get("Service / Product") == "Clarity Base, SOP Creation")

    stored_items = invoice["line_items"]
    assert len(stored_items) == 2
    assert stored_items[0]["name"] == "Clarity Base"
    assert stored_items[0]["net_amount"] == 950.0
    # SOP Creation: base = 2 * 150 = 300, 10% discount = 30, net = 270
    assert stored_items[1]["quantity"] == 2.0
    assert stored_items[1]["discount_amount"] == 30.0
    assert stored_items[1]["net_amount"] == 270.0

    # Aggregate invoice fields must be the sum across all line items.
    assert app._coerce_number(invoice["Base Net Amount (€)"]) == 950.0 + 300.0
    assert app._coerce_number(invoice["Discount (€)"]) == 30.0
    assert app._coerce_number(invoice["Net (€)"]) == 1220.0
    assert app._coerce_number(invoice["Total (€)"]) == 1220.0
    assert app._coerce_number(invoice["Balance Due (€)"]) == 1220.0


def test_invoice_line_items_required_to_submit(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    add_payload = {
        "issue_date": "2026-08-01",
        "due_date": "2026-08-31",
        "client_name": "Client A",
        "client_vat_number": "IE1234567A",
        "client_address": "1 Example Street, Dublin",
        "line_items_json": json.dumps([]),
        "status": "Draft",
    }
    response = client.post('/invoices/add', data=add_payload, follow_redirects=True)
    assert response.status_code == 200
    assert b'At least one line item is required' in response.data


def test_invoice_mark_paid_description_uses_line_item_names(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    line_items = json.dumps([
        {"name": "Clarity Base", "quantity": 1, "unit_price": 950.00, "discount_type": "€", "discount_value": 0, "vat_rate": "0%"},
        {"name": "SOP Creation", "quantity": 2, "unit_price": 150.00, "discount_type": "%", "discount_value": 10, "vat_rate": "0%"},
    ])
    client.post(
        '/invoices/add',
        data={
            "issue_date": "2026-08-01",
            "due_date": "2026-08-31",
            "client_name": "Client A",
            "client_vat_number": "IE1234567A",
            "client_address": "1 Example Street, Dublin",
            "line_items_json": line_items,
            "status": "Draft",
        },
        follow_redirects=True,
    )
    app.load_finance_data.cache_clear()
    invoice = next(row for row in app.load_finance_data()["sheets"]["Invoices"] if row.get("Service / Product") == "Clarity Base, SOP Creation")

    client.post(
        '/invoices/record-payment',
        data={"row_number": str(invoice["__row_number"]), "amount_received": "1220.00", "payment_date": "2026-08-05", "payment_method": "Cash"},
        follow_redirects=True,
    )
    app.load_finance_data.cache_clear()
    linked_income = next(row for row in app.load_finance_data()["sheets"]["Income"] if row.get("Invoice ID") == invoice["Invoice #"])
    assert linked_income["Description"] == "Clarity Base, SOP Creation"


def test_legacy_invoice_without_line_items_gets_migrated(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    legacy_invoice = {
        "Invoice #": "LEGACY-001",
        "Issue Date": "2026-01-01",
        "Due Date": "2026-01-31",
        "Client Name": "Client A",
        "Service / Product": "Legacy consulting work",
        "Net (€)": 400.0,
        "VAT Rate": "23%",
        "VAT Amount (€)": 92.0,
        "Total (€)": 492.0,
        "Balance Due (€)": 492.0,
        "Status": "Issued",
    }
    records = app._load_sheet_records_raw("Invoices")
    records.append(legacy_invoice)
    app._save_sheet_records_raw("Invoices", records)
    app.load_finance_data.cache_clear()

    app._migrate_invoice_line_items()

    app.load_finance_data.cache_clear()
    migrated = next(row for row in app.load_finance_data()["sheets"]["Invoices"] if row["Invoice #"] == "LEGACY-001")
    assert len(migrated["line_items"]) == 1
    line_item = migrated["line_items"][0]
    assert line_item["name"] == "Legacy consulting work"
    assert line_item["net_amount"] == 400.0
    assert line_item["vat_amount"] == 92.0
    assert line_item["total"] == 492.0
    # The flat Service/Product field is preserved untouched for backwards compatibility.
    assert migrated["Service / Product"] == "Legacy consulting work"


def _expense_add_payload(**overrides):
    payload = {
        "date": "2026-08-07",
        "description": "Client dinner",
        "category": "Entertainment",
        "supplier_vat_number": "IE1234567A",
        "base_net_amount": "100.00",
        "net_amount": "100.00",
        "total_amount": "123.00",
        "vat_rate": "23%",
        "vat_amount": "23.00",
        "input_vat_reclaimable": "Yes",
        "receipt_attached": "Yes",
        "bank_reconciliation": "Unreconciled",
        "status": "Pending",
        "payment_method": "Cash",
    }
    payload.update(overrides)
    return payload


def test_expense_with_unacknowledged_red_flag_is_rejected(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    flags = json.dumps([{"key": "vat", "severity": "danger", "message": "VAT cannot be reclaimed on entertainment expenses."}])
    response = client.post(
        '/expenses/add',
        data=_expense_add_payload(compliance_flags_json=flags),
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b'compliance flags that need your attention' in response.data

    app.load_finance_data.cache_clear()
    saved = [row for row in app.load_finance_data()["sheets"]["Expenses"] if row.get("Description") == "Client dinner"]
    assert saved == []


def test_expense_with_acknowledged_red_flag_saves_and_records_acknowledgement(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    flags = json.dumps([{"key": "vat", "severity": "danger", "message": "VAT cannot be reclaimed on entertainment expenses."}])
    response = client.post(
        '/expenses/add',
        data=_expense_add_payload(
            compliance_flags_json=flags,
            flags_acknowledged="1",
            flags_acknowledged_at="2026-08-07T12:00:00.000Z",
        ),
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b'Expense entry added' in response.data

    app.load_finance_data.cache_clear()
    saved = next(row for row in app.load_finance_data()["sheets"]["Expenses"] if row.get("Description") == "Client dinner")
    stored_flags = json.loads(saved["Compliance Flags"])
    assert stored_flags == [{"key": "vat", "severity": "danger", "message": "VAT cannot be reclaimed on entertainment expenses."}]
    assert saved["Flags Acknowledged"] == "Yes"
    assert saved["Flags Acknowledged At"] == "2026-08-07T12:00:00.000Z"


def test_expense_with_only_amber_flag_saves_without_acknowledgement(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    flags = json.dumps([{"key": "amount", "severity": "warning", "message": "Amounts over €1,000 may require capital allowances treatment."}])
    response = client.post(
        '/expenses/add',
        data=_expense_add_payload(
            description="New laptop",
            category="Equipment and Hardware",
            base_net_amount="1500.00",
            net_amount="1500.00",
            total_amount="1845.00",
            vat_amount="345.00",
            compliance_flags_json=flags,
        ),
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b'Expense entry added' in response.data

    app.load_finance_data.cache_clear()
    saved = next(row for row in app.load_finance_data()["sheets"]["Expenses"] if row.get("Description") == "New laptop")
    stored_flags = json.loads(saved["Compliance Flags"])
    assert stored_flags == [{"key": "amount", "severity": "warning", "message": "Amounts over €1,000 may require capital allowances treatment."}]
    assert saved["Flags Acknowledged"] == "No"
    assert saved["Flags Acknowledged At"] == ""


def test_expense_with_no_flags_saves_with_empty_flags_list(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.post(
        '/expenses/add',
        data=_expense_add_payload(description="Ordinary software cost", category="Software and Subscriptions"),
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b'Expense entry added' in response.data

    app.load_finance_data.cache_clear()
    saved = next(row for row in app.load_finance_data()["sheets"]["Expenses"] if row.get("Description") == "Ordinary software cost")
    assert json.loads(saved["Compliance Flags"]) == []
    assert saved["Flags Acknowledged"] == "No"


def test_income_btwea_category_flag_requires_acknowledgement(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    flags = json.dumps([{"key": "category", "severity": "danger", "message": "BTWEA/welfare payments are non-trading income."}])
    response = client.post(
        '/income/add',
        data={
            "date": "2026-08-07",
            "description": "Weekly BTWEA payment",
            "category": "BTWEA / Welfare Support",
            "amount": "250.00",
            "status": "Received",
            "payment_method": "Business Bank Account",
            "compliance_flags_json": flags,
        },
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b'compliance flags that need your attention' in response.data

    app.load_finance_data.cache_clear()
    saved = [row for row in app.load_finance_data()["sheets"]["Income"] if row.get("Description") == "Weekly BTWEA payment"]
    assert saved == []

    ack_response = client.post(
        '/income/add',
        data={
            "date": "2026-08-07",
            "description": "Weekly BTWEA payment",
            "category": "BTWEA / Welfare Support",
            "amount": "250.00",
            "status": "Received",
            "payment_method": "Business Bank Account",
            "compliance_flags_json": flags,
            "flags_acknowledged": "1",
            "flags_acknowledged_at": "2026-08-07T12:00:00.000Z",
        },
        follow_redirects=True,
    )
    assert b'Income entry added' in ack_response.data
    app.load_finance_data.cache_clear()
    saved_ack = next(row for row in app.load_finance_data()["sheets"]["Income"] if row.get("Description") == "Weekly BTWEA payment")
    assert saved_ack["Flags Acknowledged"] == "Yes"
    assert json.loads(saved_ack["Compliance Flags"])[0]["severity"] == "danger"


def test_tax_rules_json_loads_expected_categories():
    rules = app._load_tax_rules()
    assert "Entertainment" in rules["expense_categories"]
    assert rules["expense_categories"]["Entertainment"]["severity"] == "danger"
    assert rules["expense_categories"]["Equipment and Hardware"]["amount_threshold"] == 1000
    assert "BTWEA / Welfare Support" in rules["income_categories"]
    assert rules["field_rules"]["receipt_required_threshold"] == 50


def test_entertainment_expense_locks_deductibility_and_vat_reclaimable(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.post(
        '/expenses/add',
        data=_expense_add_payload(
            description="Client dinner - locked fields",
            category="Entertainment",
            deductibility_status="Fully Deductible",  # attempt to override — must be ignored
            input_vat_reclaimable="Yes",  # attempt to override — must be ignored
        ),
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b'Expense entry added' in response.data

    app.load_finance_data.cache_clear()
    saved = next(row for row in app.load_finance_data()["sheets"]["Expenses"] if row.get("Description") == "Client dinner - locked fields")
    assert saved["Deductibility Status"] == "Non-Deductible"
    assert saved["Input VAT Reclaimable"] == "No"


def test_home_office_expense_locks_partially_deductible(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.post(
        '/expenses/add',
        data=_expense_add_payload(
            description="Broadband share",
            category="Home Office Expenses",
            deductibility_status="Fully Deductible",
        ),
        follow_redirects=True,
    )
    assert response.status_code == 200
    app.load_finance_data.cache_clear()
    saved = next(row for row in app.load_finance_data()["sheets"]["Expenses"] if row.get("Description") == "Broadband share")
    assert saved["Deductibility Status"] == "Partially Deductible"


def test_other_category_deductibility_is_editable(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.post(
        '/expenses/add',
        data=_expense_add_payload(
            description="Consulting invoice",
            category="Professional Fees",
            deductibility_status="Partially Deductible",
        ),
        follow_redirects=True,
    )
    assert response.status_code == 200
    app.load_finance_data.cache_clear()
    saved = next(row for row in app.load_finance_data()["sheets"]["Expenses"] if row.get("Description") == "Consulting invoice")
    assert saved["Deductibility Status"] == "Partially Deductible"


def test_non_deductible_items_category_retired_from_chart_of_accounts():
    accounts = app._ensure_chart_of_accounts()
    retired = next(account for account in accounts if account["code"] == "5900")
    assert retired["active"] is False
    fines = next(account for account in accounts if account["name"] == "Fines and Penalties")
    assert fines["active"] is True
    assert "Non-Deductible Items" not in app._chart_of_accounts_category_options("Expense")
    assert "Fines and Penalties" in app._chart_of_accounts_category_options("Expense")


def test_existing_non_deductible_items_expense_gets_flagged_amber(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    records = app._load_sheet_records_raw("Expenses")
    records.append({
        "Date (Registered)": "2026-01-01",
        "Title": "Legacy non-deductible item",
        "Category": "Non-Deductible Items",
        "Net Amount (€)": 50.0,
        "Total (€)": 50.0,
        "Status": "Paid",
    })
    app._save_sheet_records_raw("Expenses", records)
    app.load_finance_data.cache_clear()

    app._migrate_flag_retired_non_deductible_category()

    app.load_finance_data.cache_clear()
    flagged = next(row for row in app.load_finance_data()["sheets"]["Expenses"] if row.get("Description") == "Legacy non-deductible item")
    flags = json.loads(flagged["Compliance Flags"])
    assert any(flag["key"] == "retired_category" and flag["severity"] == "warning" for flag in flags)


def test_saving_json_record_creates_local_backup(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post(
        '/expenses/add',
        data=_expense_add_payload(description="Backup trigger expense", category="Software and Subscriptions"),
        follow_redirects=True,
    )

    today_dir = app.BACKUPS_DIR / date.today().isoformat()
    backed_up_file = today_dir / app.EXPENSES_PATH.name
    assert backed_up_file.exists()
    backed_up_records = json.loads(backed_up_file.read_text(encoding="utf-8"))
    assert any(row.get("Description") == "Backup trigger expense" for row in backed_up_records)

    status = app._load_backup_status()
    assert status["local_ok"] is True
    assert status["timestamp"]


def test_prune_old_backups_removes_dates_past_retention(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    old_dir = app.BACKUPS_DIR / (date.today() - timedelta(days=45)).isoformat()
    recent_dir = app.BACKUPS_DIR / (date.today() - timedelta(days=1)).isoformat()
    old_dir.mkdir(parents=True)
    recent_dir.mkdir(parents=True)
    (old_dir / "expenses.json").write_text("[]", encoding="utf-8")
    (recent_dir / "expenses.json").write_text("[]", encoding="utf-8")

    app._prune_old_backups()

    assert not old_dir.exists()
    assert recent_dir.exists()


def test_restore_backup_restores_file_and_rejects_invalid_filename(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    original_expenses = json.loads(app.EXPENSES_PATH.read_text(encoding="utf-8"))

    client.post(
        '/expenses/add',
        data=_expense_add_payload(description="Will be restored away", category="Software and Subscriptions"),
        follow_redirects=True,
    )
    app.load_finance_data.cache_clear()
    assert any(row.get("Description") == "Will be restored away" for row in app.load_finance_data()["sheets"]["Expenses"])

    today_str = date.today().isoformat()
    backup_file = app.BACKUPS_DIR / today_str / app.EXPENSES_PATH.name
    assert backup_file.exists()
    # Overwrite the day's backup with the pre-addition snapshot to restore to.
    backup_file.write_text(json.dumps(original_expenses, indent=2), encoding="utf-8")

    traversal_response = client.post(
        '/settings/backups/restore',
        data={"backup_date": today_str, "filename": "../app.py"},
        follow_redirects=True,
    )
    assert traversal_response.status_code == 200
    assert b'Invalid backup file' in traversal_response.data

    restore_response = client.post(
        '/settings/backups/restore',
        data={"backup_date": today_str, "filename": app.EXPENSES_PATH.name},
        follow_redirects=True,
    )
    assert restore_response.status_code == 200

    app.load_finance_data.cache_clear()
    restored = app.load_finance_data()["sheets"]["Expenses"]
    assert not any(row.get("Description") == "Will be restored away" for row in restored)


def test_backups_view_lists_available_backup_dates(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post(
        '/expenses/add',
        data=_expense_add_payload(description="Listing check expense", category="Software and Subscriptions"),
        follow_redirects=True,
    )

    response = client.get('/settings/backups')
    assert response.status_code == 200
    assert date.today().isoformat().encode() in response.data
    assert app.EXPENSES_PATH.name.encode() in response.data


def test_non_deductible_expense_has_zero_taxable_net(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.post(
        '/expenses/add',
        data=_expense_add_payload(description="Entertainment zero net", category="Entertainment"),
        follow_redirects=True,
    )
    assert response.status_code == 200

    app.load_finance_data.cache_clear()
    saved = next(row for row in app.load_finance_data()["sheets"]["Expenses"] if row.get("Description") == "Entertainment zero net")
    assert saved["Deductibility Status"] == "Non-Deductible"
    assert saved["Net Amount (€)"] == "0.00"


def test_manually_selected_non_deductible_expense_has_zero_taxable_net(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.post(
        '/expenses/add',
        data=_expense_add_payload(
            description="Manually flagged non-deductible",
            category="Professional Fees",
            deductibility_status="Non-Deductible",
        ),
        follow_redirects=True,
    )
    assert response.status_code == 200

    app.load_finance_data.cache_clear()
    saved = next(row for row in app.load_finance_data()["sheets"]["Expenses"] if row.get("Description") == "Manually flagged non-deductible")
    assert saved["Deductibility Status"] == "Non-Deductible"
    assert saved["Net Amount (€)"] == "0.00"


def test_fully_deductible_expense_keeps_calculated_taxable_net(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.post(
        '/expenses/add',
        data=_expense_add_payload(description="Fully deductible net check", category="Software and Subscriptions", base_net_amount="80.00"),
        follow_redirects=True,
    )
    assert response.status_code == 200

    app.load_finance_data.cache_clear()
    saved = next(row for row in app.load_finance_data()["sheets"]["Expenses"] if row.get("Description") == "Fully deductible net check")
    assert saved["Deductibility Status"] == "Fully Deductible"
    assert saved["Net Amount (€)"] == "80.00"


def test_quick_add_supplier_creates_incomplete_supplier(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.post(
        '/api/suppliers/quick-add',
        json={"name": "Corner Shop Ltd"},
    )
    assert response.status_code == 200
    payload = response.get_json()
    assert payload["name"] == "Corner Shop Ltd"
    assert payload["created"] is True

    app.load_finance_data.cache_clear()
    suppliers = app.load_finance_data()["sheets"]["Suppliers"]
    created = next(row for row in suppliers if row.get("Supplier Name") == "Corner Shop Ltd")
    assert created["Needs Completion"] == "Yes"

    suppliers_response = client.get('/suppliers')
    assert suppliers_response.status_code == 200
    assert b'Incomplete' in suppliers_response.data


def test_quick_add_supplier_does_not_duplicate_existing_supplier(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    first = client.post('/api/suppliers/quick-add', json={"name": "Repeat Supplier"})
    assert first.get_json()["created"] is True

    app.load_finance_data.cache_clear()
    second = client.post('/api/suppliers/quick-add', json={"name": "Repeat Supplier"})
    assert second.status_code == 200
    assert second.get_json()["created"] is False

    app.load_finance_data.cache_clear()
    suppliers = app.load_finance_data()["sheets"]["Suppliers"]
    matches = [row for row in suppliers if row.get("Supplier Name") == "Repeat Supplier"]
    assert len(matches) == 1


def test_quick_add_supplier_requires_name(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.post('/api/suppliers/quick-add', json={"name": "  "})
    assert response.status_code == 400


def test_supplier_search_finds_matches_and_flags_incomplete(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post('/api/suppliers/quick-add', json={"name": "Widget Wholesale"})
    app.load_finance_data.cache_clear()

    too_short = client.get('/api/suppliers/search?q=W')
    assert too_short.get_json() == []

    response = client.get('/api/suppliers/search?q=widget')
    assert response.status_code == 200
    results = response.get_json()
    assert any(match["name"] == "Widget Wholesale" and match["needs_completion"] is True for match in results)


def test_pre_trading_expense_gets_pre_trading_phase_tag_and_amber_flag(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    app._save_business_profile({
        "structure": "sole_trader",
        "transition_date": "2026-06-01",
        "pre_trading_start_date": "2026-01-01",
    })
    app.load_finance_data.cache_clear()

    response = client.post(
        '/expenses/add',
        data=_expense_add_payload(
            description="Laptop before trading",
            category="Equipment and Hardware",
            date="2026-02-15",
            base_net_amount="1500.00",
            net_amount="1500.00",
            total_amount="1845.00",
        ),
        follow_redirects=True,
    )
    assert response.status_code == 200

    app.load_finance_data.cache_clear()
    saved = next(row for row in app.load_finance_data()["sheets"]["Expenses"] if row.get("Description") == "Laptop before trading")
    assert saved["Phase Tag"] == "Pre-Trading"
    flags = json.loads(saved["Compliance Flags"])
    assert any(flag["key"] == "pretrading_capex" and flag["severity"] == "warning" for flag in flags)


def test_expense_dated_after_trading_start_is_not_pre_trading(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    app._save_business_profile({
        "structure": "sole_trader",
        "transition_date": "2026-06-01",
        "pre_trading_start_date": "2026-01-01",
    })
    app.load_finance_data.cache_clear()

    client.post(
        '/expenses/add',
        data=_expense_add_payload(description="After trading start", category="Software and Subscriptions", date="2026-07-01"),
        follow_redirects=True,
    )

    app.load_finance_data.cache_clear()
    saved = next(row for row in app.load_finance_data()["sheets"]["Expenses"] if row.get("Description") == "After trading start")
    assert saved["Phase Tag"] != "Pre-Trading"


def test_expenses_page_phase_filter(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    app._save_business_profile({
        "structure": "sole_trader",
        "transition_date": "2026-06-01",
        "pre_trading_start_date": "2026-01-01",
    })
    app.load_finance_data.cache_clear()

    client.post(
        '/expenses/add',
        data=_expense_add_payload(
            description="Pre-trading filter check description",
            category="Software and Subscriptions",
            date="2026-02-01",
        ),
        follow_redirects=True,
    )

    response = client.get('/expenses?phase_filter=Pre-Trading')
    assert response.status_code == 200
    assert b'Pre-trading filter check description' in response.data

    response_phase1 = client.get('/expenses?phase_filter=Phase 1')
    assert response_phase1.status_code == 200
    assert b'Pre-trading filter check description' not in response_phase1.data


def test_expense_receipt_file_upload_stores_filename_and_file(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    data = _expense_add_payload(description="Receipt upload check", category="Software and Subscriptions")
    data['receipt_file'] = (BytesIO(b'%PDF-1.4 fake receipt content'), 'my receipt.pdf')

    response = client.post(
        '/expenses/add',
        data=data,
        content_type='multipart/form-data',
        follow_redirects=True,
    )
    assert response.status_code == 200

    app.load_finance_data.cache_clear()
    saved = next(row for row in app.load_finance_data()["sheets"]["Expenses"] if row.get("Description") == "Receipt upload check")
    filename = saved["Receipt Filename"]
    assert filename
    assert filename.endswith("my_receipt.pdf") or filename.endswith("my-receipt.pdf")
    assert (app.RECEIPTS_DIR / filename).exists()
    assert saved["Receipt Attached"] == "Yes"


def test_expense_receipt_rejects_disallowed_extension(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    data = _expense_add_payload(description="Bad receipt extension", category="Software and Subscriptions")
    data['receipt_file'] = (BytesIO(b'not a real executable'), 'malware.exe')

    client.post(
        '/expenses/add',
        data=data,
        content_type='multipart/form-data',
        follow_redirects=True,
    )

    app.load_finance_data.cache_clear()
    saved = next(row for row in app.load_finance_data()["sheets"]["Expenses"] if row.get("Description") == "Bad receipt extension")
    assert saved["Receipt Filename"] == ""


# --- Company: Documents ---------------------------------------------------

def test_company_documents_seeds_cro_certificate_on_first_load(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.get('/company/documents')
    assert response.status_code == 200
    assert b'CRO Certificate of Registration' in response.data

    documents = json.loads(app.COMPANY_DOCUMENTS_PATH.read_text(encoding="utf-8"))
    assert len(documents) == 1
    assert documents[0]["notes"] == "H-Queex business name registration, CRO No. 790968"


def test_document_upload_route_persists_document_and_file(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    data = {
        "name": "Public Liability Insurance",
        "category": "Insurance",
        "description": "Annual policy",
        "expiry_date": "2026-12-31",
        "notes": "Broker: Example Insurance Co",
        "document_file": (BytesIO(b'%PDF-1.4 fake policy content'), 'policy.pdf'),
    }
    response = client.post(
        '/company/documents/upload',
        data=data,
        content_type='multipart/form-data',
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b'Public Liability Insurance' in response.data

    documents = json.loads(app.COMPANY_DOCUMENTS_PATH.read_text(encoding="utf-8"))
    uploaded = next(doc for doc in documents if doc["name"] == "Public Liability Insurance")
    assert uploaded["category"] == "Insurance"
    assert uploaded["filename"]
    assert uploaded["file_path"] == f"documents/{uploaded['filename']}"
    assert (app.COMPANY_DOCUMENTS_DIR / uploaded["filename"]).exists()


def test_document_upload_rejects_disallowed_extension(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    data = {
        "name": "Suspicious file",
        "category": "Other",
        "document_file": (BytesIO(b'not allowed'), 'malware.exe'),
    }
    client.post(
        '/company/documents/upload',
        data=data,
        content_type='multipart/form-data',
        follow_redirects=True,
    )

    documents = json.loads(app.COMPANY_DOCUMENTS_PATH.read_text(encoding="utf-8"))
    assert not any(doc["name"] == "Suspicious file" for doc in documents)


def test_document_upload_rejects_file_over_max_size(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    oversized_content = b'0' * (app.MAX_DOCUMENT_SIZE_BYTES + 1)
    data = {
        "name": "Oversized file",
        "category": "Other",
        "document_file": (BytesIO(oversized_content), 'big.pdf'),
    }
    response = client.post(
        '/company/documents/upload',
        data=data,
        content_type='multipart/form-data',
        follow_redirects=True,
    )
    assert response.status_code == 200

    documents = json.loads(app.COMPANY_DOCUMENTS_PATH.read_text(encoding="utf-8"))
    assert not any(doc["name"] == "Oversized file" for doc in documents)


def test_document_update_route_can_archive_via_archive_route(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.get('/company/documents')
    documents = json.loads(app.COMPANY_DOCUMENTS_PATH.read_text(encoding="utf-8"))
    document_id = documents[0]["id"]

    response = client.post(
        '/company/documents/archive',
        data={"document_id": document_id},
        follow_redirects=True,
    )
    assert response.status_code == 200

    documents = json.loads(app.COMPANY_DOCUMENTS_PATH.read_text(encoding="utf-8"))
    archived = next(doc for doc in documents if doc["id"] == document_id)
    assert archived["status"] == "archived"
    assert b'CRO Certificate of Registration' not in client.get('/company/documents').data


# --- Company: Document expiry detection -----------------------------------

def test_document_expiry_severity_classifies_expired_soon_and_ok():
    today = date(2026, 8, 8)
    assert app._document_expiry_severity("2026-08-01", today=today) == "expired"
    assert app._document_expiry_severity("2026-08-20", today=today) == "soon"
    assert app._document_expiry_severity("2026-08-08", today=today) == "soon"
    assert app._document_expiry_severity("2026-12-01", today=today) == "ok"
    assert app._document_expiry_severity("", today=today) == ""


def test_documents_expiring_soon_excludes_archived_and_far_future(workbook_copy):
    today = date(2026, 8, 8)
    documents = [
        {"id": "1", "name": "Expired doc", "status": "active", "expiry_date": "2026-07-01"},
        {"id": "2", "name": "Soon doc", "status": "active", "expiry_date": "2026-08-20"},
        {"id": "3", "name": "Far future doc", "status": "active", "expiry_date": "2027-01-01"},
        {"id": "4", "name": "No expiry doc", "status": "active", "expiry_date": ""},
        {"id": "5", "name": "Archived expiring doc", "status": "archived", "expiry_date": "2026-08-10"},
    ]
    expiring = app._documents_expiring_soon(documents, today=today)
    names = {doc["name"] for doc in expiring}
    assert names == {"Expired doc", "Soon doc"}


def test_dashboard_shows_document_expiry_warning(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    data = {
        "name": "Expiring Certificate",
        "category": "Compliance",
        "expiry_date": date.today().isoformat(),
    }
    client.post('/company/documents/upload', data=data, follow_redirects=True)

    response = client.get('/')
    assert response.status_code == 200
    assert b'expiring soon' in response.data
    assert b'Expiring Certificate' in response.data


# --- Company: Compliance Calendar -----------------------------------------

def test_vat3_period_due_dates_covers_six_bi_monthly_periods():
    periods = app._vat3_period_due_dates(2026)
    assert len(periods) == 6
    assert periods[0] == (date(2026, 1, 1), date(2026, 2, 28), date(2026, 3, 23))
    assert periods[2] == (date(2026, 5, 1), date(2026, 6, 30), date(2026, 7, 23))
    assert periods[5] == (date(2026, 11, 1), date(2026, 12, 31), date(2027, 1, 23))


def test_compliance_deadline_severity_matches_red_amber_green():
    today = date(2026, 8, 8)
    assert app._compliance_deadline_severity(date(2026, 8, 1), today=today) == "red"
    assert app._compliance_deadline_severity(date(2026, 8, 20), today=today) == "amber"
    assert app._compliance_deadline_severity(date(2026, 12, 1), today=today) == "green"


def test_build_compliance_deadlines_includes_form11_and_vat3_for_sole_trader():
    today = date(2026, 8, 8)
    profile = {
        "structure": "sole_trader",
        "vat_registered": True,
        "registration_date": "2026-08-04",
        "transition_date": "",
    }
    deadlines = app._build_compliance_deadlines(profile, {"summary": {}}, [], today=today)
    names = [d["name"] for d in deadlines]
    assert any("Form 11" in name for name in names)
    assert any("Preliminary tax" in name for name in names)
    assert any(name.startswith("VAT 3 return") for name in names)
    assert not any("CT1" in name for name in names)
    assert not any("CRO annual return" in name for name in names)
    assert deadlines == sorted(deadlines, key=lambda item: item["due_date"])


def test_build_compliance_deadlines_includes_ct1_and_cro_for_limited_company_phase_2():
    today = date(2026, 8, 8)
    profile = {
        "structure": "limited_company",
        "vat_registered": False,
        "registration_date": "2020-01-15",
        "transition_date": "2026-01-01",
    }
    deadlines = app._build_compliance_deadlines(profile, {"summary": {}}, [], today=today)
    names = [d["name"] for d in deadlines]
    assert any("CT1" in name for name in names)
    assert any("CRO annual return" in name for name in names)
    assert not any("Form 11" in name for name in names)


def test_build_compliance_deadlines_excludes_ct1_when_still_phase_1():
    today = date(2026, 8, 8)
    profile = {
        "structure": "limited_company",
        "vat_registered": False,
        "registration_date": "2020-01-15",
        "transition_date": "2027-01-01",
    }
    deadlines = app._build_compliance_deadlines(profile, {"summary": {}}, [], today=today)
    names = [d["name"] for d in deadlines]
    assert not any("CT1" in name for name in names)


def test_build_compliance_deadlines_includes_manual_entries_and_excludes_completed():
    today = date(2026, 8, 8)
    profile = {"structure": "sole_trader", "vat_registered": False}
    manual_entries = [
        {"id": "m1", "name": "Renew domain", "due_date": "2026-09-01", "status": "pending", "description": ""},
        {"id": "m2", "name": "Completed task", "due_date": "2026-09-01", "status": "complete", "description": ""},
    ]
    deadlines = app._build_compliance_deadlines(profile, {"summary": {}}, manual_entries, today=today)
    names = [d["name"] for d in deadlines]
    assert "Renew domain" in names
    assert "Completed task" not in names


def test_compliance_add_route_persists_entry(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.post(
        '/company/compliance/add',
        data={"name": "Renew business insurance", "due_date": "2026-10-01", "description": "Annual renewal", "repeat_frequency": "annual"},
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b'Renew business insurance' in response.data

    entries = json.loads(app.COMPLIANCE_CALENDAR_PATH.read_text(encoding="utf-8"))
    assert entries[0]["name"] == "Renew business insurance"
    assert entries[0]["status"] == "pending"


def test_compliance_add_route_requires_name_and_due_date(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post('/company/compliance/add', data={"name": "", "due_date": ""}, follow_redirects=True)

    entries = json.loads(app.COMPLIANCE_CALENDAR_PATH.read_text(encoding="utf-8")) if app.COMPLIANCE_CALENDAR_PATH.exists() else []
    assert entries == []


def test_compliance_complete_route_marks_entry_complete(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post(
        '/company/compliance/add',
        data={"name": "One-off filing", "due_date": "2026-10-01"},
        follow_redirects=True,
    )
    entries = json.loads(app.COMPLIANCE_CALENDAR_PATH.read_text(encoding="utf-8"))
    entry_id = entries[0]["id"]

    response = client.post('/company/compliance/complete', data={"entry_id": entry_id}, follow_redirects=True)
    assert response.status_code == 200

    entries = json.loads(app.COMPLIANCE_CALENDAR_PATH.read_text(encoding="utf-8"))
    assert entries[0]["status"] == "complete"

    timeline_response = client.get('/company/compliance')
    assert b'One-off filing' not in timeline_response.data


# --- Company: Business Profile ---------------------------------------------

def test_profile_update_route_persists_all_fields(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.post(
        '/company/profile/update',
        data={
            "business_name": "H-Queex Hub",
            "owner_name": "Hevandro Martire",
            "cro_number": "790968",
            "registration_date": "2026-08-04",
            "structure": "limited_company",
            "trading_start_date": "2026-01-01",
            "pre_trading_start_date": "2025-11-01",
            "transition_date": "2026-06-01",
            "vat_registered": "1",
            "vat_threshold_basis": "goods",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200

    profile = json.loads(app.BUSINESS_PROFILE_PATH.read_text(encoding="utf-8"))
    assert profile["business_name"] == "H-Queex Hub"
    assert profile["structure"] == "limited_company"
    assert profile["trading_start_date"] == "2026-01-01"
    assert profile["vat_registered"] is True
    assert profile["vat_threshold_basis"] == "goods"


def test_profile_update_route_rejects_missing_business_name(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post(
        '/company/profile/update',
        data={"business_name": "", "structure": "sole_trader"},
        follow_redirects=True,
    )

    profile = app._load_business_profile()
    assert profile["business_name"] != ""


# --- Company: Settings redirect and navigation ------------------------------

def test_settings_redirects_to_company_settings(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.get('/settings', follow_redirects=False)
    assert response.status_code == 302
    assert response.headers["Location"] == "/company/settings"


def test_company_landing_page_and_nav_link_render(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.get('/company')
    assert response.status_code == 200
    assert b'Documents' in response.data
    assert b'Compliance Calendar' in response.data
    assert b'Business Profile' in response.data

    dashboard_response = client.get('/')
    assert b'href="/company"' in dashboard_response.data


# --- Operations: Projects ---------------------------------------------------

def _client_id_by_name(name):
    """Triggers the id migration (via load_finance_data) then resolves a client's stable id by name."""
    app.load_finance_data.cache_clear()
    data = app.load_finance_data()
    row = next(row for row in data["sheets"]["Clients"] if row.get("Client Name") == name)
    return row["id"]


def test_operations_nav_link_enabled_and_projects_created(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    dashboard_response = client.get('/')
    assert b'href="/operations"' in dashboard_response.data

    response = client.post(
        '/operations/projects/add',
        data={
            "title": "Process improvement rollout",
            "client_id": client_a_id,
            "status": "Active",
            "start_date": "2026-08-01",
            "target_end_date": "2026-09-01",
            "line_items_json": "[]",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b'Process improvement rollout' in response.data

    projects = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8"))
    assert len(projects) == 1
    project = projects[0]
    assert project["project_number"] == "HQ-PRJ-2026-001"
    assert project["client_id"] == client_a_id
    assert project["client_name"] == "Client A"
    assert project["status"] == "Active"
    assert set(app.DMAIC_PHASES) == set(project["dmaic"].keys())


def test_second_project_gets_incrementing_project_number(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    for title in ("First project", "Second project"):
        client.post(
            '/operations/projects/add',
            data={"title": title, "client_id": client_a_id, "status": "Enquiry", "line_items_json": "[]"},
            follow_redirects=True,
        )

    projects = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8"))
    numbers = sorted(p["project_number"] for p in projects)
    assert numbers == ["HQ-PRJ-2026-001", "HQ-PRJ-2026-002"]


def test_project_add_requires_title_and_client(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post('/operations/projects/add', data={"title": "", "client_name": "", "line_items_json": "[]"}, follow_redirects=True)

    projects = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8")) if app.PROJECTS_PATH.exists() else []
    assert projects == []


def test_project_status_update_route_moves_kanban_card(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    client.post(
        '/operations/projects/add',
        data={"title": "Kanban project", "client_id": client_a_id, "status": "Enquiry", "line_items_json": "[]"},
        follow_redirects=True,
    )
    projects = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8"))
    project_id = projects[0]["id"]

    response = client.post('/operations/projects/status', data={"project_id": project_id, "status": "Proposed"}, follow_redirects=True)
    assert response.status_code == 200

    projects = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8"))
    assert projects[0]["status"] == "Proposed"


def test_project_status_update_rejects_invalid_status(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    client.post(
        '/operations/projects/add',
        data={"title": "Guard project", "client_id": client_a_id, "status": "Enquiry", "line_items_json": "[]"},
        follow_redirects=True,
    )
    projects = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8"))
    project_id = projects[0]["id"]

    client.post('/operations/projects/status', data={"project_id": project_id, "status": "Not A Real Status"}, follow_redirects=True)

    projects = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8"))
    assert projects[0]["status"] == "Enquiry"


def test_project_detail_page_shows_linked_records(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    client.post(
        '/operations/projects/add',
        data={"title": "Detail page project", "client_id": client_a_id, "status": "Active", "line_items_json": "[]"},
        follow_redirects=True,
    )
    project_id = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8"))[0]["id"]

    response = client.get(f'/operations/projects/{project_id}')
    assert response.status_code == 200
    assert b'Detail page project' in response.data

    missing_response = client.get('/operations/projects/does-not-exist', follow_redirects=True)
    assert missing_response.status_code == 200
    assert b'Project not found' in missing_response.data


def test_project_archive_sets_status_cancelled(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    client.post(
        '/operations/projects/add',
        data={"title": "Archive me", "client_id": client_a_id, "status": "Active", "line_items_json": "[]"},
        follow_redirects=True,
    )
    project_id = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8"))[0]["id"]

    client.post('/operations/projects/archive', data={"project_id": project_id}, follow_redirects=True)

    projects = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8"))
    assert projects[0]["status"] == "Cancelled"


# --- Operations: DMAIC phase sequencing -------------------------------------

def test_dmaic_completion_percentage_counts_complete_phases():
    dmaic = app._default_dmaic()
    assert app._dmaic_completion_percentage(dmaic) == 0
    dmaic["Define"]["status"] = "Complete"
    assert app._dmaic_completion_percentage(dmaic) == 20
    dmaic["Measure"]["status"] = "Complete"
    assert app._dmaic_completion_percentage(dmaic) == 40


def test_dmaic_transition_blocks_completing_phase_out_of_order():
    dmaic = app._default_dmaic()
    error = app._validate_dmaic_transition(dmaic, "Measure", "Complete")
    assert "Define" in error

    dmaic["Define"]["status"] = "Complete"
    error = app._validate_dmaic_transition(dmaic, "Measure", "Complete")
    assert error == ""


def test_dmaic_transition_allows_in_progress_regardless_of_order():
    dmaic = app._default_dmaic()
    assert app._validate_dmaic_transition(dmaic, "Control", "In Progress") == ""


def test_dmaic_update_route_enforces_sequencing_server_side(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    client.post(
        '/operations/projects/add',
        data={"title": "DMAIC project", "client_id": client_a_id, "status": "Active", "line_items_json": "[]"},
        follow_redirects=True,
    )
    project_id = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8"))[0]["id"]

    # Attempt to complete Analyse before Define/Measure — must be rejected.
    response = client.post(
        '/operations/dmaic/update',
        data={"project_id": project_id, "phase": "Analyse", "status": "Complete"},
        follow_redirects=True,
    )
    assert response.status_code == 200
    projects = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8"))
    assert projects[0]["dmaic"]["Analyse"]["status"] == "Not Started"

    # Complete Define, then Measure — Measure should now be allowed.
    client.post('/operations/dmaic/update', data={"project_id": project_id, "phase": "Define", "status": "Complete"}, follow_redirects=True)
    client.post('/operations/dmaic/update', data={"project_id": project_id, "phase": "Measure", "status": "Complete", "deliverables": "Data collected\nBaseline set"}, follow_redirects=True)

    projects = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8"))
    dmaic = projects[0]["dmaic"]
    assert dmaic["Define"]["status"] == "Complete"
    assert dmaic["Measure"]["status"] == "Complete"
    assert dmaic["Measure"]["deliverables"] == ["Data collected", "Baseline set"]


# --- Operations: Delivery Log ------------------------------------------------

def test_delivery_add_route_persists_entry_with_project_client(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    client.post(
        '/operations/projects/add',
        data={"title": "Delivery project", "client_id": client_a_id, "status": "Active", "line_items_json": "[]"},
        follow_redirects=True,
    )
    project_id = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8"))[0]["id"]

    response = client.post(
        '/operations/delivery/add',
        data={
            "date": "2026-08-05",
            "project_id": project_id,
            "service_type": "Advisory Call",
            "description": "Monthly check-in call",
            "hours_spent": "1.5",
            "billing_period": "August 2026",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b'Delivery logged' in response.data

    entries = json.loads(app.DELIVERY_LOG_PATH.read_text(encoding="utf-8"))
    assert len(entries) == 1
    entry = entries[0]
    assert entry["client_name"] == "Client A"
    assert entry["project_id"] == project_id
    assert entry["hours_spent"] == 1.5
    assert entry["invoiced"] is False


def test_delivery_add_requires_description_and_client(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post(
        '/operations/delivery/add',
        data={"date": "2026-08-05", "description": "", "service_type": "Other"},
        follow_redirects=True,
    )

    entries = json.loads(app.DELIVERY_LOG_PATH.read_text(encoding="utf-8")) if app.DELIVERY_LOG_PATH.exists() else []
    assert entries == []


def test_delivery_file_attachment_stores_file(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    data = {
        "date": "2026-08-05",
        "client_id": client_a_id,
        "service_type": "Report",
        "description": "Quarterly report delivered",
        "deliverable_file": (BytesIO(b'%PDF-1.4 fake report'), 'report.pdf'),
    }
    response = client.post('/operations/delivery/add', data=data, content_type='multipart/form-data', follow_redirects=True)
    assert response.status_code == 200

    entries = json.loads(app.DELIVERY_LOG_PATH.read_text(encoding="utf-8"))
    filename = entries[0]["deliverable_filename"]
    assert filename
    assert (app.DELIVERY_FILES_DIR / filename).exists()


# --- Operations: Clarity Partner invoice generation from Delivery Log -------

def _make_clarity_partner_client(name="Partner Client", retainer_amount=500.0):
    clients = json.loads(app.CLIENTS_PATH.read_text(encoding="utf-8"))
    clients.append({
        "Client Name": name,
        "Contact Person": "",
        "Email": "",
        "Phone": "",
        "Country": "",
        "Service Tier": "Clarity Partner",
        "Retainer Frequency": "monthly",
        "Retainer Amount (€)": retainer_amount,
    })
    app.CLIENTS_PATH.write_text(json.dumps(clients), encoding="utf-8")
    app.load_finance_data.cache_clear()


def test_generate_delivery_invoice_creates_invoice_and_marks_entries_invoiced(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    _make_clarity_partner_client()
    partner_id = _client_id_by_name("Partner Client")

    for description in ("KPI review call", "Process documentation"):
        client.post(
            '/operations/delivery/add',
            data={
                "date": "2026-08-05",
                "client_id": partner_id,
                "service_type": "KPI Review",
                "description": description,
                "hours_spent": "1",
                "billing_period": "August 2026",
            },
            follow_redirects=True,
        )

    response = client.post(
        f'/operations/delivery/generate-invoice/{partner_id}/August 2026',
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b'Invoice' in response.data
    assert b'generated' in response.data

    app.load_finance_data.cache_clear()
    invoices = app.load_finance_data()["sheets"]["Invoices"]
    matching = [row for row in invoices if row.get("Client Name") == "Partner Client"]
    assert len(matching) == 1
    invoice = matching[0]
    assert invoice["Client ID"] == partner_id
    assert len(invoice["line_items"]) == 2
    assert app._coerce_number(invoice["Total (€)"]) > 0

    entries = json.loads(app.DELIVERY_LOG_PATH.read_text(encoding="utf-8"))
    partner_entries = [e for e in entries if e["client_id"] == partner_id]
    assert all(e["invoiced"] for e in partner_entries)
    assert all(e["invoice_id"] == invoice["Invoice #"] for e in partner_entries)


def test_generate_delivery_invoice_requires_retainer_amount(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    _make_clarity_partner_client(name="No Retainer Client", retainer_amount=0)
    no_retainer_id = _client_id_by_name("No Retainer Client")

    client.post(
        '/operations/delivery/add',
        data={"date": "2026-08-05", "client_id": no_retainer_id, "service_type": "Other", "description": "Some work", "billing_period": "August 2026"},
        follow_redirects=True,
    )

    response = client.post(f'/operations/delivery/generate-invoice/{no_retainer_id}/August 2026', follow_redirects=True)
    assert response.status_code == 200
    assert b'retainer amount' in response.data

    app.load_finance_data.cache_clear()
    invoices = app.load_finance_data()["sheets"]["Invoices"]
    assert not any(row.get("Client Name") == "No Retainer Client" for row in invoices)


def test_generate_delivery_invoice_no_unbilled_entries_is_noop(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    _make_clarity_partner_client()
    partner_id = _client_id_by_name("Partner Client")

    response = client.post(f'/operations/delivery/generate-invoice/{partner_id}/August 2026', follow_redirects=True)
    assert response.status_code == 200
    assert b'No unbilled delivery entries' in response.data


def test_clarity_partner_pending_billing_helper_excludes_invoiced_entries():
    entries = [
        {"client_id": "CLT-001", "client_name": "A", "billing_period": "August 2026", "invoiced": False, "hours_spent": 1},
        {"client_id": "CLT-001", "client_name": "A", "billing_period": "August 2026", "invoiced": True, "hours_spent": 1},
        {"client_id": "CLT-002", "client_name": "B", "billing_period": "", "invoiced": False, "hours_spent": 1},
    ]
    pending = app._clarity_partner_pending_billing(entries)
    assert len(pending) == 1
    assert pending[0]["client_id"] == "CLT-001"
    assert pending[0]["client_name"] == "A"
    assert pending[0]["entry_count"] == 1


# --- Operations: SOP Library version control --------------------------------

def test_sop_add_route_creates_sop_in_draft_status(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    response = client.post(
        '/operations/sops/add',
        data={"title": "Client Onboarding SOP", "client_id": client_a_id, "version": "V1.0", "process_area": "Onboarding"},
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b'SOP added' in response.data

    sops = json.loads(app.SOPS_PATH.read_text(encoding="utf-8"))
    assert len(sops) == 1
    assert sops[0]["status"] == "Draft"
    assert sops[0]["version"] == "V1.0"


def test_sop_new_version_supersedes_previous(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    client.post('/operations/sops/add', data={"title": "Expense SOP", "client_id": client_a_id, "version": "V1.0"}, follow_redirects=True)
    sops = json.loads(app.SOPS_PATH.read_text(encoding="utf-8"))
    original_id = sops[0]["id"]

    client.post(
        '/operations/sops/add',
        data={"title": "Expense SOP", "client_id": client_a_id, "version": "V1.1", "supersedes_id": original_id},
        follow_redirects=True,
    )

    sops = json.loads(app.SOPS_PATH.read_text(encoding="utf-8"))
    original = next(s for s in sops if s["id"] == original_id)
    new_version = next(s for s in sops if s["id"] != original_id)
    assert original["status"] == "Superseded"
    assert new_version["version"] == "V1.1"
    assert new_version["status"] == "Draft"


def test_sop_status_workflow_enforces_order(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    client.post('/operations/sops/add', data={"title": "Workflow SOP", "client_id": client_a_id}, follow_redirects=True)
    sop_id = json.loads(app.SOPS_PATH.read_text(encoding="utf-8"))[0]["id"]

    # Skipping straight to Approved from Draft must be rejected.
    response = client.post('/operations/sops/update', data={"sop_id": sop_id, "status": "Approved"}, follow_redirects=True)
    assert response.status_code == 200
    sops = json.loads(app.SOPS_PATH.read_text(encoding="utf-8"))
    assert sops[0]["status"] == "Draft"

    # Draft -> Review is a valid single-step transition.
    client.post('/operations/sops/update', data={"sop_id": sop_id, "status": "Review"}, follow_redirects=True)
    sops = json.loads(app.SOPS_PATH.read_text(encoding="utf-8"))
    assert sops[0]["status"] == "Review"

    # Review -> Approved is valid and stamps date_approved/approved_by.
    client.post('/operations/sops/update', data={"sop_id": sop_id, "status": "Approved", "approved_by": "Hevandro"}, follow_redirects=True)
    sops = json.loads(app.SOPS_PATH.read_text(encoding="utf-8"))
    assert sops[0]["status"] == "Approved"
    assert sops[0]["approved_by"] == "Hevandro"
    assert sops[0]["date_approved"]


def test_sop_file_upload_rejects_disallowed_extension(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    data = {
        "title": "Bad file SOP",
        "client_id": client_a_id,
        "sop_file": (BytesIO(b'not allowed'), 'malware.exe'),
    }
    client.post('/operations/sops/add', data=data, content_type='multipart/form-data', follow_redirects=True)

    sops = json.loads(app.SOPS_PATH.read_text(encoding="utf-8")) if app.SOPS_PATH.exists() else []
    assert not any(s["title"] == "Bad file SOP" for s in sops)


# --- Operations: dashboard integration --------------------------------------

def test_dashboard_shows_active_projects_kpi_and_upcoming_deadline(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    near_due_date = (date.today() + timedelta(days=5)).isoformat()
    client.post(
        '/operations/projects/add',
        data={
            "title": "Deadline soon project",
            "client_id": client_a_id,
            "status": "Active",
            "target_end_date": near_due_date,
            "line_items_json": "[]",
        },
        follow_redirects=True,
    )

    response = client.get('/')
    assert response.status_code == 200
    assert b'Active Projects' in response.data
    assert b'Deadline soon project' in response.data


# --- CRM: Leads ---------------------------------------------------------------

def test_crm_nav_link_present_and_lead_created(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    dashboard_response = client.get('/')
    assert b'href="/crm"' in dashboard_response.data

    response = client.post(
        '/crm/leads/add',
        data={
            "contact_name": "Sarah Prospect",
            "company_name": "Prospect Co",
            "email": "sarah@prospect.co",
            "source": "Referral",
            "status": "New",
            "estimated_value": "5000",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b'Sarah Prospect' in response.data

    leads = json.loads(app.LEADS_PATH.read_text(encoding="utf-8"))
    assert len(leads) == 1
    lead = leads[0]
    assert lead["lead_number"] == "HQ-LEAD-2026-001"
    assert lead["source"] == "Referral"
    assert lead["estimated_value"] == 5000.0
    assert len(lead["activity_log"]) == 1
    assert lead["activity_log"][0]["type"] == "created"


def test_lead_add_requires_contact_or_company_name(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post('/crm/leads/add', data={"contact_name": "", "company_name": "", "email": ""}, follow_redirects=True)

    leads = json.loads(app.LEADS_PATH.read_text(encoding="utf-8")) if app.LEADS_PATH.exists() else []
    assert leads == []


def test_lead_status_route_moves_kanban_card(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post('/crm/leads/add', data={"contact_name": "Kanban Lead", "source": "Direct", "status": "New"}, follow_redirects=True)
    lead_id = json.loads(app.LEADS_PATH.read_text(encoding="utf-8"))[0]["id"]

    response = client.post('/crm/leads/status', data={"lead_id": lead_id, "status": "Contacted"}, follow_redirects=True)
    assert response.status_code == 200

    leads = json.loads(app.LEADS_PATH.read_text(encoding="utf-8"))
    assert leads[0]["status"] == "Contacted"
    assert any(entry["type"] == "status_change" for entry in leads[0]["activity_log"])


def test_lead_log_contact_adds_timestamped_activity(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post('/crm/leads/add', data={"contact_name": "Contact Log Lead", "source": "Direct"}, follow_redirects=True)
    lead_id = json.loads(app.LEADS_PATH.read_text(encoding="utf-8"))[0]["id"]

    response = client.post('/crm/leads/log-contact', data={"lead_id": lead_id, "text": "Had a great call about scope"}, follow_redirects=True)
    assert response.status_code == 200

    leads = json.loads(app.LEADS_PATH.read_text(encoding="utf-8"))
    contact_entries = [e for e in leads[0]["activity_log"] if e["type"] == "contact"]
    assert len(contact_entries) == 1
    assert contact_entries[0]["text"] == "Had a great call about scope"
    assert contact_entries[0]["timestamp"]


def test_lead_schedule_followup_sets_next_action(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post('/crm/leads/add', data={"contact_name": "Followup Lead", "source": "Direct"}, follow_redirects=True)
    lead_id = json.loads(app.LEADS_PATH.read_text(encoding="utf-8"))[0]["id"]

    response = client.post(
        '/crm/leads/schedule-followup',
        data={"lead_id": lead_id, "next_action_date": "2026-09-01", "next_action": "Send updated quote"},
        follow_redirects=True,
    )
    assert response.status_code == 200

    leads = json.loads(app.LEADS_PATH.read_text(encoding="utf-8"))
    assert leads[0]["next_action_date"] == "2026-09-01"
    assert leads[0]["next_action"] == "Send updated quote"


def test_lead_convert_to_client_creates_client_record_and_links(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post(
        '/crm/leads/add',
        data={"contact_name": "Won Contact", "company_name": "Won Co", "email": "won@co.com", "source": "Direct", "status": "Negotiating"},
        follow_redirects=True,
    )
    lead_id = json.loads(app.LEADS_PATH.read_text(encoding="utf-8"))[0]["id"]

    response = client.post('/crm/leads/convert-to-client', data={"lead_id": lead_id}, follow_redirects=True)
    assert response.status_code == 200
    assert b'converted to client' in response.data

    app.load_finance_data.cache_clear()
    clients = app.load_finance_data()["sheets"]["Clients"]
    assert any(row.get("Client Name") == "Won Co" for row in clients)

    leads = json.loads(app.LEADS_PATH.read_text(encoding="utf-8"))
    assert leads[0]["status"] == "Won"
    new_client = next(row for row in clients if row.get("Client Name") == "Won Co")
    assert leads[0]["converted_client_id"] == new_client["id"]
    assert new_client["id"].startswith("CLT-")
    assert any(entry["type"] == "converted" for entry in leads[0]["activity_log"])


def test_lead_detail_page_shows_lead_and_handles_missing(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post('/crm/leads/add', data={"contact_name": "Detail Page Lead", "source": "Direct"}, follow_redirects=True)
    lead_id = json.loads(app.LEADS_PATH.read_text(encoding="utf-8"))[0]["id"]

    response = client.get(f'/crm/leads/{lead_id}')
    assert response.status_code == 200
    assert b'Detail Page Lead' in response.data

    missing_response = client.get('/crm/leads/does-not-exist', follow_redirects=True)
    assert b'Lead not found' in missing_response.data


# --- CRM: pipeline calculations (unit tests) --------------------------------

def test_leads_pipeline_value_excludes_closed_leads():
    leads = [
        {"status": "New", "estimated_value": 1000},
        {"status": "Negotiating", "estimated_value": 2000},
        {"status": "Won", "estimated_value": 5000},
        {"status": "Lost", "estimated_value": 3000},
    ]
    assert app._leads_pipeline_value(leads) == 3000.0


def test_leads_conversion_rate_computes_won_over_closed():
    leads = [
        {"status": "Won"}, {"status": "Won"}, {"status": "Lost"}, {"status": "New"},
    ]
    assert app._leads_conversion_rate(leads) == pytest.approx(66.7, rel=0.01)


def test_leads_conversion_rate_zero_when_no_closed_leads():
    assert app._leads_conversion_rate([{"status": "New"}, {"status": "Contacted"}]) == 0.0


def test_average_deal_value_only_counts_won_leads():
    leads = [
        {"status": "Won", "estimated_value": 1000},
        {"status": "Won", "estimated_value": 3000},
        {"status": "Lost", "estimated_value": 9000},
    ]
    assert app._average_deal_value(leads) == 2000.0


# --- CRM: Proposals --------------------------------------------------------------

def _proposal_line_items_json():
    return json.dumps([
        {"name": "Brand Strategy", "quantity": 1, "unit_price": 2000, "discount_type": "€", "discount_value": 0, "vat_rate": "0%"},
    ])


def test_proposal_creation_from_lead_links_and_updates_lead_status(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post('/crm/leads/add', data={"contact_name": "Proposal Lead", "company_name": "Proposal Co", "source": "Direct", "status": "Qualified"}, follow_redirects=True)
    lead_id = json.loads(app.LEADS_PATH.read_text(encoding="utf-8"))[0]["id"]

    response = client.post(
        '/crm/proposals/add',
        data={
            "lead_id": lead_id,
            "title": "Brand Strategy Proposal",
            "contact_name": "Proposal Lead",
            "company_name": "Proposal Co",
            "line_items_json": _proposal_line_items_json(),
        },
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert b'Proposal created' in response.data

    proposals = json.loads(app.PROPOSALS_PATH.read_text(encoding="utf-8"))
    assert len(proposals) == 1
    proposal = proposals[0]
    assert proposal["proposal_number"] == "HQ-PROP-2026-001"
    assert proposal["status"] == "Draft"
    assert proposal["total"] == 2000.0
    assert proposal["lead_id"] == lead_id
    assert proposal["terms"]  # standard terms auto-populated

    leads = json.loads(app.LEADS_PATH.read_text(encoding="utf-8"))
    assert leads[0]["status"] == "Proposal Sent"
    assert any(entry["type"] == "proposal_created" for entry in leads[0]["activity_log"])


def test_proposal_add_requires_title_contact_and_line_items(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post('/crm/proposals/add', data={"title": "", "contact_name": "", "line_items_json": "[]"}, follow_redirects=True)

    proposals = json.loads(app.PROPOSALS_PATH.read_text(encoding="utf-8")) if app.PROPOSALS_PATH.exists() else []
    assert proposals == []


def test_proposal_standard_terms_default_populated_in_add_form(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.get('/crm/proposals')
    assert response.status_code == 200
    assert b'50% deposit required on proposal acceptance' in response.data


def test_proposal_mark_sent_sets_status_and_dates(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post(
        '/crm/proposals/add',
        data={"title": "Sent Proposal", "contact_name": "Contact", "company_name": "Co", "line_items_json": _proposal_line_items_json()},
        follow_redirects=True,
    )
    proposal_id = json.loads(app.PROPOSALS_PATH.read_text(encoding="utf-8"))[0]["id"]

    response = client.post('/crm/proposals/mark-sent', data={"proposal_id": proposal_id}, follow_redirects=True)
    assert response.status_code == 200

    proposals = json.loads(app.PROPOSALS_PATH.read_text(encoding="utf-8"))
    assert proposals[0]["status"] == "Sent"
    assert proposals[0]["sent_date"] == date.today().isoformat()


def test_proposal_convert_to_invoice_requires_accepted_status(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post(
        '/crm/proposals/add',
        data={"title": "Draft Proposal", "contact_name": "Contact", "company_name": "Co", "line_items_json": _proposal_line_items_json()},
        follow_redirects=True,
    )
    proposal_id = json.loads(app.PROPOSALS_PATH.read_text(encoding="utf-8"))[0]["id"]

    response = client.post('/crm/proposals/convert-to-invoice', data={"proposal_id": proposal_id}, follow_redirects=True)
    assert response.status_code == 200
    assert b'Only Accepted proposals' in response.data

    app.load_finance_data.cache_clear()
    invoices = app.load_finance_data()["sheets"]["Invoices"]
    assert not any(row.get("Notes", "").startswith("Generated from Proposal") for row in invoices)


def test_proposal_convert_to_invoice_creates_invoice_and_links_back(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post(
        '/crm/proposals/add',
        data={"title": "Accepted Proposal", "contact_name": "Contact", "company_name": "Accepted Co", "line_items_json": _proposal_line_items_json()},
        follow_redirects=True,
    )
    proposal_id = json.loads(app.PROPOSALS_PATH.read_text(encoding="utf-8"))[0]["id"]

    client.post('/crm/proposals/mark-sent', data={"proposal_id": proposal_id}, follow_redirects=True)
    client.post('/crm/proposals/status', data={"proposal_id": proposal_id, "status": "Accepted"}, follow_redirects=True)

    response = client.post('/crm/proposals/convert-to-invoice', data={"proposal_id": proposal_id}, follow_redirects=True)
    assert response.status_code == 200
    assert b'created' in response.data

    app.load_finance_data.cache_clear()
    invoices = app.load_finance_data()["sheets"]["Invoices"]
    matching = [row for row in invoices if row.get("Client Name") == "Accepted Co"]
    assert len(matching) == 1
    assert matching[0]["line_items"][0]["name"] == "Brand Strategy"

    proposals = json.loads(app.PROPOSALS_PATH.read_text(encoding="utf-8"))
    assert proposals[0]["converted_invoice_id"] == matching[0]["Invoice #"]


def test_proposal_accepted_logs_lead_activity(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post('/crm/leads/add', data={"contact_name": "Accept Activity Lead", "source": "Direct"}, follow_redirects=True)
    lead_id = json.loads(app.LEADS_PATH.read_text(encoding="utf-8"))[0]["id"]

    client.post(
        '/crm/proposals/add',
        data={"lead_id": lead_id, "title": "Activity Proposal", "contact_name": "Accept Activity Lead", "line_items_json": _proposal_line_items_json()},
        follow_redirects=True,
    )
    proposal_id = json.loads(app.PROPOSALS_PATH.read_text(encoding="utf-8"))[0]["id"]

    client.post('/crm/proposals/status', data={"proposal_id": proposal_id, "status": "Accepted"}, follow_redirects=True)

    leads = json.loads(app.LEADS_PATH.read_text(encoding="utf-8"))
    assert any(entry["type"] == "proposal_accepted" for entry in leads[0]["activity_log"])


def test_proposal_expiry_severity_classification():
    today = date(2026, 8, 10)
    sent_proposal = {"status": "Sent", "expiry_date": "2026-08-05"}
    assert app._proposal_expiry_severity(sent_proposal, today=today) == "expired"
    sent_proposal["expiry_date"] = "2026-08-14"
    assert app._proposal_expiry_severity(sent_proposal, today=today) == "soon"
    sent_proposal["expiry_date"] = "2026-12-01"
    assert app._proposal_expiry_severity(sent_proposal, today=today) == "ok"
    draft_proposal = {"status": "Draft", "expiry_date": "2026-08-05"}
    assert app._proposal_expiry_severity(draft_proposal, today=today) == ""


# --- CRM: public lead intake API (website integration) ----------------------

def test_public_api_creates_lead_with_website_source(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.post(
        '/api/leads',
        json={
            "contact_name": "Web Visitor",
            "company_name": "Web Co",
            "email": "visitor@web.com",
            "phone": "0871234567",
            "service_interest": ["Brand Strategy"],
            "message": "Please get in touch",
        },
    )
    assert response.status_code == 201
    payload = response.get_json()
    assert payload["success"] is True
    assert payload["lead_number"] == "HQ-LEAD-2026-001"

    leads = json.loads(app.LEADS_PATH.read_text(encoding="utf-8"))
    assert len(leads) == 1
    assert leads[0]["source"] == "Website"
    assert leads[0]["status"] == "New"
    assert leads[0]["notes"] == "Please get in touch"


def test_public_api_requires_email(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.post('/api/leads', json={"contact_name": "No Email Visitor"})
    assert response.status_code == 400
    payload = response.get_json()
    assert payload["success"] is False

    leads = json.loads(app.LEADS_PATH.read_text(encoding="utf-8")) if app.LEADS_PATH.exists() else []
    assert leads == []


def test_public_api_cors_headers_present(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    options_response = client.options('/api/leads')
    assert options_response.status_code == 204
    assert options_response.headers["Access-Control-Allow-Origin"] == "*"
    assert "POST" in options_response.headers["Access-Control-Allow-Methods"]

    post_response = client.post('/api/leads', json={"contact_name": "CORS Test", "email": "cors@test.com"})
    assert post_response.headers["Access-Control-Allow-Origin"] == "*"


# --- CRM: dashboard integration ----------------------------------------------

def test_crm_dashboard_shows_pipeline_kpis(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post('/crm/leads/add', data={"contact_name": "Dashboard Lead", "source": "Direct", "status": "New", "estimated_value": "1500"}, follow_redirects=True)

    response = client.get('/')
    assert response.status_code == 200
    assert b'Pipeline Value' in response.data
    assert b'Proposals Awaiting Response' in response.data


def test_dashboard_upcoming_actions_flags_overdue_followup(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    overdue_date = (date.today() - timedelta(days=3)).isoformat()
    client.post(
        '/crm/leads/add',
        data={"contact_name": "Overdue Followup Lead", "source": "Direct", "status": "Contacted", "next_action_date": overdue_date, "next_action": "Call back"},
        follow_redirects=True,
    )

    response = client.get('/')
    assert response.status_code == 200
    assert b'Follow-up overdue' in response.data
    assert b'Overdue Followup Lead' in response.data


# --- Stable IDs: Clients & Suppliers ----------------------------------------

def test_new_client_gets_sequential_stable_id(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    # Client A already exists in the seeded workbook and gets migrated to CLT-001
    # the first time load_finance_data runs, so a brand-new client should be CLT-002.
    response = client.post(
        '/clients/add',
        data={"client_name": "Fresh Co", "service_tier": "None"},
        follow_redirects=True,
    )
    assert response.status_code == 200

    app.load_finance_data.cache_clear()
    clients = app.load_finance_data()["sheets"]["Clients"]
    fresh = next(row for row in clients if row.get("Client Name") == "Fresh Co")
    assert fresh["id"] == "CLT-002"

    # ids stay unique and sequential for a third client too.
    client.post('/clients/add', data={"client_name": "Third Co", "service_tier": "None"}, follow_redirects=True)
    app.load_finance_data.cache_clear()
    clients = app.load_finance_data()["sheets"]["Clients"]
    third = next(row for row in clients if row.get("Client Name") == "Third Co")
    assert third["id"] == "CLT-003"


def test_existing_clients_without_id_are_migrated_idempotently(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    # Client A was seeded with no "id" field at all.
    clients = json.loads(app.CLIENTS_PATH.read_text(encoding="utf-8"))
    assert "id" not in clients[0]

    data = app.load_finance_data()
    migrated_id = data["sheets"]["Clients"][0]["id"]
    assert migrated_id.startswith("CLT-")

    # Running the migration again (fresh cache) must not reassign a new id.
    app.load_finance_data.cache_clear()
    data_again = app.load_finance_data()
    assert data_again["sheets"]["Clients"][0]["id"] == migrated_id


def test_client_id_stable_when_client_name_changes(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    data = app.load_finance_data()
    client_a_row = next(row for row in data["sheets"]["Clients"] if row.get("Client Name") == "Client A")
    original_id = client_a_row["id"]
    row_number = client_a_row["__row_number"]

    response = client.post(
        '/clients/update',
        data={
            "row_number": row_number,
            "client_name": "Client A Renamed Ltd",
            "service_tier": "None",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200

    app.load_finance_data.cache_clear()
    data = app.load_finance_data()
    renamed_row = next(row for row in data["sheets"]["Clients"] if row.get("Client Name") == "Client A Renamed Ltd")
    assert renamed_row["id"] == original_id
    assert not any(row.get("Client Name") == "Client A" for row in data["sheets"]["Clients"])


def test_new_supplier_gets_sequential_stable_id(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    # Supplier A is seeded with no id and gets migrated to SUP-001 first.
    response = client.post(
        '/suppliers/add',
        data={"supplier_name": "Fresh Supplier Co"},
        follow_redirects=True,
    )
    assert response.status_code == 200

    app.load_finance_data.cache_clear()
    suppliers = app.load_finance_data()["sheets"]["Suppliers"]
    fresh = next(row for row in suppliers if row.get("Supplier Name") == "Fresh Supplier Co")
    assert fresh["id"] == "SUP-002"


def test_supplier_id_stable_when_supplier_name_changes(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    data = app.load_finance_data()
    supplier_a_row = next(row for row in data["sheets"]["Suppliers"] if row.get("Supplier Name") == "Supplier A")
    original_id = supplier_a_row["id"]
    row_number = supplier_a_row["__row_number"]

    client.post(
        '/suppliers/update',
        data={"row_number": row_number, "supplier_name": "Supplier A Renamed"},
        follow_redirects=True,
    )

    app.load_finance_data.cache_clear()
    data = app.load_finance_data()
    renamed_row = next(row for row in data["sheets"]["Suppliers"] if row.get("Supplier Name") == "Supplier A Renamed")
    assert renamed_row["id"] == original_id


def test_quick_add_client_and_supplier_assign_ids(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client_response = client.post('/api/clients/quick-add', json={"name": "Quick Add Client"})
    assert client_response.status_code == 200
    client_payload = client_response.get_json()
    assert client_payload["created"] is True
    assert client_payload["id"].startswith("CLT-")

    supplier_response = client.post('/api/suppliers/quick-add', json={"name": "Quick Add Supplier"})
    supplier_payload = supplier_response.get_json()
    assert supplier_payload["created"] is True
    assert supplier_payload["id"].startswith("SUP-")

    # search results also carry the id, so the smart-search hidden field can populate it.
    search_response = client.get('/api/clients/search?q=Quick')
    matches = search_response.get_json()
    assert matches[0]["id"] == client_payload["id"]


# --- Stable IDs: cross-module lookup by client_id / supplier_id -------------

def test_project_stores_client_id_and_resolves_current_name_after_rename(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    client.post(
        '/operations/projects/add',
        data={"title": "ID join test project", "client_id": client_a_id, "status": "Active", "line_items_json": "[]"},
        follow_redirects=True,
    )
    project = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8"))[0]
    assert project["client_id"] == client_a_id
    assert project["client_name"] == "Client A"

    # Rename the client — the project's stored client_name is a snapshot, but every
    # page render re-resolves the current name live via client_id.
    data = app.load_finance_data()
    row_number = next(row for row in data["sheets"]["Clients"] if row.get("id") == client_a_id)["__row_number"]
    client.post('/clients/update', data={"row_number": row_number, "client_name": "Client A Global Ltd", "service_tier": "None"}, follow_redirects=True)

    response = client.get(f"/operations/projects/{project['id']}")
    assert response.status_code == 200
    assert b'Client A Global Ltd' in response.data
    assert b'Client A</' not in response.data

    list_response = client.get('/operations/projects')
    assert b'Client A Global Ltd' in list_response.data


def test_invoice_stores_client_id_alongside_client_name(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    client.post(
        '/invoices/add',
        data={
            "issue_date": "2026-08-01",
            "due_date": "2026-08-15",
            "client_name": "Client A",
            "client_id": client_a_id,
            "line_items_json": json.dumps([{"name": "Consulting", "quantity": 1, "unit_price": 500, "discount_type": "€", "discount_value": 0, "vat_rate": "0%"}]),
        },
        follow_redirects=True,
    )

    app.load_finance_data.cache_clear()
    invoices = app.load_finance_data()["sheets"]["Invoices"]
    invoice = next(row for row in invoices if row.get("Client Name") == "Client A")
    assert invoice["Client ID"] == client_a_id


def test_invoice_client_id_resolved_from_name_when_hidden_field_missing(workbook_copy):
    """Older cached pages / any caller that only submits client_name (no client_id
    hidden field) must still resolve the correct id via a name lookup fallback."""
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    client.post(
        '/invoices/add',
        data={
            "issue_date": "2026-08-01",
            "due_date": "2026-08-15",
            "client_name": "Client A",
            "line_items_json": json.dumps([{"name": "Consulting", "quantity": 1, "unit_price": 500, "discount_type": "€", "discount_value": 0, "vat_rate": "0%"}]),
        },
        follow_redirects=True,
    )

    app.load_finance_data.cache_clear()
    invoices = app.load_finance_data()["sheets"]["Invoices"]
    invoice = next(row for row in invoices if row.get("Client Name") == "Client A")
    assert invoice["Client ID"] == client_a_id


def test_expense_stores_supplier_id_alongside_supplier_name(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    data = app.load_finance_data()
    supplier_a_id = next(row for row in data["sheets"]["Suppliers"] if row.get("Supplier Name") == "Supplier A")["id"]

    client.post(
        '/expenses/add',
        data=_expense_add_payload(description="Supplier ID check", category="Software and Subscriptions", supplier="Supplier A"),
        follow_redirects=True,
    )

    app.load_finance_data.cache_clear()
    expenses = app.load_finance_data()["sheets"]["Expenses"]
    expense = next(row for row in expenses if row.get("Description") == "Supplier ID check")
    assert expense["Supplier ID"] == supplier_a_id


def test_delivery_log_and_sop_resolve_client_id_join_key(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    client.post(
        '/operations/delivery/add',
        data={"date": "2026-08-05", "client_id": client_a_id, "service_type": "Advisory Call", "description": "Direct client delivery"},
        follow_redirects=True,
    )
    client.post('/operations/sops/add', data={"title": "Cross-module SOP", "client_id": client_a_id, "version": "V1.0"}, follow_redirects=True)

    entries = json.loads(app.DELIVERY_LOG_PATH.read_text(encoding="utf-8"))
    assert entries[0]["client_id"] == client_a_id
    assert entries[0]["client_name"] == "Client A"

    sops = json.loads(app.SOPS_PATH.read_text(encoding="utf-8"))
    assert sops[0]["client_id"] == client_a_id
    assert sops[0]["client_name"] == "Client A"


def test_cross_module_migration_backfills_client_id_on_legacy_records(workbook_copy):
    """Records written before this migration only had client_name — the migration
    must backfill client_id from the (now-migrated) Clients sheet without any
    application code path touching them, and must be a no-op the second time."""
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    app.load_finance_data()  # ensure Client A has been migrated to an id first

    legacy_project = app._normalize_project({
        "id": "legacy-1",
        "title": "Legacy project",
        "client_name": "Client A",
        "status": "Active",
    })
    legacy_project["client_id"] = ""  # simulate a pre-migration record
    app._save_json_records(app.PROJECTS_PATH, [legacy_project])

    app.load_finance_data.cache_clear()
    app.load_finance_data()

    projects = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8"))
    client_a_id = _client_id_by_name("Client A")
    assert projects[0]["client_id"] == client_a_id

    # Idempotent: running it again must not change anything or error.
    app.load_finance_data.cache_clear()
    app.load_finance_data()
    projects_again = json.loads(app.PROJECTS_PATH.read_text(encoding="utf-8"))
    assert projects_again == projects


# --- Stable IDs: remaining entities (Documents, SOPs, Delivery, Compliance, VAT, Services) ---

def test_document_gets_stable_doc_id_on_upload(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    # The seeded CRO certificate is migrated to DOC-001 on first load.
    client.get('/company/documents')
    seeded = json.loads(app.COMPANY_DOCUMENTS_PATH.read_text(encoding="utf-8"))[0]
    assert seeded["id"] == "DOC-001"

    client.post(
        '/company/documents/upload',
        data={"name": "Insurance Policy", "category": "Insurance"},
        follow_redirects=True,
    )
    documents = json.loads(app.COMPANY_DOCUMENTS_PATH.read_text(encoding="utf-8"))
    uploaded = next(doc for doc in documents if doc["name"] == "Insurance Policy")
    assert uploaded["id"] == "DOC-002"


def test_legacy_document_id_migrated_and_compliance_link_repointed(workbook_copy):
    """A document created before this migration only had a uuid4 id — the migration
    must renumber it to DOC-NNN AND repoint any compliance entry's linked_document_id
    to match, so the link doesn't silently break."""
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    legacy_doc_id = "11111111-2222-3333-4444-555555555555"
    app.COMPANY_DOCUMENTS_PATH.write_text(json.dumps([
        {"id": legacy_doc_id, "name": "Legacy Doc", "category": "Legal", "status": "active"}
    ]), encoding="utf-8")
    app.COMPLIANCE_CALENDAR_PATH.write_text(json.dumps([
        {"id": "legacy-cmp-1", "name": "Renew legacy doc", "due_date": "2026-09-01", "status": "pending", "linked_document_id": legacy_doc_id}
    ]), encoding="utf-8")

    app.load_finance_data.cache_clear()
    app.load_finance_data()

    documents = json.loads(app.COMPANY_DOCUMENTS_PATH.read_text(encoding="utf-8"))
    new_doc_id = documents[0]["id"]
    assert new_doc_id.startswith("DOC-")
    assert new_doc_id != legacy_doc_id

    entries = json.loads(app.COMPLIANCE_CALENDAR_PATH.read_text(encoding="utf-8"))
    assert entries[0]["linked_document_id"] == new_doc_id


def test_compliance_entry_gets_stable_cmp_id(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post('/company/compliance/add', data={"name": "First deadline", "due_date": "2026-09-01"}, follow_redirects=True)
    client.post('/company/compliance/add', data={"name": "Second deadline", "due_date": "2026-10-01"}, follow_redirects=True)

    entries = json.loads(app.COMPLIANCE_CALENDAR_PATH.read_text(encoding="utf-8"))
    ids = sorted(entry["id"] for entry in entries)
    assert ids == ["CMP-001", "CMP-002"]


def test_sop_gets_stable_sop_id(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    client.post('/operations/sops/add', data={"title": "First SOP", "client_id": client_a_id}, follow_redirects=True)
    client.post('/operations/sops/add', data={"title": "Second SOP", "client_id": client_a_id}, follow_redirects=True)

    sops = json.loads(app.SOPS_PATH.read_text(encoding="utf-8"))
    ids = sorted(sop["id"] for sop in sops)
    assert ids == ["SOP-001", "SOP-002"]


def test_delivery_log_entry_gets_stable_dlv_id(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()
    client_a_id = _client_id_by_name("Client A")

    client.post(
        '/operations/delivery/add',
        data={"date": "2026-08-05", "client_id": client_a_id, "service_type": "Advisory Call", "description": "First delivery"},
        follow_redirects=True,
    )
    client.post(
        '/operations/delivery/add',
        data={"date": "2026-08-06", "client_id": client_a_id, "service_type": "Advisory Call", "description": "Second delivery"},
        follow_redirects=True,
    )

    entries = json.loads(app.DELIVERY_LOG_PATH.read_text(encoding="utf-8"))
    ids = sorted(entry["id"] for entry in entries)
    assert ids == ["DLV-001", "DLV-002"]


def test_legacy_sop_and_delivery_ids_migrated_idempotently(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.SOPS_PATH.write_text(json.dumps([{"id": "old-uuid-sop", "title": "Legacy SOP", "client_name": "Client A"}]), encoding="utf-8")
    app.DELIVERY_LOG_PATH.write_text(json.dumps([{"id": "old-uuid-delivery", "description": "Legacy entry", "client_name": "Client A"}]), encoding="utf-8")

    app.load_finance_data.cache_clear()
    app.load_finance_data()

    sops = json.loads(app.SOPS_PATH.read_text(encoding="utf-8"))
    entries = json.loads(app.DELIVERY_LOG_PATH.read_text(encoding="utf-8"))
    assert sops[0]["id"] == "SOP-001"
    assert entries[0]["id"] == "DLV-001"

    # Idempotent — running again doesn't renumber already-correct ids.
    app.load_finance_data.cache_clear()
    app.load_finance_data()
    sops_again = json.loads(app.SOPS_PATH.read_text(encoding="utf-8"))
    entries_again = json.loads(app.DELIVERY_LOG_PATH.read_text(encoding="utf-8"))
    assert sops_again[0]["id"] == "SOP-001"
    assert entries_again[0]["id"] == "DLV-001"


def test_vat_period_ids_are_deterministic_and_follow_bimonthly_format():
    assert app._vat_period_id(date(2026, 1, 15)) == "VAT-2026-01"
    assert app._vat_period_id(date(2026, 3, 1)) == "VAT-2026-02"
    assert app._vat_period_id(date(2026, 5, 1)) == "VAT-2026-03"
    assert app._vat_period_id(date(2026, 7, 1)) == "VAT-2026-04"
    assert app._vat_period_id(date(2026, 9, 1)) == "VAT-2026-05"
    assert app._vat_period_id(date(2026, 11, 1)) == "VAT-2026-06"
    assert app._vat_period_id(date(2027, 1, 1)) == "VAT-2027-01"


def test_compliance_deadlines_use_stable_vat_period_ids(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    profile = {"structure": "sole_trader", "vat_registered": True, "registration_date": "2026-08-04"}
    deadlines = app._build_compliance_deadlines(profile, {"summary": {}}, [], today=date(2026, 8, 10))
    vat_deadlines = [d for d in deadlines if d["id"].startswith("VAT-")]
    assert vat_deadlines
    for deadline in vat_deadlines:
        assert deadline["id"].split("-")[1].isdigit()
        period_number = int(deadline["id"].split("-")[2])
        assert 1 <= period_number <= 6


def test_service_id_uses_svc_prefix(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post(
        '/services/add',
        data={"name": "First Service", "tier": "core", "price": "500"},
        follow_redirects=True,
    )
    client.post(
        '/services/add',
        data={"name": "Second Service", "tier": "core", "price": "750"},
        follow_redirects=True,
    )

    services = json.loads(app.SERVICES_PATH.read_text(encoding="utf-8"))
    ids = sorted(service["id"] for service in services)
    assert ids == ["SVC-001", "SVC-002"]


def test_legacy_service_id_migrated_and_invoice_line_item_repointed(workbook_copy):
    """A service created before this migration only had a uuid4 id — invoice line
    items that reference it by service_id must be repointed to the new SVC-NNN id
    so the catalogue link isn't silently lost."""
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    legacy_service_id = "aaaaaaaa-bbbb-cccc-dddd-eeeeeeeeeeee"
    app.SERVICES_PATH.write_text(json.dumps([
        {"id": legacy_service_id, "name": "Legacy Service", "tier": "core", "price": 500.0, "status": "active"}
    ]), encoding="utf-8")

    invoices = json.loads(app.INVOICES_PATH.read_text(encoding="utf-8"))
    invoices.append({
        "Invoice #": "HQ-2026-999",
        "Issue Date": "2026-08-01",
        "Client Name": "Client A",
        "Status": "Draft",
        "Total (€)": 500.0,
        "line_items": [{"service_id": legacy_service_id, "name": "Legacy Service", "quantity": 1, "unit_price": 500.0, "total": 500.0}],
    })
    app.INVOICES_PATH.write_text(json.dumps(invoices), encoding="utf-8")

    app.load_finance_data.cache_clear()
    app.load_finance_data()

    services = json.loads(app.SERVICES_PATH.read_text(encoding="utf-8"))
    new_service_id = services[0]["id"]
    assert new_service_id.startswith("SVC-")
    assert new_service_id != legacy_service_id

    invoices_after = json.loads(app.INVOICES_PATH.read_text(encoding="utf-8"))
    matching_invoice = next(row for row in invoices_after if row.get("Invoice #") == "HQ-2026-999")
    assert matching_invoice["line_items"][0]["service_id"] == new_service_id


def test_invoice_line_item_service_id_join_key_works_end_to_end(workbook_copy):
    """Confirms service_id is used correctly as the join key: adding an invoice with
    a line item that references a real service by id round-trips correctly."""
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post('/services/add', data={"name": "Consulting Package", "tier": "core", "price": "1000"}, follow_redirects=True)
    service_id = json.loads(app.SERVICES_PATH.read_text(encoding="utf-8"))[0]["id"]
    assert service_id.startswith("SVC-")

    client.post(
        '/invoices/add',
        data={
            "issue_date": "2026-08-01",
            "due_date": "2026-08-15",
            "client_name": "Client A",
            "line_items_json": json.dumps([{"service_id": service_id, "name": "Consulting Package", "quantity": 1, "unit_price": 1000, "discount_type": "€", "discount_value": 0, "vat_rate": "0%"}]),
        },
        follow_redirects=True,
    )

    app.load_finance_data.cache_clear()
    invoices = app.load_finance_data()["sheets"]["Invoices"]
    invoice = next(row for row in invoices if row.get("line_items") and row["line_items"][0].get("service_id") == service_id)
    assert invoice["line_items"][0]["service_id"] == service_id


# --- Receipt OCR ---

def test_receipt_ocr_rejects_oversized_file(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    oversized = BytesIO(b"0" * (app.MAX_OCR_FILE_SIZE_BYTES + 1))
    response = client.post(
        '/expenses/ocr',
        data={"receipt_file": (oversized, "receipt.jpg")},
        content_type='multipart/form-data',
    )
    assert response.status_code == 400
    assert "too large" in response.get_json()["error"].lower()


def test_receipt_ocr_rejects_invalid_file_type(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.post(
        '/expenses/ocr',
        data={"receipt_file": (BytesIO(b"not a receipt"), "receipt.exe")},
        content_type='multipart/form-data',
    )
    assert response.status_code == 400
    assert "unsupported" in response.get_json()["error"].lower()


def test_receipt_ocr_rejects_missing_file(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.post('/expenses/ocr', data={}, content_type='multipart/form-data')
    assert response.status_code == 400


def test_receipt_ocr_success_returns_parsed_fields(workbook_copy, monkeypatch):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    fake_fields = {
        "date": "01/08/2026",
        "supplier_name": "Test Supplier",
        "description": "Office chair",
        "net_amount": 100.0,
        "vat_amount": 23.0,
        "vat_rate": 23,
        "total_amount": 123.0,
        "currency": "EUR",
        "receipt_reference": "REC-1",
        "category_suggestion": "Office Supplies",
        "confidence": 92,
        "language_detected": "English",
        "notes": None,
        "ocr_raw_response": "{}",
    }
    monkeypatch.setattr(app, "_run_receipt_ocr", lambda file_bytes, extension, category_options: fake_fields)

    response = client.post(
        '/expenses/ocr',
        data={"receipt_file": (BytesIO(b"fake-image-bytes"), "receipt.jpg")},
        content_type='multipart/form-data',
    )
    assert response.status_code == 200
    body = response.get_json()
    assert body["supplier_name"] == "Test Supplier"
    assert body["confidence"] == 92


def test_receipt_ocr_handles_null_fields_gracefully(workbook_copy, monkeypatch):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    sparse_fields = {field: None for field in app.OCR_FIELDS}
    sparse_fields["ocr_raw_response"] = "{}"
    monkeypatch.setattr(app, "_run_receipt_ocr", lambda file_bytes, extension, category_options: sparse_fields)

    response = client.post(
        '/expenses/ocr',
        data={"receipt_file": (BytesIO(b"fake-image-bytes"), "receipt.png")},
        content_type='multipart/form-data',
    )
    assert response.status_code == 200
    body = response.get_json()
    assert body["supplier_name"] is None
    assert body["confidence"] is None


def test_receipt_ocr_failure_does_not_block_expense_saving(workbook_copy, monkeypatch):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    def raise_error(file_bytes, extension, category_options):
        raise app.OcrError("Receipt reading service unavailable")

    monkeypatch.setattr(app, "_run_receipt_ocr", raise_error)

    ocr_response = client.post(
        '/expenses/ocr',
        data={"receipt_file": (BytesIO(b"fake-image-bytes"), "receipt.jpg")},
        content_type='multipart/form-data',
    )
    assert ocr_response.status_code == 502

    payload = {
        "date": "2026-08-05",
        "title": "Manual entry after OCR failure",
        "description": "Filled in manually",
        "supplier": "Supplier A",
        "supplier_vat_number": "IE1234567A",
        "receipt_reference": "REC-2",
        "category": "Equipment and Hardware",
        "net_amount": "50.00",
        "total_amount": "61.50",
        "vat_rate": "23%",
        "vat_amount": "11.50",
        "input_vat_reclaimable": "Yes",
        "status": "Paid",
        "payment_method": "Business Bank",
    }
    save_response = client.post('/expenses/add', data=payload, follow_redirects=True)
    assert save_response.status_code == 200
    assert b'Expense entry added' in save_response.data


def test_parse_ocr_response_text_handles_markdown_fences():
    raw = '```json\n{"date": "01/08/2026", "supplier_name": "Acme"}\n```'
    parsed = app._parse_ocr_response_text(raw)
    assert parsed["date"] == "01/08/2026"
    assert parsed["supplier_name"] == "Acme"
    assert parsed["confidence"] is None


def test_parse_ocr_response_text_rejects_unparseable_text():
    with pytest.raises(app.OcrError):
        app._parse_ocr_response_text("not json at all")


def test_expense_ocr_audit_fields_saved_on_add(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    payload = {
        "date": "2026-08-05",
        "description": "Receipt-read expense",
        "supplier": "Supplier A",
        "supplier_vat_number": "IE1234567A",
        "receipt_reference": "REC-1",
        "category": "Equipment and Hardware",
        "net_amount": "50.00",
        "total_amount": "61.50",
        "vat_rate": "23%",
        "vat_amount": "11.50",
        "input_vat_reclaimable": "Yes",
        "status": "Paid",
        "payment_method": "Business Bank",
        "ocr_processed": "Yes",
        "ocr_confidence": "92",
        "ocr_language": "English",
        "ocr_raw_response": '{"supplier_name": "Supplier A"}',
    }
    response = client.post('/expenses/add', data=payload, follow_redirects=True)
    assert response.status_code == 200

    app.load_finance_data.cache_clear()
    expenses = app.load_finance_data()["sheets"]["Expenses"]
    saved = next(r for r in expenses if r.get("Description") == "Receipt-read expense")
    assert saved.get("OCR Processed") == "Yes"
    assert saved.get("OCR Confidence") == "92"
    assert saved.get("OCR Language") == "English"


# --- OCR auto-categorisation & subscription matching ---

def test_ocr_route_auto_selects_subscription_category_and_flags_match(workbook_copy, monkeypatch):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    app.SUBSCRIPTIONS_PATH.write_text(json.dumps([
        {
            "id": "sub-1",
            "title": "Adobe Creative Cloud",
            "supplier": "Adobe Inc",
            "category": "Software and Subscriptions",
            "status": "active",
            "frequency": "monthly",
            "net_amount": 50.0,
            "total_amount": 61.50,
        }
    ]), encoding="utf-8")

    fake_fields = {
        "date": "05/08/2026",
        "supplier_name": "Adobe Inc",
        "supplier_vat_number": None,
        "description": "Creative Cloud subscription",
        "net_amount": 50.0,
        "vat_amount": 11.5,
        "vat_rate": 23,
        "total_amount": 61.5,
        "currency": "EUR",
        "receipt_reference": None,
        "category_suggestion": "Office Supplies",
        "confidence": 90,
        "language_detected": "English",
        "notes": None,
        "ocr_raw_response": "{}",
    }
    monkeypatch.setattr(app, "_run_receipt_ocr", lambda file_bytes, extension, category_options: fake_fields)

    response = client.post(
        '/expenses/ocr',
        data={"receipt_file": (BytesIO(b"fake-bytes"), "receipt.jpg")},
        content_type='multipart/form-data',
    )
    assert response.status_code == 200
    body = response.get_json()
    assert body["category_suggestion"] == "Software and Subscriptions"
    assert body["subscription_match"]["title"] == "Adobe Creative Cloud"


def test_ocr_route_no_subscription_match_when_supplier_unknown(workbook_copy, monkeypatch):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    app.SUBSCRIPTIONS_PATH.write_text(json.dumps([
        {"id": "sub-1", "title": "Adobe Creative Cloud", "supplier": "Adobe Inc", "category": "Software and Subscriptions", "status": "active"}
    ]), encoding="utf-8")

    fake_fields = {
        "date": "05/08/2026", "supplier_name": "Totally Unrelated Ltd", "supplier_vat_number": None,
        "description": "Something", "net_amount": 10.0, "vat_amount": 0.0, "vat_rate": 0,
        "total_amount": 10.0, "currency": "EUR", "receipt_reference": None,
        "category_suggestion": "Office Supplies", "confidence": 80, "language_detected": "English",
        "notes": None, "ocr_raw_response": "{}",
    }
    monkeypatch.setattr(app, "_run_receipt_ocr", lambda file_bytes, extension, category_options: fake_fields)

    response = client.post(
        '/expenses/ocr',
        data={"receipt_file": (BytesIO(b"fake-bytes"), "receipt.jpg")},
        content_type='multipart/form-data',
    )
    body = response.get_json()
    assert body["subscription_match"] is None
    assert body["category_suggestion"] == "Office Supplies"


def test_ocr_route_flags_duplicate_expense_in_same_period(workbook_copy, monkeypatch):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post('/expenses/add', data={
        "date": "2026-08-01",
        "description": "Adobe Creative Cloud — August",
        "supplier": "Adobe Inc",
        "supplier_vat_number": "IE1234567A",
        "category": "Software and Subscriptions",
        "net_amount": "50.00",
        "total_amount": "61.50",
        "input_vat_reclaimable": "Yes",
        "status": "Paid",
        "payment_method": "Business Bank",
    }, follow_redirects=True)

    fake_fields = {
        "date": "15/08/2026", "supplier_name": "Adobe Inc", "supplier_vat_number": None,
        "description": "Adobe Creative Cloud subscription", "net_amount": 50.0, "vat_amount": 11.5,
        "vat_rate": 23, "total_amount": 61.5, "currency": "EUR", "receipt_reference": None,
        "category_suggestion": "Software and Subscriptions", "confidence": 95, "language_detected": "English",
        "notes": None, "ocr_raw_response": "{}",
    }
    monkeypatch.setattr(app, "_run_receipt_ocr", lambda file_bytes, extension, category_options: fake_fields)

    response = client.post(
        '/expenses/ocr',
        data={"receipt_file": (BytesIO(b"fake-bytes"), "receipt.jpg")},
        content_type='multipart/form-data',
    )
    body = response.get_json()
    assert body["duplicate_warning"] is not None
    assert "Adobe Inc" in body["duplicate_warning"]


def test_ocr_route_no_duplicate_warning_for_different_period(workbook_copy, monkeypatch):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    client.post('/expenses/add', data={
        "date": "2026-06-01",
        "description": "Adobe Creative Cloud — June",
        "supplier": "Adobe Inc",
        "supplier_vat_number": "IE1234567A",
        "category": "Software and Subscriptions",
        "net_amount": "50.00",
        "total_amount": "61.50",
        "input_vat_reclaimable": "Yes",
        "status": "Paid",
        "payment_method": "Business Bank",
    }, follow_redirects=True)

    fake_fields = {
        "date": "15/08/2026", "supplier_name": "Adobe Inc", "supplier_vat_number": None,
        "description": "Adobe Creative Cloud subscription", "net_amount": 50.0, "vat_amount": 11.5,
        "vat_rate": 23, "total_amount": 61.5, "currency": "EUR", "receipt_reference": None,
        "category_suggestion": "Software and Subscriptions", "confidence": 95, "language_detected": "English",
        "notes": None, "ocr_raw_response": "{}",
    }
    monkeypatch.setattr(app, "_run_receipt_ocr", lambda file_bytes, extension, category_options: fake_fields)

    response = client.post(
        '/expenses/ocr',
        data={"receipt_file": (BytesIO(b"fake-bytes"), "receipt.jpg")},
        content_type='multipart/form-data',
    )
    body = response.get_json()
    assert body["duplicate_warning"] is None


def test_match_subscription_by_supplier_ignores_inactive(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.SUBSCRIPTIONS_PATH.write_text(json.dumps([
        {"id": "sub-1", "title": "Cancelled Tool", "supplier": "Cancelled Co", "category": "Software and Subscriptions", "status": "cancelled"}
    ]), encoding="utf-8")

    assert app._match_subscription_by_supplier("Cancelled Co") is None


# --- Subscription <-> Expense linking ---

def test_expense_save_links_subscription_period_and_advances_next_charge(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    app.SUBSCRIPTIONS_PATH.write_text(json.dumps([{
        "id": "sub-1", "title": "Adobe Creative Cloud", "supplier": "Adobe Inc",
        "category": "Software and Subscriptions", "status": "active", "frequency": "monthly",
        "net_amount": 50.0, "total_amount": 61.50,
        "next_charge_date": "2026-08-15", "start_date": "2026-01-15",
    }]), encoding="utf-8")

    response = client.post('/expenses/add', data={
        "date": "2026-08-15",
        "description": "Adobe Creative Cloud subscription",
        "supplier": "Adobe Inc",
        "supplier_vat_number": "IE1234567A",
        "category": "Software and Subscriptions",
        "net_amount": "50.00",
        "total_amount": "61.50",
        "input_vat_reclaimable": "Yes",
        "status": "Paid",
        "payment_method": "Business Bank",
    }, follow_redirects=True)
    assert response.status_code == 200
    assert b"This matches your Adobe Creative Cloud subscription" not in response.data  # first-ever save for this period, no auto-posted duplicate to reuse

    app.load_finance_data.cache_clear()
    expenses = app.load_finance_data()["sheets"]["Expenses"]
    saved = next(r for r in expenses if r.get("Description") == "Adobe Creative Cloud subscription")
    assert saved.get("Expense Type") == "Subscription"
    assert saved.get("Subscription ID") == "sub-1"
    assert saved.get("id", "").startswith("EXP-")

    subscriptions = json.loads(app.SUBSCRIPTIONS_PATH.read_text(encoding="utf-8"))
    sub = next(s for s in subscriptions if s["id"] == "sub-1")
    assert sub["periods"]["2026-08"]["expense_id"] == saved["id"]
    assert sub["periods"]["2026-08"]["paid"] is True
    assert sub["next_charge_date"] == "2026-09-15"
    assert sub["last_posted_date"] == "2026-08-15"


def test_expense_save_reuses_auto_posted_entry_for_same_period(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    app.SUBSCRIPTIONS_PATH.write_text(json.dumps([{
        "id": "sub-1", "title": "Anthropic", "supplier": "Anthropic, PBC",
        "category": "Software and Subscriptions", "status": "active", "frequency": "monthly",
        "net_amount": 18.0, "total_amount": 22.14,
        "next_charge_date": "2026-07-26", "start_date": "2026-01-26",
    }]), encoding="utf-8")

    # Simulate the auto-sync having already posted this period's expense.
    sync_result = app._sync_subscriptions_to_expenses(today=app.date(2026, 8, 1))
    assert sync_result["posted_count"] == 1

    app.load_finance_data.cache_clear()
    auto_posted = next(r for r in app.load_finance_data()["sheets"]["Expenses"] if r.get("Subscription ID") == "sub-1")
    assert auto_posted["Status"] == "Auto-posted"
    assert auto_posted["Receipt Attached"] == "No"
    existing_row_count = len(app.load_finance_data()["sheets"]["Expenses"])

    # Now the user uploads the real receipt for that same period via the review flow.
    response = client.post('/expenses/add', data={
        "date": "2026-07-26",
        "description": "Claude Pro subscription",
        "supplier": "Anthropic, PBC",
        "supplier_vat_number": "IE1234567A",
        "category": "Software and Subscriptions",
        "net_amount": "18.00",
        "total_amount": "22.14",
        "input_vat_reclaimable": "Yes",
        "status": "Paid",
        "payment_method": "Business Bank",
    }, follow_redirects=True)
    assert response.status_code == 200
    assert b"This matches your Anthropic subscription for this period" in response.data
    assert b"receipt added to existing entry" in response.data

    app.load_finance_data.cache_clear()
    expenses_after = app.load_finance_data()["sheets"]["Expenses"]
    assert len(expenses_after) == existing_row_count  # no duplicate row created
    updated = next(r for r in expenses_after if r.get("Subscription ID") == "sub-1")
    assert updated["id"] == auto_posted["id"]
    assert updated["Description"] == "Claude Pro subscription"
    assert updated["Status"] == "Paid"


def test_subscriptions_page_shows_receipt_attached_indicator(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    app.SUBSCRIPTIONS_PATH.write_text(json.dumps([{
        "id": "sub-1", "title": "Figma", "supplier": "Figma Inc",
        "category": "Software and Subscriptions", "status": "active", "frequency": "monthly",
        "net_amount": 15.0, "total_amount": 18.45,
        "next_charge_date": "2026-08-10", "start_date": "2026-01-10",
    }]), encoding="utf-8")

    client.post('/expenses/add', data={
        "date": "2026-08-10",
        "description": "Figma subscription",
        "supplier": "Figma Inc",
        "supplier_vat_number": "IE1234567A",
        "category": "Software and Subscriptions",
        "net_amount": "15.00",
        "total_amount": "18.45",
        "input_vat_reclaimable": "Yes",
        "status": "Paid",
        "payment_method": "Business Bank",
        "receipt_file": (BytesIO(b"fake-receipt-bytes"), "figma-receipt.pdf"),
    }, content_type='multipart/form-data', follow_redirects=True)

    response = client.get('/subscriptions')
    assert response.status_code == 200
    assert b"Receipt" in response.data

    subscriptions = json.loads(app.SUBSCRIPTIONS_PATH.read_text(encoding="utf-8"))
    sub = next(s for s in subscriptions if s["id"] == "sub-1")
    assert sub["periods"]["2026-08"]["receipt_attached"] is True


def test_link_subscription_expense_does_not_advance_next_charge_for_past_period(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    subscription = {
        "id": "sub-1", "frequency": "monthly", "next_charge_date": "2026-09-10",
        "last_posted_date": "2026-08-10", "periods": {},
    }
    app._link_subscription_expense(subscription, "EXP-099", "2026-06", True)
    assert subscription["next_charge_date"] == "2026-09-10"  # unchanged — 2026-06 isn't the due period
    assert subscription["periods"]["2026-06"]["expense_id"] == "EXP-099"
    assert subscription["periods"]["2026-06"]["receipt_attached"] is True


def test_expense_review_card_has_visible_expense_type_and_status_dropdowns(workbook_copy):
    """The OCR review card's Expense Type and Status fields must be real, visible
    <select> elements wired to the same fields the full form submits — not a hidden
    input the user can't see or correct. Also verifies the subscription-match branch
    of the client-side type-assignment logic is present and wired to that field."""
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()
    client = app.app.test_client()

    response = client.get('/expenses')
    assert response.status_code == 200
    html = response.data.decode("utf-8")

    assert '<select name="expense_type" id="expense_type_field"' in html
    assert '<option value="Subscription">Subscription</option>' in html
    assert '<option value="Receipt or Invoice"' in html
    assert '<option value="Travel and Subsistence">Travel and Subsistence</option>' in html

    assert 'buildReviewRow("Expense type", "expense_type_field", "select"' in html
    assert 'buildCuratedSelectRow("Status", "expense_status", ["Paid", "Pending", "Approved"]' in html

    assert 'if (fields.subscription_match) { typeValue = "Subscription"; }' in html
    assert 'statusField.value = "Paid";' in html
